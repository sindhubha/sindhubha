from fastapi import APIRouter
from fastapi import Form,Depends
from sqlalchemy.orm import Session
from fastapi import APIRouter, FastAPI
from src.endpoints.response_json import _getReturnResponseJson,_getSuccessResponseJson,_getErrorResponseJson,get_exception_response
from fastapi.requests import Request 
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook
from pathlib import Path
import calendar
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime,date, timedelta
from log_file import createFolder
from openpyxl.drawing.image import Image
import os
from src.models.parse_date import parse_date
from src.models.check_table import check_air_table,check_air_12_table,check_air_analysis_table,check_water_table,check_water_12_table,check_water_analysis_table
import re
from sqlalchemy.ext.asyncio import AsyncSession
import wmi
import pythoncom
from concurrent.futures import ThreadPoolExecutor
import asyncio


static_dir = Path(__file__).parent / "attachments"
base_path = Path(__file__).resolve().parent / "attachments"
print("base_path",base_path) 

file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "..", "..", "database.txt"))

# Check if the file exists
if os.path.exists(file_path):
    with open(file_path, "r") as file:
        content = file.read().strip()  # Remove leading/trailing whitespace
   
    if content == 'MySQL':
        from mysql_connection import get_db
        from src.models.mysql.fuel_report_model import current_flow_dtl,month_wise_report,year_wise_report
        from src.models.mysql.master_shift_model import shift_Lists
        from src.models.mysql.master_plant_model import plant_Lists
    elif content == 'MSSQL':
        from mssql_connection import get_db
        # from src.models.mssql.report_model import month_report,daily_report,year_wise_report_print,year_report_print,holiday_report,alarmreport,hour_wise_analysis_report
    else:
        raise Exception("Database is not configured or 'database.txt' contains an unexpected value.")
else:
    raise Exception("The 'database.txt' file does not exist in the specified location.")

router = APIRouter()
app = FastAPI()

static_dir = Path(__file__).parent / "attachments"
mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
                    

@router.post("/current_flow/", tags=["Fuel Dashboard"])
async def current_flow(company_id : str = Form(''),
                        bu_id :str = Form(''),
                        campus_id :str = Form(''),
                        plant_id :str = Form(''),
                        plant_department_id :str = Form(''),
                        equipment_group_id :str = Form(''),
                        equipment_id :str = Form(''),
                        function_id : str = Form(''),
                        meter_id : str = Form (''),
                        group_for : str = Form(''),
                        groupby : str = Form(''),
                        period_id: str = Form(''),
                        from_date: str = Form(''),
                        to_date: str = Form(''),   
                        from_year: int  = Form(""),                   
                        shift_id: str = Form(''),
                        limit_report_for = Form(''),
                        limit_exception_for:str = Form(''),  
                        limit_order_by : str = Form(''),
                        limit_operation_value : str = Form(''),
                        is_critical :str = Form(''),
                        converter_id :int = Form(''),  
                        report_for : str = Form(''), 
                        is_function : str = Form(''),  
                        function_type : str = Form(''), 
                        reportfor:str = Form(''), 
                        employee_id : int = Form(''),
                        is_minmax: str = Form(''),                                
                        is_main_meter: str = Form(''),                           
                        is_demand: str = Form(''),                           
                        meter_type: str = Form(''),                           
                        is_plant_wise: str = Form(''),                           
                        source: str = Form(''),                           
                        for_android: str = Form(''),                           
                        cnx: AsyncSession = Depends(get_db)):
    try:
      
        

        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        # createFolder("Current_power_log/","group by.... "+str(groupby))
        
        if group_for == "":
            return _getErrorResponseJson("Group For Is Required")
        
        if period_id == "":
            return _getErrorResponseJson("Period Id Is Required")
        
        if groupby == "":
            return _getErrorResponseJson("Groupby Is Required")
        
        if groupby not in ['company','bu','plant', 'plant_department','equipment_group','equipment','function','meter','campus']:
            return _getErrorResponseJson("Invalid Groupby ")
        
        mill_date = date.today()
        mill_shift = 0
        no_of_shifts = 3

        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        # query = text(f'''SELECT * FROM master_shifts WHERE status = 'active' and  plant_id = '{plant_id}' ''')
        
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                no_of_shifts = shift_record["no_of_shifts"]  
    
        if reportfor == '12to12':
            if period_id != 'sel_date' and period_id != 'from_to':
                return _getErrorResponseJson("invalid period id") 
            
            if period_id == "sel_date":  
                if from_date == '':
                    return _getErrorResponseJson("from date is required") 
                    
                from_date =  await parse_date(from_date)   
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
                res = await check_air_12_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("12to12 table not available...") 
                
            if period_id == "from_to" or period_id == "#from_to":            
                if from_date == '':
                    return _getErrorResponseJson("from date is required")
                if to_date == '':
                    return _getErrorResponseJson("to_date is required") 
            
                from_date = await parse_date(from_date)
                to_date = await parse_date(to_date) 
                month_year_range = [
                    (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                    for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                ]
                union_queries = []

                for month_year in month_year_range:
                    res = check_air_12_table(cnx,month_year)
                    
                    if len(res)>0:
                        table_name = f"ems_v1_completed.{source}_{month_year}_12"
                        union_queries.append(f"{table_name}")

                if len(union_queries) == 0:
                    return _getErrorResponseJson("12to12 table not available...")  
                
        else: 
                  
            if period_id == "sel_date" or period_id == 'sel_shift':            
                if from_date == '':
                    return _getErrorResponseJson("From Date Is Required") 
                
                from_date =  await parse_date(from_date) 
                 
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
                if source == 'air':
                    res = await check_air_table(cnx,month_year)
                if source == 'water':
                    res = await check_air_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("Power Table Not Available...")   
                
                if period_id == "sel_shift":                  
                    if shift_id == '':
                        return _getErrorResponseJson("Shift Id Is Required") 
                    
            elif period_id == "#previous_shift" or period_id == "#previous_day":  
                if period_id == "#previous_shift":               
                    if int(mill_shift) == 1:
                        shift_id = no_of_shifts
                        from_date = mill_date - timedelta(days=1)
                    else:
                        shift_id = int(mill_shift) - 1
                        from_date = mill_date 

                elif period_id == "#previous_day":             
                    from_date = mill_date - timedelta(days=1)
                
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                if source == 'air':
                    res = await check_air_table(cnx,month_year)
                elif source == 'water':
                    res = await check_water_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("Power Table Not Available...")   
                       
            elif period_id == "from_to" or period_id == "#from_to":            
                if from_date == '':
                    return _getErrorResponseJson("From Date Is Required")
                if to_date == '':
                    return _getErrorResponseJson("To Date Is Required")  
                from_date = await parse_date(from_date)
                to_date = await parse_date(to_date)
              
            if period_id == '#previous_week' or period_id == "#this_week" or period_id == "#this_month" or period_id == '#previous_month' or period_id=="#previous_year" or period_id=="#this_year" or period_id=="from_to" or period_id == "#sel_year" or period_id=="#from_to" :
                if period_id  == "#this_week":
                    dt = mill_date
                    from_date=dt-timedelta(dt.weekday()+1)
                    to_date = mill_date

                elif period_id == "#previous_week":
                    dt = mill_date
                    current_week_start = dt - timedelta(days=dt.weekday())  
                    from_date = current_week_start - timedelta(weeks=1)  
                    to_date = from_date + timedelta(days=5)

                elif period_id == "#this_month":
                    from_date = mill_date.replace(day=1)
                    to_date = mill_date

                elif period_id == "#previous_month":
                    from_date = mill_date.replace(day=1)                   
                    from_date = (from_date - timedelta(days=1)).replace(day=1)
                    to_date = from_date + timedelta(days=30)   

                elif period_id=="#this_year": 
            
                    from_date = mill_date.replace(day=1,month=1) 
                    to_date = mill_date  
                    

                elif period_id=="#previous_year": 
                    from_date = mill_date.replace(day=1, month=1, year=mill_date.year - 1)
                    to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
            
                elif period_id == "#sel_year": 

                    if from_year == '':
                        return _getErrorResponseJson("From Year Is Required")
                        
                    from_date = mill_date.replace(day=1, month=1, year=from_year)
                    
                    to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
                   
            
                if from_date != '' and to_date != '':

                    month_year_range = [
                        (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                    union_queries = []

                    for month_year in month_year_range:
                        if source == 'air':
                            res = await check_air_table(cnx,month_year)
                        elif source == 'water':
                            res = await check_water_table(cnx,month_year)

                        if len(res)>0:
                            table_name = f"ems_v1_completed.{source}_{month_year}"
                            union_queries.append(f"{table_name}")

                    if len(union_queries) == 0:
                        return _getErrorResponseJson(f"{source} Table Not Available...")     

        data = await current_flow_dtl(cnx,company_id ,campus_id,bu_id ,plant_id ,plant_department_id ,equipment_group_id ,equipment_id,function_id ,meter_id ,group_for ,groupby ,period_id,from_date,to_date,shift_id,limit_report_for,limit_exception_for,limit_order_by ,limit_operation_value ,is_critical ,converter_id ,report_for,is_function , function_type ,reportfor,employee_id ,is_main_meter,is_demand,meter_type,is_plant_wise,source,is_minmax)


        if for_android == 'yes':
            response = [{
                "iserror": False,
                "message": "Data Returned Successfully.",
                "data": data
            }]
        else:
            response = {
                "iserror": False,
                "message": "Data Returned Successfully.",
                "data": data
            }

        
        return response
    except Exception as e:
        return get_exception_response(e)
    
async def month_wise_excel_report(result, month_year, report_type, report_for,shift_time,source):
    print("Function Call")
    wb = Workbook()
    ws = wb.active

    # workbook = Workbook()
    # sheet = workbook.active
    ws.title = 'EMS' 
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')  
    cell = "D2"
    if report_for == '12to12':
        data = f"MONTH WISE ENERGY CONSUMPTION REPORT - {month_year} (12:00 to 12:00)"
    else:
        data = f"MONTH WISE ENERGY CONSUMPTION REPORT  - {month_year} ({shift_time} to {shift_time})"
    ws[cell] = data
    font = Font(bold=True, name='Calibri', size=15)
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].font = font
    ws[cell].alignment = alignment 
    if report_for == '12to12':
        ws.cell(row=2, column=4, value=f"MONTH WISE ENERGY CONSUMPTION REPORT  - {month_year} (12:00 to 12:00)").fill = fill_cyan
    else:  
        ws.cell(row=2, column=4, value=f"MONTH WISE ENERGY CONSUMPTION REPORT - {month_year} ({shift_time} to {shift_time})").fill = fill_cyan
    
    # Set the border style
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    if result == []:
        print(1)
        cell = "O10"
        data = "No Data"
        ws[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment

        ws.column_dimensions[cell[0]].width = len(data) + 2  # Adjust column width

        font = Font(name='Calibri', size=25)
        ws[cell].font = font

    
    if report_type == "date":

        month, year = map(int, month_year.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        
        fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')

        
        
        cell = "B4"
        ws[cell] = "Meter Tag"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 15 
    
        cell = "C4" 
        ws[cell] = "Meter Name"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 60

        cell = "D4" 
        ws[cell] = "Meter Type"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 15

        ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
        ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
        ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
        for i in range(1, days_in_month + 1):
            header_text = f"{i:02d}-{month_year}"
            ws.merge_cells(start_row=4, start_column=4 + i, end_row=4, end_column=4 + i)
            header_cell = ws.cell(row=4, column=4 + i)
            header_cell.value = header_text
            header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
            header_cell.font = Font(bold=True, name='Calibri', size=12)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            column_width = max(len(header_text), 15)  
            ws.column_dimensions[ws.cell(row=4, column=4 + i).column_letter].width = column_width
        ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4 + i)

        for meter_data in result:
            meter_name = meter_data["meter_name"]
            meter_code = meter_data["meter_code"]
            meter_type = meter_data["meter_type"]
            plant_name = meter_data["plant_name"]
            row_data = ['', meter_code, meter_name, meter_type]
            ws.merge_cells('B2:C2')
            ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
            ws.cell(row=2, column=2).fill = fill_cyan
            font = Font(bold=True, name='Calibri', size=12)
            ws.cell(row=2, column=2).font = font

            ws.merge_cells('B3:C3')
            ws.cell(row=3, column=2, value=f"Source : {source}").alignment = Alignment(horizontal="left")
            ws.cell(row=3, column=2).fill = fill_cyan
            font = Font(bold=True, name='Calibri', size=12)
            ws.cell(row=3, column=2).font = font

            for i in range(1, days_in_month + 1):
                day = f"d{i}"
                kwh = meter_data.get(day, "")
                
                row_data.extend([kwh])
            ws.append(row_data)

        end_row = ws.max_row
        row_range = ws.iter_rows(min_row=2, max_row=end_row, min_col=2, max_col=i+4)
        for row in row_range:
            for cell in row:
                cell.border = border


    elif report_type == "shift":
        
        headers = ["Meter Tag", "Meter Name", "Meter Type"]
        for col, header in enumerate(headers, start=2):
            ws.cell(row=4, column=col, value=header).fill = fill_cyan
            ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
            ws.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = len(header) + 15

        month, year = map(int, month_year.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        # ws.merge_cells('D2:CS3')

        ws.merge_cells('B4:B5')
        cell = "B4"
        ws[cell] = "Meter Tag"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 15 

        ws.merge_cells('C4:C5')
        cell = "C4" 
        ws[cell] = "Meter Name"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 60

        ws.merge_cells('D4:D5')
        cell = "D4" 
        ws[cell] = "Meter Type"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 15

        ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
        ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
        ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
        for i in  range(1, days_in_month + 1):
            # Write the hour range header
            header_text = f"{i:02d}-{month_year}"
            ws.merge_cells(start_row=4, start_column=2 + i * 3, end_row=4, end_column=4 + i * 3)
            header_cell = ws.cell(row=4, column=2 + i * 3)
            header_cell.value = header_text

            # Apply styles to the header cell
            header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
            header_cell.font = Font(bold=True, name='Calibri', size=12)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            # shifts = ['Start CkWh', 'End Ckwh', 'kWh']
            secondary_headers = ["S1", "S2", "S3"]
            for j, header in enumerate(secondary_headers):
                secondary_header_cell = ws.cell(row=5, column=2+i*3+j)
                secondary_header_cell.value = header

                # Apply styles to the secondary header cell
                secondary_header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                secondary_header_cell.font = Font(bold=True, name='Calibri', size=12)
                secondary_header_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[secondary_header_cell.column_letter].width = len(header) + 15
        
        for meter_data in result:
            meter_name = meter_data["meter_name"]
            meter_code = meter_data["meter_code"]
            meter_type = meter_data["meter_type"]
            plant_name = meter_data["plant_name"]
            
            # Write meter information
            # ws.append(['', meter_code, meter_name, meter_type])  # Writing meter info in a row
            row_data = ['', meter_code, meter_name, meter_type]
            ws.merge_cells('B2:C2')
            ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
            ws.cell(row=2, column=2).fill = fill_cyan
            font = Font(bold=True, name='Calibri', size=12)
            ws.cell(row=2, column=2).font = font

            ws.merge_cells('B3:C3')
            ws.cell(row=3, column=2, value=f"Source : {source}").alignment = Alignment(horizontal="left")
            ws.cell(row=3, column=2).fill = fill_cyan
            font = Font(bold=True, name='Calibri', size=12)
            ws.cell(row=3, column=2).font = font

            # for shift in secondary_headers:
            row_data = ['', meter_code, meter_name, meter_type]  # Initialize row data here
            for i in range(1, days_in_month + 1):
                hour = f"ds{1}_{i}"
                kwh = meter_data.get(hour, "") 
                machine_kwh = meter_data.get(f"ds{2}_{i}", "")
                master_kwh = meter_data.get(f"ds{3}_{i}", "")
                row_data.extend([kwh, machine_kwh, master_kwh]) # Assigning a new value to kwh
            ws.merge_cells(start_row=2, start_column=4 , end_row=3, end_column=4 + i*3)
 
            ws.append(row_data)

        end_row = ws.max_row   
        row_range = ws.iter_rows(min_row=2, max_row=end_row, min_col=2, max_col=4 + i*3)

        for row in row_range:
            for cell in row:
                cell.border = border
        

    file_name = f'MonthWiseReport-{source}-{month_year}.xlsx'
    print("file_name",file_name)
    file_path = os.path.join(static_dir, file_name)
    print("file_path....",file_path)
    wb.save(file_path)

async def generate_year_wise_excel_report(result, year,next_year,source):

    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..","..", "..", "YearWiseReport_templete.xlsx"))
    print("filepath",file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS'
    cell = "B1"
    data = f"YEAR WISE ENERGY CONSUMPTION REPORT - {year}"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=13)
    sheet[cell].font = font

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    cell = "A3"
    sheet[cell] = "meter Code"
    cell = "B3"
    sheet[cell] = "meter Name"
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
    sheet.cell(row=3, column=1, value="meter Code").fill = fill_cyan
    sheet.cell(row=3, column=2, value="meter Name").fill = fill_cyan
    if result == []:
        print(1)
        cell = "E10"
        data = "No Data"
        sheet[cell] = data
        font = Font(bold=True, name='Calibri', size=13)
        sheet[cell].font = font
    else:
    # Populate month-year headers in row 3 starting from column C
        col_index = 3
        for month_year in result[0].keys():
            if month_year not in ["meter_code", "meter_name"]:
                sheet.cell(row=3, column=col_index, value=month_year).fill = fill_cyan
                col_index += 1

        for row, data in enumerate(result, start=4):
            meter_code = data["meter_code"]
            meter_name = data["meter_name"]
            sheet.cell(row=row, column=1, value=meter_code).alignment = Alignment(horizontal="center")
            alignment = Alignment(horizontal="center", wrap_text=True)
            sheet.cell(row=row, column=2, value=meter_name).alignment = alignment
            

            col_indexs = 1
            for kwh_value in data.values():
                if isinstance(kwh_value, (int, float)):
                    if kwh_value == 0:
                        kwh_value = ''
                    sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
                    cell = sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
                    cell.number_format = '0.00'

                
                col_indexs += 1
            row_range = sheet[f"A1:N{row}"]
            for row in row_range:
                for cell in row:
                    cell.border = border

    file_name = f'YearWiseReport-{source}-{year}-{next_year}.xlsx'
    file_path = os.path.join(static_dir, file_name)
    print("file_path",file_path)
    workbook.save(file_path)

@router.post("/month_year_report/", tags=["Fuel Report"])
async def month_year_report(request:Request,
                            campus_id : int = Form(''),
                            company_id : int = Form(''),
                            bu_id : int = Form(''),
                            plant_id : int = Form(''),
                            plant_department_id : int = Form(''),
                            equipment_group_id : int = Form(''),
                            meter_id: str = Form(''),
                            month_year: str = Form(''),
                            report_for: str = Form(''),
                            report_type: str = Form(''),                                                      
                            meter_type: str = Form(''),                                                                                                       
                            employee_id: str = Form(''),  
                            year: str = Form(""),                                                  
                            source: str = Form(""),                                                  
                            report: str = Form(""),                                                  
                            cnx: AsyncSession = Depends(get_db)):

    try: 
        
        if report == "" :
            return _getErrorResponseJson("report is Required...")
        
        if source == "" :
            return _getErrorResponseJson("source is Required...")
        
        if report == 'month':
            if month_year == "" :
                return _getErrorResponseJson("MonthYear is Required...")
            
            if report_for == "" :
                return _getErrorResponseJson("ReportFor is Required...")
            
            if report_type == "" :
                return _getErrorResponseJson("Report Type is Required...")
            
            if report_for == '12to12':
                formatted_month_year = month_year.replace('-', '')
                if source == 'air':
                    res = await check_air_12_table(cnx,formatted_month_year)
                elif source == 'water':
                    res = await check_water_12_table(cnx,formatted_month_year)
                # res = await f"check_{source}_12_table"(cnx,formatted_month_year)

                if report_type not in ['date']:
                    return _getErrorResponseJson("Invalid report type")

            else:
                formatted_month_year = month_year.replace('-', '')
                if source == 'air':
                    res = await check_air_table(cnx,formatted_month_year)
                elif source == 'water':
                    res = await check_water_table(cnx,formatted_month_year)  

            if report_type not in ['date', 'shift']:
                    return _getErrorResponseJson("Invalid report type")
            
            if len(res)==0:
                return _getErrorResponseJson("Table not available...")
            data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
            
            if len(data1) > 0:
                for shift_record in data1:
                    shift_time = shift_record["a_shift_start_time"]
                print(shift_time)  
            print(1111111)  
            result = await month_wise_report(cnx,campus_id,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,employee_id,meter_id,month_year,report_for,report_type,meter_type,source,request)
            await month_wise_excel_report(result, month_year,report_type, report_for,shift_time,source)
            
            file_path = os.path.join(static_dir, f"MonthWiseReport-{source}-{month_year}.xlsx")
            
            results = f"http://{request.headers['host']}/attachments/MonthWiseReport-{source}-{month_year}.xlsx"
        
        elif report == 'year':
            if year == "" :
                return _getErrorResponseJson("year is required...")
        
            if report_for == "" :
                return _getErrorResponseJson("report_for is required...")
            
            next_year = int(year) + 1
            data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
            
            if len(data1) > 0:
                for shift_record in data1:
                    shift_time = shift_record["a_shift_start_time"]
            mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}

            tables_to_union = []
            for month in range(4, 13):
                month_year = f"{mill_month[month]}{year}"
                print(month_year)
                if report_for == '12to12':
                    query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = '{source}_{month_year}_12'"""
                    result_query = await cnx.execute(query)
                    result_query = result_query.fetchall()
                    if len(result_query) > 0:
                        tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.power_{month_year}_12")
                    print(month_year)
                else:
                    query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = '{source}_{month_year}'"""
                    result_query = await cnx.execute(query)
                    result_query = result_query.fetchall()

                    if len(result_query) > 0:
                        tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.{source}_{month_year}")
            
            next_year = int(year) + 1
            mill_month = {1: "01", 2: "02", 3: "03"}

            for month in range(1, 4):
                month_year = f"{mill_month[month]}{next_year}"
                print(month_year)
                if report_for == '12to12':
                    query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = '{source}_{month_year}_12' """
                    result_query = await cnx.execute(query)
                    result_query = result_query.fetchall()

                    print("result_query",result_query)
                    if len(result_query) > 0:
                        tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.{source}_{month_year}_12")
                    tables_union_query = " UNION ALL ".join(tables_to_union)
                else:   
                    query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = '{source}_{month_year}' """
                    result_query = await cnx.execute(query)
                    result_query = result_query.fetchall()

                    print("result_query",result_query)
                    if len(result_query) > 0:
                        tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.{source}_{month_year}")
                    tables_union_query = " UNION ALL ".join(tables_to_union)
                    print("tables_union_query",tables_union_query)

            if len(tables_to_union) == 0:
                return _getErrorResponseJson("table not available")
            

            result = await year_wise_report(cnx,campus_id,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,meter_id,year,report_for,employee_id,source,request)
            await generate_year_wise_excel_report(result, year,next_year,source)
            results = f"http://{request.headers['host']}/attachments/YearWiseReport-{source}-{year}-{next_year}.xlsx"
        
        else:
            return _getErrorResponseJson("Invalid report...") 
        
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "file_url": results
                }
        return response
    except Exception as e:
        return get_exception_response(e)
