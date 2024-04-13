from fastapi import APIRouter
from fastapi import Form,Depends
from sqlalchemy.orm import Session
from fastapi import APIRouter, FastAPI
from src.endpoints.response_json import _getReturnResponseJson,_getSuccessResponseJson,_getErrorResponseJson,get_exception_response
from fastapi.requests import Request 
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook
from pathlib import Path
import calendar
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime,date, timedelta
from log_file import createFolder
from openpyxl.drawing.image import Image
import os
from src.models.parse_date import parse_date
from src.models.check_table import check_power_table,check_power_12_table,check_analysis_table,check_polling_data_tble,check_alarm_tble,check_user_count
import re
from sqlalchemy.ext.asyncio import AsyncSession
import wmi
import pythoncom
from concurrent.futures import ThreadPoolExecutor
import asyncio
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.formula.translate import Translator
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import os
from os.path import basename
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import requests
import smtplib



static_dir = Path(__file__).parent / "attachments"
base_path = Path(__file__).resolve().parent / "attachments"
print("base_path",base_path) 

file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "..", "..", "database.txt"))

# Check if the file exists
if os.path.exists(file_path):
    with open(file_path, "r") as file:
        content = file.read().strip()  # Remove leading/trailing whitespace
   
    if content == 'MySQL':
        from mysql_connection import get_db
        from src.models.mysql.report_model import month_report,daily_report,year_wise_report_print,year_report_print,holiday_report,alarmreport,hour_wise_analysis_report,get_hour_wise_report_model,import_export_dtl,current_power_dtl,function_dashboard,function_dashboard2,communication_status_query,get_equipment_cal_dtl,gateway_log,route_card_kwh,demand_report,availability_report,tnebreportdetail,avg_demand_report,minmax_kwh_dtl,manualentryhistory,transformerlossreport,submeterlossreport
        from src.models.mysql.master_shift_model import shift_Lists
        from src.models.mysql.master_plant_model import plant_Lists
    elif content == 'MSSQL':
        from mssql_connection import get_db
        # from src.models.mssql.report_model import month_report,daily_report,year_wise_report_print,year_report_print,holiday_report,alarmreport,hour_wise_analysis_report
    else:
        raise Exception("Database is not configured or 'database.txt' contains an unexpected value.")
else:
    raise Exception("The 'database.txt' file does not exist in the specified location.")

router = APIRouter()
app = FastAPI()

static_dir = Path(__file__).parent / "attachments"
mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}

app = FastAPI()

executor = ThreadPoolExecutor()

def get_serial_number_sync():
    try:
        # Initialize COM library
        pythoncom.CoInitialize()

        # Connect to the WMI service
        c = wmi.WMI()

        # Query for the serial number from the Win32_BIOS class
        for bios in c.Win32_BIOS():
            serial_number = bios.SerialNumber.strip()
            return serial_number
    except Exception as e:
        print("Error:", e)
        return None

async def get_serial_number():
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(executor, get_serial_number_sync)

@router.get("/serial-number/")
async def read_serial_number():
    serial_number = await get_serial_number()
    if serial_number:
        if serial_number == 'PF2WF6R7':
            response = {"iserror": False, "message": "Data Returned Successfully."}
        else:
            response = {"iserror": True, "message": "This CPU Serial Number is not valid, Kindly Contact Admin"}
    else:
        response = {"iserror": True, "message": "Serial number not available."}
    return response                       

@router.post("/current_power/", tags=["Dashboard"])
async def current_power(company_id : str = Form(''),
                        bu_id :str = Form(''),
                        campus_id :str = Form(''),
                        plant_id :str = Form(''),
                        plant_department_id :str = Form(''),
                        equipment_group_id :str = Form(''),
                        equipment_id :str = Form(''),
                        function_id : str = Form(''),
                        meter_id : str = Form (''),
                        group_for : str = Form(''),
                        groupby : str = Form(''),
                        period_id: str = Form(''),
                        from_date: str = Form(''),
                        to_date: str = Form(''),   
                        from_year: int  = Form(""),                   
                        shift_id: str = Form(''),
                        limit_report_for = Form(''),
                        limit_exception_for:str = Form(''),
                        limit_order_by : str = Form(''),
                        limit_operation_value : str = Form(''),
                        is_critical :str = Form(''),
                        converter_id :int = Form(''),  
                        report_for : str = Form(''), 
                        is_function : str = Form(''),  
                        function_type : str = Form(''), 
                        reportfor:str = Form(''), 
                        employee_id : int = Form(''),    
                        is_minmax: str = Form(''),                           
                        is_main_meter: str = Form(''),                           
                        is_demand: str = Form(''),                           
                        meter_type: str = Form(''),                           
                        is_plant_wise: str = Form(''),                           
                        for_android: str = Form(''),                           
                        cnx: AsyncSession = Depends(get_db)):
    try:
      
        result = await check_user_count(cnx)
        createFolder("Log/user_countlog/","UserCount.... "+str(result))

        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        # createFolder("Current_power_log/","group by.... "+str(groupby))
        if group_for != 'regular':
            createFolder("Current_power_log/","group_for by.... "+str(group_for))
        
        if group_for == "":
            return _getErrorResponseJson("Group For Is Required")
        
        if period_id == "":
            return _getErrorResponseJson("Period Id Is Required")
        
        if groupby == "":
            return _getErrorResponseJson("Groupby Is Required")
        
        if groupby not in ['company','bu','plant', 'plant_department','equipment_group','equipment','function','meter','campus']:
            return _getErrorResponseJson("Invalid Groupby ")
        
        mill_date = date.today()
        mill_shift = 0
        no_of_shifts = 3

        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        # query = text(f'''SELECT * FROM master_shifts WHERE status = 'active' and  plant_id = '{plant_id}' ''')
        
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                no_of_shifts = shift_record["no_of_shifts"]  
    
        if reportfor == '12to12':
            if period_id != 'sel_date' and period_id != 'from_to':
                return _getErrorResponseJson("invalid period id") 
            
            if period_id == "sel_date":  
                if from_date == '':
                    return _getErrorResponseJson("from date is required") 
                    
                from_date =  await parse_date(from_date)   
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
                res = await check_power_12_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("12to12 table not available...") 
                
            if period_id == "from_to" or period_id == "#from_to":            
                if from_date == '':
                    return _getErrorResponseJson("from date is required")
                if to_date == '':
                    return _getErrorResponseJson("to_date is required") 
            
                from_date = await parse_date(from_date)
                to_date = await parse_date(to_date) 
                month_year_range = [
                    (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                    for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                ]
                union_queries = []

                for month_year in month_year_range:
                    res = check_power_12_table(cnx,month_year)
                    
                    if len(res)>0:
                        table_name = f"ems_v1_completed.power_{month_year}_12"
                        union_queries.append(f"{table_name}")

                if len(union_queries) == 0:
                    return _getErrorResponseJson("12to12 table not available...")  
                
        else: 
                  
            if period_id == "sel_date" or period_id == 'sel_shift':            
                if from_date == '':
                    return _getErrorResponseJson("From Date Is Required") 
                
                from_date =  await parse_date(from_date) 
                 
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
                res = await check_power_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("Power Table Not Available...")   
                
                if period_id == "sel_shift":                  
                    if shift_id == '':
                        return _getErrorResponseJson("Shift Id Is Required") 
                    
            elif period_id == "#previous_shift" or period_id == "#previous_day":  
                if period_id == "#previous_shift":               
                    if int(mill_shift) == 1:
                        shift_id = no_of_shifts
                        from_date = mill_date - timedelta(days=1)
                    else:
                        shift_id = int(mill_shift) - 1
                        from_date = mill_date 

                elif period_id == "#previous_day":             
                    from_date = mill_date - timedelta(days=1)
                
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                res = await check_power_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("Power Table Not Available...")   
                       
            elif period_id == "from_to" or period_id == "#from_to":            
                if from_date == '':
                    return _getErrorResponseJson("From Date Is Required")
                if to_date == '':
                    return _getErrorResponseJson("To Date Is Required")  
                from_date = await parse_date(from_date)
                to_date = await parse_date(to_date)
              
            if period_id == '#previous_week' or period_id == "#this_week" or period_id == "#this_month" or period_id == '#previous_month' or period_id=="#previous_year" or period_id=="#this_year" or period_id=="from_to" or period_id == "#sel_year" or period_id=="#from_to" :
                if period_id  == "#this_week":
                    dt = mill_date
                    from_date=dt-timedelta(dt.weekday()+1)
                    to_date = mill_date

                elif period_id == "#previous_week":
                    dt = mill_date
                    current_week_start = dt - timedelta(days=dt.weekday())  
                    from_date = current_week_start - timedelta(weeks=1)  
                    to_date = from_date + timedelta(days=5)

                elif period_id == "#this_month":
                    from_date = mill_date.replace(day=1)
                    to_date = mill_date

                elif period_id == "#previous_month":
                    from_date = mill_date.replace(day=1)                   
                    from_date = (from_date - timedelta(days=1)).replace(day=1)
                    to_date = from_date + timedelta(days=30)   

                elif period_id=="#this_year": 
            
                    from_date = mill_date.replace(day=1,month=1) 
                    to_date = mill_date  
                    

                elif period_id=="#previous_year": 
                    from_date = mill_date.replace(day=1, month=1, year=mill_date.year - 1)
                    to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
            
                elif period_id == "#sel_year": 

                    if from_year == '':
                        return _getErrorResponseJson("From Year Is Required")
                        
                    from_date = mill_date.replace(day=1, month=1, year=from_year)
                    
                    to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
                   
            
                if from_date != '' and to_date != '':

                    month_year_range = [
                        (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                    union_queries = []
                    joins = []

                    for month_year in month_year_range:
                        res = await check_power_table(cnx,month_year)
                        res1 = await check_polling_data_tble(cnx,month_year)

                        if len(res)>0:
                            table_name = f"ems_v1_completed.power_{month_year}"
                            union_queries.append(f"{table_name}")

                        if len(res1) >0:
                            join_p = f"left join ems_v1_completed.dbo.polling_data_{month_year} cpd on cpd.meter_id = mm.meter_id and cpd.mill_date = cp.mill_date and cpd.mill_shift = cp.mill_shift"
                            joins.append(f"select machine_status,poll_duration, mill_date, mill_shift, meter_id from {join_p}")
                    
                    if len(union_queries) == 0:
                        return _getErrorResponseJson("Power Table Not Available...")     
                         
                    if len(joins) == 0:
                        return _getErrorResponseJson("Polling Data Table Not Available...")         
                  
        group = ''  


        data = await current_power_dtl(cnx,company_id ,campus_id,bu_id ,plant_id ,plant_department_id ,equipment_group_id ,equipment_id,function_id ,meter_id ,group_for ,groupby ,period_id,from_date,to_date,shift_id,limit_report_for,limit_exception_for,limit_order_by ,limit_operation_value ,is_critical ,converter_id ,report_for,is_function , function_type ,reportfor,employee_id ,is_minmax,is_main_meter,is_demand,meter_type,is_plant_wise)


        if for_android == 'yes':
            response = [{
                "iserror": False,
                "message": "Data Returned Successfully.",
                "data": data
            }]
        else:
            response = {
                "iserror": False,
                "message": "Data Returned Successfully.",
                "data": data
            }

        
        return response
    except Exception as e:
        return get_exception_response(e)

@router.post("/get_custom_function_dashboard_detail/", tags=["Dashboard"])
async def get_custom_function_dashboard_detail(plant_id:int = Form(''),cnx: AsyncSession = Depends(get_db)):

    try: 
        
        result = await function_dashboard(cnx,plant_id)
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": result
        }

        return response
    
    except Exception as e:
        return get_exception_response(e)
    
@router.post("/get_custom_function_dashboard_detail2/", tags=["Dashboard"])
async def get_custom_function_dashboard_detail2(plant_id:int = Form(''),cnx: AsyncSession = Depends(get_db)):

    try: 
        
        result = await function_dashboard2(cnx,plant_id)
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": result
        }

        return response
    
    except Exception as e:
        return get_exception_response(e)

@router.post("/communication_status/")
async def communication_status(plant_department_id:int=Form(''),
                               plant_id : int = Form(''),
                               campus_id : int = Form(''),
                               cnx: AsyncSession = Depends(get_db)):
    try: 

        # if plant_department_id == '':
        #     return _getErrorResponseJson("plant_department_id is required") 
        
        result = await communication_status_query(cnx,plant_department_id,plant_id,campus_id)
        
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": result
        }

        return response
    
    except Exception as e:
        return get_exception_response(e)
    
async def generate_excel_report(result, month_year, report_type, report_for,kwh,shift_time):
    print("Function Call")
    wb = Workbook()
    ws = wb.active
    border_style = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    parameter = kwh
    # workbook = Workbook()
    # sheet = workbook.active
    ws.title = 'EMS' 
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')  
    
    cell = "D2"
    data = f"MONTH WISE ENERGY CONSUMPTION REPORT FOR KWH - {month_year} ({shift_time} to {shift_time})"
    ws[cell] = data
    font = Font(bold=True, name='Calibri', size=15)
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].font = font
    ws[cell].alignment = alignment 
    ws.cell(row=2, column=4, value=f"MONTH WISE ENERGY CONSUMPTION REPORT FOR KWH - {month_year} ({shift_time} to {shift_time})").fill = fill_cyan
    
    # Set the border style
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    if result == []:
        print(1)
        cell = "O10"
        data = "No Data"
        ws[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment

        ws.column_dimensions[cell[0]].width = len(data) + 2  

        font = Font(name='Calibri', size=25)
        ws[cell].font = font

    start_rows = 6
    start_row = 4
    start_col = 5
    print("Function Start")
    
   
    
    if report_type == "date":
        start_row = 4
        start_col = 5
        month, year = map(int, month_year.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        
        fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
        if kwh == 'Energy':
            # ws.merge_cells('D2:CS3')
            ws.merge_cells('B4:B5')
            cell = "B4"
            ws[cell] = "Meter Tag"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15 

            ws.merge_cells('C4:C5')
            cell = "C4" 
            ws[cell] = "Meter Name"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 60

            ws.merge_cells('D4:D5')
            cell = "D4" 
            ws[cell] = "Meter Type"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15

            ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
            ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
            ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
            for i in  range(1, days_in_month + 1):
                # Write the hour range header
                header_text = f"{i:02d}-{month_year}"
                ws.merge_cells(start_row=4, start_column=2 + i * 3, end_row=4, end_column=4 + i * 3)
                header_cell = ws.cell(row=4, column=2 + i * 3)
                header_cell.value = header_text

                # Apply styles to the header cell
                header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                header_cell.font = Font(bold=True, name='Calibri', size=12)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                # shifts = ['Start CkWh', 'End Ckwh', 'kWh']
                secondary_headers = ["Start CkWh", "End CkWh", "kwh"]
                for j, header in enumerate(secondary_headers):
                    secondary_header_cell = ws.cell(row=5, column=2+i*3+j)
                    secondary_header_cell.value = header

                    # Apply styles to the secondary header cell
                    secondary_header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                    secondary_header_cell.font = Font(bold=True, name='Calibri', size=12)
                    secondary_header_cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.column_dimensions[secondary_header_cell.column_letter].width = len(header) + 15
            ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=2+i*3+j)
            for meter_data in result:
                meter_name = meter_data["meter_name"]
                meter_code = meter_data["meter_code"]
                meter_type = meter_data["meter_type"]
                plant_name = meter_data["plant_name"]
                row_data = ['',meter_code, meter_name, meter_type]

                ws.merge_cells('B2:C2')
                ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
                ws.cell(row=2, column=2, value=f"Plant : {plant_name}").fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=2, column=2).font = font
            
                ws.merge_cells('B3:C3')
                ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
                ws.cell(row=3, column=2, value=f"Parameter : {parameter}").fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=3, column=2).font = font

                for i in range(1, days_in_month + 1):
                    hour = f"d{i}"
                    print(hour)
                    kwh = meter_data.get(hour, "")  # Assigning a new value to kwh
                    machine_kwh = meter_data.get(f"machine_kwh_{hour}", "")
                    master_kwh = meter_data.get(f"master_kwh_{hour}", "")
                    row_data.extend([master_kwh, machine_kwh, kwh])
                ws.append(row_data)

            end_row = ws.max_row   
            row_range =ws.iter_rows(min_row=2, min_col=2, max_row=end_row, max_col=2+i*3+j)

            for row in row_range:
                for cell in row:
                    cell.border = border
        else:
            # ws.merge_cells('D2:AI3')
            
            cell = "B4"
            ws[cell] = "Meter Tag"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15 
        
            cell = "C4" 
            ws[cell] = "Meter Name"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 60

            cell = "D4" 
            ws[cell] = "Meter Type"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15

            ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
            ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
            ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
            for i in range(1, days_in_month + 1):
                header_text = f"{i:02d}-{month_year}"
                ws.merge_cells(start_row=4, start_column=4 + i, end_row=4, end_column=4 + i)
                header_cell = ws.cell(row=4, column=4 + i)
                header_cell.value = header_text
                header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                header_cell.font = Font(bold=True, name='Calibri', size=12)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                column_width = max(len(header_text), 15)  
                ws.column_dimensions[ws.cell(row=4, column=4 + i).column_letter].width = column_width
            ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4 + i)

            for meter_data in result:
                meter_name = meter_data["meter_name"]
                meter_code = meter_data["meter_code"]
                meter_type = meter_data["meter_type"]
                plant_name = meter_data["plant_name"]
                row_data = ['', meter_code, meter_name, meter_type]
                ws.merge_cells('B2:C2')
                ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
                ws.cell(row=2, column=2).fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=2, column=2).font = font

                ws.merge_cells('B3:C3')
                ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
                ws.cell(row=3, column=2).fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=3, column=2).font = font

                for i in range(1, days_in_month + 1):
                    day = f"d{i}"
                    kwh = meter_data.get(day, "")
                    
                    row_data.extend([kwh])
                ws.append(row_data)

            end_row = ws.max_row
            row_range = ws.iter_rows(min_row=2, max_row=end_row, min_col=2, max_col=i+4)
            for row in row_range:
                for cell in row:
                    cell.border = border


    elif report_type == "shift":
        
        headers = ["Meter Tag", "Meter Name", "Meter Type"]
        for col, header in enumerate(headers, start=2):
            ws.cell(row=4, column=col, value=header).fill = fill_cyan
            ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
            ws.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = len(header) + 15

        month, year = map(int, month_year.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        # ws.merge_cells('D2:CS3')

        ws.merge_cells('B4:B5')
        cell = "B4"
        ws[cell] = "Meter Tag"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 15 

        ws.merge_cells('C4:C5')
        cell = "C4" 
        ws[cell] = "Meter Name"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 60

        ws.merge_cells('D4:D5')
        cell = "D4" 
        ws[cell] = "Meter Type"
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, name='Calibri', size=12)
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws.column_dimensions[cell[0]].width = len("%") + 15

        ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
        ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
        ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
        for i in  range(1, days_in_month + 1):
            # Write the hour range header
            header_text = f"{i:02d}-{month_year}"
            ws.merge_cells(start_row=4, start_column=2 + i * 3, end_row=4, end_column=4 + i * 3)
            header_cell = ws.cell(row=4, column=2 + i * 3)
            header_cell.value = header_text

            # Apply styles to the header cell
            header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
            header_cell.font = Font(bold=True, name='Calibri', size=12)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            # shifts = ['Start CkWh', 'End Ckwh', 'kWh']
            secondary_headers = ["S1", "S2", "S3"]
            for j, header in enumerate(secondary_headers):
                secondary_header_cell = ws.cell(row=5, column=2+i*3+j)
                secondary_header_cell.value = header

                # Apply styles to the secondary header cell
                secondary_header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                secondary_header_cell.font = Font(bold=True, name='Calibri', size=12)
                secondary_header_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[secondary_header_cell.column_letter].width = len(header) + 15
        
        for meter_data in result:
            meter_name = meter_data["meter_name"]
            meter_code = meter_data["meter_code"]
            meter_type = meter_data["meter_type"]
            plant_name = meter_data["plant_name"]
            
            # Write meter information
            # ws.append(['', meter_code, meter_name, meter_type])  # Writing meter info in a row
            row_data = ['', meter_code, meter_name, meter_type]
            ws.merge_cells('B2:C2')
            ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
            ws.cell(row=2, column=2).fill = fill_cyan
            font = Font(bold=True, name='Calibri', size=12)
            ws.cell(row=2, column=2).font = font

            ws.merge_cells('B3:C3')
            ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
            ws.cell(row=3, column=2).fill = fill_cyan
            font = Font(bold=True, name='Calibri', size=12)
            ws.cell(row=3, column=2).font = font

            # for shift in secondary_headers:
            row_data = ['', meter_code, meter_name, meter_type]  # Initialize row data here
            for i in range(1, days_in_month + 1):
                hour = f"ds{1}_{i}"
                kwh = meter_data.get(hour, "") 
                machine_kwh = meter_data.get(f"ds{2}_{i}", "")
                master_kwh = meter_data.get(f"ds{3}_{i}", "")
                row_data.extend([kwh, machine_kwh, master_kwh]) # Assigning a new value to kwh
            ws.merge_cells(start_row=2, start_column=4 , end_row=3, end_column=4 + i*3)  
            ws.append(row_data)





        # start_row = 4
        # start_col = 5

        # fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
        # for data in result:  
        #     meter_code = data["meter_code"]
        #     meter_name = data["meter_name"]
        #     meter_type = data["meter_type"]
        #     plant_name = data["plant_name"]
        #     ws.merge_cells('B2:C2')
        #     ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
        #     ws.cell(row=2, column=2).fill = fill_cyan
        #     font = Font(bold=True, name='Calibri', size=12)
        #     ws.cell(row=2, column=2).font = font

        #     ws.merge_cells('B3:C3')
        #     ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
        #     ws.cell(row=3, column=2).fill = fill_cyan
        #     font = Font(bold=True, name='Calibri', size=12)
        #     ws.cell(row=3, column=2).font = font
        #     ws.cell(row=start_rows, column=2, value=meter_code).alignment = Alignment(horizontal="center")
        #     alignment = Alignment(horizontal="center", wrap_text=True)

        #     ws.cell(row=start_rows, column=4, value=meter_type).alignment = Alignment(horizontal="center")
        #     alignment = Alignment(horizontal="center", wrap_text=True)

        #     ws.cell(row=start_rows, column=3, value=meter_name).alignment = alignment
        #     meter_name_length = len(meter_name) + 40
        #     column_letter_c = get_column_letter(3)  # Column C
        #     column_width_c = max(meter_name_length, ws.column_dimensions[column_letter_c].width)
            
        #     for day in range(1, days_in_month + 1):
        #         ws.merge_cells(
        #                 start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
        #         alignment = Alignment(horizontal="center", vertical="center")
        #         cell = ws.cell(row=start_row, column=start_col)
        #         cell.alignment = alignment 
        #         cell.font = Font(bold=True, name='Calibri', size=12)
        #         cell.value = f"{day:02d}-{month_year}"
        #         cell.fill = fill_cyan

        #         shifts = ['s1', 's2', 's3']

        #         for shift in range(1, 4):
        #             # Write the shift label (s1, s2, s3) in cells D5, E5, F5 for each day
        #             ws.cell(row=5, column=start_col + shift - 1, value=shifts[shift - 1])
        #             ws.cell(row=5, column=start_col + shift - 1).fill = fill_cyan
        #             ws.cell(row=5, column=start_col + shift - 1).alignment = Alignment(horizontal="center")

        #             cell = ws.cell(row=start_rows + shift, column=start_col + shift)
        #             cell.value = data.get(f"d{day}_s{shift}", "")
        #             cell.alignment = Alignment(horizontal="center")
        #             if cell.value == 0:
        #                 cell.value = ""

        #         start_col += 3
        #     # Adjust column size based on the maximum text length vertically
        #     for shift in range(1, 4):
        #         for j in range(1, 32):
        #             column_letter = get_column_letter((shift - 1) * 31 + j + 4)  # Assuming data columns start from column E (column index 5)
        #             cell = ws.cell(row=start_rows, column=(shift - 1) * 31 + j + 4)
        #             cell.value = data.get(f"ds{shift}_{j}", "")
        #             cell.alignment = Alignment(horizontal="center")
        #             # if cell.value == 0:
        #             #     cell.value = ""

        #             # Adjust column size based on the maximum text length vertically
        #             cell_text_length = len(str(cell.value))
        #             column_width = max(cell_text_length, ws.column_dimensions[column_letter].width)
        #             ws.column_dimensions[column_letter].width = column_width
                
        #         row_range = ws[f"A{start_row}:CS{start_rows}"]
        #         for row in row_range:
        #             for cell in row:
        #                 cell.border = border
        #     start_rows +=1
        end_row = ws.max_row   
        row_range = ws.iter_rows(min_row=2, max_row=end_row, min_col=2, max_col=4 + i*3)
        for row in row_range:
            for cell in row:
                cell.border = border
        

    file_name = f'MonthWiseReport-{month_year}.xlsx'
    print("file_name",file_name)
    file_path = os.path.join(static_dir, file_name)
    print("file_path....",file_path)
    wb.save(file_path)
    
# async def generate_excel_report(result, month_year, report_type, report_for,kwh,shift_time):
#     print("Function Call")
#     wb = Workbook()
#     ws = wb.active
#     border_style = Border(left=Side(style='thin'), 
#                      right=Side(style='thin'), 
#                      top=Side(style='thin'), 
#                      bottom=Side(style='thin'))

#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = 'EMS' 
#     fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')  
    
#     cell = "D2"
#     data = f"MONTH WISE ENERGY CONSUMPTION REPORT FOR KWH - {month_year} ({shift_time} to {shift_time})"
#     sheet[cell] = data
#     font = Font(bold=True, name='Calibri', size=15)
#     alignment = Alignment(horizontal="center", vertical="center")
#     sheet[cell].font = font
#     sheet[cell].alignment = alignment 
#     sheet.cell(row=2, column=4, value=f"MONTH WISE ENERGY CONSUMPTION REPORT FOR KWH - {month_year} ({shift_time} to {shift_time})").fill = fill_cyan
    
#     # Set the border style
#     border = Border(
#         left=Side(style='thin', color='000000'),
#         right=Side(style='thin', color='000000'),
#         top=Side(style='thin', color='000000'),
#         bottom=Side(style='thin', color='000000')
#     )
    
#     sheet.merge_cells('B4:B5')
#     cell = "B4"
#     sheet[cell] = "Meter Tag"
#     alignment = Alignment(horizontal="center", vertical="center")
#     font = Font(bold=True, name='Calibri', size=12)
#     sheet[cell].alignment = alignment 
#     sheet.column_dimensions[cell[0]].width = len("%") + 15 
   
#     sheet.merge_cells('C4:C5')
#     cell = "C4" 
#     sheet[cell] = "Meter Name"
#     alignment = Alignment(horizontal="center", vertical="center")
#     font = Font(bold=True, name='Calibri', size=12)
#     sheet[cell].alignment = alignment 
#     sheet.column_dimensions[cell[0]].width = len("%") + 15

#     sheet.merge_cells('D4:D5')
#     cell = "D4" 
#     sheet[cell] = "Meter Type"
#     alignment = Alignment(horizontal="center", vertical="center")
#     font = Font(bold=True, name='Calibri', size=12)
#     sheet[cell].alignment = alignment 
#     sheet.column_dimensions[cell[0]].width = len("%") + 15

#     sheet.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
#     sheet.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
#     sheet.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
    
    
#     # border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
#     # fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
    
#     # # Set headers
#     # headers = ["Meter Tag", "Meter Name", "Meter Type"]
#     # for col, header in enumerate(headers, start=2):
#     #     sheet.cell(row=4, column=col, value=header).fill = fill_cyan
#     #     sheet.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
#     #     sheet.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center")
#     #     sheet.column_dimensions[get_column_letter(col)].width = len(header) + 15
    
#     if result == []:
#         print(1)
#         cell = "O10"
#         data = "No Data"
#         sheet[cell] = data

#         alignment = Alignment(horizontal="center", vertical="center")
#         sheet[cell].alignment = alignment

#         sheet.column_dimensions[cell[0]].width = len(data) + 2  

#         font = Font(name='Calibri', size=25)
#         sheet[cell].font = font

#     start_rows = 6
#     start_row = 4
#     start_col = 5
#     print("Function Start")
#     for data in result:  
#         meter_code = data["meter_code"]
#         meter_name = data["meter_name"]
#         meter_type = data["meter_type"]
#         plant_name = data["plant_name"]

#         sheet.merge_cells('B2:C2')
#         sheet.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
#         sheet.cell(row=2, column=2, value=f"Plant : {plant_name}").fill = fill_cyan

#         sheet.merge_cells('B3:C3')
#         sheet.cell(row=3, column=2, value=f"Parameter : {kwh}").alignment = Alignment(horizontal="left")
#         sheet.cell(row=3, column=2, value=f"Parameter : {kwh}").fill = fill_cyan

#         sheet.cell(row=start_rows, column=2, value=meter_code).alignment = Alignment(horizontal="center")
#         alignment = Alignment(horizontal="center", wrap_text=True)

#         sheet.cell(row=start_rows, column=4, value=meter_type).alignment = Alignment(horizontal="center")
#         alignment = Alignment(horizontal="center", wrap_text=True)

#         sheet.cell(row=start_rows, column=3, value=meter_name).alignment = alignment
#         meter_name_length = len(meter_name) + 40
#         column_letter_c = get_column_letter(3)  # Column C
#         column_width_c = max(meter_name_length, sheet.column_dimensions[column_letter_c].width)

#         sheet.column_dimensions[column_letter_c].width = column_width_c
#         print("Function Start111")
#         if report_type == "date":
#             start_row = 4
#             start_col = 5
#             month, year = map(int, month_year.split('-'))
#             days_in_month = calendar.monthrange(year, month)[1]
            
#             fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
#             if kwh == 'Energy':
#                 sheet.merge_cells('D2:CS3')
#                 for day in range(1, days_in_month + 1):
#                     cell = sheet.cell(row=start_row, column=start_col)
#                     cell.value = f"{day:02d}-{month_year}"
#                     cell.fill = fill_cyan
#                     sheet.merge_cells(
#                         start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
#                     alignment = Alignment(horizontal="center", vertical="center")
#                     font = Font(bold=True, name='Calibri', size=12)
#                     cell.alignment = alignment 
#                     cell.border = border_style 
#                     shifts = ['Start CkWh', 'End Ckwh', 'kWh']

#                     for shift in range(1, 4):
#                         # Write the shift label (s1, s2, s3) in cells D5, E5, F5 for each day
#                         sheet.cell(row=5, column=start_col + shift - 1, value=shifts[shift - 1])
#                         sheet.cell(row=5, column=start_col + shift - 1).fill = fill_cyan
#                         sheet.cell(row=5, column=start_col + shift - 1).alignment = Alignment(horizontal="center")
#                         cell_length = len(shifts) + 2 
#                         column_letter = get_column_letter(start_col + shift - 1)
#                         column_width = max(cell_length, sheet.column_dimensions[column_letter].width)
#                         sheet.column_dimensions[column_letter].width = column_width
#                         cell = sheet.cell(row=start_row + shift, column=start_col + shift)           
                        
#                     kwh_cell = sheet.cell(row=start_rows, column=start_col + shift-1)
#                     sheet.row_dimensions[start_rows + shift].height = 25

#                     kwh_key = f"d{day}"
#                     kwh_value = data.get(kwh_key, "")
#                     kwh_cell.value = kwh_value

#                     master_kwh_cell = sheet.cell(row=start_rows, column=start_col + shift-3)
#                     sheet.row_dimensions[start_rows + shift].height = 25

#                     master_kwh_key = f"master_kwh_d{day}"
#                     master_kwh_value = data.get(master_kwh_key, "")
#                     master_kwh_cell.value = master_kwh_value
                    
#                     machine_kwh_cell = sheet.cell(row=start_rows, column=start_col + shift-2)
#                     sheet.row_dimensions[start_rows + shift].height = 25

#                     machine_kwh_key = f"machine_kwh_d{day}"
#                     machine_kwh_value = data.get(machine_kwh_key, "")
#                     machine_kwh_cell.value = machine_kwh_value
                        
#                     start_col += 3
                    
#                 row_range = sheet[f"B2:CS{start_rows}"]
#                 for row in row_range:
#                     for cell in row:
#                         cell.border = border
#                 start_rows +=1
#             else:
#                 sheet.merge_cells('D2:AI3')
                
#                 for day in range(1, days_in_month + 1):
#                     sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row+1, end_column=start_col)
#                     cell = sheet.cell(row=start_row, column=start_col)
#                     alignment = Alignment(horizontal="center", vertical="center")
#                     font = Font(bold=True, name='Calibri', size=12)
#                     cell.alignment = alignment 
#                     cell.value = f"{day:02d}-{month_year}"
#                     cell.fill = fill_cyan
                    
                    
#                     start_col += 1

#                 for j in range(1, 32):
#                     column_letter = get_column_letter(j + 4)  
#                     cell = sheet.cell(row=start_rows, column=j + 4)
#                     cell.value = data.get(f"d{j}", "")
#                     cell.alignment = Alignment(horizontal="center")
#                     if cell.value == 0:
#                         cell.value = ""

#                 row_range = sheet[f"A{start_row}:AI{start_rows}"]
#                 for row in row_range:
#                     for cell in row:
#                         cell.border = border

#                 start_rows +=1

#         elif report_type == "shift":
#             sheet.merge_cells('D2:CS3')
#             month, year = map(int, month_year.split('-'))
#             days_in_month = calendar.monthrange(year, month)[1]

#             start_row = 4
#             start_col = 5

#             fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
            
#             for day in range(1, days_in_month + 1):
#                 sheet.merge_cells(
#                         start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
#                 alignment = Alignment(horizontal="center", vertical="center")
#                 cell = sheet.cell(row=start_row, column=start_col)
#                 cell.alignment = alignment 
#                 cell.value = f"{day:02d}-{month_year}"
#                 cell.fill = fill_cyan

#                 shifts = ['s1', 's2', 's3']

#                 for shift in range(1, 4):
#                     # Write the shift label (s1, s2, s3) in cells D5, E5, F5 for each day
#                     sheet.cell(row=5, column=start_col + shift - 1, value=shifts[shift - 1])
#                     sheet.cell(row=5, column=start_col + shift - 1).fill = fill_cyan
#                     sheet.cell(row=5, column=start_col + shift - 1).alignment = Alignment(horizontal="center")

#                     cell = sheet.cell(row=start_rows + shift, column=start_col + shift)
#                     cell.value = data.get(f"d{day}_s{shift}", "")
#                     cell.alignment = Alignment(horizontal="center")
#                     if cell.value == 0:
#                         cell.value = ""

#                 start_col += 3
#             # Adjust column size based on the maximum text length vertically
#             for shift in range(1, 4):
#                 for j in range(1, 32):
#                     column_letter = get_column_letter((shift - 1) * 31 + j + 4)  # Assuming data columns start from column E (column index 5)
#                     cell = sheet.cell(row=start_rows, column=(shift - 1) * 31 + j + 4)
#                     cell.value = data.get(f"ds{shift}_{j}", "")
#                     cell.alignment = Alignment(horizontal="center")
#                     if cell.value == 0:
#                         cell.value = ""

#                     # Adjust column size based on the maximum text length vertically
#                     cell_text_length = len(str(cell.value))
#                     column_width = max(cell_text_length, sheet.column_dimensions[column_letter].width)
#                     sheet.column_dimensions[column_letter].width = column_width
                
#                 row_range = sheet[f"A{start_row}:CS{start_rows}"]
#                 for row in row_range:
#                     for cell in row:
#                         cell.border = border
#             start_rows +=1
    
#     file_name = f'MonthWiseReport-{month_year}.xlsx'
#     print("file_name",file_name)
#     file_path = os.path.join(static_dir, file_name)
#     print("file_path....",file_path)
#     workbook.save(file_path)

@router.post("/performance_report/", tags=["Report"])
async def month_wise_report(request:Request,
                            campus_id : int = Form(''),
                            company_id : int = Form(''),
                            bu_id : int = Form(''),
                            plant_id : int = Form(''),
                            plant_department_id : int = Form(''),
                            equipment_group_id : int = Form(''),
                            meter_id: str = Form(''),
                            month_year: str = Form(''),
                            report_for: str = Form(''),
                            report_type: str = Form(''),                                                      
                            meter_type: str = Form(''),                                                      
                            kwh: str = Form(''),                                                      
                            employee_id: str = Form(''),                                                    
                            cnx: AsyncSession = Depends(get_db)):

    try: 
        # result = ''
        if month_year == "" :
            return _getErrorResponseJson("MonthYear is Required...")
        
        if report_for == "" :
            return _getErrorResponseJson("ReportFor is Required...")
        
        if report_type == "" :
            return _getErrorResponseJson("Report Type is Required...")
        
        if kwh == "" :
            return _getErrorResponseJson("Kwh is Required...")
        
        if report_for == '12to12':
            formatted_month_year = month_year.replace('-', '')
            res = await check_power_12_table(cnx,formatted_month_year)

            if report_type not in ['date']:
                 return _getErrorResponseJson("Invalid report type")

        else:
            formatted_month_year = month_year.replace('-', '')
            res = await check_power_table(cnx,formatted_month_year)    

        if report_type not in ['date', 'shift']:
                 return _getErrorResponseJson("Invalid report type")
        
        if len(res)==0:
            return _getErrorResponseJson("Table not available...")
        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        
        if len(data1) > 0:
            for shift_record in data1:
                shift_time = shift_record["a_shift_start_time"]
            print(shift_time)  
        result = await month_report(cnx,campus_id,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,employee_id,meter_id,month_year,report_for,report_type,meter_type,kwh,request)
        await generate_excel_report(result, month_year,report_type, report_for,kwh,shift_time)
        
        file_path = os.path.join(static_dir, f"MonthWiseReport-{month_year}.xlsx")
        
        results = f"http://{request.headers['host']}/attachments/MonthWiseReport-{month_year}.xlsx"
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "file_url": results
                }
        return response
    except Exception as e:
        return get_exception_response(e)
 
def dailyreport(date,day_month_year_value,report_for,dates,year,next_year):
        try:
            # file_path = f'{static_dir}\daily_report_template.xlsx'
            file_path = os.path.abspath(os.path.join(os.path.dirname(__file__),"..", "..", "..", "..","..",  "daily_report_template.xlsx"))
            print("file_path....",file_path)
            workbook = Workbook()
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            sheet.title = 'DAILY REPORT'
            border = Border(left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"))
            cell = "F5"
            data = f"Daily Energy Report"
            sheet[cell] = data
            cell = "J6"
            data = f"{date}"
            sheet[cell] = data
            bold_font = Font(bold=True,size = 12)
            sheet[cell].font = bold_font
            center_alignment = Alignment(horizontal="center")
            sheet[cell].alignment = center_alignment
            cell = "H7"
            if report_for == '6to6':
                datas = f'06:00 to 06:00'
            else:
                datas = f'Day'
            sheet[cell] = datas
            bold_font = Font(bold=True,size = 12)
            sheet[cell].font = bold_font
            cell = "J7"
            # parsed_date = datetime.datetime.strptime(date, "%d-%b-%Y")
            # year = parsed_date.year
            fiscal_year = f"{year}-{next_year}"
            data = f" FY {fiscal_year}"
            sheet[cell] = data
            sheet[cell].font = bold_font
            sheet[cell].alignment = center_alignment
            row = 8
            col = 6
            for row_item in day_month_year_value:
                energy_in_units = row_item['func_name']
                day = row_item['formula_d']
                month = row_item['formula_m']
                year = row_item['formula_y']
                roundoff_value = row_item['roundoff_value']
                print("roundoff_value",roundoff_value)
                merge_range = f'F{row}:G{row}'
                sheet[f'F{row}']=energy_in_units

                for col in range(6, 8):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = border
                if isinstance(day, (int, float)) and isinstance(month, (int, float)) and isinstance(year, (int, float)):
                    
                    day = float(day)
                    month = float(month)
                    year = float(year)
          
                    day_cell = sheet.cell(row=row, column=8)
                    if roundoff_value == 0:
                        day_cell.value = f'=({day})'
                    else:
                        day_cell.value = f'=MROUND(({day}),{roundoff_value})'
                    day_cell.number_format = '0.00'
                    day_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    day_cell.border = border

                    month_cell = sheet.cell(row=row, column=9)
                    if roundoff_value == 0:
                        month_cell.value = f'=({month})'
                    else:
                        month_cell.value = f'=MROUND(({month}),{roundoff_value})'
                    month_cell.number_format = '0.00'
                    month_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    month_cell.border = border

                    year_cell = sheet.cell(row=row, column=10)
                    if roundoff_value == 0:
                        year_cell.value = f'=({year})'
                    else:
                        year_cell.value = f'=MROUND(({year}),{roundoff_value})'
                    year_cell.number_format = '0.00'
                    year_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    year_cell.border = border

                else:
                    day_cell = sheet.cell(row=row, column=8)
                    if roundoff_value == 0:
                        day_cell.value = f'=({day})'
                    else:
                        day_cell.value = f'=MROUND(({day}),{roundoff_value})'
                    day_cell.number_format = '0.00'
                    day_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    day_cell.border = border

                    month_cell = sheet.cell(row=row, column=9)
                    if roundoff_value == 0:
                        month_cell.value = f'=({month})'
                    else:
                        month_cell.value = f'=MROUND(({month}),{roundoff_value})'
                    month_cell.number_format = '0.00'
                    month_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    month_cell.border = border

                    year_cell = sheet.cell(row=row, column=10)
                    if roundoff_value == 0:
                        year_cell.value = f'=({year})'
                    else:
                        year_cell.value = f'=MROUND(({year}),{roundoff_value})'
                    year_cell.number_format = '0.00'
                    year_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    year_cell.border = border

                sheet.merge_cells(merge_range)
                sheet.row_dimensions[row].height = 30
                row+=1

            last_row = sheet.max_row
            next_row = last_row + 1
            merge_range = f'F{next_row}:J{next_row}'

            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_row, column=6)

            sheet.row_dimensions[next_row].height = 25
            merged_cell.value = "Remarks:"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            for col in range(6, 11):  
                cell = sheet.cell(row=next_row, column=col)
                cell.border = border

            next_rows = last_row + 2
            merge_range = f'F{next_rows}:J{next_rows}'
            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_rows, column=6)
            sheet.row_dimensions[next_rows].height = 25
            for col in range(6, 11):  
                cell = sheet.cell(row=next_rows, column=col)
                cell.border = border

            next_rows1 = last_row + 3
            merge_range = f'F{next_rows1}:J{next_rows1}'
            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_rows1, column=6)
            sheet.row_dimensions[next_rows1].height = 25
            for col in range(6, 11):  
                cell = sheet.cell(row=next_rows1, column=col)
                cell.border = border

            next_rows2 = last_row + 4
            merge_range = f'F{next_rows2}:G{next_rows2}'
            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_rows2, column=6)
            sheet.row_dimensions[next_rows2].height = 15
            merged_cell.value = "Shift Incharge-(Elect)"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')

            for col in range(6, 8):  
                cell = sheet.cell(row=next_rows2, column=col)
                cell.border = border
            merge_ranged = f'H{next_rows2}'
            sheet.merge_cells(merge_ranged)
            merged_cell = sheet.cell(row=next_rows2, column=8)
            sheet.row_dimensions[next_rows2].height = 15
            merged_cell.value = "Sectional Incharge-(Elect)"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(8,9):  
                cell = sheet.cell(row=next_rows2, column=col)
                cell.border = border
            merge_rangee = f'I{next_rows2}:J{next_rows2}'
            sheet.merge_cells(merge_rangee)
            merged_cell = sheet.cell(row=next_rows2, column=9)
            sheet.row_dimensions[next_rows2].height = 15
            merged_cell.value = "HOD"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(9, 11):  
                cell = sheet.cell(row=next_rows2, column=col)
                cell.border = border
            file_name = f'DailyReport - {dates}.xlsx'
            print("111111111111",file_name)
            file_path = os.path.join(static_dir, file_name)
            workbook.save(file_path)
            return file_path
        
        except Exception as e:
            return get_exception_response(e)
   
@router.post("/custom_daily_report/", tags=["Report"])
async def custom_daily_report_api(request:Request,
                                  date: str=Form(''), 
                                  report_for : str = Form(''),          
                                  cnx: AsyncSession = Depends(get_db)):

    try: 
        if date == "" :
            return _getErrorResponseJson("date is required...")
        
        if report_for == "" :
            return _getErrorResponseJson("report_for is required...")
        
        dates=await parse_date(date)
        from_date_str = dates.strftime("%d-%m-%Y")
        datetime_obj = datetime.strptime(from_date_str, "%d-%m-%Y")
        f_date = datetime_obj.strftime("%d-%m-%Y") 
        formatted_date = datetime_obj.strftime("%d-%b-%Y") 
        month = f_date[3:5]
        given_year = formatted_date[7:]
        print("month",month)
        if int(month) >=4:
            year = formatted_date[7:]
            next_year = int(given_year) +1
        if int(month) <4:
            year = int(given_year) -1
            next_year =formatted_date[7:]

        result = await daily_report(cnx,date,report_for,request)
        await dailyreport(formatted_date,result,report_for,date,year,next_year)
        file_path = os.path.join(base_path, f"DailyReport - {date}.xlsx")
        results = f"http://{request.headers['host']}/attachments/DailyReport - {date}.xlsx"

        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "data": results
                }

        return response
    except Exception as e:
        return get_exception_response(e)

async def generate_year_wise_excel_report(result, year,next_year,report_for,shift_time):
    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..","..", "..", "YearWiseReport_templete.xlsx"))
    print("filepath",file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS'
    cell = "B1"
    data = f"YEAR WISE ENERGY CONSUMPTION REPORT FOR KWH - {year}({shift_time} to {shift_time})"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=13)
    sheet[cell].font = font

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    minimum_column_width = 15
    cell = "A3"
    sheet[cell] = "meter Code"
    cell = "B3"
    sheet[cell] = "meter Name"
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
    sheet.cell(row=3, column=1, value="meter Code").fill = fill_cyan
    sheet.cell(row=3, column=2, value="meter Name").fill = fill_cyan
    if result == []:
        print(1)
        cell = "E10"
        data = "No Data"
        sheet[cell] = data
        font = Font(bold=True, name='Calibri', size=13)
        sheet[cell].font = font
    else:
    # Populate month-year headers in row 3 starting from column C
        col_index = 3
        for month_year in result[0].keys():
            if month_year not in ["meter_code", "meter_name"]:
                sheet.cell(row=3, column=col_index, value=month_year).fill = fill_cyan
                col_index += 1

        for row, data in enumerate(result, start=4):
            meter_code = data["meter_code"]
            meter_name = data["meter_name"]
            sheet.cell(row=row, column=1, value=meter_code).alignment = Alignment(horizontal="center")
            alignment = Alignment(horizontal="center", wrap_text=True)
            sheet.cell(row=row, column=2, value=meter_name).alignment = alignment
            # meter_name_length = len(meter_name)
            # column_letter_c = get_column_letter(3)
            # column_width_c = max(meter_name_length, sheet.column_dimensions[column_letter_c].width)
            # sheet.column_dimensions[column_letter_c].width = column_width_c

            col_indexs = 1
            for kwh_value in data.values():
                if isinstance(kwh_value, (int, float)):
                    if kwh_value == 0:
                        kwh_value = ''
                    sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
                    cell = sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
                    cell.number_format = '0.00'

                
                col_indexs += 1
            row_range = sheet[f"A1:N{row}"]
            for row in row_range:
                for cell in row:
                    cell.border = border

    file_name = f'YearWiseReport-{year}-{next_year}.xlsx'
    file_path = os.path.join(static_dir, file_name)
    print("file_path",file_path)
    workbook.save(file_path)

@router.post("/year_wise_report/", tags=["Report"])
async def year_wise_report(request: Request,
                           campus_id: str = Form(''),
                           company_id: str = Form(''),
                           bu_id: str = Form(''),
                           plant_id: str = Form(''),
                           plant_department_id: str = Form(''),
                           equipment_group_id: str = Form(''),
                           meter_id: str = Form(''),
                           year: str = Form(""),
                           report_for : str  = Form(""),
                           kwh_type : str  = Form(""),
                           employee_id: str = Form(""),
                           cnx: AsyncSession = Depends(get_db)):

    try: 
        mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}

        print(meter_id)
        if year == "" :
            return _getErrorResponseJson("year is required...")
        
        if report_for == "" :
            return _getErrorResponseJson("report_for is required...")
        
        next_year = int(year) + 1
        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        
        if len(data1) > 0:
            for shift_record in data1:
                shift_time = shift_record["a_shift_start_time"]

        tables_to_union = []
        for month in range(4, 13):
            month_year = f"{mill_month[month]}{year}"
            print(month_year)
            if report_for == '12to12':
                query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = 'power_{month_year}_12'"""
                result_query = await cnx.execute(query)
                result_query = result_query.fetchall()
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.power_{month_year}_12")
                print(month_year)
            else:
                query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = 'power_{month_year}'"""
                result_query = await cnx.execute(query)
                result_query = result_query.fetchall()

                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.power_{month_year}")
        
        next_year = int(year) + 1
        mill_month = {1: "01", 2: "02", 3: "03"}

        for month in range(1, 4):
            month_year = f"{mill_month[month]}{next_year}"
            print(month_year)
            if report_for == '12to12':
                query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = 'power_{month_year}_12' """
                result_query = await cnx.execute(query)
                result_query = result_query.fetchall()

                print("result_query",result_query)
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.power_{month_year}_12")
                tables_union_query = " UNION ALL ".join(tables_to_union)
            else:   
                query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = 'power_{month_year}' """
                result_query = await cnx.execute(query)
                result_query = result_query.fetchall()

                print("result_query",result_query)
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, meter_id,mill_date from ems_v1_completed.power_{month_year}")
                tables_union_query = " UNION ALL ".join(tables_to_union)
                print("tables_union_query",tables_union_query)

        if len(tables_to_union) == 0:
            return _getErrorResponseJson("table not available")
        

        result = await year_wise_report_print(cnx,campus_id,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,meter_id,year,report_for,employee_id,kwh_type,request)
        await generate_year_wise_excel_report(result, year,next_year,report_for,shift_time)
            # process_data(month_year, result)
        file_path = os.path.join(base_path, f"YearWiseReport-{year}-{next_year}.xlsx")
        results = f"http://{request.headers['host']}/attachments/YearWiseReport-{year}-{next_year}.xlsx"
        
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "file_url": results
                }

        return response
    except Exception as e:
        return get_exception_response(e)

def year_wise_excel_report(meter_data, year,next_year,mill_date,year_record,report_type,res):
    try:
        print("res",res)
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  
        f_count = 0
        sheet = workbook.create_sheet(title="Energy Electrical Report")
        print(year_record)

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        style = {'font': Font(color='000000', size=11, name='Calibri'),
                'border': border,
                'alignment': Alignment(vertical='center',horizontal='center')}
            
        style1 = {'font': Font(bold=True,color='000000', size=12, name='Calibri'),
            'border': border,
            'alignment': Alignment(vertical='center',horizontal='center')}
        
        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=13, name='Calibri'),
            'border': border,
            'alignment': Alignment(vertical='center',horizontal='right',wrap_text=True)
        }
        style3 = {
            'fill': PatternFill(fill_type='solid', fgColor='f6c492'),
            'font': Font(bold=True, color='000000', size=13, name='Calibri'),
            'border': border,
            'alignment': Alignment(vertical='center',horizontal='center',wrap_text=True)
        }
        # script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        script_dir =  os.path.join(os.path.dirname(__file__), "..", "..","..","..", "..")

        meter_name_to_row_mapping = {}
        if len(res) == 0:
            cell = "O10"
            data = "No Data"
            sheet[cell] = data

            alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].alignment = alignment

            sheet.column_dimensions[cell[0]].width = len(data) + 2  

            font = Font(name='Calibri', size=25)
            sheet[cell].font = font
        else:
            file_name = "tnpl.png"
            image_path = os.path.join(script_dir, file_name)
            img = Image(image_path)
            img.anchor = "C3"  
            
            img_width_pixels, img_height_pixels = img.width, img.height
    
            desired_width_pixels = 165
            desired_height_pixels = 40
    
            width_scale = desired_width_pixels / img_width_pixels
            height_scale = desired_height_pixels / img_height_pixels

            img.width = int(img_width_pixels * width_scale)
            img.height = int(img_height_pixels * height_scale)
            # sheet.column_dimensions['C'].width = img.width / 10
            # sheet.row_dimensions[1].height = img.height / 2
            # sheet.row_dimensions[2].height = img.height / 2
            sheet.merge_cells('C3:D4')  
            sheet.add_image(img)
            sheet['C3'].alignment = style['alignment']

            cell = "C5"
            sheet[cell] = "Energy in Units"
            sheet[cell].border = border
            sheet.merge_cells('C5:D6')
            sheet['C5'].alignment = style['alignment']
            sheet['C5'].font = style1['font']
            row_number = 7

            start_year = int(year)
            end_year = start_year + 1
            year_str = str(start_year)
            end_str = str(end_year)
            current_year = mill_date[6:10]
            current_month = mill_date[3:5]
            cur_year = int(mill_date[6:10])
            cur_month = int(mill_date[3:5])
            cur_day= int(mill_date[:2])
            print("cur_day",cur_day)

            if current_year == year:
                    financial_year_months = [
                f"{month:02d}{year}" for month in range(4, cur_month)
            ]
            elif current_year == next_year:
                if cur_month <3:
                    financial_year_months = [
                    f"{month:02d}{year_str}" for month in range(4, 13)
                ] + [
                    f"{month:02d}{end_str}" for month in range(1, cur_month)
                ]
                else:
                    financial_year_months = [
                        f"{month:02d}{year_str}" for month in range(4, 13)
                    ] + [
                        f"{month:02d}{end_str}" for month in range(1, 3)
                    ]
            else:
                financial_year_months = [
                        f"{month:02d}{year_str}" for month in range(4, 13)
                    ] + [
                        f"{month:02d}{end_str}" for month in range(1, 3)
                    ]
                
            print("Financial Year Months:", financial_year_months)

            row_number = 6  
            col_number = 5  
            col_number1 = 3  
            row_number1 = 7 
            row_number2 = 7 
            
            # if meter_data == []:
            #     cell = "E10"
            #     data = "No Data"
            #     sheet[cell] = data
            #     font = Font(bold=True, name='Calibri', size=13)
            #     sheet[cell].font = font
            
            for i in financial_year_months:
                cell = sheet.cell(row=row_number, column=col_number, value=i)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, name='Calibri', size=12)
                column_letter = get_column_letter(col_number)
                sheet.column_dimensions[column_letter].width = 15
                col_number += 1
        
            for meter_name, date_values in meter_data["year_record"].items():
                sheet.cell(row=row_number1, column=3, value=meter_name)
                sheet.cell(row=row_number1, column=3).font = style['font']
                sheet.merge_cells(
                    start_row=row_number1, start_column=3, end_row=row_number1, end_column=3 + 1)
                meter_name_to_row_mapping[meter_name] = row_number1 
                row_number1 += 1
                col_number1 = 5  

                meter_name_length = len(meter_name) + 10
                column_letter_c = get_column_letter(3) 
                column_width_c = max(meter_name_length, sheet.column_dimensions[column_letter_c].width)
                sheet.column_dimensions[column_letter_c].width = column_width_c

                for i in financial_year_months:
                    if i in date_values:        
                        cell_value = date_values[i]["formulas"]
                        roundoff_value_month = date_values[i].get("roundoff_value_month", 0) 
                        if roundoff_value_month == 0: # Access roundoff_value_month roundoff_value_month
                            sheet.cell(row=row_number2, column=col_number1, value=f"=({cell_value})")
                            cell = sheet.cell(row=row_number2, column=col_number1, value=f"=MROUND(({cell_value}),{roundoff_value_month})")
                        else:
                            sheet.cell(row=row_number2, column=col_number1, value=f"=MROUND(({cell_value}),{roundoff_value_month})")
                            cell = sheet.cell(row=row_number2, column=col_number1, value=f"=MROUND(({cell_value}),{roundoff_value_month})")
                        cell.number_format = '0.00'
                        sheet.row_dimensions[1].height = 100
                        sheet.column_dimensions["AW"].width = 20
                
                    col_number1 += 1
                row_number2 +=1
        
            rows = row_number2+1
            rowno = row_number2+2
            month_mapping = {
                "01": "Jan",
                "02": "Feb",
                "03": "Mar",
                "04": "Apr",
                "05": "May",
                "06": "June",
                "07": "July",
                "08": "Aug",
                "09": "Sept",
                "10": "Oct",
                "11": "Nov",
                "12": "Dec"
            }
            
            updated_financial_year_months = [
                f"{month_mapping[month[:2]]}'{month[4:]+'(AVG)'}" for month in financial_year_months
            ]

            row = 6
            col = 5
            for i in updated_financial_year_months:
                cell = sheet.cell(row=row, column=col, value=i)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, name='Calibri', size=12)
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 17
                col += 1
            
            month_year = mill_date[3:]
            output_keys = [f'{day}-{month_year}' for day in range(1, cur_day+1)]
            rows= 6
            column=col_number1

            for day_date in output_keys:
                cell = sheet.cell(row=rows, column=column, value=day_date)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, name='Calibri', size=12)
                column_letter = get_column_letter(column)
                sheet.column_dimensions[column_letter].width = 17
                column += 1

            cell = sheet.cell(row=6, column=column, value="Monthly")

            cell.alignment = style['alignment']
            cell.font = Font(bold=True, name='Calibri', size=12)
            cols=column
            cell = sheet.cell(row=6, column=column+1, value=f"FY  {year}-{next_year}")
            cell.alignment = style['alignment']
            cell.font = Font(bold=True, name='Calibri', size=12)
            
            date_to_col = {date: col for col, date in enumerate(output_keys, start=col_number1)}
            
            # for meter_name, date_values in meter_data["current_month_record"].items():
            #     meter_row = meter_name_to_row_mapping[meter_name]
            #     row_formula = []  # To store SUM formula parts for each date

            #     for day_date, value in date_values.items():
            #         formula = value['formula']
            #         roundoff_value_day = value.roundoff_value_day
            #         # roundoff_value_day = date_values[next(iter(date_values))]["roundoff_value_day"]
            #         formatted_date = f"{int(day_date[:2])}-{day_date[3:]}"
            #         col_idx = date_to_col.get(formatted_date, None)

            #         if col_idx is not None:
                        
            #             sheet.cell(row=meter_row, column=col_idx, value=f"=MROUND(({formula}),{roundoff_value_day})")
            #             cell = sheet.cell(row=meter_row, column=col_idx, value=f"=MROUND(({formula}),{roundoff_value_day})")
            #             cell.number_format = '0.00'
            #             row_formula.append(f'{sheet.cell(row=meter_row, column=col_idx).coordinate}')

            #     if row_formula:
            #         sum_formula = '=SUM(' + ':'.join(row_formula) + ')'
            #         sheet.cell(row=meter_row, column=column).value = sum_formula
            #     rows += 1 
            for meter_name, date_values in meter_data["current_month_record"].items():
                meter_row = meter_name_to_row_mapping[meter_name]
                row_formula = [] 

                for day_date, value in date_values.items():
                    formula = value["formula"]
                    roundoff_value_day = value["roundoff_value_day"]  

                    formatted_date = f"{int(day_date[:2])}-{day_date[3:]}"
                    col_idx = date_to_col.get(formatted_date, None)

                    if col_idx is not None:
                        if roundoff_value_day == 0:
                            formula_with_mround = f"=({formula})"
                        else:
                            formula_with_mround = f"=MROUND(({formula}),{roundoff_value_day})"
                        sheet.cell(row=meter_row, column=col_idx, value=formula_with_mround)
                        cell = sheet.cell(row=meter_row, column=col_idx)
                        cell.number_format = '0.00'

                        row_formula.append(f'{cell.coordinate}')

                if row_formula:
                    sum_formula = '=SUM(' + ':'.join(row_formula) + ')'
                    sheet.cell(row=meter_row, column=column).value = sum_formula

                rows += 1

            row = 7
            # for meter_name, value in year_record.items():# add year record 
            #     cell = sheet.cell(row=row, column=column+1)
            #     cell.value = f"=MROUND(({value}),{roundoff_value_year})"
            #     cell.number_format = '0.00'
            #     cell.alignment = Alignment(horizontal='center', vertical='center')
            #     cell.font = Font(name='Calibri', size=11)
            #     column_letter = get_column_letter(column+1)
            #     sheet.column_dimensions[column_letter].width = 19
            #     row += 1
            for func_name, data in year_record.items():
                formula_y = data['formula_y']
                roundoff_value_year = data['roundoff_value_year']
                cell = sheet.cell(row=row, column=column+1)
                if roundoff_value_year == 0:
                    cell.value = f"=({formula_y})"            
                else:
                    cell.value = f"=MROUND({formula_y}, {roundoff_value_year})"            
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calibri', size=11)
                column_letter = get_column_letter(column+1)
                sheet.column_dimensions[column_letter].width = 19
                row += 1
            if report_type == '6to6':
                report = '06:00 to 06:00'
            else:
                report = 'Day'
            columns = 3
            cell_value = f"Daily Energy Electrical  Report-({report})"
            sheet.cell(row=4, column=columns+2).value = cell_value
            sheet.cell(row=4, column=columns+2).alignment = style2['alignment']
            sheet.merge_cells(start_row=4, start_column=5, end_row=4, end_column=column+1)
            sheet.alignment = style1['alignment']
            sheet.cell(row=4, column=columns+2).font = style2['font']

            columns = 3
            cell_value = " TNPL Unit - II   - Energy - Electrical"
            sheet.cell(row=3, column=columns+2).value = cell_value
            sheet.cell(row=3, column=columns+2).alignment = style2['alignment']
            sheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=column+1)
            sheet.cell(row=3, column=columns+2).font = style2['font']

            sheet.cell(row=5, column=column+1).value = sheet.cell(row=6, column=column-1).value
            alignment = Alignment(horizontal='center', vertical='center')
            # fill =   PatternFill(fill_type='solid', fgColor='f6c492')
            font = Font(bold = True,size=12)  

            sheet.cell(row=5, column=column + 1).alignment = alignment
            sheet.cell(row=5, column=column + 1).font = font
            column = column+2
            merge_range = f"C{row_number2}:{openpyxl.utils.get_column_letter(column-1)}{row_number2}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=row_number2, column=3)
            sheet.row_dimensions[row_number2].height = 25

            merged_cell.value = "REMARKS: "
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')

            column = column+2
            merge_range = f"C{rowno}:D{rowno}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=rowno, column=3)
            sheet.row_dimensions[rows].height = 25

            merged_cell.value = "Shift Incharge-(Elect)"
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            merge_range = f"E{rowno}:{openpyxl.utils.get_column_letter(column-5)}{rowno}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=rowno, column=5)
            sheet.row_dimensions[rowno].height = 25

            merged_cell.value = "Sectional Incharge-(Elect)"
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')

            merge_range = f"{openpyxl.utils.get_column_letter(column-4)}{rowno}:{openpyxl.utils.get_column_letter(column-3)}{rowno}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=rowno, column=column-4)
            sheet.row_dimensions[rowno].height = 25
            merged_cell.value = "HOD"
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            rows_count = 3																																										
            row_range = sheet.iter_rows(min_row=rows_count, max_row=rowno, min_col=3, max_col=column-3)
            for row in row_range:
                for cell in row:
                    cell.border = border

        file_name = f'YearReport-{year}-{next_year}.xlsx'
        file_path = os.path.join(base_path, file_name)
        workbook.save(file_path)
        
    except Exception as e:
        return get_exception_response(e)

@router.post("/year_report/", tags=["Report"])
async def year_report(request: Request,
                      year: str = Form(""),
                      report_type: str = Form(""),
                      cnx: AsyncSession = Depends(get_db)):

    try: 
        if year == "" :
            return _getErrorResponseJson("year is required...")
        
        if report_type == "" :
            return _getErrorResponseJson("report_type is required...")
        
        result = await year_report_print(cnx,year,report_type,request)
        next_year = int(year)+1
        response={}
        response = {
                "meter_data": result["meter_data"],
                "dict3": result["dict3"],
                "res_q4": result["res_q4"],
                "year": result["year"],
                "next_year": result["next_year"],
                "mill_date": result["mill_date"],
                "report_type":result["report_type"]
            }
        await year_wise_excel_report(response["meter_data"], response["year"],response["next_year"],response["mill_date"],response["dict3"],response["report_type"],response["res_q4"])

        
        file_path = os.path.join(base_path, f"YearReport-{year}-{next_year}.xlsx")
        results = f"http://{request.headers['host']}/attachments/YearReport-{year}-{next_year}.xlsx"

        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "file_url": results
                }

        return response
    except Exception as e:
        return get_exception_response(e)

@router.post("/current_power_ie/", tags=["Report"])
async def current_power_ie(meter_id : str = Form (''),
                           company_id: str = Form(''),
                           bu_id: str = Form(''),
                           period_id: str = Form(''),
                           from_date: str = Form(''),
                           to_date: str = Form(''),                      
                           shift_id: int = Form(''),
                           report_for : str = Form(''),           
                           employee_id : str = Form(''),           
                           cnx: AsyncSession = Depends(get_db)):

    try:
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}       
         
        if period_id == "":
            return _getErrorResponseJson(" period_id is required")
        
        if bu_id == "":
            return _getErrorResponseJson(" bu_id is required")
        
        if period_id == 'sel_shift' or period_id == 'sel_date':
            if from_date == '':
                return _getErrorResponseJson("date is required") 
            
            if period_id == 'sel_shift':
                if shift_id == '':
                    return _getErrorResponseJson("shift is required") 
                
            from_date=await parse_date(from_date)             
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            res = await check_power_table(cnx,month_year)
            if len(res) == 0:
                return _getErrorResponseJson("power table not available...")    

        if period_id == "from_to":            
            if from_date == '':
                return _getErrorResponseJson("from date is required")
            if to_date == '':
                return _getErrorResponseJson("to_date is required")  
            
            from_date = await parse_date(from_date)
            to_date = await parse_date(to_date)  

            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            month_year_range = [
                        (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
            union_queries_export = []
            for month_year in month_year_range:
                res = await check_power_table(cnx,month_year)
                if len(res) > 0:
                        table_name = f"ems_v1_completed.power_{month_year}"
                        union_queries_export.append(f"{table_name}")
            if len(union_queries_export) == 0:
                return _getErrorResponseJson("power table not available...")   
                
        result = await import_export_dtl(cnx,company_id,bu_id,meter_id,period_id,from_date,to_date,shift_id,report_for,employee_id)
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": result
        }

        return response
    
    except Exception as e:
        return get_exception_response(e)
       
@router.post("/holiday_report_print/", tags=["Report"])
async def holiday_report_print(company_id :str = Form(""),
                               bu_id :str = Form(""),
                               plant_id :str = Form(""),
                               plant_department_id :str = Form(""),
                               equipment_group_id :str = Form(""),
                               meter_id :str = Form(""),
                               equipment_id :str = Form(""),
                               holiday_type :str = Form(""),
                               holiday_year :int = Form(""),
                               group_by : str = Form(""),
                               limit_report_for:str = Form(''),
                               limit_exception_for:str = Form(''),
                               limit_order_by : str = Form(''),
                               limit_operation_value : str = Form(''),
                               is_critical : str = Form(''),
                               is_month_wise : str = Form(''),
                               cnx: AsyncSession = Depends(get_db)):

    try: 
        if holiday_year == "" :
            return _getErrorResponseJson("holiday_year is required...")
        
        if group_by == "" :
            return _getErrorResponseJson("group_by is required...")
        
        if group_by not in ['plant', 'plant_department','equipment_group','function','meter','equipment']:
            return _getErrorResponseJson("invalid group by ")
        
        
        data = await holiday_report(cnx,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,meter_id,equipment_id,holiday_type,holiday_year,group_by,limit_report_for,limit_exception_for,limit_order_by,limit_operation_value,is_critical,is_month_wise)
        if group_by == 'equipment' and len(data)>0:
            results = []
            group_by = "meter"
            datas  = await holiday_report(cnx,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,meter_id,equipment_id,holiday_type,holiday_year,group_by,limit_report_for,limit_exception_for,limit_order_by,limit_operation_value,is_critical,is_month_wise)
            
            dict ={}
            dict_end_kwh = {}
            dict_start_kwh = {}
            meter_id_dict = {}

            if len(datas)>0:
                for row in datas:
                    dict[row['meter_id']] = row['kWh']
                    dict_end_kwh[row["meter_id"]] = row["end_kwh"]
                    dict_start_kwh[row["meter_id"]] = row["start_kwh"]
                    meter_id_dict[row['meter_id']] = row['meter_name']

                result = await get_equipment_cal_dtl(cnx)
                equipment_id = 0
                formula = ''
                equipment_dict = {}

                if len(result)>0:
                    for rows in result:
                        equipment_id = rows['equipment_id']
                        formula = rows['formula2']
                        formula1 = rows['formula1']
                        
                        numbers = re.findall(r'\[(\d+)\]', formula)
                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict]
                        numeric_formula = formula
                        print(numeric_formula)

                        for meter_id in valid_ids:
                            numeric_value = dict.get(meter_id, 0)  # Get the value from dict2 or use 0 if not found
                            numeric_formula = numeric_formula.replace(f'[{meter_id}]', str(numeric_value))
                            formula_tooltip_kWh = {meter_id_dict[meter_id]: dict[meter_id] for meter_id in valid_ids}

                        kwh_formula = re.sub(r'dict', '', numeric_formula)
                        print(kwh_formula)
                        try:
                            kwh = eval(kwh_formula)
                        except (TypeError, ZeroDivisionError) as e:
                            print(f"An error occurred: {e}")
                            kwh = '' 
                            
                        numbers = re.findall(r'\[(\d+)\]', formula)
                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict_end_kwh]
                        numeric_formula = formula

                        for meter_id in valid_ids:
                            numeric_value = dict_end_kwh.get(meter_id, 0)  # Get the value from dict2 or use 0 if not found
                            numeric_formula = numeric_formula.replace(f'[{meter_id}]', str(numeric_value))
                            formula_tooltip_end_kwh = {meter_id_dict[meter_id]: dict_end_kwh[meter_id] for meter_id in valid_ids}
                
                        end_kwh_formula = re.sub(r'dict', '', numeric_formula)
                        try:
                            end_kwh = eval(end_kwh_formula)
                        except (TypeError, ZeroDivisionError) as e:
                            
                            print(f"An error occurred: {e}")
                            end_kwh = '' 

                        numbers = re.findall(r'\[(\d+)\]', formula)
                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict_start_kwh]
                        numeric_formula = formula

                        for meter_id in valid_ids:
                            numeric_value = dict_start_kwh.get(meter_id, 0)  # Get the value from dict2 or use 0 if not found
                            numeric_formula = numeric_formula.replace(f'[{meter_id}]', str(numeric_value))
                            formula_tooltip_start_kwh = {meter_id_dict[meter_id]: dict_start_kwh[meter_id] for meter_id in valid_ids}

                        start_kwh_formula = re.sub(r'dict', '', numeric_formula)
                        try:
                            start_kwh = eval(start_kwh_formula)
                        except (TypeError, ZeroDivisionError) as e:
                            
                            print(f"An error occurred: {e}")
                            start_kwh = '' 

                        results.append({"equipment_id": equipment_id,"kWh":kwh,"end_kwh":end_kwh,"start_kwh":start_kwh,"tooltip_start_kwh":formula_tooltip_start_kwh,"tooltip_end_kwh":formula_tooltip_end_kwh,"tooltip_kwh":formula_tooltip_kWh,"formula":formula1})
                        equipment_dict = {item["equipment_id"]: item for item in results}
                        
                        result = []
                        for row in data:
                            equipment_id = row["equipment_id"]
                            equipment_data = equipment_dict.get(equipment_id, {})
                            
                            updated_row = {
                                "company_code" :row["company_code"],
                                "company_name" :row["company_name"],
                                "bu_code" :row["bu_code"],
                                "bu_name" :row["bu_name"],
                                "plant_code" :row["plant_code"],
                                "plant_name" :row["plant_name"],
                                "plant_department_code" :row["plant_department_code"],
                                "plant_department_name" :row["plant_department_name"],
                                "equipment_group_code" :row["equipment_group_code"],
                                "equipment_group_name" :row["equipment_group_name"],
                                "equipment_code" :row["equipment_code"],
                                "equipment_name" :row["equipment_name"],
                                "company_id" :row["company_id"],
                                "bu_id" :row["bu_id"],
                                "plant_id" :row["plant_id"],
                                "plant_department_id" :row["plant_department_id"],
                                "equipment_group_id" :row["equipment_group_id"],
                                "equipment_id" :row["equipment_id"],
                                "meter_id" :row["meter_id"],
                                "month" :row["month"],
                                "mill_date" :row["mill_date"],
                                "mill_shift" :row["mill_shift"],
                                "kWh": equipment_data.get("kWh", row["kWh"]),
                                "start_kwh": equipment_data.get("start_kwh", row["start_kwh"]),
                                "end_kwh": equipment_data.get("end_kwh", row["end_kwh"]),
                                "formula": equipment_data.get("formula", row["formula"]),
                                "tooltip_start_kwh": equipment_data.get("tooltip_start_kwh", row["tooltip_start_kwh"]),
                                "tooltip_end_kwh": equipment_data.get("tooltip_end_kwh", row["tooltip_end_kwh"]),
                                "tooltip_kwh": equipment_data.get("tooltip_kwh", row["tooltip_kwh"]) 
                                }
                            result.append(updated_row)
                            data = result
           
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": data
        }

        return response
    
    except Exception as e:
        return get_exception_response(e) 
    
@router.post("/alarm_report/", tags=["Report"])
async def alarm_report(request :Request,
                       company_id : int = Form(''),
                       bu_id : int = Form(''),
                       plant_id : str = Form (''),
                       plant_department_id : str = Form (''),
                       equipment_group_id : str = Form (''),
                       meter_id : str = Form (''),
                       report_for : str = Form(''),
                       period_id: str = Form(''),
                       from_date: str = Form(''),
                       to_date: str = Form(''),                      
                       from_year: int = Form(''),                      
                       shift_id: int = Form(''),  
                       employee_id : int = Form(''),                                         
                       cnx: AsyncSession = Depends(get_db)):
    try:
        
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        if period_id == "" :
            return _getErrorResponseJson("period_id is required...")
        
        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                no_of_shifts = shift_record["no_of_shifts"] 
        
        if period_id == "sel_date" or period_id == 'sel_shift':            
            if from_date == '':
                return _getErrorResponseJson("from date is required") 
            
            from_date =  await parse_date(from_date) 
            print("from_date",from_date)  
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
            res = await check_alarm_tble(cnx,month_year)
            if len(res) == 0:
                return _getErrorResponseJson("alarm table not available...")   
            
            if period_id == "sel_shift":                  
                if shift_id == '':
                    return _getErrorResponseJson("shift_id is required") 
                
        if period_id == "#previous_shift" or period_id == "#previous_day":  
            if period_id == "#previous_shift":               
                if mill_shift == 1:
                    shift_id = no_of_shifts
                    from_date = await parse_date(mill_date) - timedelta(days=1)
                else:
                    shift_id = int(mill_shift) - 1
                    from_date = mill_date 
            if period_id == "#previous_day":             
                from_date = mill_date - timedelta(days=1)
            
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            res = await check_alarm_tble(cnx,month_year)
            if len(res) == 0:
                return _getErrorResponseJson("alarm table not available...")   
                   
        if period_id == "from_to":            
            if from_date == '':
                return _getErrorResponseJson("from date is required")
            if to_date == '':
                return _getErrorResponseJson("to_date is required")  
            from_date = await parse_date(from_date)
            to_date = await parse_date(to_date)
          
        if period_id == '#previous_week' or period_id == "#this_week" or period_id == "#this_month" or period_id == '#previous_month' or period_id=="#previous_year" or period_id=="#this_year" or period_id=="from_to" or period_id == "#sel_year":
            if period_id  == "#this_week":
                dt = mill_date
                from_date=dt-timedelta(dt.weekday()+1)
                to_date = mill_date

            if period_id == "#previous_week":
                dt = mill_date
                current_week_start = dt - timedelta(days=dt.weekday())  
                from_date = current_week_start - timedelta(weeks=1)  
                to_date = from_date + timedelta(days=5)

            if period_id == "#this_month":
                from_date = mill_date.replace(day=1)
                to_date = mill_date

            if period_id == "#previous_month":
                from_date = mill_date.replace(day=1)                   
                from_date = (from_date - timedelta(days=1)).replace(day=1)
                to_date = from_date + timedelta(days=30)
                print("to_date....................",to_date)

            if period_id=="#this_year": 
        
                from_date = mill_date.replace(day=1,month=1) 
                to_date = mill_date  
                print(from_date)
                print(to_date)
            if period_id=="#previous_year": 
                from_date = mill_date.replace(day=1, month=1, year=mill_date.year - 1)
                to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
        
            if period_id == "#sel_year": 
                if from_year == '':
                    return _getErrorResponseJson("from year is required")
                    
                from_date = mill_date.replace(day=1, month=1, year=from_year)
                print(from_date)
                to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
                print(to_date)
        
            if from_date != '' and to_date != '':
                month_year_range = [
                    (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                    for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                ]
                union_queries = []
                joins = []
                for month_year in month_year_range:
                    res = await check_alarm_tble(cnx,month_year)
                    
                    if len(res)>0:
                        table_name = f"ems_v1_completed.alarm_{month_year}"
                        union_queries.append(f"{table_name}")
                if len(union_queries) == 0:
                    return _getErrorResponseJson("Alarm Table Not Available...")     
                       
            
        result = await alarmreport(cnx,request,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,meter_id,report_for,period_id,from_date,to_date,shift_id, employee_id  )
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": result["data"],
            "data1": result["data1"]

        }
        return response
    except Exception as e:
        return get_exception_response(e) 
    
def generate_hour_report(result, from_date, parameter, shift_time,start_time):
    wb = Workbook()
    ws = wb.active
    ws.title = 'EMS' 
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')  
    border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
  
    if result["hourly_data"] == []:
        print(1)
        cell = "O10"
        data = "No Data"
        ws[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment

        ws.column_dimensions[cell[0]].width = len(data) + 2  

        font = Font(name='Calibri', size=25)
        ws[cell].font = font
    else:
        cell = "D2"
        data = f"Hour WISE ENERGY CONSUMPTION REPORT FOR KWH - {from_date} ({shift_time} to {shift_time})"
        ws[cell] = data
        font = Font(bold=True, name='Calibri', size=15)
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].font = font
        ws[cell].alignment = alignment 
        ws[cell].fill = fill_cyan
        first_hour = 0

        if start_time >0:
            first_hour = start_time
            last_hour = first_hour - 1
        
        elif start_time == 0:
            first_hour = start_time
            last_hour = 23

        print("start_time",start_time)
        if parameter == 'Energy':
            ws.merge_cells('D2:BX3')
            ws.merge_cells('B4:B5')
            cell = "B4"
            ws[cell] = "Meter Tag"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15 
        
            ws.merge_cells('C4:C5')
            cell = "C4" 
            ws[cell] = "Meter Name"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 60

            ws.merge_cells('D4:D5')
            cell = "D4" 
            ws[cell] = "Meter Type"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15

            ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
            ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
            ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
            for i in range(24):
                start_hour = (first_hour + i) % 24
                end_hour = (start_hour + 1) % 24

                # Write the hour range header
                header_text = f"{start_hour:02}:00 to {end_hour:02}:00"
                ws.merge_cells(start_row=4, start_column=5 + i * 3, end_row=4, end_column=7 + i * 3)
                header_cell = ws.cell(row=4, column=5 + i * 3)
                header_cell.value = header_text

                # Apply styles to the header cell
                header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                header_cell.font = Font(bold=True, name='Calibri', size=12)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Write secondary headers for machine_kwh, master_kwh, and kwh
                secondary_headers = ["Start CkWh", "End CkWh", "kwh"]
                for j, header in enumerate(secondary_headers):
                    secondary_header_cell = ws.cell(row=5, column=5+i*3+j)
                    secondary_header_cell.value = header

                    # Apply styles to the secondary header cell
                    secondary_header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                    secondary_header_cell.font = Font(bold=True, name='Calibri', size=12)
                    secondary_header_cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.column_dimensions[secondary_header_cell.column_letter].width = len(header) + 15


            end_column = 7+i*3
            
            for meter_data in result["hourly_data"]:
                meter_name = meter_data["meter_name"]
                meter_code = meter_data["meter_code"]
                meter_type = meter_data["meter_type"]
                plant_name = meter_data["plant_name"]
                row_data = ['',meter_code, meter_name, meter_type]
                ws.merge_cells('B2:C2')
                ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
                ws.cell(row=2, column=2, value=f"Plant : {plant_name}").fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=2, column=2).font = font
            
                ws.merge_cells('B3:C3')
                ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
                ws.cell(row=3, column=2, value=f"Parameter : {parameter}").fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=3, column=2).font = font

                for i in range(24):
                    start_hour = (first_hour + i) % 24
                    end_hour = (start_hour + 1) % 24
                    hour = f"h{start_hour}"
                    kwh = meter_data.get(hour, "")
                    machine_kwh = meter_data.get(f"machine_kwh_{hour}", "")
                    master_kwh = meter_data.get(f"master_kwh_{hour}", "")
                    row_data.extend([master_kwh, machine_kwh, kwh])

            
                ws.append(row_data)
            end_row = ws.max_row
            row_range = ws[f"B2:BX{end_row}"]
            for row in row_range:
                for cell in row:
                    cell.border = border

        else:
            ws.merge_cells('D2:AB3')
            cell = "B4"
            ws[cell] = "Meter Tag"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15 
        
            cell = "C4" 
            ws[cell] = "Meter Name"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 60

            cell = "D4" 
            ws[cell] = "Meter Type"
            alignment = Alignment(horizontal="center", vertical="center")
            font = Font(bold=True, name='Calibri', size=12)
            ws[cell].font = font
            ws[cell].alignment = alignment 
            ws.column_dimensions[cell[0]].width = len("%") + 15

            ws.cell(row=4, column=2, value="Meter Tag").fill = fill_cyan
            ws.cell(row=4, column=3, value="Meter Name").fill = fill_cyan
            ws.cell(row=4, column=4, value="Meter Type").fill = fill_cyan
            for i in range(24):
                start_hour = (first_hour + i) % 24
                end_hour = (start_hour + 1) % 24

                header_text = f"{start_hour:02}:00 to {end_hour:02}:00"
                ws.merge_cells(start_row=4, start_column=5 + i, end_row=4, end_column=5 + i)
                header_cell = ws.cell(row=4, column=5 + i)
                header_cell.value = header_text
                header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
                header_cell.font = Font(bold=True, name='Calibri', size=12)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                column_width = max(len(header_text), 15)  

                # Set the width of the column
                ws.column_dimensions[ws.cell(row=4, column=5 + i).column_letter].width = column_width


            end_column = 7 + i  # Assuming you want to track the last column

            for meter_data in result["hourly_data"]:
                meter_name = meter_data["meter_name"]
                meter_code = meter_data["meter_code"]
                meter_type = meter_data["meter_type"]
                plant_name = meter_data["plant_name"]
                row_data = ['', meter_code, meter_name, meter_type]
                ws.merge_cells('B2:C2')
                ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
                ws.cell(row=2, column=2).fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=2, column=2).font = font

                ws.merge_cells('B3:C3')
                ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
                ws.cell(row=3, column=2).fill = fill_cyan
                font = Font(bold=True, name='Calibri', size=12)
                ws.cell(row=3, column=2).font = font

                for i in range(24):
                    start_hour = (first_hour + i) % 24
                    end_hour = (start_hour + 1) % 24
                    hour = f"h{start_hour}"
                    kwh = meter_data.get(hour, "")
                    
                    row_data.extend([kwh])
                ws.append(row_data)

            end_row = ws.max_row
            row_range = ws[f"B2:AB{end_row}"]
            for row in row_range:
                for cell in row:
                    cell.border = border

        # else:
        #     ws.merge_cells('D2:BX3')
        #     for i in range(24):
        #         start_hour = (first_hour + i) % 24
        #         end_hour = (start_hour + 1) % 24

        #         # Write the hour range header
        #         header_text = f"{start_hour:02}:00 to {end_hour:02}:00"
        #         ws.merge_cells(start_row=4, start_column=5 + i * 3, end_row=4, end_column=7 + i * 3)
        #         header_cell = ws.cell(row=4, column=5 + i * 3)
        #         header_cell.value = header_text

        #         # Apply styles to the header cell
        #         header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
        #         header_cell.font = Font(bold=True, name='Calibri', size=12)
        #         header_cell.alignment = Alignment(horizontal="center", vertical="center")

        #         # Write secondary headers for machine_kwh, master_kwh, and kwh
        #         secondary_headers = ["Start CkWh", "End CkWh", "kwh"]
        #         for j, header in enumerate(secondary_headers):
        #             secondary_header_cell = ws.cell(row=5, column=5+i*3+j)
        #             secondary_header_cell.value = header

        #             # Apply styles to the secondary header cell
        #             secondary_header_cell.fill = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
        #             secondary_header_cell.font = Font(bold=True, name='Calibri', size=12)
        #             secondary_header_cell.alignment = Alignment(horizontal="center", vertical="center")
        #             ws.column_dimensions[secondary_header_cell.column_letter].width = len(header) + 15


        #     end_column = 7+i*3
            
        #     for meter_data in result["hourly_data"]:
        #         meter_name = meter_data["meter_name"]
        #         meter_code = meter_data["meter_code"]
        #         meter_type = meter_data["meter_type"]
        #         plant_name = meter_data["plant_name"]
        #         row_data = ['',meter_code, meter_name, meter_type]
        #         ws.merge_cells('B2:C2')
        #         ws.cell(row=2, column=2, value=f"Plant : {plant_name}").alignment = Alignment(horizontal="left")
        #         ws.cell(row=2, column=2, value=f"Plant : {plant_name}").fill = fill_cyan
        #         font = Font(bold=True, name='Calibri', size=12)
        #         ws.cell(row=2, column=2).font = font
            
        #         ws.merge_cells('B3:C3')
        #         ws.cell(row=3, column=2, value=f"Parameter : {parameter}").alignment = Alignment(horizontal="left")
        #         ws.cell(row=3, column=2, value=f"Parameter : {parameter}").fill = fill_cyan
        #         font = Font(bold=True, name='Calibri', size=12)
        #         ws.cell(row=3, column=2).font = font

        #         for i in range(24):
        #             start_hour = (first_hour + i) % 24
        #             end_hour = (start_hour + 1) % 24
        #             hour = f"h{start_hour}"
        #             kwh = meter_data.get(hour, "")
        #             machine_kwh = meter_data.get(f"machine_kwh_{hour}", "")
        #             master_kwh = meter_data.get(f"master_kwh_{hour}", "")
        #             row_data.extend([machine_kwh, master_kwh, kwh])

            
        #         ws.append(row_data)
        #     end_row = ws.max_row
        #     row_range = ws[f"B2:BX{end_row}"]
        #     for row in row_range:
        #         for cell in row:
        #             cell.border = border

    file_name = f"HourWiseReport-{from_date}.xlsx"
    file_path = os.path.join(static_dir, file_name)
    wb.save(file_path)

@router.post("/get_hour_wise_analysis_report/", tags=["Report"])
async def get_hour_wise_analysis_report(request : Request,
                                        campus_id : int = Form(''),
                                        company_id : int = Form(''),
                                        bu_id : int = Form(''),
                                        plant_id : int = Form(''),
                                        plant_department_id : int = Form(''),
                                        equipment_group_id : int = Form(''),
                                        meter_id: str = Form(''),
                                        from_date: str = Form(''),                                                 
                                        meter_type: str = Form(''),                                                      
                                        kwh: str = Form(''),                                                      
                                        employee_id: str = Form(''),  
                                        from_time : str = Form('') ,
                                        to_time : str = Form(""),
                                        cnx: AsyncSession = Depends(get_db)):
    try:
        if from_date == "" :
            return _getErrorResponseJson("from_date is Required...")
        
        if kwh == "" :
            return _getErrorResponseJson("Kwh is Required...")
        
        if campus_id == "" :
            return _getErrorResponseJson("Campus is Required...")
        
        if plant_id == "" :
            return _getErrorResponseJson("Plant is Required...")
        
        mill_date = from_date
        from_date = await parse_date(from_date)

        month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
        res = await check_power_table(cnx,month_year)

        if len(res)==0:
            return _getErrorResponseJson("Table not available...")
        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        shift_time = ''
        if len(data1) > 0:
            for shift_record in data1:
                shift_time = shift_record["a_shift_start_time"]
        print("start_time",shift_time)
        result = await hour_wise_analysis_report(cnx,campus_id,company_id,bu_id,plant_id,plant_department_id,equipment_group_id,employee_id,meter_id,from_date,meter_type,kwh,month_year,from_time,to_time,request,shift_time)
        from_date = from_date.date()
        print(from_date)
        generate_hour_report(result, mill_date,kwh,shift_time,result["start_time"])
       
        
        file_path = os.path.join(static_dir, f"HourWiseReport-{mill_date}.xlsx")
        
        results = f"http://{request.headers['host']}/attachments/HourWiseReport-{mill_date}.xlsx"
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "file_url": results
                }
        return response
    except Exception as e:
        return get_exception_response(e)
 

@router.post("/get_hour_wise_report/", tags=["Report"])
async def get_hour_wise_report(meter_id : str = Form (''),
                               plant_id: int = Form(''),
                               period_id: str = Form(''),
                               from_date: str = Form(''),                   
                               to_date: str = Form(''),                   
                               shift_id: str = Form(''),
                               from_time: str = Form(''),
                               to_time: str = Form(''),
                               cnx: AsyncSession = Depends(get_db)):
    
    try:
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}    
        if period_id == '':
            return _getErrorResponseJson("period_id is required") 
        
        if meter_id == '':
            return _getErrorResponseJson("meter_id is required") 
        
        if period_id == 'sel_shift' or period_id == 'sel_date':
            
            if from_date == '':
                return _getErrorResponseJson("date is required") 
            from_date = await parse_date(from_date)
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""  

            if period_id == 'sel_shift':
                    if shift_id == '':
                        return _getErrorResponseJson("shift is required")
            result = await check_analysis_table(cnx,month_year)
            if len(result) == 0:
                return _getErrorResponseJson("analysis table not available...")
            
        if period_id == "from_to":   
                     
            if from_date == '':
                return _getErrorResponseJson("from date is required")
            if to_date == '':
                return _getErrorResponseJson("to_date is required")
            from_date = await parse_date(from_date)
            to_date = await parse_date(to_date)

            union_queries = []
            month_year_range = [
                        (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
            for month_year in month_year_range:
                result = await check_analysis_table(cnx,month_year)
                if len(result) > 0:
                        table_name = f"ems_v1_completed.power_analysis_{month_year}"
                        union_queries.append(f"{table_name}")

            if len(union_queries) == 0: 
                return _getErrorResponseJson("analysis table not available...") 
            
        result = await get_hour_wise_report_model(cnx,plant_id,meter_id,period_id,from_date,to_date,shift_id,from_time,to_time)
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": result
        }
        return response
    except Exception as e:
        return get_exception_response(e) 

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from fastapi import Form, Request, Depends
from sqlalchemy.ext.asyncio import AsyncSession

@router.post("/get_gateway_log/", tags=["Report"])
async def get_gateway_log(request: Request,
                          mac: str = Form(""),
                          from_date: str = Form(""),
                          from_time: str = Form(""),
                          to_time: str = Form(""),
                          cnx: AsyncSession = Depends(get_db)):
    try:
        if from_date == "":
            return _getErrorResponseJson("from_date is required...")
        
        # from_date = await parse_date(from_date)
        result = await gateway_log(cnx, mac, from_date, from_time, to_time)

        # Create a new Excel workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        if not result:
            
            cell = "H6"
            data = "No Data"
            ws[cell] = data

            alignment = Alignment(horizontal="center", vertical="center")
            ws[cell].alignment = alignment

            ws.column_dimensions[cell[0]].width = len(data) + 2  

            font = Font(name='Calibri', size=25)
            ws[cell].font = font

        else:
            headers = ["ID", "Date&Time", "Mac", "Received Packet", "Error Code", "Slave ID","kWh"]
            ws.append(headers)
            
            ws.column_dimensions['A'].width = 5  
            ws.column_dimensions['B'].width = 20 
            ws.column_dimensions['C'].width = 20 
            ws.column_dimensions['D'].width = 100
            ws.column_dimensions['E'].width = 10 
            ws.column_dimensions['F'].width = 10 
            ws.column_dimensions['G'].width = 10 

            # Set heading alignment to center and heading color to yellow
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
                border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
                cell.border = border
          
            for index, row in enumerate(result, start=1):
                ws.append([index, row["date_time"], row["mac"], row["data"], row["is_error"], row["slave_id"],row["kwh"]])

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
                    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
        
       
                    cell.border = border

        # Save the workbook
        file_name = f'log-{from_date}.xlsx'
        file_path = os.path.join(static_dir, file_name)
        wb.save(file_path)

        results = f"http://{request.headers['host']}/attachments/{file_name}"
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": results
        }
        return response
    except Exception as e:
        return get_exception_response(e)
    
@router.post("/get_route_card_kwh/", tags=["Report"])
async def get_routecard_kwh(request: Request,
                            equipment_id: str = Form(""),
                            from_date: str = Form(""),
                            to_date: str = Form(""),
                            cnx: AsyncSession = Depends(get_db)):
    try:
        if from_date == '':
            createFolder("Log/",f"from_date is required")
        if to_date == '':
            createFolder("Log/",f"from_date is required")
            
        from_date = await  parse_date(from_date)
        to_date = await  parse_date(to_date)
        data  = await route_card_kwh(cnx,equipment_id,from_date,to_date,"equipment")
        # print(data)
        # results = []
        # datas = await route_card_kwh(cnx,equipment_id,from_date,to_date,"meter")
        # dict ={}
        # meter_id_dict = {}
        # formula_tooltip_kWh = ''
        # if len(datas)>0:
        #     for row in datas:
        #         dict[row['meter_id']] = row['kWh']
        #         meter_id_dict[row['meter_id']] = row['meter_name']
        #     result = await get_equipment_cal_dtl(cnx)
            
        #     formula = ''
        #     equipment_dict = {}
        #     if len(result)>0:
        #         for rows in result:
                    
        #             equipment_id = rows['equipment_id']
        #             formula = rows['formula2']                        
        #             formula1 = rows['formula1']
            
        #             numbers = re.findall(r'\[(\d+)\]', formula)
        #             valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict]
        #             numeric_formula = formula
        #             for meter_id in valid_ids:
        #                 numeric_value = dict.get(meter_id, 0)  # Get the value from dict2 or use 0 if not found
        #                 numeric_formula = numeric_formula.replace(f'[{meter_id}]', str(numeric_value))
        #                 formula_tooltip_kWh = {meter_id_dict[meter_id]: dict[meter_id] for meter_id in valid_ids}
                    
        #             numeric_formula = re.sub(r'dict\[\d+\]', '0.0', numeric_formula)
        #             kwh_formula = re.sub(r'dict', '', numeric_formula)
        #             try:
        #                 kwh = eval(kwh_formula)
        #                 createFolder("Log/",f"Equipment kwh -{kwh} Equipment Id -{equipment_id}")
        #             except Exception as e:
        #                 createFolder("Log/","unexpected kwh"+str(e))
        #                 kwh = 0
            
        #             results.append({"equipment_id": equipment_id,"kWh":kwh,"tooltip_kwh":formula_tooltip_kWh,"formula":formula1})
        #             equipment_dict = {item["equipment_id"]: item for item in results}
                    
        #             result = []
        #             for row in data:
        #                 print(data)
        #                 equipment_id = row["equipment_id"]
        #                 equipment_data = equipment_dict.get(equipment_id, {})
        #                 updated_row = {
        #                     "equipment_id" :row["equipment_id"],
        #                     "equipment_code" :row["equipment_code"],
        #                     "equipment_name" :row["equipment_name"],
        #                     "meter_code" :row["meter_code"],
        #                     "meter_name" :row["meter_name"],
        #                     "meter_id" :row["meter_id"],
        #                     "date_time" :row["date_time"],
        #                     "mill_date" :row["mill_date"],
        #                     "mill_shift" :row["mill_shift"],
        #                     "off_time":row["off_time"],
        #                     "idle_time":row["idle_time"],
        #                     "on_load_time":row["on_load_time"],
        #                     "on_load_kwh":row["on_load_kwh"],
        #                     "off_kwh":row["off_kwh"],
        #                     "idle_kwh":row["idle_kwh"],
        #                     "kWh": equipment_data.get("kWh", row["kWh"]),
        #                     "formula": equipment_data.get("formula", row["formula"]),
        #                     "tooltip_kwh":equipment_data.get("tooltip_kwh", row["tooltip_kwh"]) 
        #                     }
        #                 print(updated_row)
        #                 result.append(updated_row)
        #                 data = result
                        
        
        
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": data
        }
              
        return response
    except Exception as e:
        return get_exception_response(e)
    
@router.post("/get_demand_report/", tags=["Report"])
async def get_demand_report(request: Request,
                            period_id: str = Form(""),
                            campus_id: str = Form(""),
                            meter_id: str = Form(""),
                            from_date: str = Form(""),
                            to_date: str = Form(""),
                            shift_id: str = Form(""),
                            from_time: str = Form(""),
                            to_time: str = Form(""),
                            date_time: str = Form(""),
                            main_demand_meter: str = Form(""),
                            filter_type: str = Form(""),
                            parameter: str = Form(""),
                            duration: int = Form(""),
                            cnx: AsyncSession = Depends(get_db)):
    try:
        # if main_demand_meter != 'yes' and :
        #     if meter_id == '':
        #         return _getErrorResponseJson("Meter ID Required") 
        end_time = ''
        if period_id == "peak_time":
            
            if date_time == '':
                return _getErrorResponseJson("date time Is Required") 
            date_time = date_time[:10]+" "+date_time[11:]
            end_time = datetime.strptime(date_time, "%Y-%m-%d %H:%M:%S")
            end_time += timedelta(seconds=30)
            
            from_date =  await parse_date(date_time) 
        if period_id == '':
            return _getErrorResponseJson("Period Is Required") 
        
        if period_id == "sel_date" or period_id == 'sel_shift':            
            if from_date == '':
                return _getErrorResponseJson("From Date Is Required") 
            
            from_date =  await parse_date(from_date) 
            print("from_date",from_date)  
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
            res = await check_power_table(cnx,month_year)
            if len(res) == 0:
                return _getErrorResponseJson("Power Table Not Available...") 
              
            if period_id == "sel_shift":                  
                if shift_id == '':
                    return _getErrorResponseJson("Shift Id Is Required") 
                
        data1 = await shift_Lists(cnx, '','', '', '')
       
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                no_of_shifts = shift_record["no_of_shifts"]  
       
        if period_id == "#this_month":
            from_date = mill_date.replace(day=1)
            to_date = mill_date 

        if period_id == "from_to" or period_id == "#this_month":            
            if from_date == '':
                return _getErrorResponseJson("from date is required")
            if to_date == '':
                return _getErrorResponseJson("to_date is required") 
            from_date =  await parse_date(from_date) 
            to_date =  await parse_date(to_date) 
            month_year_range = [
                (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
            ]
            union_queries = []

            for month_year in month_year_range:
                res = await check_analysis_table(cnx,month_year)
                if len(res)>0:
                    table_name = f"ems_v1_completed.power_analysis_{month_year}"
                    union_queries.append(f"{table_name}")

            if len(union_queries) == 0:
                return _getErrorResponseJson("Analysis Table Not Available...")    
        # if date_time != '':
        #     date_time = date_time[:10]+" "+date_time[11:]
        data = await demand_report(period_id,campus_id,meter_id,from_date,to_date,shift_id,from_time,to_time,filter_type,parameter,duration,main_demand_meter,date_time,end_time,cnx) 
                
        response = {
            "iserror": False,
            "message": "Data Returned Successfully.",
            "data": data
        }
              
        return response
    except Exception as e:
        return get_exception_response(e)

def availabilityreport_excel(data,from_date,to_date,report_type,employee_name,month,report_method):
    wb = Workbook()
    ws = wb.active
    border_style = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    current_time = datetime.now()
    # current_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    current_time = current_time.strftime("%d-%b-%Y %H:%M:%S").upper()
    end_col = 4 
    ws.merge_cells('A1:N1')
    cell = "A1"
    
    if report_type == 'with_rate':
        try:
            month = datetime.strptime(month, "%d-%m-%Y").strftime("%b-%Y").upper()
        except:
            month = datetime.strptime(month, "%Y-%m-%d").strftime("%b-%Y").upper()
        if report_method == 'final':
            method = "Final"
        else:
            method = "Provision"
        energy_statement = f"Energy Statement With Tariff {month} ({method}) "
    else:
        
        from_date = datetime.strptime(from_date, "%d-%m-%Y").strftime("%d-%b-%Y").upper()

        to_date = datetime.strptime(to_date, "%d-%m-%Y").strftime("%d-%b-%Y").upper()
        energy_statement = f"Energy Statement Without Tariff {from_date} TO {to_date} "

    ws[cell] = energy_statement

    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment 
    ws[cell].border = border_style
    font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
    ws[cell].font = font
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill

    ws.column_dimensions[cell[0]].width = len(energy_statement) + 2  

    
    cell = "A3"
    ws[cell] = f"User : {employee_name}"
    alignment = Alignment(horizontal="left", vertical="top")
    ws[cell].alignment = alignment
    ws.column_dimensions[cell[0]].auto_size = True
    ws[cell].border = border_style 
    font = Font(name='Calibri', size=12)
    ws[cell].font = font

    cell = "A4"
    ws[cell] = f"Date&Time : {current_time}"
    alignment = Alignment(horizontal="left", vertical="top")
    ws[cell].alignment = alignment
    column_letter = cell[0]
    ws.column_dimensions[column_letter].width = 60
    ws[cell].border = border_style 
    font = Font(name='Calibri', size=12)
    ws[cell].font = font
    

    cell = "A5"
    ws[cell] = "Energy Source"
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment
    ws.column_dimensions[cell[0]].auto_size = True
    ws[cell].border = border_style 
    font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    ws[cell].font = font
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill

    if report_type =="with_rate":
        end_col = 14
        ws.merge_cells('D5:E5')
        cell = "D5"
        ws[cell] = "Rate Per Unit"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Rate Per Unit") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "D6"
        ws[cell] = "Budget"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Budget") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill
        report_name = ''
        if report_method == 'final':
            report_name = 'Actual(Final)'
        else:
            report_name = 'Actual(Provision)'
        cell = "E6"
        ws[cell] = report_name
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len('Actual(Provision)') + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        ws.merge_cells('F5:G5')
        cell = "F5"
        ws[cell] = "Total Amount	"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Total Amount	") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "F6"
        ws[cell] = "Budget"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Budget") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "G6"
        ws[cell] = "Actual"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Actual") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        ws.merge_cells('H5:I5')
        cell = "H5"
        ws[cell] = "Rate Variation"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Variation	") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        ws.merge_cells('J5:M5')
        cell = "J5"
        ws[cell] = "Mix Variation"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Variation	") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "H6"
        ws[cell] = "Rate/Unit"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Rate/Unit") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "I6"
        ws[cell] = "Total Amount"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Total Amount") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "J6"
        ws[cell] = "Mix%"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Total Amount") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "K6"
        ws[cell] = "Mix Units"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Total Amount") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        cell = "L6"
        ws[cell] = "Mix Amount"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Total Amount") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill
        cell = "M6"
        ws[cell] = "Total Amount"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        ws.column_dimensions[cell[0]].width = len("Total Amount") + 7  
        ws[cell].border = border_style 
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

    cell = "A6"
    ws[cell] = "External"
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment
    # ws.column_dimensions[cell[0]].width = len("External") + 7  
    ws[cell].border = border_style 
    font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    ws[cell].font = font
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill
    cell = "B6"
    ws[cell] = ""
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment
    ws.column_dimensions[cell[0]].width = len("External") + 7  
    ws[cell].border = border_style 
    font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    ws[cell].font = font
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill

    cell = "C6"
    ws[cell] = ""
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment
    ws.column_dimensions[cell[0]].width = len("External") + 7  
    ws[cell].border = border_style 
    font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    ws[cell].font = font
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill

    cell = "B5"
    ws[cell] = "Units"
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment
    ws.column_dimensions[cell[0]].width = 1   
    font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    ws[cell].font = font
    ws[cell].border = border_style 
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill
    

    cell = "C5"
    ws[cell] = "%"
    alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].alignment = alignment
    ws.column_dimensions[cell[0]].width = len("%") + 15  
    font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    ws[cell].font = font
    ws[cell].border = border_style 
    fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
    ws[cell].fill = fill  

    if not data or len(data) == '':
        cell = "H6"
        data = "No Data"
        ws[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment

        ws.column_dimensions[cell[0]].width = len(data) + 2  

        font = Font(name='Calibri', size=11)
        ws[cell].font = font
    else:
        cell_no = 0
        budget = 0
        actual = 0
        i_budget = 0
        i_actual = 0
        e_budget = 0
        e_actual = 0
        external_total_formula = 0
        next_row = 7
        next_row_total = next_row + len(data["data"])
        tneb_consumption = 0
        b_values = 0
        e_values = 0
        d_values = 0
        f_total = 0
        g_total = 0
        i_total = 0
        b_values_i = 0
        e_values_i = 0
        d_values_i = 0
        f_total_i = 0
        j_mix_value = 0
        mix_ieg_total = 0
        mix_ieg_grandtotal = 0
        b_mix_value = 0
        e_mix_value = 0
        g_value = 0
        g_total_i = 0
        i_total_i = 0
    
        for row in data["data"]:
            energy_source_name = row["energy_source_name"]
            if report_method == "final":
                consumption = row["consumption_total"]
            else:
                consumption = row["consumption"]
            e_budget = row["budget"]
            if report_method == "final":
                e_actual = row["actual_total"]
            else:
                e_actual = row["actual"]
            e_budget_mix = row["budget_mix"]

            # Write data to the worksheet (adjust column indices accordingly)
            ws.cell(row=next_row, column=1, value=energy_source_name)
            ws.cell(row=next_row, column=1).border = border_style

            ws.cell(row=next_row, column=2, value=consumption)
            if consumption <0:
                font = Font(name='Calibri', size=12, color='FF0000', bold=True)
                ws[cell].font = font
            cell = ws.cell(row=next_row, column=2)
            cell.number_format ="0"
            ws.cell(row=next_row, column=2).border = border_style
            if report_type == 'with_rate':
                ws.cell(row=next_row, column=4, value=e_budget)
                cell = ws.cell(row=next_row, column=4)
                cell.number_format ="0.00"
                ws.cell(row=next_row, column=4).border = border_style

                ws.cell(row=next_row, column=5, value=e_actual)
                cell = ws.cell(row=next_row, column=5)
                cell.number_format ="0.00"
                ws.cell(row=next_row, column=5).border = border_style

                ws.cell(row=next_row, column=10, value=e_budget_mix)
                cell = ws.cell(row=next_row, column=10)
                cell.number_format ="0.00"
                ws.cell(row=next_row, column=10).border = border_style

                ws.cell(row=next_row, column=6, value=f"=B{next_row}*D{next_row}")
                cell = ws.cell(row=next_row, column=6)
                cell.number_format ="0"
                ws.cell(row=next_row, column=6).border = border_style

                b_values = f"B{next_row}"
                b_values = ws[b_values].value

                d_values = f"D{next_row}"
                d_values = ws[d_values].value
                
                if (b_values * d_values)<0:
                    font = Font(color="FF0000",bold=True) 
                    ws.cell(row=next_row, column=6).font = font 
                f_total +=(b_values * d_values)
                ws.cell(row=next_row, column=7, value=f"=B{next_row}*E{next_row}")
                cell = ws.cell(row=next_row, column=7)
                cell.number_format ="0"
                ws.cell(row=next_row, column=7).border = border_style
                b_values = f"B{next_row}"
                b_values = ws[b_values].value

                e_values = f"E{next_row}"
                e_values = ws[e_values].value
                
                if (b_values * e_values)<0:
                    font = Font(color="FF0000",bold=True) 
                    ws.cell(row=next_row, column=7).font = font 
                g_total += (b_values * e_values)
                ws.cell(row=next_row, column=8, value=f"=D{next_row}-E{next_row}")
                cell = ws.cell(row=next_row, column=8)
                cell.number_format ="0.00"
                ws.cell(row=next_row, column=8).border = border_style

                if(d_values - e_values)<0:
                    font = Font(color="FF0000",bold=True) 
                    ws.cell(row=next_row, column=8).font = font 

                ws.cell(row=next_row, column=9, value=f"=round(F{next_row}-G{next_row},2)")
                cell = ws.cell(row=next_row, column=9)
                cell.number_format ="0"
                ws.cell(row=next_row, column=9).border = border_style

                ws.cell(row=next_row, column=12, value=f"=round(K{next_row}*E{next_row},0)")
                cell = ws.cell(row=next_row, column=12)
                cell.number_format ="0"
                ws.cell(row=next_row, column=12).border = border_style

                ws.cell(row=next_row, column=13, value=f"=round(L{next_row}-G{next_row},0)")
                cell = ws.cell(row=next_row, column=13)
                cell.number_format ="0"
                ws.cell(row=next_row, column=13).border = border_style
                
                if ((b_values*d_values) - (b_values*e_values)) <0:
                    font = Font(color="FF0000",bold=True)  
                    ws.cell(row=next_row, column=9).font = font
                i_total += ((b_values*d_values) - (b_values*e_values))
            if energy_source_name != 'TNEB' and energy_source_name != 'EB':
                tneb_consumption +=consumption
            next_row += 1
        
        temp_cell = f"B7"
        f_cell_value = ws[temp_cell].value
        tneb_c = f_cell_value - tneb_consumption 
        ws.cell(row=7, column=2, value=tneb_c)
        external_total_formula = tneb_c + tneb_consumption

        # Calculate and write grand total in the next row
        ws.cell(row=next_row, column=1, value="External Total")
        ws.cell(row=next_row, column=1).border = border_style
        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=next_row, column=1).font = font

        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=next_row, column=3).font = font

        ws.cell(row=next_row, column=2, value=external_total_formula)
        cell = ws.cell(row=next_row, column=2)
        cell.number_format ="0"
        ws.cell(row=next_row, column=2).border = border_style
        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=next_row, column=2).font = font

        if report_type == 'with_rate':
            ws.cell(row=next_row, column=6, value = f"=ROUND(SUM(F5:F{next_row-1}),0)")
            cell = ws.cell(row=next_row, column=6)
            cell.number_format ="0"
            ws.cell(row=next_row, column=6).border = border_style
            if f_total <0:
                font = Font(color="FF0000",bold=True)  
            else:
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=next_row, column=6).font = font

            ws.cell(row=next_row, column=7, value = f"=ROUND(SUM(G5:G{next_row-1}),0)")
            cell = ws.cell(row=next_row, column=7)
            cell.number_format ="0"
            ws.cell(row=next_row, column=7).border = border_style
            if g_total <0:
                font = Font(color="FF0000",bold=True)  
            else:
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=next_row, column=7).font = font
            

            ws.cell(row=next_row, column=9, value = f"=ROUND(SUM(I5:I{next_row-1}),0)")
            cell = ws.cell(row=next_row, column=9)
            cell.number_format ="0"
            ws.cell(row=next_row, column=9).border = border_style
            if i_total <0:
                font = Font(color="FF0000",bold=True)  
            else:
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=next_row, column=9).font = font

            ws.cell(row=next_row, column=11, value = f"=ROUND(SUM(K5:K{next_row-1}),0)")
            cell = ws.cell(row=next_row, column=11)
            cell.number_format ="0"
            ws.cell(row=next_row, column=11).border = border_style
            # if i_total <0:
            #     font = Font(color="FF0000",bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=next_row, column=11).font = font

            ws.cell(row=next_row, column=12, value = f"=ROUND(SUM(L5:L{next_row-1}),0)")
            cell = ws.cell(row=next_row, column=12)
            cell.number_format ="0"
            ws.cell(row=next_row, column=12).border = border_style
            # if i_total <0:
            #     font = Font(color="FF0000",bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=next_row, column=12).font = font

            ws.cell(row=next_row, column=13, value = f"=ROUND(SUM(M5:M{next_row-1}),0)")
            cell = ws.cell(row=next_row, column=13)
            cell.number_format ="0"
            ws.cell(row=next_row, column=13).border = border_style
            # if i_total <0:
            #     font = Font(color="FF0000",bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 13)  
            ws.cell(row=next_row, column=13).font = font

       
        ws.cell(row=next_row+1, column=1, value="Internal") 
        ws.cell(row=next_row+1, column=1).border = border_style
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        ws.cell(row=next_row+1, column=1).font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws.cell(row=next_row+1, column=1).fill = fill

        in_row = next_row + 2
        internal_total_formula = 0
        
        for internal in data["data3"]:
            source = internal["source"]
            kwh = internal["kWh"]
            i_budget = internal["budget"]
            if report_method == "final":
                i_actual = internal["actual_total"]
            else:
                i_actual = internal["actual"]
            i_budget_mix = internal["budget_mix"]
            

            ws.cell(row=in_row, column=1, value=f"{source}")
            cell = ws.cell(row=in_row, column=1)
            cell.number_format ="0"
            ws.cell(row=in_row, column=1).border = border_style

            ws.cell(row=in_row, column=2, value=kwh)
            if kwh <0: 
                font = Font(name='Calibri', size=12, color='FF0000', bold=True)
                ws[cell].font = font
            cell = ws.cell(row=in_row, column=2)
            cell.number_format ="0"
            ws.cell(row=in_row, column=2).border = border_style
            if report_type == 'with_rate':
                ws.cell(row=in_row, column=4, value=i_budget)
                cell = ws.cell(row=in_row, column=4)
                cell.number_format ="0"
                ws.cell(row=in_row, column=4).border = border_style

                ws.cell(row=in_row, column=5, value=i_actual)
                cell = ws.cell(row=in_row, column=5)
                cell.number_format ="0"
                ws.cell(row=in_row, column=5).border = border_style

                ws.cell(row=in_row, column=10, value=i_budget_mix)
                cell = ws.cell(row=in_row, column=10)
                cell.number_format ="0"
                ws.cell(row=in_row, column=10).border = border_style

                ws.cell(row=in_row, column=6, value=f"=B{in_row}*D{in_row}")
                cell = ws.cell(row=in_row, column=6)
                cell.number_format ="0"
                ws.cell(row=in_row, column=6).border = border_style
                
                b_values_i = f"B{in_row}"
                b_values_i = ws[b_values_i].value

                d_values_i = f"D{in_row}"
                d_values_i = ws[d_values_i].value
                
                if(b_values_i * d_values_i)<0:
                    font = Font(color="FF0000",bold=True) 
                    ws.cell(row=in_row, column=6).font = font 
                f_total_i +=(b_values_i * d_values_i)

                ws.cell(row=in_row, column=7, value=f"=B{in_row}*E{in_row}")
                cell = ws.cell(row=in_row, column=7)
                cell.number_format ="0"
                ws.cell(row=in_row, column=7).border = border_style

                b_values_i = f"B{in_row}"
                b_values_i = ws[b_values_i].value

                e_values_i = f"E{in_row}"
                e_values_i = ws[e_values_i].value

                if(b_values_i * e_values_i)<0:
                    font = Font(color="FF0000",bold=True)  
                    ws.cell(row=in_row, column=7).font = font
                g_total_i +=(b_values_i * e_values_i)

                ws.cell(row=in_row, column=8, value=f"=D{in_row}-E{in_row}")
                cell = ws.cell(row=in_row, column=8)
                cell.number_format ="0.00"
                ws.cell(row=in_row, column=8).border = border_style

                if(d_values_i - e_values_i)<0:
                    font = Font(color="FF0000",bold=True)  
                    ws.cell(row=in_row, column=8).font = font

                ws.cell(row=in_row, column=9, value=f"=F{in_row}-G{in_row}")
                cell = ws.cell(row=in_row, column=9)
                cell.number_format ="0"
                ws.cell(row=in_row, column=9).border = border_style

                ws.cell(row=in_row, column=12, value=f"=K{in_row}*E{in_row}")
                cell = ws.cell(row=in_row, column=12)
                cell.number_format ="0"
                ws.cell(row=in_row, column=12).border = border_style

                ws.cell(row=in_row, column=13, value=f"=L{in_row}-G{in_row}")
                cell = ws.cell(row=in_row, column=13)
                cell.number_format ="0"
                ws.cell(row=in_row, column=13).border = border_style

                if((b_values_i*d_values_i) - (b_values_i*e_values_i))<0:
                    font = Font(color="FF0000",bold=True)  
                    ws.cell(row=in_row, column=9).font = font
                i_total_i += ((b_values_i*d_values_i) - (b_values_i*e_values_i))
            internal_total_formula +=kwh
           
            in_row +=1

        ws.cell(row=in_row, column=1, value="Internal Total")
        ws.cell(row=in_row, column=1).border = border_style
        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=in_row, column=1).font = font

        ws.cell(row=in_row, column=2, value = internal_total_formula)
        cell = ws.cell(row=in_row, column=2)
        cell.number_format ="0"
        ws.cell(row=in_row, column=2).border = border_style
        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=in_row, column=2).font = font
        if report_type == 'with_rate':
            ws.cell(row=in_row, column=6, value = f"=ROUND(SUM(F{next_row + 2}:F{in_row-1}),0)")
            cell = ws.cell(row=in_row, column=6)
            cell.number_format ="0"
            ws.cell(row=in_row, column=6).border = border_style
            if f_total_i <0:
                font = Font(color="FF0000",bold=True)  
            else:
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row, column=6).font = font

            ws.cell(row=in_row, column=7, value = f"=ROUND(SUM(G{next_row + 2}:G{in_row-1}),0)")
            cell = ws.cell(row=in_row, column=7)
            cell.number_format ="0"
            ws.cell(row=in_row, column=7).border = border_style
            if g_total_i <0:
                font = Font(color="FF0000",bold=True)  
            else:
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row, column=7).font = font

            ws.cell(row=in_row, column=9, value = f"=ROUND(SUM(I{next_row + 2}:I{in_row-1}),0)")
            cell = ws.cell(row=in_row, column=9)
            cell.number_format ="0"
            ws.cell(row=in_row, column=9).border = border_style
            if i_total_i <0:
                font = Font(color="FF0000",bold=True)  
            else:
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row, column=9).font = font

            ws.cell(row=in_row, column=11, value = f"=ROUND(SUM(K{next_row + 2}:K{in_row-1}),0)")
            cell = ws.cell(row=in_row, column=11)
            cell.number_format ="0"
            ws.cell(row=in_row, column=11).border = border_style
            # if i_total_i <0:
            #     font = Font(color="FF0000",bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row, column=11).font = font

            ws.cell(row=in_row, column=12, value = f"=ROUND(SUM(L{next_row + 2}:L{in_row-1}),0)")
            cell = ws.cell(row=in_row, column=12)
            cell.number_format ="0"
            ws.cell(row=in_row, column=12).border = border_style
            # if i_total_i <0:
            #     font = Font(color="FF0000",bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row, column=12).font = font

            ws.cell(row=in_row, column=13, value = f"=ROUND(SUM(M{next_row + 2}:M{in_row-1}),0)")
            cell = ws.cell(row=in_row, column=13)
            cell.number_format ="0"
            ws.cell(row=in_row, column=13).border = border_style
            # if i_total_i <0:
            #     font = Font(color="FF0000",bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row, column=13).font = font

        font = Font(color='0710ba',bold=True,size = 13)  
        
        ws.cell(row=in_row, column=3).font = font
       

        ws.cell(row=in_row+1, column=1, value="Grand Total")
        ws.cell(row=in_row+1, column=1).border = border_style
        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=in_row+1, column=1).font = font
        grand_total = round(external_total_formula + internal_total_formula)

        ws.cell(row=in_row+1, column=2, value = f"={grand_total}")
        cell = ws.cell(row=in_row, column=2)
        cell.number_format ="0"
        ws.cell(row=in_row+1, column=2).border = border_style
        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=in_row+1, column=2).font = font

        if report_type =="with_rate" :
            ws.cell(row=in_row+1, column=6, value = f"=F{next_row}+F{in_row}")
            cell = ws.cell(row=in_row, column=6)
            cell.number_format ="0"
            ws.cell(row=in_row+1, column=6).border = border_style

            if(f_total + f_total_i)<0:
                font = Font(color="FF0000",bold=True)  
            else :   
                font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=in_row+1, column=6).font = font

            ws.cell(row=in_row+1, column=7, value = f"=G{next_row}+G{in_row}")
            cell = ws.cell(row=in_row, column=7)
            cell.number_format ="0"
            ws.cell(row=in_row+1, column=7).border = border_style

            if(g_total + g_total_i)<0:
                font = Font(color="FF0000",bold=True)  
            else :   
                font = Font(color='0710ba',bold=True,size = 12)  

            ws.cell(row=in_row+1, column=7).font = font

            ws.cell(row=in_row+1, column=9, value = f"=I{next_row}+I{in_row}")
            cell = ws.cell(row=in_row, column=9)
            cell.number_format ="0"
            ws.cell(row=in_row+1, column=9).border = border_style
            i_values = f"I{next_row}"
            i_values = ws[i_values].value

            if(i_total + i_total_i)<0:
                font = Font(color="FF0000",bold=True)  
            else :   
                font = Font(color='0710ba',bold=True,size = 12)   
            ws.cell(row=in_row+1, column=9).font = font


            ws.cell(row=in_row+1, column=11, value = f"=round(k{next_row}+k{in_row},0)")
            cell = ws.cell(row=in_row, column=11)
            cell.number_format ="0"
            font = Font(color='0710ba',bold=True,size = 12)   
            ws.cell(row=in_row+1, column=11).font = font


            ws.cell(row=in_row+1, column=12, value = f"=round(L{next_row}+L{in_row},0)")
            cell = ws.cell(row=in_row, column=12)
            cell.number_format ="0"
            font = Font(color='0710ba',bold=True,size = 12)   
            ws.cell(row=in_row+1, column=12).font = font

            ws.cell(row=in_row+1, column=13, value = f"=round(M{next_row}+M{in_row},0)")
            cell = ws.cell(row=in_row, column=13)
            cell.number_format ="0"
            font = Font(color='0710ba',bold=True,size = 12)   
            ws.cell(row=in_row+1, column=13).font = font

            

        font = Font(color='0710ba',bold=True,size = 12)  
        ws.cell(row=in_row+1, column=3).font = font

        row1 = in_row + 3
       
        for per_row in range(7,row1 - 1):
            if per_row == next_row+1:
                continue
            ws.cell(row=per_row, column=3, value=f"=B{per_row}/{grand_total}*100")
            cell = ws.cell(row=per_row, column=3)
            cell.number_format ="0.00"
            ws.cell(row=per_row, column=3).border = border_style
            a_value = f"A{per_row}"
            a_value = ws[a_value].value
            # print(per_row)
            
            if report_type == 'with_rate':
                if a_value not in  ["External Total","Internal Total","Grand Total"]:
                    ws.cell(row=per_row, column=11, value=f"=J{per_row}/100*{grand_total}")
                    cell = ws.cell(row=per_row, column=11)
                    cell.number_format ="0"
                    ws.cell(row=per_row, column=11).border = border_style
                    font = Font(color='0710ba',bold=True,size = 12)  
                    ws.cell(row=in_row+1, column=11).font = font
                
                    g_value = f"G{per_row}"
                    g_value = ws[g_value].value    

                    e_mix_value = f"E{per_row}"
                    e_mix_value = ws[e_mix_value].value            

                    b_mix_value = f"B{per_row}"
                    b_mix_value = ws[b_mix_value].value            

                    j_mix_value = f"J{per_row}"
                    j_mix_value = ws[j_mix_value].value
                    createFolder("Log/",f"{a_value}--{mix_ieg_total}")

                    if (((j_mix_value/100*grand_total)*e_mix_value) - (b_mix_value*e_mix_value))<0 :
                        font = Font(color='FF0000',bold=True)      
                        ws.cell(row=per_row, column=13).font = font
                    mix_ieg =(((j_mix_value/100*grand_total)*e_mix_value) - (b_mix_value*e_mix_value))
                    
                    mix_ieg_total +=mix_ieg
                    mix_ieg_grandtotal +=mix_ieg

                if a_value in ["External Total","Internal Total"]:
                    if mix_ieg_total <0:
                        font = Font(color='FF0000',bold=True)    
                    else:  
                        font = Font(color='0710ba',bold=True)
                    ws.cell(row=per_row, column=13).font = font

                    mix_ieg_total = 0  
                if a_value == "Grand Total":
                    if mix_ieg_grandtotal <0:
                        font = Font(color='FF0000',bold=True)    
                    else:  
                        font = Font(color='0710ba',bold=True)
                    ws.cell(row=per_row, column=13).font = font

        for row_num in range(5,per_row+1):
                for col_num in range(1, end_col):  
                    border_style = Border(left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin'))

                    ws.cell(row=row_num, column=col_num).border = border_style
                    
        cell=ws.cell(row=row1, column=1, value="Plant")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10
        

        cell = ws.cell(row=row1, column=2, value="Reporting Department")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 

        column_letter = get_column_letter(2)  
        ws.column_dimensions[column_letter].width = 25

        cell=ws.cell(row=row1, column=3, value="Units (EB+DG)")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        ws.merge_cells(start_row=row1, start_column=6, end_row=row1, end_column=7)
        cell=ws.cell(row=row1, column=6, value="Dis Loss")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        ws.merge_cells(start_row=row1, start_column=4, end_row=row1, end_column=5)
        cell=ws.cell(row=row1, column=4, value="Utility	")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

            
        cell=ws.cell(row=row1, column=8, value="  Total units  ")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 20

        ws.merge_cells(start_row=row1, start_column=9, end_row=row1, end_column=12)
        cell = ws.cell(row=row1, column=9, value="Dispatch T")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10
    

        cell=ws.cell(row=row1, column=13, value="Bud U/T")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=13, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1, column=14, value="Act U/T")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name='Calibri', size=14, color='FFFFFF', bold=True)
        cell.font = font
        cell.border = border_style 
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        cell.fill = fill 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=6, value="Units")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style 
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=7, value="%")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10) 
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=4, value="Units ")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=5, value="%")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=9, value="Cust + WH")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=10, value="Within Plant")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=11, value="To Other Plants")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        cell=ws.cell(row=row1+1, column=12, value="Total")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        column_letter = get_column_letter(10)  
        ws.column_dimensions[column_letter].width = 10

        current_plant = ''
        current_subtotal = 0
        budget_subtotal = 0
        budget_grandtotal = 0
        common_grandtotal = 0
        common_subtotal = 0
        overall_grand_total = 0
        subtotal_row_index = row1 + 2

        start_cell_no = subtotal_row_index
        end_cell_no = subtotal_row_index
        index = subtotal_row_index
        utility_cell = {}
        utility_cell_start = subtotal_row_index
        start = index -2
        for i in data["data2"]:
            plant_name = i["plant_name"]
            reporting_department = i["reporting_department"]
            pm_kwh = i["pm_kwh"]
            pm_common_kwh = i["pm_common_kwh"]
            budget = i["budget"]
            is_corporate = i["is_corporate"]
            campus_name = i["campus_name"]
            # Check if the plant has changed
            cell = "A2"
            ws[cell] = f"Campus : {campus_name}"
            alignment = Alignment(horizontal="left", vertical="top")
            ws[cell].alignment = alignment
            ws.column_dimensions[cell[0]].auto_size = True
            ws[cell].border = border_style 
            font = Font(name='Calibri', size=12)
            ws[cell].font = font
            if current_plant != plant_name:
                if is_corporate == 'no':
                
                    ws.cell(row=index, column=1, value=plant_name)
                    ws.cell(row=index, column=1).border = border_style

                    if current_plant != '':
                        tempt_celldata = subtotal_row_index-1
                        utility_cell[subtotal_row_index] = [utility_cell_start, tempt_celldata]
                        utility_cell_start = subtotal_row_index + 1
                        ws.cell(row=subtotal_row_index, column=2, value="Total")
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=2).font = font   
                        ws.cell(row=subtotal_row_index, column=3, value=current_subtotal)
                        cell = ws.cell(row=subtotal_row_index, column=3)
                        cell.number_format ="0"    
                        # if  current_subtotal<0:
                        #     font = Font(color='FF0000',bold=True)  
                        # else:
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=3).font = font 
                         

                        ws.cell(row=subtotal_row_index, column=6, value=0)
                        cell = ws.cell(row=subtotal_row_index, column=6)
                        cell.number_format ="0"
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=6).font = font   
                        ws.cell(row=subtotal_row_index, column=7, value=f"=D{subtotal_row_index}/C{subtotal_row_index}*100")
                        cell = ws.cell(row=subtotal_row_index, column=7)
                        cell.number_format ="0.00"
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=5).font = font 

                        ws.cell(row=subtotal_row_index, column=4, value=common_subtotal)
                        cell = ws.cell(row=subtotal_row_index, column=4)
                        cell.number_format ="0"
                        # if  common_subtotal<0:
                        #     font = Font(color='FF0000',bold=True)  
                        # else:
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=4).font = font 

                        ws.cell(row=subtotal_row_index, column=5, value=f"=F{subtotal_row_index}/C{subtotal_row_index}*100")
                        cell = ws.cell(row=subtotal_row_index, column=5)
                        cell.number_format ="0.00"
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=5).font = font 
                        ws.cell(row=subtotal_row_index, column=8, value=f"=C{subtotal_row_index}+F{subtotal_row_index}+D{subtotal_row_index}")
                        cell = ws.cell(row=subtotal_row_index, column=8)
                        cell.number_format ="0"
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=subtotal_row_index, column=8).font = font 
                        
                        common_grandtotal +=common_subtotal
                        print(common_grandtotal)
                        current_subtotal = 0
                        budget_subtotal = 0
                        common_subtotal = 0
                        subtotal_row_index = subtotal_row_index + 1
            if is_corporate =='no':
                
                ws.cell(row=subtotal_row_index, column=2, value=reporting_department)
                ws.cell(row=subtotal_row_index, column=2).border = border_style
                ws.cell(row=subtotal_row_index, column=3, value=pm_kwh)
                # if  pm_kwh<0:
                #     font = Font(color='FF0000',bold=True)  
                #     ws.cell(row=subtotal_row_index, column=3).font = font
                cell = ws.cell(row=subtotal_row_index, column=3)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=3).border = border_style
                ws.cell(row=subtotal_row_index, column=6, value=0)
                cell = ws.cell(row=subtotal_row_index, column=6)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=6).border = border_style
                ws.cell(row=subtotal_row_index, column=7, value=f"=D{subtotal_row_index}/C{subtotal_row_index}*100")
                cell = ws.cell(row=subtotal_row_index, column=7)
                cell.number_format ="0.00"
                ws.cell(row=subtotal_row_index, column=7).border = border_style
                ws.cell(row=subtotal_row_index, column=4, value=pm_common_kwh)
                cell = ws.cell(row=subtotal_row_index, column=4)
                cell.number_format ="0"
                # if  pm_common_kwh<0:
                #     font = Font(color='FF0000',bold=True)  
                #     ws.cell(row=subtotal_row_index, column=6).font = font
                ws.cell(row=subtotal_row_index, column=4).border = border_style
                ws.cell(row=subtotal_row_index, column=5, value=f"=F{subtotal_row_index}/C{subtotal_row_index}*100")
                cell = ws.cell(row=subtotal_row_index, column=5)
                cell.number_format ="0.00"
                ws.cell(row=subtotal_row_index, column=5).border = border_style
                ws.cell(row=subtotal_row_index, column=8, value=f"=C{subtotal_row_index}+F{subtotal_row_index}+D{subtotal_row_index}")
                cell = ws.cell(row=subtotal_row_index, column=8)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=8).border = border_style

                ws.cell(row=subtotal_row_index, column=9, value=0)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=9).border = border_style

                ws.cell(row=subtotal_row_index, column=10, value=0)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=10).border = border_style

                ws.cell(row=subtotal_row_index, column=11, value=0)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=11).border = border_style
               
                ws.cell(row=subtotal_row_index, column=12, value=f"=I{subtotal_row_index}+J{subtotal_row_index}+K{subtotal_row_index}")
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=12).border = border_style
               
                ws.cell(row=subtotal_row_index, column=13, value=budget)
                cell = ws.cell(row=subtotal_row_index, column=13)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=13).border = border_style
                ws.cell(row=subtotal_row_index, column=14, value=f"=H{subtotal_row_index}/L{subtotal_row_index}")
                cell = ws.cell(row=subtotal_row_index, column=14)
                cell.number_format ="0"
                ws.cell(row=subtotal_row_index, column=14).border = border_style

                current_subtotal += pm_kwh
                overall_grand_total += pm_kwh
                common_subtotal =pm_common_kwh
                budget_subtotal += budget
                budget_grandtotal += budget
                subtotal_row_index += 1
                # Set column width for column 3 based on the maximum length of pm_kwh
                ws.column_dimensions[get_column_letter(3)].width = max(
                    ws.column_dimensions[get_column_letter(3)].width,
                    len(str(pm_kwh)) + 2
                )
                current_plant = plant_name

                index = subtotal_row_index +1
        
        if current_plant != '':
            common_grandtotal +=common_subtotal
            
            tempt_celldata = subtotal_row_index-1
            utility_cell[subtotal_row_index] = [utility_cell_start, tempt_celldata]
            for i in utility_cell:
                for row in range(utility_cell[i][0],utility_cell[i][1]+1):
                    ws.cell(row=row, column=6, value=f"=C{row}/C{i}*F{i}")
            
            ws.cell(row=subtotal_row_index, column=2, value="Total")
            ws.cell(row=subtotal_row_index, column=2).border = border_style
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index, column=2).font = font   

            ws.cell(row=subtotal_row_index, column=3, value=current_subtotal)
            cell = ws.cell(row=subtotal_row_index, column=3)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index, column=3).border = border_style
            # if  current_subtotal<0:
            #     font = Font(color='FF0000',bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index, column=3).font = font            

            ws.cell(row=subtotal_row_index, column=6, value=0)
            cell = ws.cell(row=subtotal_row_index, column=6)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index, column=6).border = border_style
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index, column=6).font = font

            ws.cell(row=subtotal_row_index, column=7, value=f"=C{subtotal_row_index}/{overall_grand_total}*100")
            # ws.cell(row=subtotal_row_index, column=5).border = border_style
            # font = Font(color='FF0000',bold=True)  
            # ws.cell(row=subtotal_row_index, column=5).font = font
            cell = ws.cell(row=subtotal_row_index, column=7)
            cell.number_format = '0.00'
            border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell.border = border_style
            font = Font(color='0710ba', bold=True)
            cell.font = font

            ws.cell(row=subtotal_row_index, column=4, value=common_subtotal)
            cell = ws.cell(row=subtotal_row_index, column=4)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index, column=4).border = border_style
            # if  common_subtotal<0:
            #     font = Font(color='FF0000',bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
                   
            ws.cell(row=subtotal_row_index, column=4).font = font

            ws.cell(row=subtotal_row_index, column=5, value=f"=F{subtotal_row_index}/C{subtotal_row_index}*100")
            # ws.cell(row=subtotal_row_index, column=7).border = border_style
            # font = Font(color='FF0000',bold=True)  
            # ws.cell(row=subtotal_row_index, column=7).font = font
            cell = ws.cell(row=subtotal_row_index, column=5)
            cell.number_format = '0.00'
            border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell.border = border_style
            font = Font(color='0710ba', bold=True)
            cell.font = font

            ws.cell(row=subtotal_row_index, column=8, value=f"=C{subtotal_row_index}+F{subtotal_row_index}+D{subtotal_row_index}")
            cell = ws.cell(row=subtotal_row_index, column=8)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index, column=8).border = border_style
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index, column=8).font = font

            
            ws.cell(row=subtotal_row_index + 1, column=2, value="Grand Total")
            ws.cell(row=subtotal_row_index + 1, column=2).border = border_style
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index + 1, column=2).font = font

            ws.cell(row=subtotal_row_index + 1, column=3, value=overall_grand_total)
            cell = ws.cell(row=subtotal_row_index + 1, column=3)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index + 1, column=3).border = border_style
            # if  overall_grand_total<0:
            #     font = Font(color='FF0000',bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
           
            ws.cell(row=subtotal_row_index + 1, column=3).font = font

            ws.cell(row=subtotal_row_index + 1, column=6, value=0)
            cell = ws.cell(row=subtotal_row_index + 1, column=6)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index + 1, column=6).border = border_style
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index + 1, column=6).font = font

            ws.cell(row=subtotal_row_index + 1, column=5, value=f"=C{subtotal_row_index+1}/C{subtotal_row_index+1}*100")
            # ws.cell(row=subtotal_row_index + 1, column=5).border = border_style
            # font = Font(color='008000',bold=True)  
            # ws.cell(row=subtotal_row_index + 1, column=5).font = font
            cell = ws.cell(row=subtotal_row_index + 1, column=7)
            cell.number_format = '0.00'
            border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell.border = border_style
            font = Font(color='0710ba', bold=True)
            cell.font = font

            ws.cell(row=subtotal_row_index + 1, column=4, value=common_grandtotal)
            cell = ws.cell(row=subtotal_row_index + 1, column=4)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index + 1, column=4).border = border_style
            # if  common_grandtotal<0:
            #     font = Font(color='FF0000',bold=True)  
            # else:
            font = Font(color='0710ba',bold=True,size = 12)  
             
            ws.cell(row=subtotal_row_index + 1, column=4).font = font

            ws.cell(row=subtotal_row_index + 1, column=5, value=f"=F{subtotal_row_index+1}/C{subtotal_row_index+1}*100")
            # ws.cell(row=subtotal_row_index + 1, column=7).border = border_style
            # font = Font(color='008000',bold=True)  
            # ws.cell(row=subtotal_row_index + 1, column=7).font = font
            cell = ws.cell(row=subtotal_row_index + 1, column=5)
            cell.number_format = '0.00'
            border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            cell.border = border_style
            font = Font(color='0710ba', bold=True)
            cell.font = font

            ws.cell(row=subtotal_row_index + 1, column=8, value=f"=C{subtotal_row_index+1}+F{subtotal_row_index+1}+D{subtotal_row_index+1}")
            cell = ws.cell(row=subtotal_row_index + 1, column=8)
            cell.number_format ="0"
            ws.cell(row=subtotal_row_index + 1, column=8).border = border_style
            font = Font(color='0710ba',bold=True,size = 12)  
            ws.cell(row=subtotal_row_index + 1, column=8).font = font

            temp_cell = f"C{subtotal_row_index+1}"
            c_cell_value = ws[temp_cell].value

            temp_cell = f"H{subtotal_row_index+1}"
            h_cell_value = ws[temp_cell].value
            

            temp_cell = f"D{subtotal_row_index+1}"
            d_cell_value = ws[temp_cell].value
            # print("d_cell_value",d_cell_value)
            temp_cell = f"F{subtotal_row_index+1}"
            f_cell_value = ws[temp_cell].value

            diff_value = grand_total - (c_cell_value + d_cell_value + f_cell_value)
            
            end_cell_no = subtotal_row_index + 1
            crop_total = 0
            crop_grand_total= 0
            cell_no = end_cell_no +1
            
            for i in data["data2"]:
                plant_name = i["plant_name"]
                reporting_department = i["reporting_department"]
                pm_kwh = i["pm_kwh"]
                pm_common_kwh = i["pm_common_kwh"]
                budget = i["budget"]
                is_corporate = i["is_corporate"]
                
                if is_corporate == 'yes':
                    if reporting_department != '':
                        ws.cell(row=cell_no, column=2, value=reporting_department)  
                        ws.cell(row=cell_no, column=8, value=f"={pm_kwh}")
                        # if  common_grandtotal<0:
                        #     font = Font(color='FF0000',bold=True)  
                        #     ws.cell(row=cell_no, column=8,).font = font
                        cell = ws.cell(row=subtotal_row_index, column=8)
                        cell.number_format ="0" 
                        crop_total +=pm_kwh
           
                        cell_no +=1
            
                    if reporting_department != '':  
                        # crop_grand_total = grand_total +  crop_total  
                        
                        ws.cell(row=cell_no, column=2, value="Total")
                        ws.cell(row=cell_no, column=2).border = border_style
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=cell_no, column=2).font = font

                        ws.cell(row=cell_no, column=8, value=crop_total)
                        cell = ws.cell(row=cell_no, column=8)
                        cell.number_format ="0"
                        ws.cell(row=cell_no, column=8).border = border_style
                        # if  crop_total<0:
                        #     font = Font(color='FF0000',bold=True)  
                        # else:          
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=cell_no, column=8).font = font

                        ws.cell(row=cell_no + 1, column=2, value="Grand Total")
                        ws.cell(row=cell_no + 1, column=2).border = border_style
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=cell_no + 1, column=2).font = font

                        ws.cell(row=cell_no + 1, column=8, value=grand_total)
                        cell = ws.cell(row=cell_no + 1, column=8)
                        cell.number_format ="0"
                        ws.cell(row=cell_no + 1, column=8).border = border_style   
                        font = Font(color='0710ba',bold=True,size = 12)  
                        ws.cell(row=cell_no + 1, column=8).font = font

            
            diff_value = grand_total - (c_cell_value + d_cell_value + f_cell_value + crop_total) 

            for i in range(start_cell_no, end_cell_no + 1):
                c_cell = f"C{i}"
                c_value = ws[c_cell].value
                
                ws.cell(row=i, column=4, value=f"=C{i}/{c_cell_value}*{common_grandtotal}")
                ws.cell(row=i, column=6, value=f"=C{i}/{c_cell_value}*{diff_value}")
                d_cell = f"D{i}"
                d_value = ws[d_cell].value
                # print("d_value",d_value)
                ws.cell(row=i, column=7, value=f"=F{i}/{c_cell_value}*100")
                f_cell = f"F{i}"
                f_value = ws[f_cell].value
                # print("f_value",f_value)
                ws.cell(row=i, column=5, value=f"=D{i}/{c_cell_value}*100")
        
                if c_value/c_cell_value*diff_value< 0:
                    ws.cell(row=i, column=6).font = Font(color="FF0000",bold=True) 
    
                if (c_value/c_cell_value*diff_value)/c_cell_value*100< 0:
                    ws.cell(row=i, column=7).font = Font(color="FF0000",bold=True) 
                
                if (c_value/c_cell_value*common_grandtotal)/c_cell_value*100<0:
                    ws.cell(row=i, column=5).font = Font(color="FF0000",bold=True)  

                if c_value/c_cell_value*common_grandtotal< 0:
                    ws.cell(row=i, column=4).font = Font(color="FF0000",bold=True)  
                
                if (c_value+(c_value/c_cell_value*diff_value)+(c_value/c_cell_value*common_grandtotal))< 0:
                    ws.cell(row=i, column=8).font = Font(color="FF0000",bold=True)  

            for row_num in range(start, cell_no + 2):
                for col_num in range(1, 15):  
                    border_style = Border(left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin'))

                    ws.cell(row=row_num, column=col_num).border = border_style
    
        for row_num in range(1, cell_no + 2):
            for col_num in range(1, 14):

                cell_display_value = ws.cell(row=row_num, column=col_num).value

                if isinstance(cell_display_value, str) :
                    cell_display_value = ws.cell(row=row_num, column=col_num).value
                    # print(cell_display_value)
                if isinstance(cell_display_value, (int, float)) and cell_display_value < 0:
                    ws.cell(row=row_num, column=col_num).font = Font(color='FF0000',bold=True)
    
    file_name = f'availability_report.xlsx'
    file_path = os.path.join(static_dir, file_name)
    wb.save(file_path)
    
def availabilityreport_pdf(data, from_date, to_date,report_type,employee_name,month,report_method):
    file_name = 'availability_report.pdf'
    file_path = os.path.join(static_dir, file_name)
    doc = SimpleDocTemplate(file_path, pagesize=letter, leftMargin=10, rightMargin=10, topMargin=20, bottomMargin=10)  # Adjust margin size as needed

    styles = getSampleStyleSheet()

    # Build the document content
    content = []
    current_date = datetime.now()
    current_date = current_date.strftime("%d-%b-%Y %H:%M:%S").upper()
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]
    title_style.alignment = 1  # Center alignment

    # Create content
    content = []
    from_date = from_date.strftime("%d-%b-%Y").upper()

    to_date = to_date.strftime("%d-%b-%Y").upper()
    # Add title
    if report_type == 'with_rate':
        try:
            month = datetime.strptime(month, "%d-%m-%Y").strftime("%b-%Y").upper()
        except:
            month = datetime.strptime(month, "%Y-%m-%d").strftime("%b-%Y").upper()
        if report_method == 'final':
            method = "Final"
        else:
            method = "Provision"
        energy_statement = f"Energy Statement With Tariff {month} ({method})"
    else:
        energy_statement = f"Energy Statement Without Tariff {from_date} TO {to_date}"

    title = Paragraph(energy_statement, title_style)
    content.append(title)
    normal_style = styles["Normal"]
    
    content.append(Paragraph("<br/>", getSampleStyleSheet()["Normal"]))
    custom_style = ParagraphStyle(
    name='CustomStyle',
    fontSize=12,)
    user_info = f"      User : {employee_name}"
    content.append(Paragraph(user_info, custom_style))
    print("111",data)
    if data!= []:
        for row in data["data2"]:
            campus_name = row["campus_name"]
            break
        user_info = f"      Campus : {campus_name}"
        content.append(Paragraph(user_info, custom_style))

        date_info = f"      Date&Time : {current_date}"
        content.append(Paragraph(date_info, custom_style))
        
        content.append(Paragraph("<br/><br/>", getSampleStyleSheet()["Normal"]))
        external_total = 0
        tneb_consumption = 0
        bal_consumption = 0
        internal_total = 0
        EI_total = 0
        report_name = ''
        if report_method == "final":
            report_name = "Actual(Final)"
        else:
            report_name = "Actual(Provision)"
        if report_type == 'with_rate':
            table_data = [
        
                ["Energy Source", "Units", "%", "Rate Per Unit", "", "Total Amount", "", "Rate Variation", "","",'Mix Variation','',''],
                ["External", "", "", "Budget", f"{report_name}", "Budget", "Actual", "Rate/Unit", "Total Amount",'Mix%','Mix Units','Mix Amount','Total Amount']
            ]
            
            for row in data["data"]:
                energy_source_name = row["energy_source_name"]
                if report_method == 'final':
                    consumption = row["consumption_total"]
                else:
                    consumption = row["consumption"]

                e_budget = row["budget"]
                if report_method == "final":
                    e_actual = row["actual_total"]
                else:
                    e_actual = row["actual"]
                e_budget_mix = row["budget_mix"]
                
                table_data.append([energy_source_name, consumption, "",e_budget,e_actual,"","","","",round(e_budget_mix,2),0,0,0])

                if energy_source_name != "TNEB" and energy_source_name != "EB":
                    bal_consumption += consumption

                if energy_source_name == "TNEB" or energy_source_name == "EB":
                    external_total = consumption

            tneb_consumption = external_total - bal_consumption
            ie_budget_total = 0
            ie_actual_total = 0
            ie_unit_total = 0
            ie_amount_total = 0
            g_budget_total = 0
            g_actual_total = 0
            g_unit_total = 0
            g_amount_total = 0
            external_end_row = 0
            
            # for i, row in enumerate(table_data):
            #     if row[0] != 'Energy Source'  and row[1]!="Total Amount":
            #         if row[1] != '':
            #             if row[3] != '':
            #                 table_data[i][1] = round(table_data[i][1])
            #                 table_data[i][5] = row[1]*row[3]
            #                 external_budget_total +=(float(table_data[i][1])*float(table_data[i][3]))
            #                 table_data[i][6] = round(float(table_data[i][1])*float(table_data[i][4]))
            #                 external_actual_total +=(float(table_data[i][1])*float(table_data[i][4]))
            #                 table_data[i][7] = round(float(table_data[i][3])-float(table_data[i][4]),2)
            #                 external_unit_total +=(float(table_data[i][3])-float(table_data[i][4]))
            #                 table_data[i][8] = round(float(table_data[i][5])-float(table_data[i][6]))
            #                 external_amount_total +=(float(table_data[i][5])-float(table_data[i][6]))
            # external_end_row = i
            table_data.append(["External Total", round(external_total), "","","",0,0,'',0,'',0,0,0])
            
            table_data.append(["Internal", '', "","","","","","",""])
            for internal in data["data3"]:
                source = internal["source"]
                kwh = internal["kWh"]
                i_budget = internal["budget"]
                if report_method == "final":
                    i_actual = internal["actual_total"]
                else:
                    i_actual = internal["actual"]
                i_budget_mix = internal["budget_mix"]
                table_data.append([source, kwh, "",i_budget,i_actual,"","","","",round(i_budget_mix,2),0,0,0])
                internal_total += kwh

            internal_budget_total = 0
            internal_actual_total = 0
            internal_unit_total = 0
            internal_amount_total = 0
            # for i, row in enumerate(table_data):
            #     if row[0] != 'Energy Source'  and row[1]!="Total Amount"and  row[1] != '':
            #        if row[3] != '' and i>external_end_row:
            #             table_data[i][1] = round((table_data[i][1]))
            #             table_data[i][5] = round(row[1]*row[3])
            #             internal_budget_total +=(float(table_data[i][1])*float(table_data[i][3]))
            #             table_data[i][6] = round(float(table_data[i][1])*float(table_data[i][4]))
            #             internal_actual_total +=(float(table_data[i][1])*float(table_data[i][4]))
            #             table_data[i][7] = round(float(table_data[i][3])-float(table_data[i][4]),2)
            #             internal_unit_total +=(float(table_data[i][3])-float(table_data[i][4]))
            #             table_data[i][8] = round(float(table_data[i][5])-float(table_data[i][6]))
            #             internal_amount_total +=(float(table_data[i][5])-float(table_data[i][6]))

            table_data.append(["Internal Total", round(internal_total), "","","",round(internal_budget_total),round(internal_actual_total),'',round(internal_amount_total),'',0,0,0])
            EI_total = external_total + internal_total
            # budget_total = internal_budget_total + external_budget_total
            # actual_total = internal_actual_total + external_actual_total
            # unit_total = internal_unit_total + external_unit_total
            # amount_total = internal_amount_total + external_amount_total
            table_data.append(["Grand Total",round(EI_total), "","","",0,0,'',0,'',0,0,0])
            mix_unit_total=0
            mix_amount_total=0
            total_amount_total=0
            total_total = 0
            amount_total = 0
            unit_total = 0

            for i, row in enumerate(table_data):
                if row[0] == 'TNEB':
                    table_data[i][1] = round(tneb_consumption)
                    table_data[i][2] = round((tneb_consumption/EI_total)* 100,2)
                if row[0] != 'Energy Source'  and row[1]!="Total Amount":
                    if row[1] != '':
                        numeric_value = float(row[1])
                        table_data[i][2] = round((numeric_value / EI_total) * 100,2)
                        if row[3] != '':
                            
                            table_data[i][5] = round(row[1]*row[3])
                            ie_budget_total +=row[1]*row[3]
                            table_data[i][6] = round(row[1]*row[4])
                            ie_actual_total +=row[1]*row[4]
                            table_data[i][7] = round(row[3]-row[4])
                            ie_unit_total +=row[3]-row[4]
                            table_data[i][8] = round(row[5]-row[6])
                            ie_amount_total +=row[5]-row[6]
                            table_data[i][1] = round(row[1])

                            table_data[i][10] = round((row[9]/100)*EI_total)
                            mix_unit_total +=(row[9]/100)*EI_total
                            unit_total +=(row[9]/100)*EI_total
                            table_data[i][11] = round(row[10]*row[4])
                            mix_amount_total +=row[10]*row[4]
                            amount_total +=row[10]*row[4]
                            table_data[i][12] = round(row[11]-row[6])
                            total_amount_total +=row[11]-row[6]
                            total_total +=row[11]-row[6]

                    if row[0]=="External Total" or row[0]=="Internal Total":
                            table_data[i][5] = round(ie_budget_total)
                            table_data[i][6] = round(ie_actual_total)
                            # table_data[i][7] = ie_unit_total
                            table_data[i][8] = round(ie_amount_total)
                            table_data[i][10] = round(mix_unit_total)
                            table_data[i][11] = round(mix_amount_total)
                            table_data[i][12] = round(total_amount_total)
                            
                            mix_unit_total = 0
                            mix_amount_total = 0
                            total_amount_total = 0
                            g_budget_total +=ie_budget_total 
                            g_actual_total +=ie_actual_total 
                            g_unit_total +=ie_unit_total 
                            g_amount_total +=ie_amount_total
                            ie_budget_total = 0
                            ie_actual_total = 0
                            ie_unit_total = 0
                            ie_amount_total =0

                    if row[0]=="Grand Total":
                        table_data[i][5] = round(g_budget_total)
                        table_data[i][6] = round(g_actual_total)
                        # table_data[i][7] = g_unit_total
                        table_data[i][8] = round(g_amount_total)
                        table_data[i][10] = round(unit_total)
                        table_data[i][11] = round(amount_total)
                        table_data[i][12] = round(total_total)

                    
            table_style = [
            ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor("#1e98a3")),  
            ('TEXTCOLOR', (0, 0), (-1, 1), colors.white), 
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
            ('FONTSIZE', (0, 0), (-1, -1), 5.7),  
            ('FONTWEIGHT', (0, 0), (-1, -1), 'BOLD'),  
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  
            ('GRID', (0, 0), (-1, -1), 1, colors.black), 
            ('SPAN', (3, 0), (4, 0)),  
            ('SPAN', (5, 0), (6, 0)),  
            ('SPAN', (7, 0), (8, 0)),  
            ('SPAN', (9, 0), (10, 0)),  
            ('SPAN', (10, 0), (11, 0)),  
            ('SPAN', (11, 0), (12, 0)),  
            ]
        else:
            table_data = [
                ["Energy Source", "Units", "%"],
                ["External", "", ""]
            ]

            for row in data["data"]:
                energy_source_name = row["energy_source_name"]
                consumption = row["consumption"]
                e_budget = row["budget"]
                e_actual = row["actual"]
                
                table_data.append([energy_source_name, round(consumption), ""])

                if energy_source_name != "TNEB" and energy_source_name != "EB":
                    bal_consumption += consumption

                if energy_source_name == "TNEB" or energy_source_name == "EB":
                    external_total = consumption

            tneb_consumption = external_total - bal_consumption

            table_data.append(["External Total", round(external_total), ""])
            table_data.append(["Internal", '', ""])
            for internal in data["data3"]:
                source = internal["source"]
                kwh = internal["kWh"]
                table_data.append([source, round(kwh), ""])
                internal_total += kwh

            
            table_data.append(["Internal Total", round(internal_total), ""])
            EI_total = external_total + internal_total
            table_data.append(["Grand Total",round(EI_total), ""])

            for i, row in enumerate(table_data):
                if row[0] == 'TNEB':
                    table_data[i][1] = round(tneb_consumption)
                    table_data[i][2] = round((tneb_consumption/EI_total)* 100,2)
                if row[0] != "Energy Source":
                    if row[1] != '':
                        numeric_value = float(row[1])
                        table_data[i][2] = round((numeric_value / EI_total) * 100,2)
            table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e98a3")),  
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), 
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
            ('FONTSIZE', (0, 0), (-1, -1), 9),  
            ('FONTWEIGHT', (0, 0), (-1, -1), 'BOLD'),  
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  
            ('GRID', (0, 0), (-1, -1), 1, colors.black), 
            
            ]
        for i, row in enumerate(table_data):
            if row[0] == "External Total" or row[0] == "Internal Total":
                table_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#0710ba')))
            if row[0] == "Grand Total":
                table_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#0710ba')))
            if row[0] == "Internal" or row[0] == "External":
                table_style.append(('BACKGROUND', (0, i), (0, i), colors.HexColor('#1e98a3')))
                table_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#FFFFFF')))

        for i, row in enumerate(table_data):
            for j, value in enumerate(row):
                if j < 1:  
                    continue

                if isinstance(value, str):
                    continue  

                if value < 0:
                    table_style.append(('TEXTCOLOR', (j, i), (j, i), colors.red))

                                
        table = Table(table_data)
        table.setStyle(TableStyle(table_style))
        
        content.append(table)

        content.append(Paragraph("<br/><br/>", getSampleStyleSheet()["Normal"]))
        table_data2 = []
        table_data2 = [
            ["   Plant   ", "Reporting Department", "Units (EB+DG)", ' Utility ', '', ' Dis Loss ', '', 'Total units', '','Dispatch T','','', 'Bud U/T', 'Act U/T'],
            ["        ", "", "", 'Units', ' %  ', 'Units', '%', '','Cust + WH','Within Plant','To Other Plants','Total', '', '']
        ]
        
    
        
        current_plant = None
        current_total = 0
        unit_total = 0
        current_com_total = 0
        grand_total = 0
        grand_com_total = 0
        corporate_total = 0
        grand_total_total = 0
        for i in data["data2"]:
            plant_name = i["plant_name"]
            reporting_department = i["reporting_department"]
            pm_kwh = i["pm_kwh"]
            pm_common_kwh = i["pm_common_kwh"]
            budget = i["budget"]
            is_corporate = i["is_corporate"]
            
            # Check if the plant name has changed
            if plant_name != current_plant:
                # If it's not the first iteration, add a "Total" row for the previous plant
                if current_plant is not None:
                    table_data2.append(["", "Total", current_total, 0,0, 0,0,0])
                    grand_com_total += current_com_total
                    current_total = 0
                    current_com_total = 0
                    
                current_plant = plant_name
                table_data2.append([plant_name, reporting_department, pm_kwh, 0,0, 0, 0,0, 0,0,0,0,round(budget),0])
                current_total += pm_kwh
                grand_total += pm_kwh
                current_com_total = pm_common_kwh
            else:
                if is_corporate == 'no':
                    
                    table_data2.append(['', reporting_department, pm_kwh, 0,0, 0, 0,0, 0,0,0,0,round(budget),0])
                    current_total += pm_kwh
                    grand_total += pm_kwh
                
                    current_com_total = pm_common_kwh

        if current_plant is not None:
            table_data2.append(["", "Total", current_total, 0,'', 0, "",0,'' ,""])
            grand_com_total += current_com_total
            print("grand_com_total",grand_com_total)
            table_data2.append(["", "Grand Total", grand_total, grand_com_total,'', 0, "", ""])
            for row in table_data2:
                if row[1] != "Reporting Department" and row[3] != 'Units':
                    row[3] = round((row[2] / grand_total) * grand_com_total)
                    row[7] = round(row[2] +row[3]+ row[5])
        
        for i in data["data2"]: 
            plant_name = i["plant_name"]
            reporting_department = i["reporting_department"]
            pm_kwh = i["pm_kwh"]
            pm_common_kwh = i["pm_common_kwh"]
            budget = i["budget"]
            is_corporate = i["is_corporate"] 
            if is_corporate == 'yes':      
                table_data2.append(["", reporting_department, '', "",'', '', "", round(pm_kwh),""])
                corporate_total +=pm_kwh
                table_data2.append(["", "Total", '', "",'', '', "", round(corporate_total),""])
                table_data2.append(["", "Grand Total", '', "",'', '', "", round(unit_total),""])
        for row in table_data2:
            if row[1] == "Grand Total" and row[2]!= '':
                
                total_dis_loss =EI_total - grand_total - grand_com_total - corporate_total
                print(total_dis_loss)
                row[5] = round(total_dis_loss)
                
        for row in table_data2:
            if row[1] != "Reporting Department" and row[3] != 'Units':
                if row[2]=='':
                    break
                else:
                    if row[1]!='Grand Total':   
                        row[5] = round((row[2] / grand_total) * total_dis_loss)
                    row[7] = row[2] +row[3]+ row[5]
                if row[1]=='Grand Total' and row[2]!='':
                    grand_total_total = row[7]

        for row in table_data2 :
            if row[1] != "Reporting Department" and row[2] != '':
                row[6]=round((row[5]/grand_total)*100,2)
                row[4]=round((row[3]/grand_total)*100,2)
                row[2] = round(row[2])
                row[5] = round(row[5])
                row[3] = round(row[3])
                row[7] = round(row[7])
            if row[1]!='Total' and row[1] != 'Grand Total'and row[2] != ''and row[1] != "Reporting Department":
                if row[8]!=0:
                    row[10] = round(row =[7]/row[8])

            if row[1] =='Grand Total' and row[2] == '':
                row[7] = round(grand_total_total+corporate_total)
        table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e98a3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FFFFFF')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.8),  
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.black), 
        ('SPAN', (3, 0), (4, 0)),  
        ('SPAN', (5, 0), (6, 0)),  
        ('SPAN', (8, 0), (9, 0)),  
        ('SPAN', (9, 0), (10, 0)),  
        ('SPAN', (10, 0), (11, 0)),  
    ]
        
        for i, row in enumerate(table_data2):
            if row[1] == "Total" or row[1] == "Grand Total":
                table_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#0710ba')))
        for i, row in enumerate(table_data2):
            for j, value in enumerate(row):
                if j < 2:  
                    continue

                if isinstance(value, str):
                    continue  

                if value < 0:
                    table_style.append(('TEXTCOLOR', (j, i), (j, i), colors.red))

        
    else:
        table_data2 = []     
        table_data2 = [["NoData"]]  
        table_style = [
        
       ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Set text color of all cells to black
    ('FONTSIZE', (0, 0), (-1, -1), 20) 
    ] 
    table2 = Table(table_data2)
    table2.setStyle(table_style)  
    content.append(table2)   
    doc.build(content)

async def manual_mail(data, campus_id,report_type,cnx):
    if report_type == 'with_rate':
        report = "Energy Statement With Tariff"
    else:
        report = "Energy Statement"
    sql = f'''SELECT  * FROM ems_v1.master_mail where campus_id = '{campus_id}' and report = '{report}' '''
    print(sql)
    data1 = await cnx.execute(sql)
    automail_list = data1.fetchall()
    if len(automail_list)>0  :
     for rows in automail_list:
        from_mail = rows["from_mail"]
        password = rows["pwd"]
        to_mail = rows["to_mail"]
        cc_mail = rows["cc_mail"]
        bcc_mail = rows["bcc_mail"]
        subject = rows["subject"]
        compose_textarea = rows["compose_textarea"]
        compose_textarea1 = rows["compose_textarea1"]
        compose_textarea2 = rows["compose_textarea2"]
        location = rows["report_location"] 
        campus_id = rows["campus_id"]        
        if data!=None:
            mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}                
            
            try:
                # Create a multipart message container
                message = MIMEMultipart()
            
                message['From'] = from_mail
                message['Subject'] = subject
                print(from_mail)
                attachment = location+'availability_report.pdf'
                attachment_name = os.path.basename(attachment)
                with open(attachment, 'rb') as file:
                    attachment_part = MIMEApplication(file.read())
                    attachment_part.add_header('Content-Disposition', 'attachment', filename=attachment_name)
                    message.attach(attachment_part)
                
                email_content = compose_textarea + '<br><br>'  +compose_textarea1 + '<br><br>'+ compose_textarea2

            
                html_part = MIMEText(email_content, 'html')
                message.attach(html_part)

                # Connect to the SMTP server
                smtp_server = "chnbmsgw3.netaccess-india.com"
                smtp_port = 587
        
                with smtplib.SMTP(smtp_server, smtp_port) as smtp:
                    smtp.starttls()
                    smtp.login(from_mail, password)
                    message['To'] = to_mail
                    message['Cc'] = cc_mail
                    recipients = to_mail.split(',') + cc_mail.split(',')
                    smtp.send_message(message, from_mail, recipients)
                createFolder(f"Log/",f"Mail Send Sucessfully!")  
            except Exception as e:
                createFolder("ErrorLog/", "Issue in Sending Mail " + str(e))
    else:
        createFolder(f"Log/",f"API response is empty")  
    
@router.post("/get_availability_report/", tags=["Report"])
async def get_availability_report(request: Request,
                                 campus_id: str = Form(""), 
                                 from_date: str = Form(""),
                                 to_date: str = Form(""),
                                 month: str = Form(""),
                                 report_for: str = Form(""),
                                 report_type: str = Form(""),
                                 send_mail: str = Form(""),
                                 report_method: str = Form(""),
                                 user_login_id: str = Form(""),
                                 cnx: AsyncSession = Depends(get_db)):
    try:

        if campus_id == '':
            return _getErrorResponseJson("Campud ID is Required")
        employee_name = ''
        sql = f"select employee_name from master_employee where employee_id = '{user_login_id}'"
        employee = await cnx.execute(sql)
        employee = employee.mappings().all()
        if len(employee)>0:
            for row in employee:
                employee_name = row["employee_name"]
            print(employee_name)
        start_date = from_date
        end_date = to_date
        
        if report_type == "with_rate":
            if month == '':
                return _getErrorResponseJson("Month is Required")
            from_date = await parse_date(month)
            to_date = from_date.replace(day=1, month=from_date.month + 1)
            to_date = to_date - timedelta(days=1)
            
        else:
            if from_date == '':
                return _getErrorResponseJson("from_date is Required")
        
            if to_date == '':
                return _getErrorResponseJson("to_date is Required")
            from_date = await parse_date(from_date)
            to_date = await parse_date(to_date)

        from_year = from_date.year
        to_year = to_date.year

        month_year_range = [
                    (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                    for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                ]
        union_queries = []
        error_messages = []
      
        current_date = from_date
        while current_date <= to_date:
            month_year = current_date.strftime("%m%Y")
           
            
            # Your function to check if there's data for this date in the database
            result_query = await check_power_table(cnx, month_year)
            
            if len(result_query) > 0:
                table_name = f"ems_v1_completed.power_{month_year}"
                union_queries.append(f"SELECT * FROM {table_name}")
            
                sql = f"select * from master_source_entry_date where mill_date = '{current_date}' and campus_id = '{campus_id}' "
                c_data = await cnx.execute(sql)
                c_data = c_data.mappings().all()
                if len(c_data)>0:
                    pass
                else:
                    date = current_date
                    date = date.strftime("%d-%m-%Y")
                    error_messages.append(f"{date}")
                    
            current_date += timedelta(days=1)
        if error_messages != []:
            error_data = ", ".join(error_messages)
            return _getErrorResponseJson(f"Kindly enter the EB consumption for the following days - ({error_data})")
                
        if len(union_queries) == 0:
            return _getErrorResponseJson("Power table not available...")

        data = await availability_report(cnx,campus_id,from_date,to_date,from_year,to_year,report_type)
        results = ''
        if report_for == "pdf":
            availabilityreport_pdf(data, from_date, to_date,report_type,employee_name,month,report_method)
            file_name = f'availability_report.pdf'
            results = f"http://{request.headers['host']}/attachments/{file_name}"
        else:
            availabilityreport_excel(data,start_date,end_date,report_type,employee_name,month,report_method)
            file_name = f'availability_report.xlsx'
            results = f"http://{request.headers['host']}/attachments/{file_name}"
        
        if send_mail == "yes":
            sql = f"select * from master_mail where campus_id = '{campus_id}'"
            mail_data = await cnx.execute(sql)
            mail_data = mail_data.mappings().all()
            if  len(mail_data)==0:
                return _getErrorResponseJson("Kindly Enter the Mail details for this Campus...")
            
            await manual_mail(data,campus_id,report_type,cnx)

            response = {
                        "iserror": False,
                        "message": "Mail Send Sucessfully!",
                        "file_url":''
                    }
        else:
            response = {
                        "iserror": False,
                        "message": "Data Returned Successfully.",
                        "file_url":results
                    }

        return response
    except Exception as e:
        return get_exception_response(e) 
    
@router.post("/get_tneb_report/", tags=["Report"])
async def get_tneb_report(request: Request,
                          campus_id: str = Form(""), 
                          period_id: str = Form(""), 
                          from_date: str = Form(""),
                          to_date: str = Form(""),
                          cnx: AsyncSession = Depends(get_db)):
    try:

        if campus_id == '':
            return _getErrorResponseJson("Campus ID is Required")
         
        if period_id == '':
            return _getErrorResponseJson("Period is Required") 
        
        if from_date == '':
            return _getErrorResponseJson("From Date is Required") 
        
        if period_id == 'from_to':
            if to_date == '':
                return _getErrorResponseJson("To Date is Required") 
        data = await tnebreportdetail(cnx,campus_id,period_id,from_date,to_date)
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "data":data
                }

        return response
    except Exception as e:
        return get_exception_response(e) 

@router.post("/get_avg_demand/", tags=["Report"])
async def get_avg_demand(request: Request,
                         campus_id: str = Form(""), 
                         meter_id: str = Form(""), 
                         plant_id: str = Form(""), 
                         main_demand_meter: str = Form(""), 
                         period_id: str = Form(""), 
                         from_date: str = Form(""),
                         to_date: str = Form(""),
                         cnx: AsyncSession = Depends(get_db)):
    try:

        if main_demand_meter != 'yes':
            if meter_id == '':
                return _getErrorResponseJson("Meter Id is Required") 
            
        if period_id == '':
            return _getErrorResponseJson("Period is Required") 
        
        if period_id !='#this_month':
            return _getErrorResponseJson("Invalid Period") 
        
        data1 = await shift_Lists(cnx, '',plant_id, '', '')
       
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                no_of_shifts = shift_record["no_of_shifts"]  
       
        if period_id == "#this_month":
            from_date = mill_date.replace(day=1)
            to_date = mill_date 
        
        data = await avg_demand_report(cnx,campus_id,period_id,from_date,to_date,meter_id,main_demand_meter)
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "data":data
                }

        return response
    except Exception as e:
        return get_exception_response(e) 

@router.post("/get_minmax_dtl/", tags=["Report"])
async def get_minmax_dtl(company_id : str = Form(''),
                         bu_id :str = Form(''),
                         plant_id :str = Form(''),
                         plant_department_id :str = Form(''),
                         equipment_group_id :str = Form(''),
                         equipment_id :str = Form(''),
                         meter_id : str = Form (''),
                         groupby : str = Form(''),
                         period_id: str = Form(''),
                         from_date: str = Form(''),
                         to_date: str = Form(''),   
                         from_year: int  = Form(""),                   
                         shift_id: str = Form(''),                           
                         reportfor: str = Form(''),                           
                         for_android: str = Form(''),                           
                         cnx: AsyncSession = Depends(get_db)):
    try:
      
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        
        if period_id == "":
            return _getErrorResponseJson("Period Id Is Required")
        
        if groupby == "":
            return _getErrorResponseJson("Groupby Is Required")
        
        if groupby not in ['plant_department','equipment_group','equipment','meter']:
            return _getErrorResponseJson("Invalid Groupby ")
        
        mill_date = date.today()
        mill_shift = 0
        no_of_shifts = 3

        data1 = await shift_Lists(cnx, '',plant_id, bu_id, company_id)
        # query = text(f'''SELECT * FROM master_shifts WHERE status = 'active' and  plant_id = '{plant_id}' ''')
        
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                no_of_shifts = shift_record["no_of_shifts"]  
    
        if reportfor == '12to12':
            if period_id != 'sel_date' and period_id != 'from_to':
                return _getErrorResponseJson("invalid period id") 
            
            if period_id == "sel_date":  
                if from_date == '':
                    return _getErrorResponseJson("from date is required") 
                    
                from_date =  await parse_date(from_date)   
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
                res = await check_power_12_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("12to12 table not available...") 
                
            if period_id == "from_to" or period_id == "#from_to":            
                if from_date == '':
                    return _getErrorResponseJson("from date is required")
                if to_date == '':
                    return _getErrorResponseJson("to_date is required") 
            
                from_date = await parse_date(from_date)
                to_date = await parse_date(to_date) 
                month_year_range = [
                    (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                    for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                ]
                union_queries = []

                for month_year in month_year_range:
                    res = check_power_12_table(cnx,month_year)
                    
                    if len(res)>0:
                        table_name = f"ems_v1_completed.power_{month_year}_12"
                        union_queries.append(f"{table_name}")

                if len(union_queries) == 0:
                    return _getErrorResponseJson("12to12 table not available...")  
                
        else: 
                  
            if period_id == "sel_date" or period_id == 'sel_shift':            
                if from_date == '':
                    return _getErrorResponseJson("From Date Is Required") 
                
                from_date =  await parse_date(from_date) 
                 
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}""" 
                res = await check_power_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("Power Table Not Available...")   
                
                if period_id == "sel_shift":                  
                    if shift_id == '':
                        return _getErrorResponseJson("Shift Id Is Required") 
                    
            elif period_id == "#previous_shift" or period_id == "#previous_day":  
                if period_id == "#previous_shift":               
                    if int(mill_shift) == 1:
                        shift_id = no_of_shifts
                        from_date = mill_date - timedelta(days=1)
                    else:
                        shift_id = int(mill_shift) - 1
                        from_date = mill_date 

                elif period_id == "#previous_day":             
                    from_date = mill_date - timedelta(days=1)
                
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                res = await check_power_table(cnx,month_year)
                if len(res) == 0:
                    return _getErrorResponseJson("Power Table Not Available...")   
                       
            elif period_id == "from_to" or period_id == "#from_to":            
                if from_date == '':
                    return _getErrorResponseJson("From Date Is Required")
                if to_date == '':
                    return _getErrorResponseJson("To Date Is Required")  
                from_date = await parse_date(from_date)
                to_date = await parse_date(to_date)
              
            if period_id == '#previous_week' or period_id == "#this_week" or period_id == "#this_month" or period_id == '#previous_month' or period_id=="#previous_year" or period_id=="#this_year" or period_id=="from_to" or period_id == "#sel_year" or period_id=="#from_to" :
                if period_id  == "#this_week":
                    dt = mill_date
                    from_date=dt-timedelta(dt.weekday()+1)
                    to_date = mill_date

                elif period_id == "#previous_week":
                    dt = mill_date
                    current_week_start = dt - timedelta(days=dt.weekday())  
                    from_date = current_week_start - timedelta(weeks=1)  
                    to_date = from_date + timedelta(days=5)

                elif period_id == "#this_month":
                    from_date = mill_date.replace(day=1)
                    to_date = mill_date

                elif period_id == "#previous_month":
                    from_date = mill_date.replace(day=1)                   
                    from_date = (from_date - timedelta(days=1)).replace(day=1)
                    to_date = from_date + timedelta(days=30)   

                elif period_id=="#this_year": 
            
                    from_date = mill_date.replace(day=1,month=1) 
                    to_date = mill_date  
                    

                elif period_id=="#previous_year": 
                    from_date = mill_date.replace(day=1, month=1, year=mill_date.year - 1)
                    to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
            
                elif period_id == "#sel_year": 

                    if from_year == '':
                        return _getErrorResponseJson("From Year Is Required")
                        
                    from_date = mill_date.replace(day=1, month=1, year=from_year)
                    
                    to_date = from_date.replace(day=1, month=12) + timedelta(days=30)
                   
            
                if from_date != '' and to_date != '':

                    month_year_range = [
                        (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                    union_queries = []
                    joins = []

                    for month_year in month_year_range:
                        res = await check_power_table(cnx,month_year)
                        if len(res)>0:
                            table_name = f"ems_v1_completed.power_{month_year}"
                            union_queries.append(f"{table_name}")

                    if len(union_queries) == 0:
                        return _getErrorResponseJson("Power Table Not Available...")     

        data = await  minmax_kwh_dtl(cnx, plant_department_id ,equipment_group_id ,equipment_id,meter_id  ,groupby ,period_id,from_date,to_date,shift_id,reportfor) 
    
              
        if for_android == 'yes':
            response = [{
                "iserror": False,
                "message": "Data Returned Successfully.",
                "data": data
            }]
        else:
            response = {
                "iserror": False,
                "message": "Data Returned Successfully.",
                "data": data
            }
              
        return response
    except Exception as e:
        return get_exception_response(e)

@router.post("/manualentry_history_report/", tags=["Report"])
async def manualentry_history_report(campus_id: str = Form(""), 
                                     company_id: str = Form(""),  
                                     bu_id: str = Form(""),  
                                     plant_id: str = Form(""),  
                                     meter_id: str = Form(""),  
                                     from_date: str = Form(""),
                                     to_date: str = Form(""),
                                     cnx: AsyncSession = Depends(get_db)):
    try:
        if from_date == '':
            return _getErrorResponseJson("From Date is required") 
        
        if to_date == '':
            return _getErrorResponseJson("To Date is required") 
 
                    
        data = await manualentryhistory(cnx,campus_id,company_id,bu_id,plant_id,meter_id,from_date,to_date)
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "data":data
                }

        return response
    except Exception as e:
        return get_exception_response(e) 

async def transformerlossreportexcel(data,from_date,to_date,employee_name,campus_id):
    month_year=f"""{mill_month[from_date.month]}-{str(from_date.year)}""" 

    from_date = from_date.strftime("%d-%b-%Y").upper()
    to_date = to_date.strftime("%d-%b-%Y").upper()
    wb = Workbook()
    ws = wb.active
    border_style = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    current_time = datetime.now()
    current_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    print(current_time)
    if data == []:
        
       
        cell = "O10"
        data = "No Data"
        ws[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment

        ws.column_dimensions[cell[0]].width = len(data) + 2  

        font = Font(name='Calibri', size=25)
        ws[cell].font = font
    else:
        cell = "B2"
        
        Heading = f"Main Transformer wise Distribution Loss Report {from_date} TO {to_date}"
        ws[cell] = Heading
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill
        
        cell = "B7"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        ws.column_dimensions[cell[0]].width = 5  
        
        cell = "C7"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill
        ws.column_dimensions[cell[0]].width = 25  

        cell = "B4"
        ws[cell]= f"User:{employee_name}"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=12, bold=True)
        ws[cell].font = font
        ws.column_dimensions[cell[0]].width = 25  
        current_date = datetime.now()
        current_date = current_date.strftime("%d-%b-%Y %H:%M:%S").upper()

        cell = "B6"
        ws[cell]= f"Date&Time:{current_date}"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=12, bold=True)
        ws[cell].font = font
        ws.column_dimensions[cell[0]].width = 25  
        
        month_year = "03-2024"  # Example month_year
        month, year = map(int, month_year.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        current_plant_name = ''
        i=0
        cols = 0
        grand_row_kwh_value = 0
        grand_row_cal_kwh_value = 0
        grand_row_total_percentage_value = 0
        grand_col_cal_kwh_value = 0
        grand_col_kwh_value = 0
        grand_col_total_percentage_value = 0
        headers = ['', "slno","date"]
        for item in data:
            meter_code = item["meter_code"]
            meter_total_kwh = 0 
            

            if item["plant_name"] != current_plant_name and current_plant_name != '':
                header_text = f"Total"
                ws.merge_cells(start_row=7, start_column=4 + i * 3, end_row=7, end_column=6 + i * 3)
                header_cell = ws.cell(row=7, column=4 + i * 3)
                header_cell.value = header_text

                # Apply styles to the header cell
                header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
                header_cell.font = Font(bold=True, name='Calibri', size=10)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                i+=1
                # set_row = i * 3
                # cols = set_row+4
                headers.extend(["kwh", "dist_loss", "%"])
                # ws.append(headers)
            

            header_text = meter_code
            ws.merge_cells(start_row=7, start_column=4 + i * 3, end_row=7, end_column=6 + i * 3)
            header_cell = ws.cell(row=7, column=4 + i * 3)
            header_cell.value = header_text

            header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
            header_cell.font = Font(bold=True, name='Calibri', size=9,color = "FFFFFF")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            headers.extend(["kwh", "dist_loss", "%"])
            # ws.append(headers)
            i+=1
            current_plant_name = item["plant_name"]

        header_text = f"Total"
        ws.merge_cells(start_row=7, start_column=4 + i * 3, end_row=7, end_column=6 + i * 3)
        header_cell = ws.cell(row=7, column=4 + i * 3)
        header_cell.value = header_text

        # Apply styles to the header cell
        header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        header_cell.font = Font(bold=True, name='Calibri', size=10)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        set_row = i * 3
        cols = set_row+4
        headers.extend(["kwh", "dist_loss", "%"])

        header_text = f"Grand Total"
        ws.merge_cells(start_row=7, start_column=7 + i * 3, end_row=7, end_column=9 + i * 3)
        header_cell = ws.cell(row=7, column=7 + i * 3)
        header_cell.value = header_text

        # Apply styles to the header cell
        header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        header_cell.font = Font(bold=True, name='Calibri', size=10)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        set_row = i * 3
        cols = set_row+7
        
        headers.extend(["kwh", "dist_loss", "%"])
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):  
            cell = ws.cell(row=8, column=col)
            cell.font = Font(bold=True, name='Calibri', size=10)
        
        for row in range(1, days_in_month + 1):
            mill_date = f"{row:02d}-{month_year}"
           
            row_data = ['', row, mill_date]
            # ws.append(row_data)
            current_plant_name = ''
            row_kwh_value = 0
            row_cal_kwh_value = 0
            row_total_percentage_value = 0
            grand_row_kwh_value = 0
            grand_row_cal_kwh_value = 0
            grand_row_total_percentage_value = 0
            row_number = 5
            for item in data:
                meter_code = item["meter_code"]
                campus_name = item["campus_name"]
                meter_total_kwh = 0 
                
                if item["plant_name"] != current_plant_name and current_plant_name != '':
                    row_data.append(round(row_kwh_value))
                    row_data.append(round(row_cal_kwh_value))
                    row_data.append(row_total_percentage_value)
                    row_kwh_value = 0
                    row_cal_kwh_value = 0
                    cell = ws.cell(row=row, column=len(row_data))  
                    cell.font = Font(bold=True, color='0710ba', name='Calibri')
                    cell = ws.cell(row=row, column=len(row_data)-1)  
                    cell.font = Font(bold=True, color='0710ba', name='Calibri')
                    cell = ws.cell(row=row, column=len(row_data)-2)  
                    cell.font = Font(bold=True, color='0710ba', name='Calibri')
                current_plant_name = item["plant_name"]

                kwh_key = f"d{row}"
                cal_kwh_key = f"cal_kwh_d{row}"
                if kwh_key in item and cal_kwh_key in item:
                    kwh_value = item[kwh_key]
                    cal_kwh_value = item[cal_kwh_key]
                    if kwh_value != 0 and kwh_value is not None:
                        percentage_value = round(cal_kwh_value / kwh_value * 100,2)
                    else:
                        percentage_value = 0  # Handle division by zero
                    row_data.append(round(kwh_value))
                    row_data.append(round(cal_kwh_value))
                    row_data.append(percentage_value)
                    row_kwh_value +=kwh_value
                    grand_row_kwh_value +=kwh_value
                    row_cal_kwh_value +=cal_kwh_value
                    grand_row_cal_kwh_value +=cal_kwh_value
                    if row_kwh_value != 0 and row_kwh_value is not None:
                        row_total_percentage_value = round(row_cal_kwh_value / row_kwh_value * 100,2)
                    else:
                        row_total_percentage_value = 0

                    if grand_row_kwh_value != 0 and grand_row_kwh_value is not None:
                        grand_row_total_percentage_value = round(grand_row_cal_kwh_value / grand_row_kwh_value * 100,2)
                    else:
                        grand_row_total_percentage_value = 0
                    
                    # meter_totals_kwh[meter_code] += kwh_value
                    # meter_totals_dist_loss[meter_code] += cal_kwh_value
                    row_number +=1
            row_data.append(round(row_kwh_value))
            row_data.append(round(row_cal_kwh_value))
            row_data.append(row_total_percentage_value)   
            cell = ws.cell(row=row, column=len(row_data))  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-1)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-2)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            
            row_data.append(round(grand_row_kwh_value))
            row_data.append(round(grand_row_cal_kwh_value))
            row_data.append(grand_row_total_percentage_value)   
            cell = ws.cell(row=row, column=len(row_data))  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-1)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-2)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            ws.append(row_data)
       
        if campus_id == '':
            campus_name = 'ALL'
        cell = "B5"
        ws[cell]= f"Campus:{campus_name}" 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=12, bold=True)
        ws[cell].font = font
        ws.column_dimensions[cell[0]].width = 25  
        current_plant_name = ''
        totals_row = ['', 'Total','']
        col_kwh_value = 0
        col_cal_kwh_value = 0
        col_total_percentage_value = 0
        for item in data:
            meter_code = item["meter_code"]
            campus_name = item["campus_name"]
            total_kwh = item["total_kwh"]
            total_cal = item["total_cal"]
            
            if item["plant_name"] != current_plant_name and current_plant_name != '':
                totals_row.append(round(col_kwh_value))
                totals_row.append(round(col_cal_kwh_value))
                totals_row.append(col_total_percentage_value)
                col_kwh_value = 0
                col_cal_kwh_value = 0
                

            current_plant_name = item["plant_name"]
            if total_kwh != 0 and total_kwh is not None:
                total_percentage_value = round(total_cal / total_kwh * 100,2)
            else:
                total_percentage_value = 0  
            totals_row.append(round(total_kwh))
            totals_row.append(round(total_cal))
            totals_row.append(total_percentage_value) 
            col_kwh_value  +=total_kwh
            grand_col_kwh_value +=total_kwh
            grand_col_cal_kwh_value +=total_cal
            col_cal_kwh_value +=total_cal
            if col_kwh_value != 0 and col_kwh_value is not None:
                col_total_percentage_value = round(col_cal_kwh_value / col_kwh_value * 100,2)
            else:
                col_total_percentage_value = 0

            if grand_col_kwh_value != 0 and grand_col_kwh_value is not None:
                grand_col_total_percentage_value = round(grand_col_cal_kwh_value / grand_col_kwh_value * 100,2)
            else:
                grand_col_total_percentage_value = 0

        totals_row.append(round(col_kwh_value))
        totals_row.append(round(col_cal_kwh_value))
        totals_row.append(col_total_percentage_value)
        totals_row.append(round(grand_col_kwh_value))
        totals_row.append(round(grand_col_cal_kwh_value))
        totals_row.append(grand_col_total_percentage_value)
        ws.append(totals_row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
                cell.border = border
          
        ws.merge_cells(start_row=2, start_column=2 , end_row=3, end_column=cols+2)
        ws.merge_cells(start_row=4, start_column=2 , end_row=4, end_column=cols+2)
        ws.merge_cells(start_row=5, start_column=2 , end_row=5, end_column=cols+2)
        ws.merge_cells(start_row=6, start_column=2 , end_row=6, end_column=cols+2)
        
        for col in range(1, cols+3): 
            cell = ws.cell(row=7, column=col)
            cell.font = Font(bold=True, name='Calibri', size=12)
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
            column_letter = openpyxl.utils.get_column_letter(col)

            ws.column_dimensions[column_letter].width = 15
        
            cell = ws.cell(row=8, column=col)
            cell.font = Font(bold=True, name='Calibri', size=12)
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment

        max_row=ws.max_row 
        for col in range(1, cols): 
            cell = ws.cell(row=max_row, column=col)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        
        for row in range(9,max_row+1): 
            cell = ws.cell(row=row, column=cols)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
            cell = ws.cell(row=row, column=cols+1)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
            cell = ws.cell(row=row, column=cols+2)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment

        cell = "B2"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        for row_num in range(1, ws.max_row+2):
            for col_num in range(1, ws.max_column+2):

                cell_display_value = ws.cell(row=row_num, column=col_num).value

                if isinstance(cell_display_value, str) :
                    cell_display_value = ws.cell(row=row_num, column=col_num).value
                    # print(cell_display_value)
                if isinstance(cell_display_value, (int, float)) and cell_display_value < 0:
                    ws.cell(row=row_num, column=col_num).font = Font(color='FF0000',bold=True)

    file_name = f'MainTransformer_Wise_Distribution_Loss.xlsx'
    file_path = os.path.join(static_dir, file_name)
    wb.save(file_path)

@router.post("/transformer_wise_loss_report/", tags=["Report"])
async def transformer_wise_loss_report(request:Request,
                                       campus_id: str = Form(""),   
                                       from_date: str = Form(""),
                                       to_date: str = Form(""),
                                       user_login_id: str = Form(""),
                                       cnx: AsyncSession = Depends(get_db)):
    try:
        # if campus_id == '':
        #     return _getErrorResponseJson("Campus is required") 
        
        if from_date == '':
            return _getErrorResponseJson("From Date is required") 
        
        if to_date == '':
            return _getErrorResponseJson("To Date is required") 
        
        if user_login_id == '':
            return _getErrorResponseJson("User Login ID is required") 
        
        employee_name = ''
        sql = f"select employee_name from master_employee where employee_id = '{user_login_id}'"
        employee = await cnx.execute(sql)
        employee = employee.mappings().all()
        if len(employee)>0:
            for row in employee:
                employee_name = row["employee_name"]
        from_date = await parse_date(from_date)     
        to_date = await parse_date(to_date) 
        if from_date != '' and to_date != '':

            month_year_range = [
                (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
            ]
            union_queries = []
            for month_year in month_year_range:
                res = await check_power_table(cnx,month_year)
                if len(res)>0:
                    table_name = f"ems_v1_completed.power_{month_year}"
                    union_queries.append(f"{table_name}")

            if len(union_queries) == 0:
                return _getErrorResponseJson("Power Table Not Available...")     

        data = await transformerlossreport(cnx,campus_id,from_date,to_date)
        await transformerlossreportexcel(data,from_date,to_date,employee_name,campus_id)
        file_name = f"MainTransformer_Wise_Distribution_Loss.xlsx"
        results = f"http://{request.headers['host']}/attachments/{file_name}"
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "data":results
                }

        return response
    except Exception as e:
        return get_exception_response(e) 

async def submeterlossreportexcel(data,from_date,to_date,employee_name,campus_id):
    month_year=f"""{mill_month[from_date.month]}-{str(from_date.year)}""" 

    from_date = from_date.strftime("%d-%b-%Y").upper()
    to_date = to_date.strftime("%d-%b-%Y").upper()
    wb = Workbook()
    ws = wb.active
    border_style = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    current_time = datetime.now()
    current_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    print(current_time)
    if data == []:
        
       
        cell = "O10"
        data = "No Data"
        ws[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment

        ws.column_dimensions[cell[0]].width = len(data) + 2  

        font = Font(name='Calibri', size=25)
        ws[cell].font = font
    else:
        cell = "B2"
        
        Heading = f"Sub Transformer wise Distribution Loss Report {from_date} TO {to_date}"
        ws[cell] = Heading
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill
        
        cell = "B7"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill

        ws.column_dimensions[cell[0]].width = 5  
        
        cell = "C7"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=18, color='FFFFFF', bold=True)
        ws[cell].font = font
        fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        ws[cell].fill = fill
        ws.column_dimensions[cell[0]].width = 25  

        cell = "B4"
        ws[cell]= f"User:{employee_name}"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=12, bold=True)
        ws[cell].font = font
        ws.column_dimensions[cell[0]].width = 25  
        current_date = datetime.now()
        current_date = current_date.strftime("%d-%b-%Y %H:%M:%S").upper()

        cell = "B6"
        ws[cell]= f"Date&Time:{current_date}"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=12, bold=True)
        ws[cell].font = font
        ws.column_dimensions[cell[0]].width = 25  
        
        month_year = "03-2024"  # Example month_year
        month, year = map(int, month_year.split('-'))
        days_in_month = calendar.monthrange(year, month)[1]
        current_transformer_id = ''
        i=0
        cols = 0
        grand_row_kwh_value = 0
        grand_row_cal_kwh_value = 0
        grand_row_total_percentage_value = 0
        grand_col_cal_kwh_value = 0
        grand_col_kwh_value = 0
        grand_col_total_percentage_value = 0
        headers = ['', "slno","date"]
        for item in data:
            meter_code = item["meter_code"]
            meter_total_kwh = 0 
            

            if item["main_transformer_meter_id"] != current_transformer_id and current_transformer_id != '':
                header_text = f"Total"
                ws.merge_cells(start_row=7, start_column=4 + i * 3, end_row=7, end_column=6 + i * 3)
                header_cell = ws.cell(row=7, column=4 + i * 3)
                header_cell.value = header_text

                # Apply styles to the header cell
                header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
                header_cell.font = Font(bold=True, name='Calibri', size=10)
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                i+=1
                # set_row = i * 3
                # cols = set_row+4
                headers.extend(["kwh", "dist_loss", "%"])
                # ws.append(headers)
            

            header_text = meter_code
            ws.merge_cells(start_row=7, start_column=4 + i * 3, end_row=7, end_column=6 + i * 3)
            header_cell = ws.cell(row=7, column=4 + i * 3)
            header_cell.value = header_text

            header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
            header_cell.font = Font(bold=True, name='Calibri', size=9,color = "FFFFFF")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            headers.extend(["kwh", "dist_loss", "%"])
            # ws.append(headers)
            i+=1
            current_transformer_id = item["main_transformer_meter_id"]

        header_text = f"Total"
        ws.merge_cells(start_row=7, start_column=4 + i * 3, end_row=7, end_column=6 + i * 3)
        header_cell = ws.cell(row=7, column=4 + i * 3)
        header_cell.value = header_text

        # Apply styles to the header cell
        header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        header_cell.font = Font(bold=True, name='Calibri', size=10)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        set_row = i * 3
        cols = set_row+4
        headers.extend(["kwh", "dist_loss", "%"])

        header_text = f"Grand Total"
        ws.merge_cells(start_row=7, start_column=7 + i * 3, end_row=7, end_column=9 + i * 3)
        header_cell = ws.cell(row=7, column=7 + i * 3)
        header_cell.value = header_text

        # Apply styles to the header cell
        header_cell.fill = PatternFill(start_color='1e98a3', end_color='1e98a3', fill_type='solid')
        header_cell.font = Font(bold=True, name='Calibri', size=10)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        set_row = i * 3
        cols = set_row+7
        
        headers.extend(["kwh", "dist_loss", "%"])
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):  
            cell = ws.cell(row=8, column=col)
            cell.font = Font(bold=True, name='Calibri', size=10)
        
        for row in range(1, days_in_month + 1):
            mill_date = f"{row:02d}-{month_year}"
           
            row_data = ['', row, mill_date]
            # ws.append(row_data)
            current_transformer_id = ''
            row_kwh_value = 0
            row_cal_kwh_value = 0
            row_total_percentage_value = 0
            grand_row_kwh_value = 0
            grand_row_cal_kwh_value = 0
            grand_row_total_percentage_value = 0
            row_number = 5
            for item in data:
                meter_code = item["meter_code"]
                campus_name = item["campus_name"]
                meter_total_kwh = 0 
                
                if item["main_transformer_meter_id"] != current_transformer_id and current_transformer_id != '':
                    row_data.append(round(row_kwh_value))
                    row_data.append(round(row_cal_kwh_value))
                    row_data.append(row_total_percentage_value)
                    row_kwh_value = 0
                    row_cal_kwh_value = 0
                    cell = ws.cell(row=row, column=len(row_data))  
                    cell.font = Font(bold=True, color='0710ba', name='Calibri')
                    cell = ws.cell(row=row, column=len(row_data)-1)  
                    cell.font = Font(bold=True, color='0710ba', name='Calibri')
                    cell = ws.cell(row=row, column=len(row_data)-2)  
                    cell.font = Font(bold=True, color='0710ba', name='Calibri')
                current_transformer_id = item["main_transformer_meter_id"]

                kwh_key = f"d{row}"
                cal_kwh_key = f"cal_kwh_d{row}"
                if kwh_key in item and cal_kwh_key in item:
                    kwh_value = item[kwh_key]
                    cal_kwh_value = item[cal_kwh_key]
                    if kwh_value != 0 and kwh_value is not None:
                        percentage_value = round(cal_kwh_value / kwh_value * 100,2)
                    else:
                        percentage_value = 0  # Handle division by zero
                    row_data.append(round(kwh_value))
                    row_data.append(round(cal_kwh_value))
                    row_data.append(percentage_value)
                    row_kwh_value +=kwh_value
                    grand_row_kwh_value +=kwh_value
                    row_cal_kwh_value +=cal_kwh_value
                    grand_row_cal_kwh_value +=cal_kwh_value
                    if row_kwh_value != 0 and row_kwh_value is not None:
                        row_total_percentage_value = round(row_cal_kwh_value / row_kwh_value * 100,2)
                    else:
                        row_total_percentage_value = 0

                    if grand_row_kwh_value != 0 and grand_row_kwh_value is not None:
                        grand_row_total_percentage_value = round(grand_row_cal_kwh_value / grand_row_kwh_value * 100,2)
                    else:
                        grand_row_total_percentage_value = 0
                    
                    # meter_totals_kwh[meter_code] += kwh_value
                    # meter_totals_dist_loss[meter_code] += cal_kwh_value
                    row_number +=1
            row_data.append(round(row_kwh_value))
            row_data.append(round(row_cal_kwh_value))
            row_data.append(row_total_percentage_value)   
            cell = ws.cell(row=row, column=len(row_data))  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-1)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-2)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            
            row_data.append(round(grand_row_kwh_value))
            row_data.append(round(grand_row_cal_kwh_value))
            row_data.append(grand_row_total_percentage_value)   
            cell = ws.cell(row=row, column=len(row_data))  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-1)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            cell = ws.cell(row=row, column=len(row_data)-2)  
            cell.font = Font(bold=True, color='0710ba', name='Calibri') 
            ws.append(row_data)
       
        if campus_id == '':
            campus_name = 'ALL'
        cell = "B5"
        ws[cell]= f"Campus:{campus_name}" 
        ws[cell].border = border_style
        font = Font(name='Calibri', size=12, bold=True)
        ws[cell].font = font
        ws.column_dimensions[cell[0]].width = 25  
        current_transformer_id = ''
        totals_row = ['', 'Total','']
        col_kwh_value = 0
        col_cal_kwh_value = 0
        col_total_percentage_value = 0
        for item in data:
            meter_code = item["meter_code"]
            campus_name = item["campus_name"]
            total_kwh = item["total_kwh"]
            total_cal = item["total_cal"]
            
            if item["main_transformer_meter_id"] != current_transformer_id and current_transformer_id != '':
                totals_row.append(round(col_kwh_value))
                totals_row.append(round(col_cal_kwh_value))
                totals_row.append(col_total_percentage_value)
                col_kwh_value = 0
                col_cal_kwh_value = 0
                

            current_transformer_id = item["main_transformer_meter_id"]
            if total_kwh != 0 and total_kwh is not None:
                total_percentage_value = round(total_cal / total_kwh * 100,2)
            else:
                total_percentage_value = 0  
            totals_row.append(round(total_kwh))
            totals_row.append(round(total_cal))
            totals_row.append(total_percentage_value) 
            col_kwh_value  +=total_kwh
            grand_col_kwh_value +=total_kwh
            grand_col_cal_kwh_value +=total_cal
            col_cal_kwh_value +=total_cal
            if col_kwh_value != 0 and col_kwh_value is not None:
                col_total_percentage_value = round(col_cal_kwh_value / col_kwh_value * 100,2)
            else:
                col_total_percentage_value = 0

            if grand_col_kwh_value != 0 and grand_col_kwh_value is not None:
                grand_col_total_percentage_value = round(grand_col_cal_kwh_value / grand_col_kwh_value * 100,2)
            else:
                grand_col_total_percentage_value = 0

        totals_row.append(round(col_kwh_value))
        totals_row.append(round(col_cal_kwh_value))
        totals_row.append(col_total_percentage_value)
        totals_row.append(round(grand_col_kwh_value))
        totals_row.append(round(grand_col_cal_kwh_value))
        totals_row.append(grand_col_total_percentage_value)
        ws.append(totals_row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
                cell.border = border
          
        ws.merge_cells(start_row=2, start_column=2 , end_row=3, end_column=cols+2)
        ws.merge_cells(start_row=4, start_column=2 , end_row=4, end_column=cols+2)
        ws.merge_cells(start_row=5, start_column=2 , end_row=5, end_column=cols+2)
        ws.merge_cells(start_row=6, start_column=2 , end_row=6, end_column=cols+2)
        
        for col in range(1, cols+3): 
            cell = ws.cell(row=7, column=col)
            cell.font = Font(bold=True, name='Calibri', size=12)
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
            column_letter = openpyxl.utils.get_column_letter(col)

            ws.column_dimensions[column_letter].width = 15
        
            cell = ws.cell(row=8, column=col)
            cell.font = Font(bold=True, name='Calibri', size=12)
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment

        max_row=ws.max_row 
        for col in range(1, cols): 
            cell = ws.cell(row=max_row, column=col)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
        
        for row in range(9,max_row+1): 
            cell = ws.cell(row=row, column=cols)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
            cell = ws.cell(row=row, column=cols+1)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment
            cell = ws.cell(row=row, column=cols+2)
            cell.font = Font(bold=True, color='0710ba', name='Calibri')
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment

        cell = "B2"
        alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].alignment = alignment
        for row_num in range(1, ws.max_row+2):
            for col_num in range(1, ws.max_column+2):

                cell_display_value = ws.cell(row=row_num, column=col_num).value

                if isinstance(cell_display_value, str) :
                    cell_display_value = ws.cell(row=row_num, column=col_num).value
                    # print(cell_display_value)
                if isinstance(cell_display_value, (int, float)) and cell_display_value < 0:
                    ws.cell(row=row_num, column=col_num).font = Font(color='FF0000',bold=True)

    file_name = f'SubTransformer_Wise_Distribution_Loss.xlsx'
    file_path = os.path.join(static_dir, file_name)
    wb.save(file_path)
   
@router.post("/submeter_wise_loss_report/", tags=["Report"])
async def submeter_wise_loss_report(request:Request,
                                    campus_id: str = Form(""),   
                                    main_transformer_meter_id: str = Form(""),   
                                    from_date: str = Form(""),
                                    to_date: str = Form(""),
                                    user_login_id: str = Form(""),
                                    cnx: AsyncSession = Depends(get_db)):
    try:
        if campus_id == '':
            return _getErrorResponseJson("Campus is required") 
        
        if from_date == '':
            return _getErrorResponseJson("From Date is required") 
        
        if to_date == '':
            return _getErrorResponseJson("To Date is required") 
        
        if user_login_id == '':
            return _getErrorResponseJson("User Login ID is required") 
        
        employee_name = ''
        sql = f"select employee_name from master_employee where employee_id = '{user_login_id}'"
        employee = await cnx.execute(sql)
        employee = employee.mappings().all()
        if len(employee)>0:
            for row in employee:
                employee_name = row["employee_name"]
        from_date = await parse_date(from_date)     
        to_date = await parse_date(to_date) 
        if from_date != '' and to_date != '':

            month_year_range = [
                (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
            ]
            union_queries = []
            for month_year in month_year_range:
                res = await check_power_table(cnx,month_year)
                if len(res)>0:
                    table_name = f"ems_v1_completed.power_{month_year}"
                    union_queries.append(f"{table_name}")

            if len(union_queries) == 0:
                return _getErrorResponseJson("Power Table Not Available...")     

        data = await submeterlossreport(cnx,campus_id,main_transformer_meter_id,from_date,to_date)
        await submeterlossreportexcel(data,from_date,to_date,employee_name,campus_id)
        file_name = f"SubTransformer_Wise_Distribution_Loss.xlsx"
        results = f"http://{request.headers['host']}/attachments/{file_name}"
        response = {
                    "iserror": False,
                    "message": "Data Returned Successfully.",
                    "data":results
                }

        return response
    except Exception as e:
        return get_exception_response(e) 

