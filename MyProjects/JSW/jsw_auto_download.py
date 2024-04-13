
from typing import Optional, List, Dict, Any
from datetime import datetime, date

import traceback
import time
import pyodbc

from datetime import datetime
import os

from dateutil import parser
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pathlib import Path
import os   
 
     
from dateutil.relativedelta import relativedelta  
import re
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

import xlsxwriter
import sys

from datetime import timedelta

def createFolder(directory, data):
    date_time = datetime.datetime.now()
    curtime1 = date_time.strftime("%d/%m/%Y %H:%M:%S")
    curtime2 = date_time.strftime("%d-%m-%Y")

    try:
        # Get the path of the current script
        base_path = os.path.abspath(os.path.dirname(sys.argv[0]))

        # Create the directory inside the user's file directory
        directory = os.path.join(base_path, directory)
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Create the log file inside the directory
        file_path = os.path.join(directory, f"{curtime2}.txt")
        with open(file_path, "a+") as f:
            f.write(f"{curtime1} {data}\r\n")
    except OSError as e:
        print(f"Error: Creating directory. {directory} - {e}")

# @app.post("/send_jsw_email/")
# async def send_jsw_email(cursor: Session = Depends(get_db)):  
#     try:
#         query = f'''
#         SELECT
#         	FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
#         	FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
#         	FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
#         FROM
#         	ems_v1.dbo.master_shifts
#         '''
#         shifttimings = cursor.execute(query).fetchall()
#         shift_1 = ''
#         previous_mill_date = date.today()
		
#         for row in shifttimings:
#             shift_1 = row["shift_1"]
#             previous_mill_date = row["previous_mill_date"]
#         sql = f'''SELECT TOP 1 * FROM ems_v1.dbo.master_mail ORDER BY mail_id DESC'''
#         automail_list = cursor.execute(sql).fetchall()
#         from_mail = ''
#         password = ''
#         to_mail = ''
#         subject = ''
#         compose_textarea = ''

#         for rows in automail_list:
#             from_mail = rows["from_mail"]
#             password = rows["pwd"]
#             to_mail = rows["to_mail"]
#             subject = rows["subject"]
#             compose_textarea = rows["compose_textarea"]
#         from_date = previous_mill_date
#         mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
       
#         sql= text(f"select * from [ems_v1].[dbo].[master_production_entry] where FORMAT(entry_date,'dd-MM-yyyy')='{from_date}'")
#         print(sql)
#         datas1 = cursor.execute(sql).fetchall()    
#         res = []
#         datas = ''
#         function_id = ''
#         function_id1 = 0
#         where = ''
#         from_date = parse_date(previous_mill_date)
#         month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
#         tbl_name_month = f"{month_year}"
#         if function_id != '':
#             where = f'and function_id = {function_id}'
#         query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
#         query1 = cursor.execute(query1).fetchall()
#         if len(query1)>0:  
#             query = f'''select * from ems_v1.dbo.master_function where status = 'active'{where} '''
#             dt = cursor.execute(query).fetchall()
#             if len(dt)>0:
#                 for i in dt:
#                     function_id1 = i["function_id"]
                
#                     sql1=text(f'''  SELECT
#                                 min(fd.function_name)as function_name,
#                                 min(fd.function_id)as function_id,
#                                 min(fd.function_code)as function_code,
#                                 min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
#                                 min(fd.function_id)as opt_id,
#                                 min(fd.function_name)as opt_actual,
#                                 min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
#                                 count(cp.machine_id)as meter_count,
#                                 min(fd.image)as image
#                             from
#                                 [ems_v1].[dbo].[master_function] fd
#                                 left join [ems_v1].[dbo].[master_machine] mm ON mm.function_id= fd.function_id and mm.status='active'
#                                 left join [ems_v1].[dbo].[current_power] cp ON cp.machine_id=mm.machine_id
#                             where fd.status<>'delete' and fd.function_id  = '{function_id1}'
#                             group by fd.function_id
#                             order by fd.function_id''')
#                     data1= cursor.execute(sql1).fetchall()
                    
#                     where_1 = f"where FORMAT(cp.mill_date, 'yyyy-MM-dd HH:mm:ss') = '{from_date}'"
        
#                     sql2 = text(f"""
#                             SELECT 
#                         isnull((select department_name from ems_v1.dbo.master_department where department_id=m1.department_id),'') as zone1,
#                         isnull((select department_name from ems_v1.dbo.master_department where department_id=m2.department_id),'') as zone2,
#                         isnull((select shed_name from ems_v1.dbo.master_shed where shed_id=m1.shed_id),'') as area1,
#                         isnull((select shed_name from ems_v1.dbo.master_shed where shed_id=m2.shed_id),'') as area2,
#                         isnull((select function_name from ems_v1.dbo.master_function where function_id=m1.function_id),'') as function1,
#                         isnull((select function_name from ems_v1.dbo.master_function where function_id=m2.function_id),'') as function2,
#                         isnull(m1.capacity_id,'')as capacity1,
#                         isnull(m2.capacity_id,'') as capacity2,
#                         c.*
#                     from
#                     (
#                         select
#                             min(case when p.department_id = 1 then p.machine_name  end) as machine1,
#                             sum(case when p.department_id = 1 then p.kwh/1000  end) as kwh1,
#                             min(case when p.department_id = 2 then p.machine_name  end) as machine2,
#                             sum(case when p.department_id = 2 then p.kwh/1000  end) as kwh2
#                         from (
#                             select
#                                 min(md.department_id) as department_id,
#                                 min(md.department_name) as department_name,
#                                 min(mf.function_id) as function_id,
#                                 min(mf.function_name) as function_name,
#                                 min(mm.machine_id) as machine_id,
#                                 min(mm.machine_name) as machine_name,
#                                 sum(cp.kwh) as kwh,
#                                 ROW_NUMBER() over (partition by min(md.department_name) order by min(mm.machine_name)) as sno
#                             from 
#                                 ems_v1_completed.dbo.power_{tbl_name_month} as cp
#                                 inner join ems_v1.dbo.master_machine mm on mm.machine_id = cp.machine_id
#                                 inner join ems_v1.dbo.master_department md on md.department_id = mm.department_id
#                                 inner join ems_v1.dbo.master_function mf on mf.function_id = mm.function_id
#                             {where_1} and cp.status = '0' and  mm.function_id = {function_id1}
#                             group by mm.machine_id
                            
#                             ) as p
#                         group by p.sno
#                     ) as c
#                     left join ems_v1.dbo.master_machine m1 on m1.machine_name=c.machine1 and m1.status='active'
#                     left join ems_v1.dbo.master_machine m2 on m2.machine_name=c.machine2 and m2.status='active'
#                     """)
#                     print(sql2)
                    
#                     data2= cursor.execute(sql2).fetchall()
                    
#                     if function_id == 'all' or function_id == '':

#                         for row in data1:
#                             res.append({"name": row["function_name"], "data": data2})
#                     else:
                                
#                         res.append({"name":"", "data": data2})
#                     curtime1 = from_date.strftime("%d-%m-%Y %H:%M:%S")
#                     from_date1 = curtime1[:11]
#                     print("from_date1",curtime1)
#                     datas = {
#                             "productionLists": datas1,
#                             "res": res,
#                             "from_date": from_date1,
#                             "excel_filename": ""
#                         }

#         functionality_wise_report(datas)
#         file_path = os.path.join(base_path, "functionality_wise_report.xlsx")
#         where = ''
#         select = '*'
#         entry_date = previous_mill_date
#         if entry_date is not None :
#             data = {}
#             data = {'entry_date': entry_date}
#             # entry_date = entry_date[6:10] + '-' + entry_date[3:5] + '-' + entry_date[0:2]
#             where = text(f"WHERE FORMAT(entry_date, 'dd-MM-yyyy') = '{entry_date}'")
#             print(entry_date)

#             diesel_sql = text(f"SELECT {select} FROM [ems_v1].[dbo].[master_diesel_entry] {where}")
#             production_sql = text(f"SELECT {select} FROM [ems_v1].[dbo].[master_production_entry] {where}")
#             load_sql = text(f"SELECT {select} FROM [ems_v1].[dbo].[master_load_entry] {where}")
#             heat_sql = text(f"SELECT {select} FROM [ems_v1].[dbo].[master_heat_entry] {where}")

#             data['dieselLists'] = cursor.execute(diesel_sql).fetchall()
#             data['productionLists'] = cursor.execute(production_sql).fetchall()
#             data['loadLists'] = cursor.execute(load_sql).fetchall()
#             data['heatLists'] = cursor.execute(heat_sql).fetchall()
#             print(diesel_sql)
#         else:  
#             return JSONResponse({"iserror": True, "message": "entry_date is required"})
#         manual_wise_report(data, entry_date)
#         file_path = os.path.join(base_path, "manual_wise_report.xlsx")
#         time.sleep(5)
#         try:
#             # Create a multipart message container
#             message = MIMEMultipart()

#             # Set the email attributes
#             message['From'] = from_mail
#             message['Subject'] = subject
#             print(from_mail)
#             # Add attachments
#             attachments = ['functionally_wise_report.xlsx', 'manual_wise_report.xlsx']
#             for attachment in attachments:
#                 with open(attachment, 'rb') as file:
#                     attachment_part = MIMEApplication(file.read())
#                     attachment_part.add_header('Content-Disposition', 'attachment', filename=attachment)
#                     message.attach(attachment_part)
#             # Set the email body
#             message.attach(MIMEText(compose_textarea, 'html'))

#             # Connect to the SMTP server
#             smtp_server = "smtpout.secureserver.net"
#             smtp_port = 587
#             print(smtp_server)
#             with smtplib.SMTP(smtp_server, smtp_port) as smtp:
#                 smtp.starttls()
#                 print(password)
#                 smtp.login(from_mail, password)
#                 print(password)
#                 # Add recipients
#                 addr = to_mail.split(',')
#                 message['To'] = ", ".join(addr)
#                 # Send the email
#                 smtp.send_message(message)
#             print("Success!")
#         except Exception as e:
#             print("Fail :(", str(e))
#             createFolder("Log/", "Issue in returning data " + str(e))
#             return JSONResponse({"iserror": True, "message": str(e)}) 

#     except Exception as e:
#         error_type = type(e).__name__
#         error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
#         error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
#         error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
#         createFolder("Log/", "Issue in returning data " + error_message)


def parse_date(from_date):
    date_from = from_date.split("-")
    if len(date_from[0]) <= 2:
        if int(date_from[0]) > 12:
            from_date = parser.parse(from_date).strftime("%Y-%m-%d %H:%M:%S")
        else:
            from_date = parser.parse(from_date).strftime("%Y-%d-%m %H:%M:%S")
        from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S")
    else:
        from_date = parser.parse(from_date).strftime("%Y-%m-%d %H:%M:%S")
        from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S")
    return from_date 
base_path = Path(__file__).parent

# static_dir = Path(__file__).parent / "attachments"
mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
 
def power_report(data,previous_mill_date,filename):
    try:
        # file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "power_report_template.xlsx"))
        
        script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))

        # Combine the script directory with the file name to create the file path
        file_name = "power_report_template.xlsx"
        file_path = os.path.join(script_dir, file_name)
                
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        sheet.title = 'EMS'# Remove the default sheet created
  
        style = {'alignment': Alignment(horizontal='center')}
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        styles = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        style1 = {
            
            'font': Font(bold=True, color='006400', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }

        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        row_count = 5

        # Create a set to store unique function names
        function_names = set()

        # Iterate over the data to collect unique function names
        for row in data["res"]:
            function_names.add(row.function_name)

        # Iterate over each unique function name
        for function_name in function_names:
            # Display the function name in the sheet
            sheet['A' + str(row_count)].value = function_name
            sheet['A' + str(row_count)].alignment = style1['alignment']
            sheet['A' + str(row_count)].font = style1['font']
            row_count += 1
            
            # Iterate over the data again to find matching records for the current function
            for row in data["res"]:
                if row.function_name == function_name:
                    sheet['A' + str(row_count)].value = row.mill_date
                    sheet['B' + str(row_count)].value = row.machine_name
                    sheet['C' + str(row_count)].value = row.c_kwh
                    sheet['D' + str(row_count)].value = row.kwh
                    row_count += 1

            row_count += 1  # Add an empty row between each function's details

        file_name = 'power_report.xlsx'
        file_path = os.path.join(filename, file_name)
        workbook.save(file_path)
        print(file_name)
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)

def power_report_dtl(groupby,groupids,machine_ids,report_type,from_date,to_date,shift_id,report_type_for):
    try:
        sql = "select * from ems_v1.dbo.master_shifts"
        result = cursor.execute(sql).fetchall()      
        mill_date = ''
        mill_shift = 0
        no_of_shifts = 0        
        for row in result:
            mill_date = row.mill_date
            mill_shift = row.mill_shift
            no_of_shifts = row.no_of_shifts

        # if mill_date == "":
        #     return JSONResponse({"iserror": True, "message": "Mill Date, Shift Missing. Check Current Shifts Table"})       
        day_kwh_tbl = f'''select machine_id,mill_date,kwh from ems_v1.dbo.current_power'''
        
        month_year=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""       
        tbl_name_month = f"{month_year}"
        if len(result)>0:

            day_kwh_tbl = f" union all select machine_id,mill_date,kwh from ems_v1_completed.dbo.power_{tbl_name_month}"

        is_cur = 0

        if datetime.datetime.strptime(from_date, "%d-%m-%Y") <= mill_date and datetime.datetime.strptime(to_date, "%d-%m-%Y") >= mill_date:
            is_cur = mill_shift
         
        group_by = ''
        order_by = ''
        where = ''
        tbl_name = ''
	
        if shift_id != '':
	
            where += " and cp.mill_shift=$shift_id"

        if len(groupids) == len(groupids.replace('all', '')) and len(groupids) != 0:
            if groupby == "function":
                where += f" and mm.{groupby}_id in ({groupids})"
            else:
                where += f" and cp.{groupby}_id in ({groupids})"


        if len(machine_ids) == len(machine_ids.replace('all', '')) and len(machine_ids) != 0:
            where += f" and cp.machine_id in ({machine_ids})"

        if report_type == "detail":
            if groupby=="function":
			
                group_by = f"mm.{groupby}_id ,cp.machine_id"
                order_by = f"mm.{groupby}_id ,min(mm.machine_order)"
            else:
                group_by = f"cp.mill_date ,cp.{groupby}_id ,cp.machine_id"
                order_by = f"cp.mill_date ,cp.{groupby}_id ,min(mm.machine_order)"
        else:
            if groupby=="function":
			
                group_by = f"mm.{groupby}_id ,cp.machine_id"
                order_by = f"mm.{groupby}_id ,min(mm.machine_order)"            
            else:
                if report_type_for == 'machine':
                    group_by = f"cp.{groupby}_id ,cp.machine_id"
                    order_by = f"cp.{groupby}_id ,min(mm.machine_order)"
                else:
                    group_by = f"cp.{groupby}_id ,cp.mill_date"
                    order_by = f"cp.{groupby}_id ,cp.mill_date"                  
	
        tbl_name += f'(SELECT company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,kwh,machine_kwh as start_kwh,master_kwh as end_kwh,mill_date,mill_shift FROM ems_v1.dbo.current_power '

        tmpdt = datetime.datetime.strptime(from_date, '%d-%m-%Y').strftime('%Y-%m-%d')
        j = 0

        for i in range(6):  # For Tmp Loop
            if i == 5:
                i = 1
            tmpdt = (datetime.datetime.strptime(tmpdt, '%Y-%m-%d') + relativedelta(months=j)).strftime('%Y-%m-%d')
            tbl_name += f''' UNION ALL 
                SELECT company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,kwh,machine_kwh as start_kwh,master_kwh as end_kwh,mill_date,mill_shift FROM ems_v1_completed.dbo.power_{tmpdt[5:7]}{tmpdt[0:4]}'''
            j = 1
            print(tmpdt)
            print(to_date)
            if tmpdt[5:7] == to_date[3:5] and tmpdt[0:4] == to_date[6:10]:
                break

        tbl_name += " ) as cp"
        sql = f'''SELECT 
			min(md.department_id) as department_id,
			min(md.department_code) as department_code,
			min(md.department_name) as department_name,
			min(ms.shed_id) as shed_id,
			min(ms.shed_code) as shed_code,
			min(ms.shed_name) as shed_name,
			min(mt.machinetype_id) as machinetype_id,
			min(mt.machinetype_code) as machinetype_code,
			min(mt.machinetype_name) as machinetype_name,
			FORMAT(min(cp.mill_date),'dd-MM-yyyy') as mill_date,
			min(mm.machine_code) as machine_code,
			min(mm.machine_name) as machine_name,
			min(mf.function_id) as function_id,
			min(mf.function_name) as function_name,
			min(mf.function_code) as function_code,
			SUM(case WHEN mm.energy_selection = 'wh' then cp.kwh/1000 else cp.kwh end) AS kwh,
			ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case WHEN mm.energy_selection = 'wh' then cp.kwh/1000 else cp.kwh end ELSE 0 END),2) AS kwh_1,
			ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case WHEN mm.energy_selection = 'wh' then cp.kwh/1000 else cp.kwh end ELSE 0 END),2) AS kwh_2,
			ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case WHEN mm.energy_selection = 'wh' then cp.kwh/1000 else cp.kwh end ELSE 0 END),2) AS kwh_3,
			ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case WHEN mm.energy_selection = 'wh' then cp.start_kwh/1000 else cp.start_kwh end ELSE 0 END),2) AS start_kwh_1,
			ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case WHEN mm.energy_selection = 'wh' then cp.start_kwh/1000 else cp.start_kwh end ELSE 0 END),2) AS start_kwh_2,
			ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case WHEN mm.energy_selection = 'wh' then cp.start_kwh/1000 else cp.start_kwh end ELSE 0 END),2) AS start_kwh_3,
			ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case WHEN mm.energy_selection = 'wh' then cp.end_kwh/1000 else cp.end_kwh end ELSE 0 END),2) AS end_kwh_1,
			ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case WHEN mm.energy_selection = 'wh' then cp.end_kwh/1000 else cp.end_kwh end ELSE 0 END),2) AS end_kwh_2,
			ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case WHEN mm.energy_selection = 'wh' then cp.end_kwh/1000 else cp.end_kwh end ELSE 0 END),2) AS end_kwh_3,
			case when {is_cur} <> 0 and '{mill_date}' = FORMAT(min(cp.mill_date),'yyyy-MM-dd') then
				ROUND(SUM(CASE WHEN cp.mill_shift = {is_cur} THEN case WHEN mm.energy_selection = 'wh' then cp.end_kwh/1000 else cp.end_kwh end ELSE 0 END),2)
			else
				ROUND(SUM(CASE WHEN cp.mill_shift = {no_of_shifts} THEN case WHEN mm.energy_selection = 'wh' then cp.end_kwh/1000 else cp.end_kwh end ELSE 0 END),2)
			end as c_kwh
		FROM 
			{tbl_name} ,
			ems_v1.dbo.master_machine mm,
			ems_v1.dbo.master_machinetype mt,
			ems_v1.dbo.master_shed ms,
			ems_v1.dbo.master_department md,
			ems_v1.dbo.master_function mf
		WHERE cp.machine_id = mm.machine_id and mm.machinetype_id=mt.machinetype_id and mm.shed_id=ms.shed_id and mm.department_id=md.department_id and mm.function_id=mf.function_id and mm.status='active' and FORMAT(cp.mill_date ,'dd-MM-yyyy')>='{from_date}' and FORMAT(cp.mill_date , 'dd-MM-yyyy')<='{to_date}' {where}
		GROUP BY {group_by}
		ORDER BY {order_by} '''
        
        createFolder("Log/", "query " + str(sql))

        result = cursor.execute(sql).fetchall()
        return result
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)

def auto_download_main(filename):
    try:
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            ems_v1.dbo.master_shifts'''
        shifttimings = cursor.execute(sql).fetchall()
        previous_mill_date = date.today()
        for row in shifttimings:
            previous_mill_date = row.previous_mill_date
        from_date = previous_mill_date
        to_date = previous_mill_date
        shift_id = ''
        pdf1 = 'excel'
        group_by = 'department'
        group_ids = ''
        machine_ids = ''
        report_type = 'detail'
        report_type_for = 'machine'
        rpt_field = 'kwh,c_kwh'
        pdf = 'print'
        data = {}
        head = ""
        if group_by == "shed":
            head = "Area"
        elif group_by == "machinetype":
            head = "Location"
        elif group_by == "department":
            head = "Department"       
        elif group_by == "function":
            head = "Function"
        title = head + "wise Summary Report"
        if report_type == "detail":
            title = head + "wise Detail Report"

        title += " From " + from_date + " To " + to_date
        if shift_id != '':
            title += " Shift " + shift_id

        res = power_report_dtl(group_by,group_ids,machine_ids,report_type,from_date,to_date,shift_id,report_type_for)
        createFolder("Log/","res"+str(res))
        data['res'] = res
        data['title'] = title
        data['head'] = head
        data['pdf'] = pdf
        data['group_by'] = group_by
        data['rpt_field'] = rpt_field

        data['res'] = res
        power_report(data,previous_mill_date,filename)
        
        data['excel_filename'] = filename
        createFolder("Log/", "data " + str(data)) 


    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)    

def functionality_wise_report(datas,filename):
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet created
        f_count = 0
        style = {'alignment': Alignment(horizontal='center')}
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        styles = style1 = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        style1 = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }

        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        
        
        if len(datas) == 0:
            row_count = 7
            sheet = workbook.create_sheet()
            # image_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "jsw.jpeg"))
            image_path = os.path.abspath(__file__)+"//jsw.jpeg"
            img = Image(image_path)
            img.anchor = "A1"  # Set the anchor point to A1
            
            img_width_pixels, img_height_pixels = img.width, img.height
            # Set the desired image size
            desired_width_pixels = 150
            desired_height_pixels = 45
            # Calculate the scaling factor for width and height
            width_scale = desired_width_pixels / img_width_pixels
            height_scale = desired_height_pixels / img_height_pixels
            # Scale the image size
            img.width = int(img_width_pixels * width_scale)
            img.height = int(img_height_pixels * height_scale)
            sheet.column_dimensions['A'].width = img.width / 10
            sheet.row_dimensions[1].height = img.height / 2
            sheet.row_dimensions[2].height = img.height / 2
            sheet.merge_cells('A1:A2')  # Merge cells A1 to A2
            sheet.add_image(img)
            sheet['I' + str(row_count)].value = 'No Data'
            sheet['I' + str(row_count)].alignment = styles['alignment'] 
        else:
            for row in datas["res"]:
                date = datas["from_date"]
                print(date)
                sheet = workbook.create_sheet(title=row['name'])
                # image_path = os.path.abspath(os.path.join(os.path.dirname(__file__),"jsw.jpeg"))
                script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))

                # Combine the script directory with the file name to create the file path
                file_name = "jsw.jpeg"
                image_path = os.path.join(script_dir, file_name)
                img = Image(image_path)
                img.anchor = "A1"  # Set the anchor point to A1
                
                img_width_pixels, img_height_pixels = img.width, img.height
                # Set the desired image size
                desired_width_pixels = 150
                desired_height_pixels = 45
                # Calculate the scaling factor for width and height
                width_scale = desired_width_pixels / img_width_pixels
                height_scale = desired_height_pixels / img_height_pixels
                # Scale the image size
                img.width = int(img_width_pixels * width_scale)
                img.height = int(img_height_pixels * height_scale)
                sheet.column_dimensions['A'].width = img.width / 10
                sheet.row_dimensions[1].height = img.height / 2
                sheet.row_dimensions[2].height = img.height / 2

                sheet.merge_cells('A1:A2')  # Merge cells A1 to A2
                sheet.add_image(img)
                
                sheet['B2'].value = 'SMS Power Consumption Report - ' + row['name'] + ' Wise '
                sheet.merge_cells('B2:H2')
                sheet['B2'].alignment = style['alignment']
                sheet.merge_cells('A3:D3')
                sheet['A3'].value = 'SMS#1'
                sheet['A3'].alignment = style['alignment']
                sheet['A4'].value = 'Area'
                sheet['A4'].alignment = style['alignment']
                sheet['B4'].value = 'Feeder'
                sheet['B4'].alignment = style['alignment']
                sheet['C4'].value = 'Capacity'
                sheet['C4'].alignment = style['alignment']
                sheet['D4'].value = 'Consumption(kWh)'
                sheet['D4'].alignment = style['alignment']
                sheet.merge_cells('E3:H3')
                sheet['E3'].value = 'SMS#2'
                sheet['E3'].alignment = style['alignment']
                sheet['E4'].value = 'Area'
                sheet['E4'].alignment = style['alignment']
                sheet['F4'].value = 'Feeder'
                sheet['F4'].alignment = style['alignment']
                sheet['G4'].value = 'Capacity'
                sheet['G4'].alignment = style['alignment']
                sheet['H4'].value = 'Consumption(kWh)'
                sheet['H4'].alignment = style['alignment']
                sheet['B1'].value = f'date:{date}'
                sheet['B1'].alignment = style['alignment']
                kwh1 = 0
                kwh2 = 0
                capacity1 = 0
                capacity2 = 0
                
                row_count = 5
                for rows in row['data']:
                    sheet['A' + str(row_count)].value = rows.area1
                    sheet['A' + str(row_count)].alignment = style['alignment'] 
                    sheet['B' + str(row_count)].value = rows.machine1
                    sheet['B' + str(row_count)].alignment = style['alignment']
                    
                    if rows.machine1:
                        sheet['D' + str(row_count)].value = rows.kwh1
                        sheet['D' + str(row_count)].alignment = style['alignment']
                        
                        kwh1= rows.kwh1
                    if rows.machine1:
                        sheet['C' + str(row_count)].value = rows.capacity1
                        sheet['C' + str(row_count)].alignment = style['alignment']

                    sheet['E' + str(row_count)].value = rows.area2
                    sheet['E' + str(row_count)].alignment = style['alignment']
                    
                    sheet['F' + str(row_count)].value = rows.machine2
                    sheet['F' + str(row_count)].alignment = style['alignment']

                    if rows.machine2:
                        sheet['H' + str(row_count)].value = rows.kwh2
                        sheet['H' + str(row_count)].alignment = style['alignment']
                        
                        kwh2 = rows.kwh2
                    if rows.machine2:
                        sheet['G' + str(row_count)].value = rows.capacity2
                        sheet['G' + str(row_count)].alignment = style['alignment']
                        
                    row_count += 1

                production1 = 0
                production2 = 0

                for production in datas["productionLists"]:
                    if production.production_name == 'SMS1 Production':
                        production1 = production.production_value
                    elif production.production_name == 'SMS2 Production':
                        production2 = production.production_value

                add_row = row_count + 4
                sheet['A' + str(add_row)].value = 'SMS1 Emergency Power Consumption'
                
                sheet['D' + str(add_row)].value = '=SUM(D4:D' + str(row_count) + ')'
                sheet['D' + str(add_row)].alignment = style['alignment']
                
                sheet['E' + str(add_row)].value = 'kWh'
                sheet['E' + str(add_row)].alignment = style['alignment']
            
                sheet['F' + str(add_row)].value = 'SMS1 Production'
                
                sheet['G' + str(add_row)].value = int(production1)
                sheet['G' + str(add_row)].alignment = style['alignment']
            
                sheet['H' + str(add_row)].value = 'MT'
                
                add_row1 = add_row + 1
                sheet['A' + str(add_row1)].value = 'SMS2 Emergency Power Consumption' 
                sheet['D' + str(add_row1)].value = '=SUM(H4:H' + str(row_count) + ')'
                sheet['D' + str(add_row1)].alignment = style['alignment']
                sheet['E' + str(add_row1)].value = 'kWh'
                sheet['E' + str(add_row1)].alignment = style['alignment'] 
                sheet['F' + str(add_row1)].value = 'SMS2 Production'
                sheet['G' + str(add_row1)].value = int(production2)
                sheet['G' + str(add_row1)].alignment = style['alignment'] 
                sheet['H' + str(add_row1)].value = 'MT'
                
                total_production = int(production1)+ int(production2)
                print("total_production",total_production)
                add_row2 = add_row1 + 1
                sheet['A' + str(add_row2)].value = 'Total Emergency Power Consumption' 
                total_kwh1 = sheet['D' + str(add_row)].value
                total_kwh2 = sheet['D' + str(add_row1)].value
                sheet['D' + str(add_row2)].value = '=SUM(D4:D' + str(row_count) + ')+SUM(H4:H' + str(row_count) + ')'
                sheet['D' + str(add_row2)].alignment = style['alignment'] 
                sheet['E' + str(add_row2)].value = 'kWh'
                sheet['E' + str(add_row2)].alignment = style['alignment']
                sheet['F' + str(add_row2)].value = 'Total Production'
                sheet['G' + str(add_row2)].value = total_production
                sheet['G' + str(add_row2)].alignment = style['alignment']
                sheet['H' + str(add_row2)].value = 'MT'
                print("production1",production1)
                print("production2",production2)
                print("kwh1",kwh1)
                add_row3 = add_row2 + 1
                add_row3 = add_row2 + 1
                if production1 == 0:
                    sms1_cons = 0
                else:
                    sms1_cons = kwh1 / int(production1)

                if production2 == 0:
                    sms2_cons = 0
                else:
                    sms2_cons = kwh2 / int(production2)

                sheet['A' + str(add_row3)].value = 'SMS1 Emergency Power Consumption(kWh/MT)' 
                sheet['D' + str(add_row3)].value = sms1_cons
                sheet['D' + str(add_row3)].alignment = style['alignment']
                sheet['D' + str(add_row3)].fill = style1['fill']
                sheet['D' + str(add_row3)].font = style1['font']
                sheet['E' + str(add_row3)].value = 'SMS2 Emergency Power Consumption(kWh/MT)'
                sheet['H' + str(add_row3)].value = sms2_cons
                sheet['H' + str(add_row3)].alignment = style['alignment']
                sheet['H' + str(add_row3)].fill = style1['fill']
                sheet['H' + str(add_row3)].font = style1['font']

                add_row4 = add_row3 + 1
                
                # Evaluate the formula and get the calculated value for total_consumption1
                total_consumption1 = sheet['D' + str(add_row2)].value
                print("total_consumption1",total_consumption1)
                            # Get the value of total_consumption2 directly
                total_consumption2 = sheet['G' + str(add_row2)].value
                print("total_consumption2:", total_consumption2)

                if total_consumption2 == 0:
                    total_cons = 0
                else:
                    total_cons = f"=(SUM(D4:D{row_count})+SUM(H4:H{row_count}))/{total_consumption2}"

                sheet['A' + str(add_row4)].value = 'SMS Power Consumption (kWh/MT) = '
                sheet.merge_cells('E' + str(add_row4) + ':H' + str(add_row4))

                merged_cell = sheet['E' + str(add_row4)]
                merged_cell.alignment = style['alignment']
                merged_cell.fill = style2['fill']
                merged_cell.font = style2['font']

                sheet['E' + str(add_row4)].value = total_cons
                sheet['E' + str(add_row4)].alignment = style['alignment'] 
                rows_count = 1
                row_range = sheet.iter_rows(min_row=rows_count, max_row=add_row4, min_col=1, max_col=9)
                for row in row_range:
                    for cell in row:
                        cell.border = border

                sheet.column_dimensions['A'].width = 40
                sheet.column_dimensions['B'].width = 32
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width = 25
                sheet.column_dimensions['F'].width = 20
                sheet.column_dimensions['H'].width = 20
                
                f_count += 1
            
        file_name = 'functionality_wise_report.xlsx'
        file_path = os.path.join(filename, file_name)
        workbook.save(file_path)
        print(file_name)
       
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)

def auto_download_functionality(filename):
    try:
                
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            master_shifts
        '''
        shifttimings = cursor.execute(sql).fetchall()
        previous_mill_date = date.today()
        for row in shifttimings:
            previous_mill_date = row.previous_mill_date
		
        function_id = ''
        from_date = previous_mill_date
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
       
        sql1= f'''select * from [ems_v1].[dbo].[master_production_entry] where FORMAT(entry_date,'dd-MM-yyyy')='{from_date}' '''
        print("sql1",sql1)
        datas1 = cursor.execute(sql1).fetchall() 
        print(datas1)  

        res = []
        datas = ''
    
        function_id1 = 0
        where = ''
        from_date = parse_date(from_date)
        month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
        tbl_name_month = f"{month_year}"
        # if function_id != '':
        #     where = f'and function_id = {function_id}'
        # query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
        # query1 = cursor.execute(query1).fetchall()
        # if len(query1)>0:  
        query = f'''select * from ems_v1.dbo.master_function where status = 'active'{where} '''
        dt = cursor.execute(query).fetchall()
        if len(dt)>0:
            for i in dt:
                function_id1 = i.function_id
                
                sql1=f'''  SELECT
                                min(fd.function_name)as function_name,
                                min(fd.function_id)as function_id,
                                min(fd.function_code)as function_code,
                                min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
                                min(fd.function_id)as opt_id,
                                min(fd.function_name)as opt_actual,
                                min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
                                count(cp.machine_id)as meter_count,
                                min(fd.image)as image
                            from
                                [ems_v1].[dbo].[master_function] fd
                                left join [ems_v1].[dbo].[master_machine] mm ON mm.function_id= fd.function_id and mm.status='active'
                                left join [ems_v1].[dbo].[current_power] cp ON cp.machine_id=mm.machine_id
                            where fd.status<>'delete' and fd.function_id  = '{function_id1}'
                            group by fd.function_id
                            order by fd.function_id'''
                data1= cursor.execute(sql1).fetchall()
                
                where_1 = f"where FORMAT(cp.mill_date, 'yyyy-MM-dd HH:mm:ss') = '{from_date}'"
    
                sql2 = f"""
                            SELECT 
                        isnull((select department_name from ems_v1.dbo.master_department where department_id=m1.department_id),'') as zone1,
                        isnull((select department_name from ems_v1.dbo.master_department where department_id=m2.department_id),'') as zone2,
                        isnull((select shed_name from ems_v1.dbo.master_shed where shed_id=m1.shed_id),'') as area1,
                        isnull((select shed_name from ems_v1.dbo.master_shed where shed_id=m2.shed_id),'') as area2,
                        isnull((select function_name from ems_v1.dbo.master_function where function_id=m1.function_id),'') as function1,
                        isnull((select function_name from ems_v1.dbo.master_function where function_id=m2.function_id),'') as function2,
                        isnull(m1.capacity_id,'')as capacity1,
                        isnull(m2.capacity_id,'') as capacity2,
                        c.*
                    from
                    (
                        select
                            min(case when p.department_id = 1 then p.machine_name  end) as machine1,
                            sum(case when p.department_id = 1 then p.kwh/1000  end) as kwh1,
                            min(case when p.department_id = 2 then p.machine_name  end) as machine2,
                            sum(case when p.department_id = 2 then p.kwh/1000  end) as kwh2
                        from (
                            select
                                min(md.department_id) as department_id,
                                min(md.department_name) as department_name,
                                min(mf.function_id) as function_id,
                                min(mf.function_name) as function_name,
                                min(mm.machine_id) as machine_id,
                                min(mm.machine_name) as machine_name,
                                sum(cp.kwh) as kwh,
                                ROW_NUMBER() over (partition by min(md.department_name) order by min(mm.machine_name)) as sno
                            from 
                                ems_v1_completed.dbo.power_{tbl_name_month} as cp
                                inner join ems_v1.dbo.master_machine mm on mm.machine_id = cp.machine_id
                                inner join ems_v1.dbo.master_department md on md.department_id = mm.department_id
                                inner join ems_v1.dbo.master_function mf on mf.function_id = mm.function_id
                            {where_1} and cp.status = '0' and  mm.function_id = {function_id1}
                            group by mm.machine_id
                            
                            ) as p
                        group by p.sno
                    ) as c
                    left join ems_v1.dbo.master_machine m1 on m1.machine_name=c.machine1 and m1.status='active'
                    left join ems_v1.dbo.master_machine m2 on m2.machine_name=c.machine2 and m2.status='active'
                    """
                print("sql2",sql2)
                    
                data2= cursor.execute(sql2).fetchall()
                    
                if function_id == 'all' or function_id == '':

                    for row in data1:
                        res.append({"name": row.function_name, "data": data2})
                else:
                                
                    res.append({"name":"", "data": data2})
                curtime1 = from_date.strftime("%d-%m-%Y %H:%M:%S")
                from_date1 = curtime1[:11]
                print("from_date1",curtime1)
                datas = {
                        "productionLists": datas1,
                        "res": res,
                        "from_date": from_date1,
                        "excel_filename": filename
                    }
                    
        print("datas...",datas)
        functionality_wise_report(datas,filename)
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
def manual_wise_report(data, entry_date,filename):
    # file_path = f'{static_dir}/manual_wise_report_template.xlsx'
    # file_path = os.path.abspath(os.path.join(os.path.dirname(__file__),"manual_wise_report_template.xlsx"))
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))

    # Combine the script directory with the file name to create the file path
    file_name = "manual_wise_report_template.xlsx"
    file_path = os.path.join(script_dir, file_name)     
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS'

    # Set the entry date
    sheet['B3'] = f"Date: {entry_date}"
    sheet['B3'].font = Font(bold=True, name='Calibri', size=11)

    # Set the headers
    sheet['B6'] = 'DG Name'
    sheet['C6'] = 'Consumption (Litre)'

    sheet['E6'] = 'Production Zone'
    sheet['F6'] = 'Production(MT)'

    sheet['H6'] = 'Production Zone'
    sheet['I6'] = 'Production(MT)'

    sheet['E14'] = 'Area'
    sheet['F14'] = 'Heats(Nos)'

    sheet['B14'] = 'DG Name'
    sheet['C14'] = 'Time(Min)'

    # Populate the data
    total_consumption = 0
    total_production = 0 
    total_power = 0
    total_heats = 0
    total_running_time  = 0

    row_index = 7

    for item in data['dieselLists']:
        dg_name = str(item.dg_name)
        if item.diesel_litre == '':
            diesel_litre = 0
        else:
            diesel_litre = float(item.diesel_litre)
        total_consumption += diesel_litre

        sheet.cell(row=row_index, column=2).value = dg_name
        sheet.cell(row=row_index, column=3).value = diesel_litre
        row_index += 1

    # Add total consumption row
    total_row_index = row_index
    sheet.cell(row=11, column=2).value = 'Total Consumption'
    sheet.cell(row=11, column=3).value = total_consumption

    # Auto-adjust column widths
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

    row_index = 7
    for item in data['productionLists']:
       
        production_name = str(item.production_name)
        if item.production_value=='':
            production_value = 0
        else:
            production_value = float(item.production_value)
        total_production += production_value

        sheet.cell(row=row_index, column=5).value = production_name
        sheet.cell(row=row_index, column=6).value = production_value
        row_index += 1

    # Add total production row
    total_row_index = row_index
    sheet.cell(row=11, column=5).value = 'Total Production'
    sheet.cell(row=11, column=6).value = total_production
    # Auto-adjust column widths
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20

    row_index = 7
    for item in data['loadLists']:
        
        load_name = str(item.load_name)
        if item.load_value =='':
            load_value = 0
        else:
            load_value = float(item.load_value)
        total_power += load_value
        sheet.cell(row=row_index, column=8).value = load_name
        sheet.cell(row=row_index, column=9).value = load_value
        row_index += 1
        
        sheet.cell(row=23, column=8).value = 'Total Power'
        sheet.cell(row=23, column=9).value = total_power

        # Auto-adjust column widths
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20

    row_index = 15
    for item in data['heatLists']:
        
        heat_name = str(item.heat_name)
        if item.heat_value == '':
            heat_value = 0
        else:
            heat_value = float(item.heat_value)
        total_heats += heat_value
        sheet.cell(row=row_index, column=5).value = heat_name
        sheet.cell(row=row_index, column=6).value = heat_value
        row_index += 1
        
        sheet.cell(row=23, column=5).value = 'Total Heats'
        sheet.cell(row=23, column=6).value = total_heats

        # Auto-adjust column widths
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20

    row_index = 15
    for item in data['dieselLists']:
        
        dg_name = str(item.dg_name)
        if item.dg_runtime == '':
            dg_runtime = 0
        else:
            dg_runtime = float(item.dg_runtime)
        total_running_time += dg_runtime
        sheet.cell(row=row_index, column=2).value = dg_name
        sheet.cell(row=row_index, column=3).value = dg_runtime
        row_index += 1

        sheet.cell(row=23, column=2).value = 'Total Running Time '
        sheet.cell(row=23, column=3).value = total_running_time

        # Auto-adjust column widths
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20

    file_name = 'manual_wise_report.xlsx'
    file_path = os.path.join(filename, file_name)
    workbook.save(file_path)

def auto_download_manual(filename):
    try:
        data = {}
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            master_shifts
        '''
        shifttimings = cursor.execute(sql).fetchall()
        from_date = date.today()
        for row in shifttimings:
            from_date = row.previous_mill_date
        pdf = 'excel'
        data['entry_date'] = from_date
        
        where = f"WHERE FORMAT(entry_date, 'dd-MM-yyyy') = '{from_date}'"

        diesel_sql = f"SELECT * FROM [ems_v1].[dbo].[master_diesel_entry] {where}"
        production_sql = f"SELECT * FROM [ems_v1].[dbo].[master_production_entry] {where}"
        load_sql = f"SELECT * FROM [ems_v1].[dbo].[master_load_entry] {where}"
        heat_sql = f"SELECT * FROM [ems_v1].[dbo].[master_heat_entry] {where}"
    
        data['dieselLists'] = cursor.execute(diesel_sql).fetchall()
        data['productionLists'] = cursor.execute(production_sql).fetchall()
        data['loadLists'] = cursor.execute(load_sql).fetchall()
        data['heatLists'] = cursor.execute(heat_sql).fetchall()
        manual_wise_report(data,from_date,filename)

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
def auto_download_main_manual_functionality():
    try:
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            master_shifts'''
        shifttimings = cursor.execute(sql).fetchall()
        previous_mill_date = date.today()
        shift_1 = ''
        for row in shifttimings:
            previous_mill_date = row.previous_mill_date
            shift_1 = row.shift_1
        drive_name = os.path.abspath(__file__)[0]
        createFolder("Log/","drive_name...... "+str(drive_name))
        if not os.path.isdir(drive_name + ":\\AIC"):
            os.mkdir(drive_name + ":\\AIC")
        if not os.path.isdir(drive_name + ":\\AIC\\Power_Report_v2"):
            os.mkdir(drive_name + ":\\AIC\\Power_Report_v2")
        if not os.path.isdir(drive_name + ":\\AIC\\Power_Report_v2\\" + previous_mill_date):
            os.mkdir(drive_name + ":\\AIC\\Power_Report_v2\\" + previous_mill_date)
        
        filename = drive_name + ":\\AIC\\Power_Report_v2\\" + previous_mill_date
        print("filename",filename)
        sy_date = '13:23' 
        if sy_date == datetime.datetime.now().strftime("%H:%M"):
            auto_download_main(filename)
            auto_download_functionality(filename)
            auto_download_manual(filename)

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)

while True:
    try:
        driver = 'ODBC Driver 17 for SQL Server'
        server_name = 'WIN-KSF3D686G23'
        port = '1433'  
        database_name = 'ems_v1'
        user_name = 'sa'     
        password = 'admin@2022'

        conn = (
            f'DRIVER={driver};'
            f'SERVER={server_name},{port};' 
            f'DATABASE={database_name};'
            f'UID={user_name};'
            f'PWD={password}'
        )    
        # driver = 'ODBC Driver 17 for SQL Server'
        # server_name = 'DESKTOP-92NNMNK'
        # database_name = 'ems_v1'
        # user_name = 'sa'     
        # password = 'admin@2023'

        # conn = (
        #     f'DRIVER={driver};'
        #     f'SERVER={server_name};' 
        #     f'DATABASE={database_name};'
        #     f'UID={user_name};'
        #     f'PWD={password}'
        # )    
        conn = pyodbc.connect(conn)
        cursor = conn.cursor()
        auto_download_main_manual_functionality()
        conn.close()
        time.sleep(60)
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
