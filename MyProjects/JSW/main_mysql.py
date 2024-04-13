from fastapi import FastAPI, Depends, Form, File, UploadFile
from typing import Optional, List
from datetime import datetime,date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy.sql import text
from sqlalchemy.exc import SQLAlchemyError
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse
import random
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pathlib import Path
import json
from fastapi.staticfiles import StaticFiles
from log_folder import createFolder
import os
import traceback
from dateutil import parser
import shutil
from mysql_connection import get_db
from fastapi.middleware.cors import CORSMiddleware   
from fastapi import Request
import datetime
import calendar
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image
from dateutil.relativedelta import relativedelta 
import sys
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins= ['*'],
    allow_credentials = True,
    allow_methods = ["*"],
    allow_headers = ["*"]
)

def parse_date(from_date):
    date_from = from_date.split("-")
    if len(date_from[0]) <= 2:
        if int(date_from[0]) > 12:
            from_date = parser.parse(from_date).strftime("%Y-%m-%d %H:%M:%S")
        else:
            from_date = parser.parse(from_date).strftime("%Y-%d-%m %H:%M:%S")
        from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S")
    else:
        from_date = parser.parse(from_date).strftime("%Y-%m-%d %H:%M:%S")
        from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S")
    return from_date   

@app.post("/login/")
async def employee_login(username:str=Form(),
                         password:str=Form(),
                         cnx: Session = Depends(get_db)
                        ):
    
    try:
        squery=text(f'''
                SELECT 
                    me.*,
                    mc.company_name as company_name,
                    mc.company_code as company_code,
                    mb.branch_name as branch_name,
                    mb.branch_code as branch_code
                FROM 
                    master_employee me
                    INNER JOIN master_company mc on me.company_id=mc.company_id
                    INNER JOIN master_branch mb on me.branch_id=mb.branch_id
                WHERE 
                    employee_code='{username}' AND password_login=md5('{password}') AND is_login='yes' ''')
        data = cnx.execute(squery).fetchall()
        print(data)
        print(squery)
        if len(data) > 0:
            createFolder("Log/","Response Sent Successfully ")
            return JSONResponse({"iserror": False, "message": "Data Returned Succesfully", "data": jsonable_encoder(data)})
        else:
            createFolder("Log/","username and password incorrect ")
            return JSONResponse({"iserror": True, "message": "Invalid username or password"})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
base_path = os.path.abspath(os.path.dirname(__file__)) + "/attachment"
if not os.path.exists(base_path):
    os.makedirs(base_path)  
# Mount static directory for serving image files
app.mount("/attachment", StaticFiles(directory=base_path), name="attachment")

@app.post("/get_company_list/{company_id}")
@app.post("/get_company_list/")
async def get_company_list(request:Request,
                           company_id: Optional[str] = None,
                           cnx: Session = Depends(get_db)):
    try:
       
                 
        where = ""
        if company_id is not None:
            where = f" and company_id = {company_id}"
        query = text(f'''
                     SELECT
                    	mc.*, 
                    	CONCAT('/attachment/company_logo/',mc.logo) AS logo,
                    	IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                    	IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM
                    	master_company as mc 
                    	left join master_employee cu on cu.employee_id=mc.created_by
                    	left join master_employee mu on mu.employee_id=mc.modified_by
                    WHERE mc.status!='delete' {where}
                     ''')
        data = cnx.execute(query).fetchall()
        results = []
        for i in data:
            new_logo = dict(i)
            if new_logo["logo"] is not None:
                new_logo["logo"] = "http://" + request.headers["host"] + new_logo['logo']
            results.append(new_logo)
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(results)})
            
    except SQLAlchemyError as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_company_details/")
async def save_company_details(company_id: int = Form(None),
                               company_name: str = Form(None),
                               company_code: str = Form(None),
                               address: str = Form(None),
                               city: str = Form(None),
                               state:str = Form(None),
                               country:str= Form(None),
                               pincode:str = Form(None),
                               phone_number :str=Form(None),
                               mobile_number : str=Form(None),
                               email_id: str=Form(None),
                               website:str=Form(None),
                               logo:UploadFile=File(None),
                               user_login_id: str=Form(None),
                               cnx: Session = Depends(get_db)):
    if company_code == None:  
        return JSONResponse({"iserror": False, "message": " company_code is required"})

    if company_name == None:  
        return JSONResponse({"iserror": False, "message": " company_name is required"})

    try:
        if logo is  None:    
            filename = '' 
        else:   
        # get base path to save image
            base_path = os.path.abspath(os.path.dirname(__file__)) + "/attachment/company_logo/"
            if not os.path.exists(base_path):
                os.makedirs(base_path)
            random_number = random.randint(1, 100000) 
            extension = os.path.splitext(logo.filename)[1].lower()
            # generate filename
            filename = f"{random_number}_{datetime.datetime.now().strftime('%y_%m_%d_%H_%M_%S')}{extension}"
            # save image to disk
            with open(base_path + filename, "wb") as f:
                f.write(await logo.read())
            
        if company_id is not None:
            query = text(f"""
                UPDATE master_company 
                    SET company_name = '{company_name}', company_code = '{company_code}',address = '{address}',
                    city = '{city}', state = '{state}', country = '{country}', pincode = '{pincode}',
                    phone_number = '{phone_number}', mobile_number = '{mobile_number}', email_id = '{email_id}',
                    website = '{website}',modified_by = '{user_login_id}', modified_on = NOW(), logo='{filename}'
                WHERE company_id = {company_id}
                """)
            
        else:
            
            select_query = text(f'''select * from master_company where company_code = '{company_code}' and status!='delete' ''')
            data =cnx.execute(select_query).fetchall()            

            if len(data)>0:
                return JSONResponse({"iserror":True,"message":"country_code already exists "})
            
            query = text(f"""
                INSERT INTO master_company (company_name, company_code, address, city, state,
                country, pincode, phone_number, mobile_number, email_id, website, created_on, created_by, logo)
                
                VALUES ('{company_name}', '{company_code}','{address}','{city}','{state}',
                '{country}','{pincode}','{phone_number}','{mobile_number}','{email_id}','{website}', NOW(), 
                '{user_login_id}','{filename}')
            """)        
        cnx.execute(query)
        company_id = cnx.execute(text("SELECT LAST_INSERT_ID()")).first()[0]       
        cnx.commit()

        sql = text(f'''
        INSERT INTO ems_v1.power_report_fields_original (report_id, field_code, field_name, is_show, slno, field_name_display, company_id)
        SELECT pr.report_id, pfo.field_code, pfo.field_name, pfo.is_show, pfo.slno, pfo.field_name_display, {company_id}
        FROM (SELECT DISTINCT report_id FROM power_report) AS pr
        CROSS JOIN power_report_fields_original AS pfo
        ORDER BY pr.report_id
        ''')
        cnx.execute(sql)
        cnx.commit()
    
        return JSONResponse({"iserror": False, "message": "sata saved successfully", "data":" "})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_company/")
async def remove_company(company_id: int = Form(None),
                         status : str = Form(None),
                         cnx: Session = Depends(get_db)):    
    try:
        if company_id is not None:
            if status is not None:
                query = text(f"UPDATE master_company SET status = '{status}' WHERE company_id = {company_id}")        

            else:
                query = text(f"UPDATE master_company SET status = 'delete' WHERE company_id = {company_id}")         
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror": False, "message": "status update successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_branch_list/{branch_id}")
@app.post("/get_branch_list/")
async def branch_list(branch_id: Optional[int] = None, cnx: Session = Depends(get_db)):    
    try:        
        where = ""
        if branch_id is not None:
            where = f" and branch_id = {branch_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.*,
                       IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM master_branch mb
                       left join master_employee cu on cu.employee_id=mb.created_by
                       left join master_employee mu on mu.employee_id=mb.modified_by
                       INNER JOIN master_company mc ON mb.company_id = mc.company_id
                    WHERE mb.status != 'delete' {where}
                    ''')
        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_branch_details/")
async def save_branch_details(company_id: int = Form(None),
                              branch_id:int = Form(None),
                              branch_name: str = Form(None),
                              branch_code: str = Form(None),
                              address: str = Form(None),
                              city: str = Form(None),
                              state:str = Form(None),
                              country:str= Form(None),
                              pincode:str = Form(None),
                              phone_number :str=Form(None),
                              mobile_number : str=Form(None),
                              email_id: str=Form(None),
                              website:str=Form(None),
                              user_login_id : str = Form(None),
                              cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": False, "message": " company_id is required"})
     
    if branch_code == None:  
        return JSONResponse({"iserror": False, "message": " branch_code is required"})

    if branch_name == None:  
        return JSONResponse({"iserror": False, "message": " branch_name is required"})

    try:        
        if branch_id is not None:
            query = text(f"""
                UPDATE master_branch 
                SET company_id = {company_id}, branch_name = '{branch_name}', branch_code = '{branch_code}',
                address = '{address}', city = '{city}', state = '{state}', country = '{country}', pincode = '{pincode}',
                phone_number = '{phone_number}', mobile_number = '{mobile_number}', email_id = '{email_id}',
                website = '{website}', modified_on = NOW(), modified_by = '{user_login_id}'
                WHERE branch_id = {branch_id}

            """)
        else:
            
            select_query = text(f'''select * from master_branch where branch_code = '{branch_code}'  and status != 'delete'  ''')
            data =cnx.execute(select_query).fetchall()
            
            if len(data)>0:
                return JSONResponse({"iserror":True,"message":"branch_code already exists "})
            
            query = text(f"""
                INSERT INTO master_branch (company_id, branch_name, branch_code, created_on, 
                created_by, address, city, state, country, pincode, phone_number, mobile_number, email_id, website)
                
                VALUES ({company_id},'{branch_name}', '{branch_code}', NOW(), '{user_login_id}', '{address}',
                '{city}','{state}','{country}','{pincode}','{phone_number}',
                '{mobile_number}','{email_id}','{website}')
            """)        
        cnx.execute(query)
        cnx.commit()
        
        sql = text(f''' select * from master_branch where company_id = '{company_id}' ''')
        data1 = cnx.execute(sql).fetchall()
        if len(data1)>0:
            sql1= text(f''' update master_branch set is_assign = 'yes' where company_id = '{company_id}' ''')
            cnx.execute(sql1)
            cnx.commit()
    
        return JSONResponse({"iserror": False, "message": "data saved successfully", "data":" "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_branch/")
async def remove_branch(branch_id: int = Form(None), 
                        status : str = Form(None),
                        cnx: Session = Depends(get_db)):    
    try: 
        if branch_id is not None:
            if status is not None:
                query = text(f'''UPDATE master_branch SET status = '{status}' WHERE branch_id = {branch_id}''')      

            else:
                query = text(f'''UPDATE master_branch SET status = 'delete' WHERE branch_id = {branch_id}''')
            cnx.execute(query)
            cnx.commit()
        query = text(f'''SELECT * FROM master_branch WHERE company_id = (SELECT company_id FROM master_branch WHERE branch_id = '{branch_id}') AND status != 'delete' ''')
        result = cnx.execute(query).fetchall()           
        # If no active branches are left, update the is_assign status of the company to "no"
        if result == []:
            query = text(f'''UPDATE master_company SET is_assign = 'no' WHERE company_id = (SELECT company_id FROM master_branch WHERE branch_id = '{branch_id}')''')
            cnx.execute(query)
            cnx.commit()  
       
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_department_list/{department_id}")
@app.post("/get_department_list/")
async def department_list(department_id: Optional[int] = None , cnx: Session = Depends(get_db)):    
    try:        
        where = ""
        if department_id is not None:
            where = f" and department_id = {department_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_name AS branch_name,
                       mb.branch_code AS branch_code,
                       md.*,
                       IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM master_department md
                       left join master_employee cu on cu.employee_id=md.created_by
                       left join master_employee mu on mu.employee_id=md.modified_by
                       INNER JOIN master_company mc ON md.company_id = mc.company_id
                       INNER JOIN master_branch mb ON  md.branch_id = mb.branch_id
                    WHERE md.status != 'delete' {where}
                    ''')      
        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_department_details/")
async def save_department_details(company_id: int = Form(None),
                                  branch_id: int = Form(None),
                                  department_id: Optional[int] = Form(None),
                                  department_name: str = Form(None),
                                  department_code: str = Form(None),                                                                 
                                  user_login_id: str = Form(None),
                                  cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": False, "message": " company_id is required"})
     
    if branch_id == None:  
        return JSONResponse({"iserror": False, "message": " branch_id is required"})
    
    if department_code == None:  
        return JSONResponse({"iserror": False, "message": " department_code is required"})

    if department_name == None:  
        return JSONResponse({"iserror": False, "message": " department_name is required"})

    try:        
        if department_id is not None:
            query = text(f"""
                UPDATE master_department 
                SET company_id = {company_id}, department_name = '{department_name}', department_code = '{department_code}',
                 modified_on = NOW(), modified_by = '{user_login_id}', branch_id = {branch_id}
                WHERE department_id = {department_id}
            """)
            
        else:            
            select_query = text(f'''select * from master_department where department_code = '{department_code}' and status != 'delete'  ''')
            data = cnx.execute(select_query).fetchall()

            if len(data)>0:
                return JSONResponse({"iserror": True, "message": "department code already exists"})

            query = text(f"""
                INSERT INTO master_department (company_id, department_name, department_code,
                created_on, created_by, branch_id)
                VALUES ({company_id},'{department_name}', '{department_code}', NOW(), '{user_login_id}', 
                {branch_id})
            """)
        cnx.execute(query)
        cnx.commit()
        sql = text(f''' select * from master_department where branch_id = '{branch_id}' ''')
        data1 = cnx.execute(sql).fetchall()
        if len(data1)>0:
            sql1= text(f''' update master_branch set is_assign = 'yes' where branch_id = '{branch_id}' ''')
            cnx.execute(sql1)
            cnx.commit()

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_department/")
async def remove_department(department_id: int = Form(None),
                            status : str = Form(None),
                            cnx: Session = Depends(get_db)):    
    try: 
        if department_id is not None:

            if status is not None:       
                query = text(f'''UPDATE master_department SET status = '{status}' WHERE department_id = {department_id}''')

            else:
                query = text(f'''UPDATE master_department SET status = 'delete' WHERE department_id = {department_id}''')            
            cnx.execute(query)
            cnx.commit()
        
        query = text(f'''SELECT * FROM master_department WHERE branch_id = (SELECT branch_id FROM master_department WHERE department_id = '{department_id}') AND status != 'delete' ''')
        result = cnx.execute(query).fetchall()
                   
        if result == []:
            query = text(f'''UPDATE master_branch SET is_assign = 'no' WHERE branch_id = (SELECT branch_id FROM master_department WHERE department_id = '{department_id}')''')
            cnx.execute(query)
            cnx.commit() 
       
        return JSONResponse({"iserror":False,"message":" status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_shed_list/{shed_id}")
@app.post("/get_shed_list/")
async def shed_list(shed_id: Optional[int] = None,
                    department_id : int = Form(None),
                    cnx: Session = Depends(get_db)):    
    try:        
        where = ""
        if shed_id is not None:
            where = f" and ms.shed_id = {shed_id}"
        
        if department_id is not None:
            where = f" and ms.department_id = {department_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_name AS branch_name,
                       mb.branch_code AS branch_code,
                       md.department_name AS department_name ,
                       md.department_code AS department_code,
                       ms.*,
                       IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM master_shed ms
                       left join master_employee cu on cu.employee_id=ms.created_by
                       left join master_employee mu on mu.employee_id=ms.modified_by
                       INNER JOIN master_company mc ON ms.company_id = mc.company_id
                       INNER JOIN master_branch mb ON  ms.branch_id = mb.branch_id
                       INNER JOIN master_department md ON ms.department_id = md.department_id
                    WHERE ms.status != 'delete' {where}
                    ''')       
        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_shed_details/")
async def save_shed_details(company_id: int = Form(None),  
                            branch_id:int = Form(None),
                            shed_id : int = Form(None),
                            department_id: int = Form(None),        
                            shed_name: str = Form(None),
                            shed_code: str = Form(None),
                            user_login_id : str = Form(None),                                                   
                            cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": False, "message": " company_id is required"})
     
    if branch_id == None:  
        return JSONResponse({"iserror": False, "message": " branch_id is required"})
    
    if department_id == None:  
        return JSONResponse({"iserror": False, "message": " department_id is required"}) 
    
    if shed_code == None:  
        return JSONResponse({"iserror": False, "message": " shed_code is required"})

    if shed_name == None:  
        return JSONResponse({"iserror": False, "message": " shed_name is required"})
    
    try:        
        if shed_id is not None:
            query =text(f"""
                UPDATE master_shed
                SET company_id = {company_id}, shed_name = '{shed_name}', shed_code = '{shed_code}',
                branch_id = {branch_id}, department_id = {department_id}, modified_on = NOW(),
                modified_by = '{user_login_id}'
                WHERE shed_id = {shed_id}
            """)         
  
        else:            
            select_query = text(f'''select * from master_shed where shed_code = '{shed_code}' and status != 'delete' ''')
            data = cnx.execute(select_query).fetchall()

            if len(data) > 0:
              return JSONResponse({"iserror": True, "message": "shed_code already exists" })

            query = text(f"""
                INSERT INTO master_shed (
                    company_id, shed_name, shed_code, branch_id, department_id, created_on, created_by
                )
                VALUES (
                    {company_id},'{shed_name}', '{shed_code}', {branch_id}, {department_id}, NOW(), '{user_login_id}'
                )
            """)    
        cnx.execute(query)
        cnx.commit()
        sql = text(f''' select * from master_shed where department_id = '{department_id}' ''')
        data1 = cnx.execute(sql).fetchall()
        if len(data1)>0:
            sql1= text(f''' update master_department set is_assign = 'yes' where department_id = '{department_id}' ''')
            cnx.execute(sql1)
            cnx.commit()

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_shed/")
async def remove_shed(shed_id: int = Form(None),
                      status : str = Form(None), 
                      cnx: Session = Depends(get_db)):
    try:
        if shed_id is not None:
            if status is not None:
                query = text(f'''UPDATE master_shed  SET status = '{status}' WHERE shed_id = {shed_id}''')   

            else:
                query = text(f'''UPDATE master_shed  SET status = 'delete' WHERE shed_id = {shed_id}''')     
            cnx.execute(query)
            cnx.commit()
        query = text(f'''SELECT * FROM master_shed WHERE department_id = (SELECT department_id FROM master_shed WHERE shed_id = '{shed_id}') AND status != 'delete' ''')
        result = cnx.execute(query).fetchall()
        if result == []:
            query = text(f'''UPDATE master_department SET is_assign = 'no' WHERE department_id = (SELECT department_id FROM master_shed WHERE shed_id = '{shed_id}')''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_machinetype_list/{machinetype_id}")
@app.post("/get_machinetype_list/")
async def machinetype_list(machinetype_id: Optional[int] = None,
                           department_id : int = Form(None),
                           cnx: Session = Depends(get_db)):
    try:        
        where = ""
        if machinetype_id is not None:
            where = f" and machinetype_id = {machinetype_id}"
            
        if department_id is not None:
            where = f" and mmt.department_id = {department_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_name AS branch_name,
                       mb.branch_code AS branch_code,
                       md.department_name AS department_name ,
                       md.department_code AS department_code,
                       mmt.*,
                       IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM master_machinetype  mmt
                       left join master_employee cu on cu.employee_id=mmt.created_by
                       left join master_employee mu on mu.employee_id=mmt.modified_by
                       INNER JOIN master_company  mc ON mmt.company_id = mc.company_id
                       INNER JOIN master_branch  mb ON  mmt.branch_id = mb.branch_id
                       INNER JOIN master_department  md ON mmt.department_id = md.department_id
                    WHERE mmt.status != 'delete'{where}
                    ''')        
        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_machinetype_details/")
async def save_machinetype_details(company_id: int = Form(None),
                                   branch_id:int = Form(None),
                                   department_id: int = Form(None),
                                   machinetype_id : int = Form(None),
                                   machinetype_name: str = Form(None),
                                   machinetype_code: str = Form(None),
                                   user_login_id : str = Form(None),                                                     
                                   cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": False, "message": " company_id is required"})
     
    if branch_id == None:  
        return JSONResponse({"iserror": False, "message": " branch_id is required"})
    
    if department_id == None:  
        return JSONResponse({"iserror": False, "message": " department_id is required"}) 
    
    if machinetype_code == None:  
        return JSONResponse({"iserror": False, "message": " machinetype_code is required"}) 
    
    if machinetype_name == None:  
        return JSONResponse({"iserror": False, "message": " machinetype_name is required"}) 
        
    try:        
        if machinetype_id is not None:
            query =text(f"""
                UPDATE master_machinetype
                SET company_id = {company_id}, machinetype_name = '{machinetype_name}', machinetype_code = '{machinetype_code}',
                branch_id = {branch_id}, department_id = {department_id}, modified_on = NOW(),
                modified_by = '{user_login_id}'
                WHERE machinetype_id = {machinetype_id}
            """)
        else:            
            select_query = text(f'''select * from master_machinetype where machinetype_code = '{machinetype_code}' and status != 'delete' ''')
            data = cnx.execute(select_query).fetchall()

            if len(data) > 0:
              return JSONResponse({"iserror": True, "message": "machinetype_code already exists" })

            query = text(f"""
                INSERT INTO master_machinetype (
                    company_id, machinetype_name, machinetype_code, branch_id, department_id, created_on, created_by
                )
                VALUES (
                    {company_id},'{machinetype_name}', '{machinetype_code}', {branch_id}, {department_id}, NOW(), '{user_login_id}'
                )
            """)    
        cnx.execute(query)
        cnx.commit()
        sql = text(f''' select * from master_machinetype where department_id = '{department_id}' ''')
        data1 = cnx.execute(sql).fetchall()
        
        if len(data1)>0:
            sql1= text(f''' update master_department set is_assign = 'yes' where department_id = '{department_id}' ''')
            cnx.execute(sql1)
            cnx.commit()

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_machinetype/")
async def remove_machinetype(machinetype_id: int = Form(None), 
                             status : str = Form(None),
                             cnx: Session = Depends(get_db)):
    try:
        if machinetype_id is not None:
            if status is not None:
                query = text(f'''UPDATE master_machinetype  SET status = '{status}' WHERE machinetype_id = {machinetype_id}''')        

            else:
                 query = text(f'''UPDATE master_machinetype  SET status = 'delete' WHERE machinetype_id = {machinetype_id}''')        
            cnx.execute(query)
            cnx.commit()
        query = text(f'''SELECT * FROM master_machinetype WHERE department_id = (SELECT department_id FROM master_machinetype WHERE machinetype_id = '{machinetype_id}') AND status != 'delete' ''')
        result = cnx.execute(query).fetchall()
        if result == []:
            query = text(f'''UPDATE master_department SET is_assign = 'no' WHERE department_id = (SELECT department_id FROM master_machinetype WHERE machinetype_id = '{machinetype_id}')''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_function_list/{function_id}")
@app.post("/get_function_list/")
async def function_list(request:Request,
                        function_id:int = None, 
                        cnx:Session = Depends(get_db)):
    try:                                     
        where=" "
        if function_id is not None:
            where = f"and function_id = {function_id}"
            
        query =text( f''' 
                    SELECT 
                    	mf.* , 
                    	CONCAT('/attachment/images/',mf.image) AS image,
                        IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                    	IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM
                    	master_function as mf 
                    	left join master_employee cu on cu.employee_id=mf.created_by
                    	left join master_employee mu on mu.employee_id=mf.modified_by
                    WHERE
                    	mf.status !='delete' {where} 
                    ''')
        data=cnx.execute(query).fetchall()
        results = []
        for i in data:
            new_img = dict(i)
            if new_img["image"] is not None:
                new_img["image"] = "http://" + request.headers["host"] + new_img['image']
            results.append(new_img)   
        
        return JSONResponse({"iserror":False, "message": " data return successfully","data":jsonable_encoder(results)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_function_details/")
async def save_function_details(function_id: int = Form(None),
                                function_name: str = Form(None),
                                function_code: str = Form(None),
                                image: UploadFile = File(None),
                                user_login_id: str = Form(None),
                                cnx: Session = Depends(get_db)):
    if function_code == None:  
        return JSONResponse({"iserror": False, "message": " function_code is required"}) 
    if function_name == None:  
        return JSONResponse({"iserror": False, "message": " function_name is required"})    
           
    try:
        if image is not None:           
            # get base path to save image
            base_path = os.path.abspath(os.path.dirname(__file__)) + "/attachment/images/"
            if not os.path.exists(base_path):
                os.makedirs(base_path)
            random_number = random.randint(1, 100000) 
            extension = os.path.splitext(image.filename)[1].lower()
            # generate filename
            filename = f"{random_number}_{datetime.datetime.now().strftime('%y_%m_%d_%H_%M_%S')}{extension}"

            # save image to disk
            with open(base_path + filename, "wb") as f:
                f.write(await image.read())
        else:
            filename = ''
        if function_id is not None:            
            query = text(f"""
                UPDATE master_function SET function_name = '{function_name}', 
                function_code = '{function_code}',
                image = '{filename}',
                modified_on = NOW(), 
                modified_by = '{user_login_id}' WHERE function_id = {function_id}
            """)
            
        else:           
            select_query = text(f'''select * from master_function where function_code = '{function_code}' and status != 'delete'  ''')
            data = cnx.execute(select_query).fetchall()
            if len(data) > 0:
                return JSONResponse({"iserror": True, "message": "function code already exists"})
            
            query = text(f"""
                INSERT INTO master_function (function_name, function_code, image, created_on, created_by)
                VALUES ('{function_name}', '{function_code}', '{filename}', NOW(), '{user_login_id}')
            """)
        cnx.execute(query)
        cnx.commit()        
        
        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_function/")
async def remove_function(function_id: int = Form(None), 
                          status : str = Form(None),
                          cnx: Session = Depends(get_db)):    
    try:
        if function_id is not None:
            if status is not None:        
                query = text(f'''UPDATE master_function SET status = '{status}' WHERE function_id = {function_id}''')

            else:
                 query = text(f'''UPDATE master_function SET status = 'delete' WHERE function_id = {function_id}''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_converter_list/{converter_id}")
@app.post("/get_converter_list/")
async def converter_list(converter_id:int = None, cnx:Session = Depends(get_db)):    
    try:        
        where=" "
        if converter_id is not None:
            where = f"and converter_id = {converter_id}"
            
        query =text( f''' 
                    SELECT 
                    	mcd.*,
                    	IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                    	IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                    FROM  
                    	master_converter_detail mcd
                    	left join master_employee cu on cu.employee_id=mcd.created_by
                    	left join master_employee mu on mu.employee_id=mcd.modified_by
                    WHERE 
                    	mcd.status !='delete'
                    	{where} 
                    ''')
        data=cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror":False, "message": " data return successfully","data":jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
 
@app.post("/save_converter_details/")
async def save_converter_details(converter_id:int=Form(None),
                                 converter_name:str = Form(None),
                                 ip_address:str=Form(None),
                                 port_no:str=Form(None),
                                 user_login_id:str=Form(None),
                                 cnx:Session=Depends(get_db)):
    if converter_name == None:  
        return JSONResponse({"iserror": False, "message": " converter_name is required"}) 
      
    if ip_address == None:  
        return JSONResponse({"iserror": False, "message": " ip_address is required"})
    
    if port_no == None:  
        return JSONResponse({"iserror": False, "message": " port_no is required"})    
    try:        
        if converter_id is not None:
            query =text(f"""
                UPDATE master_converter_detail
                SET converter_name = '{converter_name}', ip_address = '{ip_address}',
                port_no = {port_no},  modified_on = NOW(),
                modified_by = '{user_login_id}'
                WHERE converter_id = {converter_id}
            """)
        else:        
            query = text(f"""
                INSERT INTO master_converter_detail (
                     converter_name, ip_address, port_no, created_on, created_by
                )
                VALUES (
                    '{converter_name}', '{ip_address}', {port_no},  NOW(), '{user_login_id}'
                )
            """)    
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_converter_details/")
async def remove_converter_detail(converter_id: int = Form(None),
                                  status : str = Form(None),
                                  cnx: Session = Depends(get_db)):    
    try:
        if converter_id is not None:
            if status is not None:
                query = text(f'''UPDATE master_converter_detail  SET status = '{status}' WHERE converter_id = {converter_id}''')

            else:
                query = text(f'''UPDATE master_converter_detail  SET status = 'delete' WHERE converter_id = {converter_id}''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_machine_list/{machine_id}")
@app.post("/get_machine_list/")
async def machine_list(company_id : int= Form(None),
                       machine_id: Optional[int] = None, 
                       type_value: str = Form(None),
                       type_id : str = Form(None), 
                       is_critical : str = Form(None),
                       model_name: str = Form(None),
                       department_id : int = Form(None),
                       shed_id : int = Form(None),
                       machinetype_id : int = Form(None),
                       function_id : int = Form(None),
                       cnx: Session = Depends(get_db)):
    try:        
       
        if type_id is not None:
            value = type_id.split(",")
            if len(value) > 1:
                values = tuple(value)
                type_id = ",".join(values)
            else:
                type_id = value[0]
                
        where = ""
        if company_id is not None:
            where += f" and mm.company_id = {company_id}" 
        if machine_id is not None:
            where += f" and mm.machine_id = {machine_id}"  
        if type_value is not None and type_id is not None:
            if type_value == 'zone':
                where += f" and mm.department_id in ({','.join(str(x) for x in value)})"
          
            elif type_value == 'area':
                where += f" and mm.shed_id in ({','.join(str(x) for x in value)})"

            elif type_value == 'location':
                where += f" and mm.machinetype_id in ({','.join(str(x) for x in value)})"

            elif type_value == 'function':
                where += f" and mm.function_id in ({','.join(str(x) for x in value)})"
                
        if is_critical == "yes" or is_critical == "no"  :
            where += f" and mm.major_nonmajor = '{is_critical}' "   
        
        if model_name is not None:
            where += f" and mm.model_name"  
        
        if department_id is not None:
            where += f" and mm.department_id = {department_id}"
            
        if shed_id is not None:
            where +=f" and mm.shed_id = {shed_id}"
        
        if machinetype_id is not None:
            where += f" and mm.machinetype_id = {machinetype_id}"
            
        if function_id is not None:
            where += f" and mm.function_id = {function_id}"                         
        
        query = text(f"""SELECT mm.*,
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_code AS branch_code,
                       mb.branch_name AS branch_name,
                       md.department_code AS department_code,
                       md.department_name AS department_name,
                       ms.shed_code AS shed_code,
                       ms.shed_name AS shed_name,
                       mmt.machinetype_code AS machinetype_code,
                       mmt.machinetype_name AS machinetype_name,
                       mcd.converter_name AS converter_name,
                       mf.function_code AS function_code,
                       mf.function_name AS function_name,
                       IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                   IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                    FROM 
                        master_machine mm
                        left join master_employee cu on cu.employee_id=mm.created_by
	                    left join master_employee mu on mu.employee_id=mm.modified_by                    
                        INNER JOIN master_company mc ON mm.company_id = mc.company_id
                        INNER JOIN master_branch mb ON mm.branch_id = mb.branch_id
                        INNER JOIN master_department md ON mm.department_id = md.department_id
                        INNER JOIN master_shed ms ON mm.shed_id = ms.shed_id
                        INNER JOIN master_machinetype mmt ON mm.machinetype_id = mmt.machinetype_id
                        INNER JOIN master_converter_detail mcd ON mm.converter_id = mcd.converter_id
                        INNER JOIN master_function mf ON mm.function_id = mf.function_id
                    WHERE 
                        mm.status != 'delete' {where}""")
        data = cnx.execute(query).fetchall()

        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_machine_details/")
async def save_machine_details(company_id: int = Form(None),
                               branch_id:int = Form(None),
                               department_id: int = Form(None),
                               shed_id: int = Form(None), 
                               machinetype_id: int = Form(None),
                               function_id:int = Form(None),
                               converter_id: int = Form(None),
                               machine_id : str = Form(None),
                               machine_code: str = Form(None),
                               machine_name: str = Form(None),
                               ip_address : str = Form(None),
                               port :int = Form(None),
                               major_nonmajor : str = Form(None), 
                               model_name : str = Form(None),                                                
                               energy_selection : str = Form(None), 
                               IMEI : int = Form(None),                                                 
                               user_login_id : str = Form(None),                             
                               cnx: Session = Depends(get_db)):

    if company_id == None:  
        return JSONResponse({"iserror": False, "message": " company_id is required"}) 
    
    if branch_id == None:  
        return JSONResponse({"iserror": False, "message": " branch_id is required"})    
    
    if department_id == None:  
        return JSONResponse({"iserror": False, "message": " department_id is required"})   
    
    if shed_id == None:  
        return JSONResponse({"iserror": False, "message": " shed_id is required"})
    
    if machinetype_id == None:  
        return JSONResponse({"iserror": False, "message": " machinetype_id is required"}) 
    
    if function_id == None:  
        return JSONResponse({"iserror": False, "message": " function_id is required"})    
    
    if converter_id == None:  
        return JSONResponse({"iserror": False, "message": " converter_id is required"}) 
       
    if machine_name == None:
        return JSONResponse({"iserror": False, "message": " machine_name is required"})
     
    if machine_code == None:  
        return JSONResponse({"iserror": False, "message": " machine_code is required"}) 
    
    if ip_address == None:  
        return JSONResponse({"iserror": False, "message": " ip_address is required"}) 
    
    if port == None:  
        return JSONResponse({"iserror": False, "message": " port is required"})    

    if major_nonmajor == None:  
        return JSONResponse({"iserror": False, "message": " major_nomajor is required"})
    
    if model_name == None:  
        return JSONResponse({"iserror": False, "message": " model_name is required"})
    
    if IMEI == None:  
        return JSONResponse({"iserror": False, "message": " IMEI is required"})
     
    mill_date = ""
    mill_shift = ""
    try:        
        if machine_id is  None:
            select_query = text(f'''select * from master_machine where machine_code = '{machine_code}' and status != 'delete' ''')
            data = cnx.execute(select_query).fetchall()

            if len(data) > 0:
              return JSONResponse({"iserror": True, "message": "machine_code already exists" })

            query = text(f"""
                INSERT INTO master_machine (
                    company_id, machine_name, machine_code, branch_id, department_id, shed_id, converter_id, function_id,machinetype_id,
                    ip_address, port, created_on, created_by, major_nonmajor, model_name, energy_selection, IMEI
                )
                VALUES (
                    {company_id},'{machine_name}', '{machine_code}', {branch_id}, {department_id}, {shed_id}, {converter_id}, {function_id},
                    {machinetype_id}, '{ip_address}',{port}, NOW(), '{user_login_id}', '{major_nonmajor}', '{model_name}','{energy_selection}', {IMEI}
                )
            """)
            
        else:

            sql = text(f'''INSERT INTO ems_v1.master_machine_history] (
                    company_id, machine_name, machine_code, branch_id, department_id, shed_id, converter_id, function_id,machinetype_id,
                    ip_address, port, modified_on, modified_by, major_nonmajor, model_name, energy_selection, IMEI
                )
                VALUES (
                    {company_id},'{machine_name}', '{machine_code}', {branch_id}, {department_id}, {shed_id}, {converter_id}, {function_id},
                    {machinetype_id}, '{ip_address}',{port}, GETDATE(), '{user_login_id}', '{major_nonmajor}', '{model_name}','{energy_selection}',{IMEI}
                ) 
                ''')
            cnx.execute(sql)
            cnx.commit()

            query =text(f"""
                UPDATE master_machine
                SET company_id = {company_id}, machinetype_id = {machinetype_id}, machine_code = '{machine_code}',machine_name = '{machine_name}',
                branch_id = {branch_id}, shed_id = {shed_id},converter_id = {converter_id}, department_id = {department_id},function_id = {function_id},
                ip_address = '{ip_address}', port = '{port}', modified_on = NOW(), modified_by = '{user_login_id}',
                major_nonmajor = '{major_nonmajor}', model_name = '{model_name}', energy_selection = '{energy_selection}', IMEI = {IMEI}
                WHERE machine_id = {machine_id}
            """)            
            
        cnx.execute(query)
        if machine_id is not None:
            insert_id = machine_id
        else:
            insert_id = cnx.execute(text("SELECT LAST_INSERT_ID()")).first()[0]        
        cnx.commit()  

        if insert_id is not None:  
             
            sql = text(f''' select * from current_power where machine_id = {insert_id} ''') 
            data = cnx.execute(sql).fetchall()
            print(data)

            if len(data)==0:                       
                sql1 = text(f"select * from master_machine where machine_id = {insert_id}")

                data1 = cnx.execute(sql1).fetchall()
                for row in data1:
                    machine_id = row["machine_id"]
                    company_id = row["company_id"]
                    branch_id = row["branch_id"]
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]  

                sql2= text(f" select * from master_shifts  where company_id = {company_id} and branch_id = {branch_id} AND status = 'active' ")
                data2 = cnx.execute(sql2).fetchall()
                print(sql2)
                
                if len(data2)>0:
                    for row in data2:
                        mill_date = row["mill_date"]
                        mill_shift = row["mill_shift"]  
                        
                    sql3 = text(f'''
                                INSERT INTO current_power (machine_id, date_time, date_time1,
                                mill_date, mill_shift,company_id, branch_id, department_id, shed_id, machinetype_id)
                                VALUES ({machine_id}, now(), now(), '{mill_date}', '{mill_shift}',{company_id},
                                {branch_id}, {department_id}, {shed_id}, {machinetype_id})
                                ''')  
                    cnx.execute(sql3)
                    cnx.commit()
                    createFolder("Log/","insert query executed successfully")        

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_machine_details/")
async def remove_converter_detail(machine_id: int = Form(None), 
                                  status : str = Form(None),
                                  cnx: Session = Depends(get_db)):    
    try:
        if machine_id is not None:
            if status is not None:        
                query = text(f'''UPDATE master_machine  SET status = '{status}' WHERE machine_id = {machine_id}''')

            else:
                query = text(f'''UPDATE master_machine  SET status = 'delete' WHERE machine_id = {machine_id}''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":" status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_meter_group_list/")
async def meter_group_list(meter_group_id: int = None, cnx: Session = Depends(get_db)):
    try:
        where = ''
        if meter_group_id is not None:
            where = f" AND meter_group_id = {meter_group_id}"
        query = text(f"""
            SELECT 
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                (CASE 
                WHEN group_type='Zone' THEN (SELECT department_name FROM master_department WHERE department_id=type_id)
                WHEN group_type='Area' THEN (SELECT shed_name FROM master_shed WHERE shed_id=type_id)
                WHEN group_type='Location' THEN (SELECT machinetype_name FROM master_machinetype WHERE machinetype_id=type_id)
                END) AS type_name,                
                mmg.*,
                '' AS machine_dtl,
                IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	            IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

            FROM 
                master_meter_group mmg
                left join master_employee cu on cu.employee_id=mmg.created_by
	            left join master_employee mu on mu.employee_id=mmg.modified_by
                INNER JOIN master_machine mm ON mmg.machine_id = mm.machine_id
            WHERE 
                mmg.status != 'delete'{where}
        """)
        data = cnx.execute(query).fetchall()
        result = []
        for row in data:
            machine_id_list = row["machine_id"].split(",")  # Split comma-separated machine IDs into a list
            machine_dtl = ""
            for machine_id in machine_id_list:
                sub_query = text(f"SELECT * FROM master_machine WHERE machine_id = {machine_id}")
                sub_data = cnx.execute(sub_query).fetchall()
                for sub_row in sub_data:
                    if machine_dtl != "":
                        machine_dtl += "\n"
                    machine_dtl += f"{sub_row['machine_name']}"
            new_row = dict(row)
            new_row["machine_dtl"] = machine_dtl
            result.append(new_row)
            
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(result)})             
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
   
@app.post("/save_meter_group/")
async def save_meter_group(meter_group_id:int=Form(None),
                           group_type:str = Form(None),
                           type_id:int=Form(None),
                           machine_id:str=Form(None),
                           user_login_id:str=Form(None),
                           cnx:Session=Depends(get_db)):
    try:
        if machine_id is not None:
                value = machine_id.split(",")
                if len(value) > 1:
                    values = tuple(value)
                    machine_id = ",".join(values)
                else:
                    machine_id = value[0]
        
        if meter_group_id is not None:
            query =text(f"""
                UPDATE master_meter_group
                SET group_type = '{group_type}', type_id = '{type_id}',
                machine_id = '{machine_id}',  modified_on = NOW(),
                modified_by = '{user_login_id}'
                WHERE meter_group_id = {meter_group_id}
            """)
         
        else:
            if machine_id is not None:
                  value = machine_id.split(",")
                  if len(value) > 1:
                      values = tuple(value)
                      machine_id = ",".join(values)
                  else:
                      machine_id = value[0]                

            query = text(f"""
                INSERT INTO master_meter_group (
                     group_type, type_id, machine_id, created_on, created_by
                )
                VALUES (
                    '{group_type}', '{type_id}', '{machine_id}',  NOW(), '{user_login_id}'
                )
            """)   
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_meter_group/")
async def remove_meter_group(meter_group_id: int = Form(None), 
                             status : str = Form(None),
                             cnx: Session = Depends(get_db)):    
    try:
        if meter_group_id is not None:
            if status is not None:
                query = text(f'''UPDATE master_meter_group  SET status = '{status}' WHERE meter_group_id = {meter_group_id}''')

            else:
                query = text(f'''UPDATE master_meter_group  SET status = 'delete' WHERE meter_group_id = {meter_group_id}''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_shift_list/{shift_id}")
@app.post("/get_shift_list/")
async def shift_list(shift_id: Optional[int] = None, cnx: Session = Depends(get_db)):
    try:
        where = ""
        if shift_id is not None:
            where = f" and shift_id = {shift_id}"

        query = text(f"""
                SELECT 
                    mc.company_code AS company_code,
                    mc.company_name AS company_name,
                    mb.branch_code AS branch_code,
                    mb.branch_name AS branch_name,
                    DATE_FORMAT(ms.shift1_start_time,'%r') AS shift1_time,
                    DATE_FORMAT(ms.shift2_start_time,'%r') AS shift2_time,
                    DATE_FORMAT(ms.shift3_start_time,'%r') AS shift3_time,
                    ms.*,
                    IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                FROM 
                    master_shifts ms
                    left join master_employee cu on cu.employee_id=ms.created_by
	                left join master_employee mu on mu.employee_id=ms.modified_by
                    INNER JOIN master_company mc ON ms.company_id = mc.company_id
                    INNER JOIN master_branch mb ON ms.branch_id = mb.branch_id
                WHERE ms.status != 'delete' {where}""")
        data = cnx.execute(query).fetchall()

        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:       
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_shifts_details/")
async def save_shifts_details(company_id: int = Form(None),
                              branch_id:int = Form(None),
                              shift_id : str = Form(None),
                              no_of_shifts:int= Form(None),
                              shift1_start_time:str=Form(None),
                              shift2_start_time:str=Form(None),
                              shift3_start_time:str=Form(None),
                              user_login_id : str = Form(None),                             
                              cnx: Session = Depends(get_db)):
    
    if company_id == None:  
        return JSONResponse({"iserror": False, "message": " company_id is required"})
    
    if branch_id == None:  
        return JSONResponse({"iserror": False, "message": " branch_id is required"}) 
    
    if no_of_shifts == None:
         
        return JSONResponse({"iserror": False, "message": " no_of_shifts is required"}) 
    
    if no_of_shifts == 1 and (shift1_start_time is None):
            return JSONResponse({"iserror": True, "message": "shift1_start_time is required"})        
        
    if no_of_shifts == 2 and (shift1_start_time is None or shift2_start_time is None) :
        return JSONResponse({"iserror": True, "message": "shift1_start_time and shift2_start_time is required"})
    
    if no_of_shifts == 3 and (shift1_start_time is None or shift2_start_time is None or shift3_start_time is None):
        return JSONResponse({"iserror": True, "message": "shift1_start_time and shift2_start_time and shift3_start_time are required"})
      
    try:                              
        if shift_id is not None:              
            query =text(f"""
                UPDATE master_shifts
                SET company_id = {company_id}, branch_id = {branch_id}, no_of_shifts = {no_of_shifts},
                shift1_start_time=concat(date('1900-01-01'), 'T', time('{shift1_start_time}')),
                shift2_start_time=concat(date('1900-01-01'), 'T', time('{shift2_start_time}')),
                shift3_start_time=concat(date('1900-01-01'), 'T', time('{shift3_start_time}')),
                modified_on = NOW(),modified_by = '{user_login_id}' 
                WHERE shift_id = {shift_id}
            """)           
  
        else:            
            query = text(f"""
                INSERT INTO master_shifts (
                     company_id, branch_id, no_of_shifts, shift1_start_time, shift2_start_time, shift3_start_time, created_on, created_by
                )
                VALUES (
                   {company_id}, {branch_id}, {no_of_shifts}, 
                   concat(date('1900-01-01'), 'T', time('{shift1_start_time}')),
                   concat(date('1900-01-01'), 'T', time('{shift2_start_time}')),
                   concat(date('1900-01-01'), 'T', time('{shift3_start_time}')),
                   NOW(), '{user_login_id}'
                )
            """)
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data saved successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_shift_details/")
async def remove_shift_detail(shift_id: int = Form(None),
                              status : str = Form(None),
                              cnx: Session = Depends(get_db)):    
    try:
        if shift_id is not None:
            if status is not None:        
                query = text(f'''UPDATE master_shifts  SET status = '{status}' WHERE shift_id = {shift_id}''')

            else:
                query = text(f'''UPDATE master_shifts  SET status = 'delete' WHERE shift_id = {shift_id}''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_employee_list/")
async def get_employeelist(request:Request,
                           employee_id:str=Form(None),
                           cnx: Session = Depends(get_db)):
    
    try:    
        where = ""
        if employee_id is not None:
            where = f"and me.employee_id = '{employee_id}'"

        query=text(f'''
                    SELECT
                        me.*,
                        concat('/attachment/employee_images/',me.employee_image) AS employee_image,
                        mc.company_name AS company_name,
                        mb.branch_name AS branch_name,
                        IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                    IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM 
                        master_employee me
                        left join master_employee cu on cu.employee_id=me.created_by
	                    left join master_employee mu on mu.employee_id=me.modified_by
                        INNER JOIN master_company mc ON mc.company_id=me.company_id
                        INNER JOIN master_branch mb ON mb.branch_id=me.branch_id                        
                        WHERE me.status!='delete' {where} ''')      
        
        data = cnx.execute(query).fetchall()
        results = []
        for i in data:
            new_img= dict(i)
            if new_img["employee_image"] is not None:
                new_img["employee_image"] = "http://" + request.headers["host"] + new_img['employee_image']
            results.append(new_img)
        createFolder("Log/","Query executed successfully for  employee list")
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(results)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_employee_detail/")
async def save_employee_detail(employee_id:str=Form(None),
                               company_id:str=Form(None),
                               branch_id:str=Form(None),
                               employee_code:str=Form(None),
                               employee_name:str=Form(None),
                               employee_type:str=Form(None),
                               mobileno:str=Form(None),
                               email:str=Form(None),
                               password_login:str=Form(None),
                               is_login:str=Form(None),
                               employee_image : UploadFile = File(None),
                               login_id:str=Form(None),
                               cnx: Session = Depends(get_db)):
    
    if company_id == None:
        return JSONResponse({"iserror":True,"message":"company id is required"}) 
    
    if branch_id == None:
        return JSONResponse({"iserror":True,"message":"branch id is required"})

    if employee_code == None:
        return JSONResponse({"iserror":True,"message":"employee code is required"})
    
    if employee_name == None:
        return JSONResponse({"iserror":True,"message":"employee name is required"})
    
    if employee_type == None:
        return JSONResponse({"iserror":True,"message":"employee type is required"})

    if password_login == None:
        return JSONResponse({"iserror":True,"message":" login password is required"})
    
    if is_login == None:
        return JSONResponse({"iserror":True,"message":" is login is required"})    
    
    if mobileno is None:
        mobileno = ''
    
    if email is None:
        email = ''
    
    
    try:
        if employee_image is None:
            filename = ''
        else:
            base_path = os.path.abspath(os.path.dirname(__file__)) + "/attachment/employee_images/"
            if not os.path.exists(base_path):
                os.makedirs(base_path)
            random_number = random.randint(1, 100000)
            extension = os.path.splitext(employee_image.filename)[1].lower()
            # generate filename
            filename = f"{random_number}_{datetime.datetime.now().strftime('%y_%m_%d_%H_%M_%S')}{extension}"
            # save image to disk
            with open(base_path + filename, "wb") as f:
                f.write(await employee_image.read())
        
        if employee_id is not None:
            query =text(f'''update  master_employee set company_id = '{company_id}',branch_id = '{branch_id}',  
                       employee_name = '{employee_name}',employee_code = '{employee_code}',employee_type = '{employee_type}',
                       mobileno = '{mobileno}',email = '{email}',password_login= MD5('{password_login}'),is_login='{is_login}',
                       employee_image = '{filename}', modified_on = now(),modified_by='{login_id}' where employee_id = '{employee_id}'
                       ''')
            
        else:
            select_query = text(f'''select * from master_employee where employee_code = '{employee_code}' and status != 'delete' ''')
            data1 = cnx.execute(select_query).fetchall()

            if len(data1)>0:
                return JSONResponse({"iserror":True,"message":"employee code already exists "})

            query= text(f'''insert into master_employee (company_id,branch_id,employee_name,
                       employee_code,employee_type,mobileno,email,password_login,is_login,employee_image,created_on,created_by )
                       values('{company_id}','{branch_id}' ,'{employee_name}','{employee_code}','{employee_type}',
                       '{mobileno}','{email}', MD5('{password_login}'),'{is_login}','{filename}',now(),'{login_id}') ''')
        cnx.execute(query)
        cnx.commit()            

        createFolder("Log/","Query executed successfully for save employee")
        return JSONResponse({"iserror":False,"message":"data saved successfully","data":""})
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_employee_detail/")
async def remove_employee_detail(employee_id:str=Form(None),
                                 status : str = Form(None),
                                 cnx: Session = Depends(get_db)):
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})    
    
    try:
        if employee_id is not None:
            if status is not None:
                query = text(f''' update master_employee set status = '{status}' where employee_id = '{employee_id}' ''')

            else:
                 query = text(f''' update master_employee set status = 'delete' where employee_id = '{employee_id}' ''')
            cnx.execute(query)
            cnx.commit()
        
        createFolder("Log/","query executed successfully for remove employee")
        return JSONResponse({"iserror":False,"message":"data remove successfully","data":""}) 
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})


@app.post("/current_power/")
async def current_power(company_id : str = Form(None),
                        branch_id : str = Form(None),
                        department_id :str = Form(None),
                        shed_id :str = Form(None),
                        machinetype_id :str = Form(None),
                        function_id : str = Form(None),
                        machine_id : str = Form (None),
                        group_for : str = Form(None),
                        groupby : str = Form(None),
                        period_id: str = Form(None),
                        from_date: str = Form(None),
                        to_date: str = Form(None),                      
                        shift_id: int = Form(None),
                        limit_report_for = Form(None),
                        limit_exception_for:str = Form(None),
                        limit_order_by : str = Form(None),
                        limit_operation_value : str = Form(None),
                        is_critical :str = Form(None),
                        converter_id :int = Form(None),  
                        report_for : str = Form(None),                     
                        cnx: Session = Depends(get_db)):
    
    if period_id is None:
            return JSONResponse({"iserror": True, "message": "period id is required"})
    
    if groupby is None:
        return JSONResponse({"iserror":True, "message": "groupby is required"}) 
    
    if group_for is None:
        return JSONResponse({"iserror":True, "message": "group_for is required"}) 
    try:
        def id(machine_id):
            if machine_id is not None:
                value = machine_id.split(",")
                if machine_id == 'all':
                    pass
                if len(value) > 1:
                    values = tuple(value)
                    machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
     
        machine_id = id(machine_id)
        company_id = id(company_id)
        department_id = id(department_id)
        shed_id = id(shed_id)
        machinetype_id = id(machinetype_id)
        function_id = id(function_id)
        
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="ems_v1_completed"           
        where = "" 
        group_by = ""
        order_by = ""  
        
        on_where_day = ""
        where_day = ""
        group_by_day = ""
        table_name_day = ""  
        
        if company_id is not None and company_id != 'all':
            where += f" and  mm.company_id in ({company_id})" 
            where_day += f" and  mm.company_id in ({company_id})" 

        if department_id is not None and department_id != 'all':
            where += f" and  mm.department_id in ({department_id})"          
            where_day += f" and  mm.department_id in ({department_id})"          
            
        if shed_id  is not None and shed_id != 'all':
            where += f" and mm.shed_id in ({shed_id})"
            where_day += f" and mm.shed_id in ({shed_id})"
            
        if machinetype_id is not None and machinetype_id != 'all':
            where += f" and mm.machinetype_id in ({machinetype_id})"
            where_day += f" and mm.machinetype_id in ({machinetype_id})"
            
        if function_id is not None and function_id != 'all':
            where += f" and mm.function_id in ({function_id})"
            where_day += f" and mm.function_id in ({function_id})"
            
        if machine_id is not None and machine_id != 'all':
            where += f" and mm.machine_id in ({machine_id})"
            where_day += f" and mm.machine_id in ({machine_id})"
            
        if converter_id is not None and converter_id != 'all':
            where += f" and mm.converter_id = {converter_id}"
            where_day += f" and mm.converter_id = {converter_id}"
            
        query = text(f'''SELECT * FROM master_shifts WHERE status = 'active' ''')
        data1 = cnx.execute(query).fetchall()
        mill_date = date.today()
        mill_shift = 0
        no_of_shifts = 3
        group_id = ""
        group_code = ""
        group_name = ""
    
        is_day_wise = ""
        month_year_day = ""
        table_name_day = ""
        name = ""
        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]  
              no_of_shifts = shift_record["no_of_shifts"]          
        
        if period_id == "cur_shift":       
            where += f''' and cp.mill_date = '{mill_date}' AND cp.mill_shift = '{mill_shift}' '''              
            table_name = "ems_v1.current_power "  

        elif period_id == "#cur_shift":
            where += f''' and cp.mill_date = '{mill_date}' AND cp.mill_shift = '{mill_shift}' '''              
            table_name = "ems_v1.current_power cp"

            sql = f'select * from master_branch where branch_id = {branch_id} '
            data2= cnx.execute(sql).mappings().all()
            if len(data2) > 0:
                for record in data2:
                    is_day_wise = record["is_day_wise"]

            if is_day_wise == "yes":
                month_year_day=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""            
                table_name_day=f"  {completed_db}.power_{month_year_day} as cp_d "
                where_day += f''' and cp_d.mill_date = '{mill_date}' '''
            
        elif period_id == "sel_shift":                  
            if from_date is None:
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if shift_id is None:
                return JSONResponse({"iserror": True, "message": "shift_id is required"}) 
            
            from_date = parse_date(from_date)          
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}.power_{month_year} " 
    
            where += f''' and cp.mill_date = '{from_date}' AND cp.mill_shift = '{shift_id}' '''   

        elif period_id == "#sel_shift":                 
            if mill_shift == 1:
                shift_id = no_of_shifts
                from_date = parse_date(mill_date) - timedelta(days=1)

            else:
                shift_id = int(mill_shift) - 1
                from_date = mill_date                      
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}.power_{month_year} " 
        
            where += f''' and cp.mill_date = '{from_date}' AND cp.mill_shift = '{shift_id}' '''   
        
        elif period_id == "sel_date":            
            if from_date is None:
                 return JSONResponse({"iserror": True, "message": "from date is required"})    
              
            from_date = parse_date(from_date)
            
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}.power_{month_year}  "           
            where += f''' and cp.mill_date = '{from_date}' '''

        elif period_id == "#sel_date":             
            from_date = mill_date - timedelta(days=1)
            # from_date = parse_date(from_date)
            
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}.power_{month_year}  "           
            where += f''' and cp.mill_date = '{from_date}' '''
            
        elif period_id  == "#this_week":
            dt = mill_date
            from_date=dt-timedelta(dt.weekday()+1)
            to_date = mill_date
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
            table_name=f"  {completed_db}.power_{month_year}  "
            
            where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''
        
        elif period_id == "#this_month":
            from_date = mill_date.replace(day=1)
            to_date = mill_date
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
            table_name=f"  {completed_db}.power_{month_year}  "
            
            where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''
        
        elif period_id == "from_to":            
            if from_date is None:
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date is None:
                 return JSONResponse({"iserror": True, "message": "to_date is required"})  
                       
            from_date = parse_date(from_date)
            to_date =  parse_date(to_date)
              
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
            table_name=f"  {completed_db}.power_{month_year}  "
            if from_date.month!= to_date.month:           
                month_year=f"""{mill_month[to_date.month]}{str(to_date.year)}"""            
                name=f"  {completed_db}.power_{month_year} as cp "
                field_name = 'power_id,company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,design_id,beam_id,date_time,date_time1,mill_date,mill_shift,vln_avg,r_volt,y_volt,b_volt,vll_avg,ry_volt,yb_volt,br_volt,t_current,r_current,y_current,b_current,t_watts,r_watts,y_watts,b_watts,t_var,r_var,y_var,b_var,t_voltampere,r_voltampere,y_voltampere,b_voltampere,avg_powerfactor,r_powerfactor,y_powerfactor,b_powerfactor,powerfactor,kWh,kvah,kw,kvar,power_factor,kva,frequency,machine_status,status,created_on,created_by,modified_on,modified_by,machine_kWh,master_kwh'
                table_name = f'(select {field_name} from {table_name} UNION All select {field_name} from {name}) as cp'

            where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''
            
            if shift_id is not None:                
                where += f''' and cp.mill_shift = '{shift_id}' '''             
            
        else:
             return JSONResponse({"iserror": True, "message": "invalid period id"}) 
        
        if limit_report_for == "exception" :
            if limit_exception_for == "kwh":
                order_by += "sum(cp.kwh)"
                
            if limit_order_by == "asc":
                order_by += " "+limit_order_by +","
            else:
                order_by += " "+limit_order_by +","

        if groupby is not None and groupby == "company":
            group_by_day += " mm.company_id "
            on_where_day += f" and mm.company_id = cp_d.company_id"
            
            group_by += " mm.company_id "
            order_by += " mm.company_id "
            group_id = '''mc.company_id AS group_id '''
            group_code = '''mc.company_code AS group_code'''
            group_name = '''mc.company_name AS group_name'''       
            
        if groupby is not None and groupby == "department":
            group_by_day += " mm.department_id "
            on_where_day += f" and mm.department_id = cp_d.department_id"

            group_by += " mm.department_id "
            order_by += " mm.department_id "
            group_id = '''md.department_id AS group_id '''
            group_code = '''md.department_code AS group_code'''
            group_name = '''md.department_name AS group_name'''        
            
        if groupby is not None and groupby == "shed":
            group_by_day += "  mm.shed_id "
            on_where_day += f" and mm.shed_id = cp_d.shed_id"

            group_by += "  mm.shed_id "
            order_by += "  mm.shed_id "
            group_id = ''' ms.shed_id AS group_id '''
            group_code = ''' ms.shed_code AS group_code'''
            group_name = ''' ms.shed_name AS group_name'''
            
        if groupby is not None and groupby == "machinetype":
            group_by_day += " mm.machinetype_id"
            on_where_day += f" and mm.machinetype_id = cp_d.machinetype_id"

            group_by += " mm.machinetype_id"
            order_by += " mm.machinetype_id"
            group_id = '''mmt.machinetype_id AS group_id '''
            group_code = '''mmt.machinetype_code AS group_code'''
            group_name = '''mmt.machinetype_name AS group_name'''
            
        if groupby is not None and groupby == "function": 
            group_by_day += " mm.function_id"
            on_where_day += f" and mm.function_id = cp_d.function_id"
          
            group_by += " mm.function_id"
            order_by += " mm.function_id"
            group_id = '''mf.function_id AS group_id '''
            group_code = '''mf.function_code AS group_code'''
            group_name = '''mf.function_name AS group_name'''
            
        if groupby is not None and groupby == "machine":  
            group_by_day += " mm.machine_id"
            on_where_day += f" and mm.machine_id = cp_d.machine_id"

            group_by += " mm.machine_id"
            order_by += " mm.machine_order"
            group_id = '''mm.machine_id AS group_id '''
            group_code = '''mm.machine_code AS group_code'''
            group_name = '''mm.machine_name AS group_name'''  
           
        if limit_operation_value is not None and limit_operation_value != '0':           
            order_by += ' LIMIT '+str(limit_operation_value) 

        if is_critical == "yes" or is_critical == "no"  :
            where += f" and mm.major_nonmajor = '{is_critical}' "   
            
        where_group_for = ""  
        # if group_for == "exception" and machine_id != 'all' and machine_id!= "":
        if group_for == "exception":
            if groupby == "company":
                where_group_for += "and group_type = 'company' " 
                if company_id != 'all' and company_id is not None:
                    where_group_for += f"and type_id = '{company_id}'"

            if groupby == "department":
                where_group_for += "and group_type = 'zone' " 
                if department_id != 'all' and department_id is not None:
                    where_group_for += f"and type_id = '{department_id}'"
                    
            if groupby == "shed":
                where_group_for += "and group_type = 'area' "
                if shed_id != 'all' and shed_id is not None:
                    where_group_for += f"and type_id = '{shed_id}'"
                    
            if groupby == "machinetype":
                where_group_for += "and group_type = 'location' "
                if machinetype_id != 'all' and machinetype_id is not None:
                    where_group_for += f"and type_id = '{machinetype_id}'"
                    
            sql = text(f'''SELECT * FROM ems_v1.master_meter_group] where 1=1 {where_group_for} ''') 
            data2 = cnx.execute(sql).fetchall()
            machine_id = []  
            
            if len(data2) > 0:
                for record in data2:
                    machine_id.append(record["machine_id"]) 
                    
                where += f" and mm.machine_id in ({','.join(str(x) for x in machine_id)})"

        if report_for == 'detail' or report_for is None:
            group_by = " cp.mill_date , cp.mill_shift," + group_by
            order_by = " cp.mill_date, cp.mill_shift," + order_by
        
        if report_for == 'summary':
            group_by = " cp.mill_date," + group_by
            order_by = " cp.mill_date," + order_by  
                       
        if group_by != "":
            group_by = f"group by {group_by} "    
        if order_by != "":
            order_by = f"order by {order_by}"
        
        if group_by_day != "":
            group_by_day = f"group by {group_by_day} "    

        select_day = ""
        
        if is_day_wise == 'yes':
            if period_id == "#cur_shift":
                
                table_name_day = text(f'''(
                                  select
                                    mm.company_id as company_id,
                                    mm.branch_id as branch_id,
                                    mm.department_id as department_id,
                                    mm.shed_id as shed_id,
                                    mm.machinetype_id as machinetype_id,
                                    mm.machine_id as machine_id,
                                    sum(cp_d.kwh) as kwh
                                  from
                                    {table_name_day}
                                    INNER JOIN ems_v1.master_machine mm ON cp_d.machine_id = mm.machine_id 
                                  where 1=1 {where_day}
                                  {group_by_day}
                                ) as cp_d on 1=1 {on_where_day}''')
                table_name_day = text(f'''left join {table_name_day}''')
                print(table_name_day)
                select_day = " min(ROUND(CASE WHEN mm.energy_selection = 'wh' THEN CAST(IFNULL(cp_d.kWh,0) AS DECIMAL(18, 2))/1000 ELSE IFNULL(cp_d.kWh,0) END, 2)) + "
        
        query = text(f'''
                SELECT                       
                    mc.company_code AS company_code,
                    mc.company_name AS company_name,
                    mb.branch_code AS branch_code,
                    mb.branch_name AS branch_name,
                    md.department_code AS department_code,
                    md.department_name As department_name,
                    ms.shed_code AS shed_code,
                    ms.shed_name AS shed_name,
                    mmt.machinetype_code AS machinetype_code,
                    mmt.machinetype_name AS machinetype_name,
                    mf.function_code AS function_code,
                    mf.function_name As function_name,
                    mm.machine_code AS machine_code,
                    mm.machine_name AS machine_name,
                    count(mm.machine_name) AS machine_count,
                    cp.power_id as power_id,
                    cp.company_id as company_id,
                    cp.branch_id as branch_id,
                    cp.department_id as department_id,
                    cp.shed_id as shed_id,
                    cp.machinetype_id as machinetype_id,
                    cp.machine_id as machine_id,
                    cp.design_id as design_id,
                    cp.beam_id as beam_id,
                    cp.date_time as date_time,
                    cp.date_time1 as date_time1,
                    cp.mill_date as mill_date,
                    cp.mill_shift as mill_shift,
                    ROUND(cp.vln_avg,2) as vln_avg,
                    ROUND(cp.r_volt,2) as r_volt,
                    ROUND(cp.y_volt,2) as y_volt,
                    ROUND(cp.b_volt,2) as b_volt,
                    ROUND(cp.vll_avg,2) as vll_avg,
                    ROUND(cp.ry_volt,2) as ry_volt,
                    ROUND(cp.yb_volt,2) as yb_volt,
                    ROUND(cp.br_volt,2) as br_volt,
                    ROUND(cp.t_current,2) as t_current,
                    ROUND(cp.r_current,2) as r_current,
                    ROUND(cp.y_current,2) as y_current,
                    ROUND(cp.b_current,2) as b_current,
                    ROUND(cp.t_watts,2) as t_watts,
                    ROUND(cp.r_watts,2) as r_watts,
                    ROUND(cp.y_watts,2) as y_watts,
                    ROUND(cp.b_watts,2) as b_watts,
                    ROUND(cp.t_var,2) as t_var,
                    ROUND(cp.r_var,2) as r_var,
                    ROUND(cp.y_var,2) as y_var,
                    ROUND(cp.b_var,2) as b_var,
                    ROUND(cp.t_voltampere,2) as t_voltampere,
                    ROUND(cp.r_voltampere,2) as r_voltampere,
                    ROUND(cp.y_voltampere,2) as y_voltampere,
                    ROUND(cp.b_voltampere,2) as b_voltampere,
                    ROUND(cp.avg_powerfactor,2) as avg_powerfactor,
                    ROUND(cp.r_powerfactor,2) as r_powerfactor,
                    ROUND(cp.y_powerfactor,2) as y_powerfactor,
                    ROUND(cp.b_powerfactor,2) as b_powerfactor,
                    ROUND(cp.powerfactor,2) as powerfactor,
                    {select_day} sum(ROUND(CASE WHEN mm.energy_selection = 'wh' THEN CAST(cp.kWh AS DECIMAL(18, 2))/1000 ELSE cp.kWh END, 2)) AS kWh,
                    ROUND(cp.kvah,2) as kvah,
                    sum(ROUND(cp.t_watts/1000,2)) as kw,
                    ROUND(cp.kvar,2) as kvar,
                    ROUND(cp.power_factor,2) as power_factor,
                    ROUND(cp.kva,2) as kva,
                    ROUND(cp.frequency,2) as frequency,
                    cp.machine_status as machine_status,
                    cp.status as status,
                    cp.created_on as created_on,
                    cp.created_by as created_by,
                    cp.modified_on as modified_on,
                    cp.modified_by as modified_by,
                    sum(ROUND(cp.machine_kWh,2)) as machine_kWh,
                    sum(ROUND(cp.master_kwh,2)) as master_kwh,
                    mm.ip_address as ip_address,
                    mm.port as port,
                        
                    CASE WHEN cp.date_time <= DATE_SUB(NOW(), INTERVAL 2 MINUTE) THEN 'S' ELSE 'N' END AS nocom,
                    {group_id},
                    {group_code},
                    {group_name}                       
                FROM 
                    {table_name}                 
                    INNER JOIN ems_v1.master_machine mm ON cp.machine_id = mm.machine_id
                    INNER JOIN ems_v1.master_company mc ON mm.company_id = mc.company_id
                    INNER JOIN ems_v1.master_branch mb ON mm.branch_id = mb.branch_id
                    INNER JOIN ems_v1.master_department md ON mm.department_id = md.department_id
                    INNER JOIN ems_v1.master_shed ms ON mm.shed_id = ms.shed_id
                    INNER JOIN ems_v1.master_machinetype mmt ON mm.machinetype_id = mmt.machinetype_id 
                    INNER JOIN ems_v1.master_function mf ON mm.function_id = mf.function_id      
                    {table_name_day}                 
                WHERE  
                    cp.status = '0' and mm.status = 'active'
                    {where}                        
                    {group_by}
                    {order_by}
                ''')
        print(query)
        data = cnx.execute(query).fetchall()
                 
        return JSONResponse({"iserror":False, "message":"data return successfully", "data" : jsonable_encoder(data)})    
    except Exception as e:      
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
       
@app.post("/alarm_report/")
async def alarm_report(company_id :str = Form(None),
                       machine_id : str = Form (None),
                       report_for : str = Form(None),
                       period_id: str = Form(None),
                       from_date: str = Form(None),
                       to_date: str = Form(None),                      
                       shift_id: int = Form(None),                                           
                       cnx: Session = Depends(get_db)):
    if period_id is None:
            return JSONResponse({"iserror": True, "message": "period id is required"})
    
    start_time = date.today()
    try:       
        where = ""          
        query = text(f'''SELECT * FROM master_shifts WHERE status = 'active' ''')
        data1 = cnx.execute(query).fetchall()
        mill_date = date.today()
        mill_shift = 0        
    
        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]            
        
        if period_id == "cur_shift":          

            where += f''' pa.mill_date = '{mill_date}' AND pa.mill_shift = '{mill_shift}' '''  
            
        elif period_id == "sel_shift":            
            if from_date is None:
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if shift_id is None:
                return JSONResponse({"iserror": True, "message": "shift_id is required"})                       
         
            where += f''' pa.mill_date = '{parse_date(from_date)}' AND pa.mill_shift = '{shift_id}' ''' 

        elif period_id == "sel_date":            
            if from_date is None:
                 return JSONResponse({"iserror": True, "message": "from date is required"})
            
            where += f''' pa.mill_date = '{parse_date(from_date)}' '''
            
        elif period_id == "from_to":            
            if from_date is None:
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date is None:
                 return JSONResponse({"iserror": True, "message": "to_date is required"})
        
            where += f'''  pa.mill_date  >= '{parse_date(from_date)}' and pa.mill_date <= '{parse_date(to_date)}' '''
            
            if shift_id is not None:                
                where += f''' and pa.mill_shift = '{shift_id}' ''' 
                            
        elif period_id == "live_alarm":

            where += f''' pa.start_time <> 0 and stop_time = 0'''   
            sql = text(f'SELECT start_time FROM present_alarm ORDER BY start_time DESC LIMIT 1')
            data = cnx.execute(sql).fetchall()
            for i in data:
                start_time = i['start_time']

            sql1= text(f''' UPDATE master_company
            SET alarm_status = CASE WHEN alarm_last_time = '{start_time}' THEN alarm_status ELSE 1 END,
                alarm_last_time = '{start_time}'
            WHERE company_id = '{company_id}'

        ''')
            cnx.execute(sql1)
            cnx.commit()   

            query2=text(f'''select * from master_company where company_id = '{company_id}' and alarm_status = 1''')
            data1 = cnx.execute(query2).fetchall()
        else:
             return JSONResponse({"iserror": True, "message": "invalid period id"})   
        
        if machine_id is not None and machine_id != 'all':
            where += f" and mm.machine_id = '{machine_id}' "       
                 
        duration = f'''TIMESTAMPDIFF(SECOND, pa.start_time, pa.stop_time) AS duration'''
       
        groupby = ""  
        if report_for == "summary" and machine_id != 'all' and machine_id != "":
            sql = text(f'''
                SELECT 
                    ma.alarm_name as alarm_name,		    
                    pa.parameter_name as parameter_name,
                    mm.machine_name as machine_name,
                    {duration}
                ''')
            groupby = f'''group by ma.alarm_name '''
        else:
            sql = text(f'''
                SELECT 
                    mm.machine_code,
                    mm.machine_name,
                    ma.alarm_name,
                    ma.parameter_name,
                    pa.start_time,
                    pa.stop_time,
                    {duration}
                ''')
        query = text(f''' 
                      {sql}
                FROM              
                    ems_v1.present_alarm pa,
                    ems_v1.master_alarm_target ma,
                    ems_v1.master_machine mm                      
                WHERE  {where} {groupby}                     
                        
                ''')
        print(query)
        data = cnx.execute(query).fetchall() 
         
        return JSONResponse({"iserror":False, "message":"data return successfully", "data" : jsonable_encoder(data), "data1":jsonable_encoder(data1)})    
    except Exception as e:       
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
                
@app.post("/alarm_list/")
async def get_alarmlist(company_id : int = Form(None),
                        alarm_target_id: str=Form(None),
                        alarm_type : str = Form(None),
                        cnx: Session = Depends(get_db)):
    try:       
        where = ""

        if alarm_target_id is not  None:
            where += f" and ma.alarm_target_id = '{alarm_target_id}' "
        if alarm_type is not None:
            where += f" and ma.alarm_type= '{alarm_type}' "
        if company_id is not None:
            where += f" and mm.company_id= '{company_id}' "
        
        query=text(f''' 
        SELECT
            mc.company_id,
            ma.*,
            '' AS machine_dtl,
            CONCAT_WS('-', cu.employee_code, cu.employee_name) AS created_user,
            CONCAT_WS('-', mu.employee_code, mu.employee_name) AS modified_user
        FROM
            master_alarm_target ma
            LEFT JOIN master_employee cu ON cu.employee_id = ma.created_by
            LEFT JOIN master_employee mu ON mu.employee_id = ma.modified_by
            INNER JOIN master_machine mm ON ma.machine_id LIKE CONCAT('%,', CAST(mm.machine_id AS CHAR)) OR ma.machine_id = CAST(mm.machine_id AS CHAR)
            INNER JOIN master_company mc ON mm.company_id = mc.company_id
        WHERE
            ma.status <> 'delete'
            {where}
        ''')
        print(query)
        data = cnx.execute(query).fetchall()
        result = []
        for row in data:
            machine_id_list = row["machine_id"].split(",")   # Split comma-separated machine IDs into a list
            machine_dtl = ""
            for machine_id in machine_id_list:                             
                sub_query = text(f"SELECT * FROM master_machine WHERE machine_id = {machine_id}")
                sub_data = cnx.execute(sub_query).fetchall()
                for sub_row in sub_data:
                    if machine_dtl != "":
                        machine_dtl += '\n' 
                    machine_dtl += f'''{sub_row['machine_name']}''' 
                    print(machine_dtl)           
            new_row = dict(row)
            new_row["machine_dtl"] = machine_dtl
            result.append(new_row)            
        
        createFolder("Log/","Query executed successfully for alarm list")
        
        return JSONResponse({"iserror":False,"message":"data returned succesfully","data":jsonable_encoder(result)})        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_alarm_detail/")
async def save_alarm_detail(alarm_target_id:str=Form(None),
                            parameter_name:str=Form(None),
                            machine_id:str=Form(None),
                            alarm_name:str=Form(None),
                            alarm_type:str = Form(None),
                            alarm_duration:int = Form(None),
                            color_1:str=Form(None),
                            color_2:str=Form(None),
                            color_3:str=Form(None),
                            login_id:str=Form(None),
                            cnx: Session = Depends(get_db)):
 
    if machine_id == None:
        return JSONResponse({"iserror":True,"message":"equipment id is required"}) 
    
    if parameter_name == None:
        return JSONResponse({"iserror":True,"message":"parameter name is required"}) 
      
    if alarm_name == None:
        return JSONResponse({"iserror":True,"message":"alarm name is required"}) 
    
    if alarm_type == None:
        return JSONResponse({"iserror":True,"message":"alarm type is required"}) 
    
    if alarm_type == "time_based":
        if alarm_duration == None:
            return JSONResponse({"iserror":True,"message":"alarm duration is required"})
        color_1 = 0
        color_2 = 0
        color_3 = 0
    else:
        alarm_duration = 0
        
        if color_1 == None:
            return JSONResponse({"iserror":True,"message":"color_1 is required"})
        
        if color_2 == None:
            return JSONResponse({"iserror":True,"message":"color_2 is required"})
        
        if color_3 == None:
            return JSONResponse({"iserror":True,"message":"color_3 is required"})
        
    try:        
        if machine_id is not None:
                value = machine_id.split(",")
                if len(value) > 1:
                    values = tuple(value)
                    machine_id = ",".join(values)
                else:
                    machine_id = value[0]  
                    
        if alarm_target_id is not None:
            query =text(f'''UPDATE  master_alarm_target SET machine_id='{machine_id}',parameter_name='{parameter_name}',
                       alarm_name='{alarm_name}',color_1='{color_1}',color_2='{color_2}',color_3='{color_3}',
                       modified_on = now(),modified_by='{login_id}', alarm_duration = '{alarm_duration}', alarm_type = '{alarm_type}' where alarm_target_id = '{alarm_target_id}'
                         ''')
            cnx.execute(query)
            cnx.commit()

        else:
            if machine_id is not None:
                values = machine_id.split(",")
                if len(values) > 1:
                    machine_id = "'" + ",".join(values) + "'"
                    print(machine_id)
                else:
                    machine_id = values[0]
                
            select_query = text(f'''select * from master_alarm_target where alarm_name = '{alarm_name}' and status != 'delete' ''')
            data1 = cnx.execute(select_query).fetchall()

            if len(data1)>0:
                return JSONResponse({"iserror":True,"message":"alarm name already exists "})
            
            query= text(f'''INSERT INTO master_alarm_target (machine_id,parameter_name,alarm_name,color_1,color_2,color_3,
                       created_on,created_by, alarm_duration, alarm_type )
                       VALUES ({machine_id},'{parameter_name}','{alarm_name}','{color_1}', '{color_2}','{color_3}',
                       now(),'{login_id}', '{alarm_duration}', '{alarm_type}') ''')
            cnx.execute(query)
            cnx.commit()

        createFolder("Log/","Query executed successfully for save alarm")
        return JSONResponse({"iserror":False,"message":"data saved succesfully","data":""})
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})    
    
@app.post("/remove_alarm_detail/")
async def remove_alarm_detail(alarm_target_id:str=Form(None),
                              status : str = Form(None),
                              cnx: Session = Depends(get_db)):

    try:
        if alarm_target_id is not None:
            if status is not None:
                query = text(f''' UPDATE master_alarm_target SET status = '{status}' where alarm_target_id = '{alarm_target_id}' ''')

            else:
                query = text(f''' UPDATE master_alarm_target SET status = 'delete' where alarm_target_id = '{alarm_target_id}' ''')             
            cnx.execute(query)
            cnx.commit()
        
        createFolder("Log/","Query executed successfully for remove alarm")
        return JSONResponse({"iserror":False,"message":"status update succesfully","data":""}) 
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/get_power_report_name/")
async def get_power_report_name(cnx: Session = Depends(get_db)):

    try:        
        query= text(f'''select * from power_report where status = 'active' ''')
        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror":False,"message":"data return succesfully","data":jsonable_encoder(data)})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
# @app.post("/get_power_report_fields/")
# async def get_power_report_fields(company_id : int = Form(None),
#                                   report_id:int=Form(None),
#                                   cnx: Session = Depends(get_db)):
#     if report_id == None:
#         return JSONResponse({"iserror":True,"message":"report_id is required"}) 
    
#     try:
#         query = text(f'''SELECT * FROM power_report_fields_original WHERE report_id = '{report_id}' order by slno''')
#         data = cnx.execute(query).fetchall()
        
#         return JSONResponse({"iserror":False,"message":"data return succesfully","data":jsonable_encoder(data)}) 
#     except Exception as e :
#         error_type = type(e).__name__
#         error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
#         error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
#         error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
#         createFolder("Log/","Issue in returning data "+error_message)
#         return JSONResponse({"iserror":True,"message":error_message})  

@app.post("/get_power_report_fields/")
async def get_power_report_fields(company_id : int = Form(None),
                                  report_id:int=Form(None),
                                  cnx: Session = Depends(get_db)):
    if report_id == None:
        return JSONResponse({"iserror":True,"message":"report_id is required"}) 
    
    try:
        if report_id is not None:
            if report_id == 0:
                query = text(f'''
                            SELECT 
                                report_field_id as report_field_id,
                                report_id as report_id,
                                field_code as field_code,
                                field_name as field_name,
                                is_show as is_show,
                                slno as slno,
                                field_name_display as field_name_display,
                                company_id as company_id
                            FROM 
                                power_report_fields_original
                            WHERE company_id = {company_id} 
                            group by 
                                field_code 
                            order by
                                 slno
                             ''')

            else:
                query = text(f'''SELECT * FROM power_report_fields_original where report_id = {report_id} and company_id = {company_id} order by slno''')

        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)}) 
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})  

@app.post("/update_power_report_fields/")
async def update_power_report_fields(company_id : int = Form(None),
                                     report_id:int=Form(None),
                                     obj: str = Form(None),
                                     cnx: Session = Depends(get_db)):
    
    if company_id == None:
        return JSONResponse({"iserror":True,"message":"company_id is required"}) 
    
    if report_id == 0:
        query = text(f'''UPDATE ems_v1.power_report_fields_original
                        SET is_show = 'no' 
                        WHERE company_id = {company_id}  ''')
        
        cnx.execute(query)
        cnx.commit()

    else:
        query = text(f'''UPDATE ems_v1.power_report_fields_original
                        SET is_show = 'no' 
                        WHERE company_id = {company_id} and report_id = '{report_id}' ''')
        cnx.execute(query)
        cnx.commit()                     
                    
    try:
        if obj is not None:
            obj_data = json.loads(obj)
            for row in obj_data:
                field_name_display = row["field_name_display"]
                report_field_id = row["report_field_id"]
                slno = row["slno"]
                field_code = row["field_code"]
        
                if report_id == 0:                   
                    
                    sql = text(f'''UPDATE ems_v1.power_report_fields_original
                                   SET is_show = 'yes', field_name_display = '{field_name_display}', slno = '{slno}' 
                                   WHERE company_id = {company_id} and field_code = '{field_code}' ''')
                
                else:                    
                    
                    sql = text(f'''UPDATE ems_v1.power_report_fields_original
                                   SET is_show = 'yes', field_name_display = '{field_name_display}', slno = '{slno}' 
                                   WHERE company_id = {company_id} and report_field_id = '{report_field_id}' ''')
                
                cnx.execute(sql)
                cnx.commit()
        return JSONResponse({"iserror":False,"message":"data save successfully","data":''}) 
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})    
    
@app.post("/communication_status/")
async def communication_status(machinetype_id:int=Form(None),
                               cnx: Session = Depends(get_db)):
    
    try:
        query = text(f'''
                    select
                        c.converter_id as converter_id,
                        c.converter_name as converter_name
                    from
                        master_machine mm,
                        master_converter_detail c
                    where
                        mm.converter_id = c.converter_id and mm.machinetype_id = '{machinetype_id}'
                        group by c.converter_id
                     ''')
        data = cnx.execute(query).fetchall()
        print(query)
        return JSONResponse({"iserror":False,"message":"data retrun succesfully","data":jsonable_encoder(data)}) 
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/model/")
async def model_list(cnx: Session = Depends(get_db)):

    try:        
        query= text(f'''
                     SELECT
                        mm.*,
                        IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                        IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                    FROM
                        master_model mm

                        left join master_employee cu on cu.employee_id=mm.created_by
                        left join master_employee mu on mu.employee_id=mm.modified_by
                    WHERE
                        mm.status != 'delete'
                    ''')
        print(query)
        data = cnx.execute(query).fetchall()
        
        return JSONResponse({"iserror":False,"message":"data return succesfully","data":jsonable_encoder(data)})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/save_model/")
async def save_model(model_id :int =Form(None),
                     model_name : str = Form(None),
                     user_login_id : str = Form(None),
                     cnx: Session = Depends(get_db)):
    try:
        
        if model_id is None:
            select_query = text(f'''SELECT * FROM master_model WHERE model_name = '{model_name}' and status != 'delete' ''')
            data1 = cnx.execute(select_query).fetchall()

            if len(data1)>0:
                return JSONResponse({"iserror":True,"message":"model name already exists "})
            
            query = text(f'''
                        INSERT INTO master_model (model_name, created_on, created_by)
                        VALUES ('{model_name}', now() , '{user_login_id}')
                        ''')
        else:
            
            query = text(f'''
                        UPDATE master_model SET model_name = '{model_name}', modified_on = now(),
                        modified_by = '{user_login_id}'
                        WHERE model_id = {model_id}
                        ''')
        cnx.execute(query)
        cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"data save succesfully","data":""})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/remove_model/")
async def remove_model(model_id :int =Form(None),
                       status : str = Form(None),
                       cnx: Session = Depends(get_db)):
    try:
        if model_id is not None:
            if status is not None: 
                query = text(f'''
                            UPDATE master_model SET status = '{status}'
                            WHERE model_id = {model_id}
                            ''')
            else:
                query = text(f'''
                            UPDATE master_model SET status = 'delete'
                            WHERE model_id = {model_id}
                            ''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update succesfully","data":""})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/employeelist_userrights/")
async def employeelistuser(employee_id:str=Form(None),
                           is_login:str=Form(None),
                           cnx: Session = Depends(get_db)):    
    
    if is_login == None:
        return JSONResponse({"iserror":True,"message":" is login is required"})
    
    try: 
        where = ""
        if employee_id is not None:
            where = f"and employee_id = '{employee_id}' "

        query=text(f''' SELECT
                         *
                   FROM 
                        master_employee 
                        WHERE status='active' and employee_type <> 'admin' {where} ''')
        
        
        data1 = cnx.execute(query).fetchall()
        createFolder("Log/","Query executed successfully for  user rights employee list")
        return JSONResponse({"iserror":False,"message":"data return succesfully","data":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/user_menu_list/")
async def employeelistuser(employee_id:str=Form(None),
                           cnx: Session = Depends(get_db)):
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})
    
    try:

        query1=text(f''' SELECT 
                        ms.*,
                        IFNULL (u.id,0) AS u_r_id,
                        IFNULL (u.add_op,'')AS add_opp,
                        IFNULL (u.edit_op,'')AS edit_opp,
                        IFNULL (u.delete_op,'')AS delete_opp
                    FROM
                        menu_mas ms
                        LEFT JOIN 
                        (select * from user_rights where userid={employee_id}) As u
                        ON u.menu_id=ms.menu_id
                        WHERE STATUS='active' 
                        ORDER BY ms.slno
			  ''')
        print(query1)  
        data1 = cnx.execute(query1).fetchall()
        createFolder("Log/","Query executed successfully for  user rights ")
        return JSONResponse({"iserror":False,"message":"data return successfully","data1":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    

# menu='[{"menu_id":1,"add_op": "yes","edit_op": "yes","delete_op": "yes"},{"menu_id":2,"add_op": "yes","edit_op": "yes","delete_op": "yes"}]'
@app.post("/save_userrights_detail/")
async def save_userrights_detail(employee_id:str=Form(None),
                                 menu:str=Form(None),
                                 cnx: Session = Depends(get_db)):
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"user id is required"}) 
    
    if menu == None:
        return JSONResponse({"iserror":True,"message":"menu is required"})
    
    try:

        if employee_id is not None:
                  del_query=text(f'''DELETE FROM user_rights WHERE userid='{employee_id}' ''')
                  cnx.execute(del_query)
                  cnx.commit()

                  user_dict = json.loads(menu)
                  for i in user_dict:
                        menu_id=i['menu_id']
                        add_op = i['add_op']
                        edit_op = i['edit_op']
                        delete_op=i['delete_op']
                        query=text(f'''insert into user_rights(menu_id,add_op,edit_op,delete_op,userid)
                                values('{menu_id}','{add_op}','{edit_op}','{delete_op}','{employee_id}') ''')        
                        cnx.execute(query)
                        cnx.commit()    
        createFolder("Log/","Query executed successfully for save user rights")
        return JSONResponse({"iserror":False,"message":"Data Saved Succesfully","data":""})
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/sidebar_list/")
async def sidebarlistuser(employee_id:str=Form(None),
                           cnx: Session = Depends(get_db)):
    
    employee_type=''
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})
    
    try:
        query=text(f''' select * from master_employee where employee_id={employee_id}''')
        data=cnx.execute(query).fetchall()
        # print(data)

        if len(data) > 0 :
            for record in data:
                employee_type=record['employee_type']

        if employee_type == 'Admin':
            query1=text(f'''
                        select * 
                        from menu_mas 
                        where status='active' 
                        order by slno
                        ''')
            # print(query1)
        else:
            query1=text(f''' SELECT 
                            ms.*,
                            IFNULL(u.id, 0) AS u_r_id,
                            IFNULL(u.add_op, '') AS add_opp,
                            IFNULL(u.edit_op, '') AS edit_opp,
                            IFNULL(u.delete_op, '') AS delete_opp
                        FROM
                            menu_mas ms,
                            user_rights u
                        WHERE
                            ms.status = 'active'
                            AND ms.menu_id = u.menu_id
                            AND u.userid = {employee_id}
                        ORDER BY ms.slno
                            
			  ''')
        print(query1)  
        data1 = cnx.execute(query1).fetchall()
        createFolder("Log/","Query executed successfully for  sidebar list")
        return JSONResponse({"iserror":False,"message":"data return successfully","data1":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/operations/")
async def operations(employee_id:str=Form(None),
                     menu_id:str=Form(None),
                     cnx: Session = Depends(get_db)):
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})
    
    where=""
    if menu_id is not None:
        where+=f''' and menu_id='{menu_id}' '''
    
    try:
        query=text(f'''SELECT 
                        u.*,
                        e.employee_type
                  FROM 
                        user_rights u
                        inner join master_employee e on e.employee_id=u.userid
                  WHERE
                        u.userid={employee_id} {where}
                        
        ''')
        data1 = cnx.execute(query).fetchall()
        createFolder("Log/","Query executed successfully for  sidebar list")
        return JSONResponse({"iserror":False,"message":"data return successfully","data1":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/load_analysis/")
async def load_analysis(period_id: str = Form(None),
                        machine_id: str=Form(None), 
                        from_date: str = Form(None),
                        shift_id :int = Form(None),
                        from_time: str=Form(None), 
                        to_time: str=Form(None),                        
                        cnx: Session = Depends(get_db)):

    if machine_id == None:
        return JSONResponse({"iserror":True,"message":"machine id is required"}) 
    
    if period_id == None:
        return JSONResponse({"iserror":True,"message":"period_id is required"}) 
    
    mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}    
    
    try: 
        where = ''
        table_name = ''
        
        if period_id == 'cur_shift':
            query=text(f'''SELECT * FROM master_shifts WHERE status='active' ''')
            data1 = cnx.execute(query).fetchall()
            mill_date = date.today()
            mill_shift = 0       
    
            if len(data1) > 0:
                for shift_record in data1:
                    mill_date = shift_record["mill_date"]
                    mill_shift = shift_record["mill_shift"]            

            table_name = 'current_power_analysis '
            where = f" cp.mill_date = '{mill_date}' and cp.mill_shift ='{mill_shift}' "

        elif period_id == 'sel_shift' or period_id == 'sel_date':
            if from_date == None:
                return JSONResponse({"iserror":True,"message":"date is required"}) 
            
            mill_date=parse_date(from_date)             
            month_year=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""
            table_name=f"ems_v1_completed.power_analysis_{month_year}" 
            where = f" cp.mill_date = '{mill_date}' "

            field_name = 'machine_id, date_time, mill_date, mill_shift, t_current, r_current, y_current, b_current, vll_avg, ry_volt, yb_volt, br_volt, vln_avg, r_volt, y_volt, b_volt, t_watts, kWh, kvah, kw, kvar, power_factor, r_watts, kva, y_watts, b_watts, avg_powerfactor, r_powerfactor, y_powerfactor, b_powerfactor, powerfactor, kwh_actual, frequency, t_voltampere, r_voltampere, y_voltampere, b_voltampere, t_var, r_var, y_var, b_var, master_kwh ,machine_kWh'
            table_name = f'(select {field_name} from current_power_analysis UNION All select {field_name} from {table_name})'

            if period_id == 'sel_shift':
                if shift_id == None:
                    return JSONResponse({"iserror":True,"message":"shift is required"}) 
                where += f" and cp.mill_shift ='{shift_id}' " 
                
        if from_time is not None:
            where += f" and FORMAT(cp.date_time ,'HH:mm')>='{from_time}' "
        if to_time is not None:
            where += f" and FORMAT(cp.date_time ,'HH:mm')<='{to_time}' "

        query=text(f'''
            SELECT
			    mm.machine_code,
			    mm.machine_name,
			    cp.machine_id,
			    cp.date_time,
			    cp.mill_date,
			    cp.mill_shift,
			    ROUND(cp.t_current,2)as t_current,
			    ROUND(cp.r_current,2)as r_current,
			    ROUND(cp.y_current,2)as y_current,
			    ROUND(cp.b_current,2)as b_current,
			    ROUND(cp.vll_avg,2)as vll_avg,
			    ROUND(cp.ry_volt,2)as ry_volt,
			    ROUND(cp.yb_volt,2)as yb_volt,
			    ROUND(cp.br_volt,2)as br_volt,
			    ROUND(cp.vln_avg,2)as vln_avg,
			    ROUND(cp.r_volt,2)as r_volt,
			    ROUND(cp.y_volt,2)as y_volt,
			    ROUND(cp.b_volt,2)as b_volt,
			    ROUND(cp.t_watts,2)as t_watts,
			    ROUND((case WHEN mm.energy_selection = 'wh' then cp.kWh/1000 else cp.kWh end),2) AS kWh,
			    ROUND(cp.kvah,2)as kvah,
                ROUND((cp.t_watts/1000),2) as kw,
			    ROUND(cp.kvar,2)as kvar,
			    ROUND(cp.power_factor,2)as power_factor,
			    ROUND(cp.r_watts,2)as r_watts,
			    ROUND(cp.kva,2)as kva,
			    ROUND(cp.y_watts,2)as y_watts,
			    ROUND(cp.b_watts,2)as b_watts,
			    ROUND(cp.avg_powerfactor,2)as avg_powerfactor,
			    ROUND(cp.r_powerfactor,2)as r_powerfactor,
			    ROUND(cp.y_powerfactor,2)as y_powerfactor,
			    ROUND(cp.b_powerfactor,2)as b_powerfactor,
			    ROUND(cp.powerfactor,2)as powerfactor,
			    ROUND(cp.kwh_actual,2)as kwh_actual,
			    ROUND(cp.frequency,2)as frequency,
			    ROUND(cp.t_voltampere,2)as t_voltampere,
			    ROUND(cp.r_voltampere,2)as r_voltampere,
			    ROUND(cp.y_voltampere,2)as y_voltampere,
			    ROUND(cp.b_voltampere,2)as b_voltampere,
			    ROUND(cp.t_var,2)as t_var,
			    ROUND(cp.r_var,2)as r_var,
			    ROUND(cp.y_var,2)as y_var,
			    ROUND(cp.b_var,2) as b_var,
                ROUND(cp.master_kwh,2) as master_kwh,
                ROUND(cp.machine_kWh,2) as machine_kWh
		    from
                {table_name} as cp 
		        inner join master_machine mm on mm.machine_id=cp.machine_id
		    where 
                cp.machine_id in ({machine_id}) and {where}
		    order 
                by mm.machine_id, cp.date_time                                
            ''')  
        print(query)
        data=cnx.execute(query).fetchall()
        label = {}
        machine_data = {}
        org_data = []
        for d in data:
            machine_id = d['machine_id']
            machine_name = d['machine_name']
            if machine_id not in label:        
                label[machine_id] = machine_name
            if machine_id not in machine_data:
                machine_data[machine_id] = []
            # set machine_data for machine_id
            temp = {
                'date_time': d['date_time'],
                't_current': d['t_current'],
                'r_current': d['r_current'],
                'y_current': d['y_current'],
                'b_current': d['b_current'],
                'vll_avg': d['vll_avg'],
                'ry_volt': d['ry_volt'],
                'yb_volt': d['yb_volt'],
                'br_volt': d['br_volt'],
                'vln_avg': d['vln_avg'],
                'r_volt': d['r_volt'],
                'y_volt': d['y_volt'],
                'b_volt': d['b_volt'],
                't_watts': d['t_watts'],
                'kWh': d['kWh'],
                'kvah': d['kvah'],
                'kw': d['kw'],
                'kvar': d['kvar'],
                'power_factor': d['power_factor'],
                'r_watts': d['r_watts'],
                'kva': d['kva'],
                'y_watts': d['y_watts'],
                'b_watts': d['b_watts'],
                'avg_powerfactor': d['avg_powerfactor'],
                'r_powerfactor': d['r_powerfactor'],
                'y_powerfactor': d['y_powerfactor'],
                'b_powerfactor': d['b_powerfactor'],
                'powerfactor': d['powerfactor'],
                'kwh_actual': d['kwh_actual'],
                'frequency': d['frequency'],
                't_voltampere': d['t_voltampere'],
                'r_voltampere': d['r_voltampere'],
                'y_voltampere': d['y_voltampere'],
                'b_voltampere': d['b_voltampere'],
                't_var': d['t_var'],
                'r_var': d['r_var'],
                'y_var': d['y_var'],
                'b_var': d['b_var'],
                'master_kwh':d['master_kwh'],
                'machine_kWh':d['machine_kWh']
            }

            machine_data[machine_id].append(temp)

        for key, value in machine_data.items():
            org_data.append({'label': label[key], 'data': value})

        createFolder("Log/","query executed successfully in load analysis")
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(org_data),"data1":jsonable_encoder(data)}) 

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/change_password/")
async def change_password(employee_id:str=Form(None),
                          old_password:str=Form(None),
                          new_password:str=Form(None),
                          retype_password:str=Form(None),
                          cnx: Session = Depends(get_db)):  
    
    try:
        sql = text(f'''select * from master_employee where employee_id = {employee_id} and password_login = md5('{old_password}') ''')
        data = cnx.execute(sql).fetchall()          

        if len(data) == 0:            
           return JSONResponse({"iserror":True,"message":"incorrect user id or password"})

        else:
            if new_password != retype_password:
                return JSONResponse({"iserror":True,"message":"retype password is incorrect"})

            query=text(f'''update master_employee set password_login = md5('{new_password}') where employee_id ='{employee_id}' ''')
            cnx.execute(query)
            cnx.commit()

        return JSONResponse({"iserror":False,"message":"password changed successfully","data":""}) 
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
static_dir = Path(__file__).parent 

def generate_excel_report(result, month_year, report_type):

    if report_type == "date":
        file_path = f'{static_dir}/performanceReport_templete.xlsx'
    elif report_type == "shift":
        file_path = f'{static_dir}/performanceReport_shift_templete.xlsx'
    workbook = Workbook()
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS' 
            
    cell = "D2"
    data = f"PERFORMANCE REPORT FOR KWH - {month_year}"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=25)
    sheet[cell].font = font
    # Set the border style
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Set the minimum column width
    minimum_column_width = 15
    # Iterate over the data and set the values and formatting
    for i, data in enumerate(result, start=6):  # Assuming machine_code starts from row 7
        machine_code = data["machine_code"]
        machine_name = data["machine_name"]
        print(machine_name)
        # if machine_code  is None and machine_name is None:
        #     sheet.delete_rows(6, sheet.max_row - 5)  # Clear all rows starting from row 6
        #     sheet.cell(row=6, column=6).value = "No Data"        
        sheet.cell(row=i, column=2, value=machine_code).alignment = Alignment(horizontal="center")
        # Create a new Alignment object with text wrapping enabled for machine_name
        alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.cell(row=i, column=3, value=machine_name).alignment = alignment
        # Adjust column size based on the maximum text length in columns C and D
        machine_name_length = len(machine_name) + 40
        # print(machine_name_length)
        column_letter_c = get_column_letter(3)  # Column C
        column_width_c = max(machine_name_length, sheet.column_dimensions[column_letter_c].width)
        # print(column_width_c)
        sheet.column_dimensions[column_letter_c].width = column_width_c
        if report_type == "date":
            for j in range(1, 32):
                column_letter = get_column_letter(j + 3)  # Assuming data columns start from column E (column index 5)
                cell = sheet.cell(row=i, column=j + 3)
                cell.value = data.get(f"d{j}", "")
                cell.alignment = Alignment(horizontal="center")
                if cell.value == 0:
                    cell.value = ""
            # Apply border to cells
            row_range = sheet[f"A{i}:AH{i}"]
            for row in row_range:
                for cell in row:
                    cell.border = border
        elif report_type == "shift":
            for shift in range(1, 4):
                for j in range(1, 32):
                    column_letter = get_column_letter((shift-1)*31 + j + 3)  # Assuming data columns start from column E (column index 5)
                    cell = sheet.cell(row=i, column=(shift-1)*31 + j + 3)
                    cell.value = data.get(f"ds{shift}_{j}", "")
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == 0:
                        cell.value = ""
            # Apply border to cells
            row_range = sheet[f"A{i}:CR{i}"]
            for row in row_range:
                for cell in row:
                    cell.border = border
            # Adjust column size based on the maximum text length vertically
        cell_text_length = len(str(cell.value))
        column_width = max(cell_text_length, sheet.column_dimensions[column_letter].width)
        sheet.column_dimensions[column_letter].width = column_width
    if result == []:
        # sheet.delete_rows(6, sheet.max_row - 5)
        cell = "O10"
        data = "No Data"
        sheet[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        sheet[cell].alignment = alignment

        sheet.column_dimensions[cell[0]].width = len(data) + 2  # Adjust column width

        font = Font(name='Calibri', size=25)
        sheet[cell].font = font

    file_name = 'PerformanceReport.xlsx'
    file_path = os.path.join(base_path, file_name)
    workbook.save(file_path)

@app.post("/performance_report/")
async def performance_report(request:Request,
                             machine_id: str = Form(None),
                             month_year: str = Form(None),
                             report_type: str = Form(None),                             
                             cnx: Session = Depends(get_db)):
    
    if month_year is None:
        return JSONResponse({"iserror": True, "message": "month year is required"})

    if report_type is None:
        return JSONResponse({"iserror": True, "message": "report type is required"})
    
    try:
        groupby = ""
        where = ""
        if machine_id is not None:
            machine_id = machine_id.split(',')
            where += f" and mm.machine_id IN ({','.join(machine_id)})"
        if month_year is not None:
            month, year = month_year.split('-')
            tbl_name = f"ems_v1_completed.power_{month}{year} cp"

        if report_type not in ['date', 'shift']:
            return JSONResponse({"iserror": True, "message": "Invalid report type"})

        if report_type == 'date':
            day = "CONCAT('d', DAY(cp.mill_date))"

        elif report_type == 'shift':
            day = "CONCAT(CONCAT('ds', cp.mill_shift), '_', DAY(cp.mill_date))"
            groupby = ",cp.mill_shift"
            
        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                {day} AS day,
                SUM(cp.kwh) AS kwh
            FROM
                {tbl_name}
                INNER JOIN ems_v1.master_machine mm ON mm.machine_id = cp.machine_id
            WHERE
                1=1 {where} and FORMAT(cp.mill_date, 'MM-yyyy') = '{month_year}' 
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                DAY(cp.mill_date)
                {groupby}               
        ''')
        rslt = cnx.execute(query).fetchall()
        print(query)
        if rslt is not None:
            output = {}
            
            if report_type == 'date':
                output_keys = [f'd{day}' for day in range(1, 32)]
            elif report_type == 'shift':
                output_keys = [f'ds{shift}_{day}' for day in range(1, 32) for shift in range(1, 4)]
            
            for row in rslt:
                machine_code = row.machine_code
                machine_name = row.machine_name
                day = row.day
                kwh = row.kwh
            
                if machine_code not in output:
                    output[machine_code] = {
                        'machine_code': machine_code,
                        'machine_name': machine_name
                    }
                    for key in output_keys:
                        output[machine_code][key] = 0
            
                output[machine_code][day] = kwh
                result=list(output.values())
            if rslt == []:
                result = []
               
            # machine = {"result" : data["machine_name"],
            #            "result1" : data[]}
            generate_excel_report(result, month_year,report_type)
            # process_data(month_year, result)
        file_path = os.path.join(base_path, "PerformanceReport.xlsx")
        results = f"http://{request.headers['host']}/attachment/PerformanceReport.xlsx"

        if os.path.exists(file_path):

            return {"file_url": results}
        else:
            return {"file_url": None}   
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

# [{"machine_id":"1","kWh": "34658","mill_date": "2023-05-26 ","mill_shift": "2"}]
@app.post("/manual_entry/")
async def save_manual_entry(obj: str = Form(None),                             
                            cnx: Session = Depends(get_db)):
    try:
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="ems_v1_completed."    
        if obj == None:
            return JSONResponse({"iserror":True,"message":"obj is required"})
        
        if obj != '':
            user_dict = json.loads(obj)
            for i in user_dict:
                machine_id = i['machine_id']
                mill_date = i['mill_date']
                mill_shift = i['mill_shift']
                kWh = i['kWh']
                month_year=f"""{mill_month[parse_date(mill_date).month]}{str(parse_date(mill_date).year)}"""
                table_name=f"  {completed_db}power_{month_year}"           
                query = text(f'''
                    UPDATE {table_name}
                    SET kWh = {kWh}
                    WHERE machine_id = '{machine_id}' and mill_date = '{mill_date}' and mill_shift = '{mill_shift}' ''')
                cnx.execute(query)
                cnx.commit()

        return JSONResponse({"iserror": False, "message": "parameter saved sucessfully", "data": ''})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/machine_history_list/")
async def save_manual_entry(machine_id: str = Form(None),                             
                            cnx: Session = Depends(get_db)):
    try:
        where = ''
        if machine_id is not None:
            where = f' and machine_id = {machine_id}'

        query = text(f'''
                    SELECT 
                        mc.company_code AS company_code,
                        mc.company_name AS company_name,
                        mb.branch_code AS branch_code,
                        mb.branch_name AS branch_name,
                        md.department_code AS department_code,
                        md.department_name AS department_name,
                        ms.shed_code AS shed_code,
                        ms.shed_name AS shed_name,
                        mmt.machinetype_code AS machinetype_code,                        
                        mmt.machinetype_name AS machinetype_name,
                        mf.function_code AS function_code,
                        mf.function_name AS function_name,
                        mcd.converter_name AS converter_name,
                        mh.*,
                        IFNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                    IFNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM 
                        master_machine_history mh
                        left join master_employee cu on cu.employee_id=mh.created_by
	                    left join master_employee mu on mu.employee_id=mh.modified_by
                        INNER JOIN ems_v1.master_company mc ON mh.company_id = mc.company_id
                        INNER JOIN ems_v1.master_branch mb ON mh.branch_id = mb.branch_id
                        INNER JOIN ems_v1.master_department md ON mh.department_id = md.department_id
                        INNER JOIN ems_v1.master_shed ms ON mh.shed_id = ms.shed_id
                        INNER JOIN ems_v1.master_machinetype mmt ON mh.machinetype_id = mmt.machinetype_id
                        INNER JOIN ems_v1.master_function mf ON mh.function_id = mf.function_id
                        INNER JOIN ems_v1.master_converter_detail mcd ON mh.converter_id = mcd.converter_id
                    WHERE mh.status = 'active'  {where}''')
        
        data = cnx.execute(query).fetchall()

        return JSONResponse({"iserror": False, "message": "data return sucessfully", "data": jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/update_alarm_popup_status/")
async def save_manual_entry(company_id: str = Form(None),                             
                            cnx: Session = Depends(get_db)):
    
    try:

        query = text(f'''Update ems_v1.master_company set alarm_status = 0 where company_id = {company_id}''')
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "sucessfully updated", "data": ''})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/manualentry_list/")
async def manualentry_list(entry_date: str = Form(''), 
                           cnx: Session = Depends(get_db)):
    
    mill_date = date.today()
    try:

        entry_date = entry_date.replace('%20', ' ')
        data = {'entry_date': entry_date}
        print(data)
        query = text(f'''SELECT * FROM ems_v1.master_shifts WHERE status = 'active' ''')
        data1 = cnx.execute(query).fetchall()
        
        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
        mill_date = mill_date.strftime('%d-%m-%Y') 
      
        if entry_date != '':
            mill_date = entry_date
            data['mill_date'] = mill_date
         
        else:
            data['mill_date'] = mill_date

        where = ''
        select = '*'

        if entry_date != '':
            where = text(f"WHERE FORMAT(entry_date, 'dd-MM-yyyy') = '{entry_date}'")
        else:
            select = "entry_date,FORMAT(entry_date,'dd-MM-yyyy') as m_date"
            where = "GROUP BY entry_date"

        diesel_sql = text(f"SELECT {select} FROM ems_v1.master_diesel_entry {where}")
        production_sql = text(f"SELECT {select} FROM ems_v1.master_production_entry {where}")
        load_sql = text(f"SELECT {select} FROM ems_v1.master_load_entry {where}")
        heat_sql = text(f"SELECT {select} FROM ems_v1.master_heat_entry {where}")
        print(diesel_sql)
        data['dieselLists'] = cnx.execute(diesel_sql).fetchall()
        data['productionLists'] = cnx.execute(production_sql).fetchall()
        data['loadLists'] = cnx.execute(load_sql).fetchall()
        data['heatLists'] = cnx.execute(heat_sql).fetchall()
        print(load_sql)

        return JSONResponse({"iserror": False, "message": "successfully updated", "data": jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})
# 10-07-2023
# obj_runtime [{"dg_runtime":"88"},{"dg_runtime":""},{"dg_runtime":""}]
# obj_litre [{"dg_name":"SMS-1 DG-1 (625kVA)","diesel_litre":"23"},{"dg_name":null,"diesel_litre":"56"},{"dg_name":"SMS-1 DG-2 (625kVA)","diesel_litre":"11"},{"dg_name":"SMS-2 DG (1250kVA)","diesel_litre":""},{"dg_name":"CCM-3 DG (1010kVA)","diesel_litre":""}]
# obj_production [{"production_name":"SMS1 Production","production_value":"56"},{"production_name":"SMS2 Production","production_value":""}]
# obj_load [{"load_name":"LRF-1 HT Transformer","load_value":"34"},{"load_name":"LRF-2 HT Transformer","load_value":"56"},{"load_name":"LRF-3 HT Transformer","load_value":"78"},{"load_name":"LRF-4 HT Transformer","load_value":"90"},{"load_name":"VD1 Power Consumption","load_value":""},{"load_name":"VD2 Power Consumption","load_value":""},{"load_name":"CCM-1 LCSS Transformer-7A","load_value":""},{"load_name":"CCM-1 LCSS Transformer-7B","load_value":""},{"load_name":"LRF Power Consumption-7A","load_value":""},{"load_name":"LRF Power Consumption-7B","load_value":""},{"load_name":"CCM-2 Transformer-33","load_value":""},{"load_name":"CCM-2 Transformer-34","load_value":""},{"load_name":"SMS-2 PH Transformer-35","load_value":""},{"load_name":"SMS-2 PH Transformer-36","load_value":""},{"load_name":"EOF Power Consumption-35","load_value":""},{"load_name":"EOF Power Consumption-36","load_value":""}]
# obj_heat [{"heat_name":"EOF1","heat_value":"12"},{"heat_name":"EOF2","heat_value":""}]
@app.post("/save_manualentry/")
async def save_manualentry(mill_date: str = Form(''), 
                           entry_date:str = Form(''),
                           obj_runtime:str = Form(''),
                           obj_litre : str = Form(''),
                           obj_production : str = Form(''),
                           obj_load : str = Form(''),
                           obj_heat : str = Form(''),
                           user_login_id: str=Form(None),
                           cnx: Session = Depends(get_db)):
    try:
        createFolder("Log/", "mill_date....." + str(mill_date))
        createFolder("Log/", "entry_date....." + str(entry_date))
        createFolder("Log/", "obj_runtime....." + str(obj_runtime))
        createFolder("Log/", "obj_litre..."+str(obj_litre))
        createFolder("Log/", "obj_production..."+str(obj_production))
        createFolder("Log/", "obj_load..."+str(obj_load))
        createFolder("Log/", "obj_heat..."+str(obj_heat))
        dg_runtime = 0
        dg_name = ''
        diesel_litre = ''
        production_name= ''
        production_value = ''
        load_name = ''
        load_value = ''
        heat_name = ''
        heat_value = ''
        mill_date = parse_date(mill_date)
        if entry_date == '':
            query = f"select * from ems_v1.master_diesel_entry where entry_date = '{mill_date}'"
            data = cnx.execute(query).fetchall()
            if len(data)>0:
                return JSONResponse({"iserror": False, "message": "entry date already exists"})
        
        if obj_litre is not None:
            obj_data_litre = json.loads(obj_litre)
        if obj_runtime is not None:
            obj_data_runtime = json.loads(obj_runtime)
            for idx, row_runtime in enumerate(obj_data_runtime):
                dg_runtime = row_runtime["dg_runtime"]  # Accessing 'dg_runtime' within 'obj_data_runtime'
                row_litre = obj_data_litre[idx]
                dg_name = row_litre["dg_name"]
                diesel_litre = row_litre["diesel_litre"]
                

                if entry_date == "":
                    sql1 = text(f"""INSERT INTO ems_v1.master_diesel_entry (
                        dg_name,
                        diesel_litre,
                        dg_runtime,
                        entry_date,
                        created_on,
                        created_by
                    ) VALUES (
                        '{dg_name}',
                        '{diesel_litre}',
                        '{dg_runtime}',
                        '{mill_date}',
                        getdate(),
                        '{user_login_id}'
                    )""")
                    createFolder("Log/", "sql1..." + str(sql1))
                    cnx.execute(sql1)
                    cnx.commit()
                else:
                    sql1 = text(f'''UPDATE ems_v1.master_diesel_entry SET 
                            diesel_litre = '{diesel_litre}',
                            dg_runtime = '{dg_runtime}',
                            modified_on = getdate(),
                            modified_by = '{user_login_id}'
                        WHERE
                            dg_name = '{dg_name}' AND 
                            entry_date = '{mill_date}' ''')
                    createFolder("Log/", "sql1..." + str(sql1))
                    cnx.execute(sql1)
                    cnx.commit()
        if obj_production is not None:
            obj_data = json.loads(obj_production)
            for row in obj_data: 
                production_name = row["production_name"]
                production_value = row["production_value"]
                if entry_date == '':
                    sql2 = text(f"""INSERT INTO ems_v1.master_production_entry (
                        production_name,
                        production_value,
                        entry_date,
                        created_on,
                        created_by
                    ) VALUES (
                        '{production_name}',
                        '{production_value}',
                        '{mill_date}',
                        getdate(),
                        '{user_login_id}'
                    )""")
                    createFolder("Log/", "sql2..." + str(sql2))
                    cnx.execute(sql2)
                    cnx.commit()
                else:
                    sql2=text(f'''update ems_v1.master_production_entry set 
                            production_value = '{production_value}',
                            modified_on = getdate(),
                            modified_by = '{user_login_id}'
                        where
                            production_name = '{production_name}' and 
                            entry_date = '{mill_date}' ''')
                    createFolder("Log/", "sql2..." + str(sql2))
                    cnx.execute(sql2)
                    cnx.commit()
        if obj_load is not None:
            obj_data = json.loads(obj_load)
            for row in obj_data: 
                load_name = row["load_name"]
                load_value = row["load_value"]
                if entry_date =='':
                    sql3 = text(f"""INSERT INTO ems_v1.master_load_entry (
                        load_name,
                        load_value,
                        entry_date,
                        created_on,
                        created_by
                    ) VALUES (
                        '{load_name}',
                        '{load_value}',
                        '{mill_date}',
                        getdate(),
                        '{user_login_id}'
                    )""")
                    createFolder("Log/", "sql3..." + str(sql3))
                    cnx.execute(sql3)
                    cnx.commit()
                else:
                    sql3=f'''update ems_v1.master_load_entry set 
                            load_value = '{load_value}',
                            modified_on = getdate(),
                            modified_by = '{user_login_id}'
                        where
                            load_name = '{load_name}' and 
                            entry_date = '{mill_date}' '''
                    createFolder("Log/", "sql3..." + str(sql3))
                    cnx.execute(sql3)
                    cnx.commit()
                
        if obj_heat is not None:
            obj_data = json.loads(obj_heat)
            for row in obj_data: 
                heat_name = row["heat_name"]
                heat_value = row["heat_value"]
                if entry_date =='':              
                    sql4 = text(f"""INSERT INTO ems_v1.master_heat_entry (
                        heat_name,
                        heat_value,
                        entry_date,
                        created_on,
                        created_by
                    ) VALUES (
                        '{heat_name}',
                        '{heat_value}',
                        '{mill_date}',
                        getdate(),
                        '{user_login_id}'
                    )""")
                    createFolder("Log/", "sql4..." + str(sql4))
                    cnx.execute(sql4)
                    cnx.commit()
                else:                

                    sql4=text(f'''update ems_v1.master_heat_entry set 
                            heat_value = '{heat_value}',
                            modified_on = getdate(),
                            modified_by = '{user_login_id}'
                        where
                            heat_name = '{heat_name}' and 
                            entry_date = '{mill_date}' ''')
                    createFolder("Log/", "sql4..." + str(sql4))
                    
                    cnx.execute(sql4)
                    cnx.commit()
        return JSONResponse({"iserror": False, "message": "successfully updated", "data": ""})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})      

@app.post("/remove_manualentry/")
async def remove_manualentry(entry_date: str = Form(''), 
                             cnx: Session = Depends(get_db)):
    
    try:
        
        if entry_date == '':
            return JSONResponse({"iserror": False, "message": "entry_date not Passed."})
        
        else:
            sql1= text(f"delete from ems_v1.master_diesel_entry where entry_date = '{entry_date}'")
            cnx.execute(sql1)
            cnx.commit()

            sql2= text(f"delete from ems_v1.master_production_entry where entry_date = '{entry_date}'")
            cnx.execute(sql2)
            cnx.commit()

            sql3= text(f"delete from ems_v1.master_load_entry where entry_date = '{entry_date}'")
            cnx.execute(sql3)
            cnx.commit()

            sql4= text(f"delete from ems_v1.master_heat_entry where entry_date = '{entry_date}'")
            cnx.execute(sql4)
            cnx.commit()

        return JSONResponse({"iserror": False, "message": "deleted successfully", "data": ""})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message}) 

def manual_wisereport(data, entry_date):
    try:
        # file_path = f'{static_dir}/manual_wise_report_template.xlsx'
        file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "manual_wise_report_template.xlsx"))
        createFolder("Log/", "file path " + str(file_path))
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        sheet.title = 'EMS'
        createFolder("Log/", "Issue in returning data ..")
        # Set the entry date
        sheet['B3'] = f"Date: {entry_date}"
        sheet['B3'].font = Font(bold=True, name='Calibri', size=11)

        # Set the headers
        sheet['B6'] = 'DG Name'
        sheet['C6'] = 'Consumption (Litre)'

        sheet['E6'] = 'Production Zone'
        sheet['F6'] = 'Production(MT)'

        sheet['H6'] = 'Production Zone'
        sheet['I6'] = 'Production(MT)'

        sheet['E14'] = 'Area'
        sheet['F14'] = 'Heats(Nos)'

        sheet['B14'] = 'DG Name'
        sheet['C14'] = 'Time(Min)'

        # Populate the data
        total_consumption = 0
        total_production = 0 
        total_power = 0
        total_heats = 0
        total_running_time  = 0

        row_index = 7

        for item in data['dieselLists']:
            dg_name = str(item["dg_name"])
            if item["diesel_litre"] == '':
                diesel_litre = 0
            else:
                diesel_litre = float(item["diesel_litre"])
            total_consumption += diesel_litre

            sheet.cell(row=row_index, column=2).value = dg_name
            sheet.cell(row=row_index, column=3).value = diesel_litre
            row_index += 1

        # Add total consumption row
        total_row_index = row_index
        sheet.cell(row=11, column=2).value = 'Total Consumption'
        sheet.cell(row=11, column=3).value = total_consumption

        # Auto-adjust column widths
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20

        row_index = 7
        for item in data['productionLists']:
        
            production_name = str(item["production_name"])
            if item["production_value"]=='':
                production_value = 0
            else:
                production_value = float(item["production_value"])
            total_production += production_value

            sheet.cell(row=row_index, column=5).value = production_name
            sheet.cell(row=row_index, column=6).value = production_value
            row_index += 1

        # Add total production row
        total_row_index = row_index
        sheet.cell(row=11, column=5).value = 'Total Production'
        sheet.cell(row=11, column=6).value = total_production
        # Auto-adjust column widths
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20

        row_index = 7
        for item in data['loadLists']:
            
            load_name = str(item["load_name"])
            if item["load_value"] =='':
                load_value = 0
            else:
                load_value = float(item["load_value"])
            total_power += load_value
            sheet.cell(row=row_index, column=8).value = load_name
            sheet.cell(row=row_index, column=9).value = load_value
            row_index += 1
            
            sheet.cell(row=23, column=8).value = 'Total Power'
            sheet.cell(row=23, column=9).value = total_power

            # Auto-adjust column widths
            sheet.column_dimensions['H'].width = 20
            sheet.column_dimensions['I'].width = 20

        row_index = 15
        for item in data['heatLists']:
            
            heat_name = str(item["heat_name"])
            if item["heat_value"] == '':
                heat_value = 0
            else:
                heat_value = float(item["heat_value"])
            total_heats += heat_value
            sheet.cell(row=row_index, column=5).value = heat_name
            sheet.cell(row=row_index, column=6).value = heat_value
            row_index += 1
            
            sheet.cell(row=23, column=5).value = 'Total Heats'
            sheet.cell(row=23, column=6).value = total_heats

            # Auto-adjust column widths
            sheet.column_dimensions['H'].width = 20
            sheet.column_dimensions['I'].width = 20

        row_index = 15
        for item in data['dieselLists']:
            dg_name = str(item["dg_name"])
            if item["dg_runtime"] == '':
                dg_runtime = 0
            else:
                dg_runtime = float(item["dg_runtime"])
            total_running_time += dg_runtime
            sheet.cell(row=row_index, column=2).value = dg_name
            sheet.cell(row=row_index, column=3).value = dg_runtime
            row_index += 1

            sheet.cell(row=23, column=2).value = 'Total Running Time '
            sheet.cell(row=23, column=3).value = total_running_time

            # Auto-adjust column widths
            sheet.column_dimensions['H'].width = 20
            sheet.column_dimensions['I'].width = 20

        file_name = f'manual_wise_report-{entry_date}.xlsx'
        print(file_name)
        file_path = os.path.join(base_path, file_name)
        workbook.save(file_path)

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)

@app.post("/manualentry_report/")
async def manualentry_report(request:Request,
                             entry_date: str = Form(None), 
                             pdf : str = Form(''),
                             cnx: Session = Depends(get_db)):
   
    try:       
        where = ''
        select = '*'

        if entry_date is not None :
            data = {}
            data = {'entry_date': entry_date}
            # entry_date = entry_date[6:10] + '-' + entry_date[3:5] + '-' + entry_date[0:2]
            where = text(f"WHERE FORMAT(entry_date, 'dd-MM-yyyy') = '{entry_date}'")
            print(where)

            diesel_sql = text(f"SELECT {select} FROM ems_v1.master_diesel_entry {where}")
            production_sql = text(f"SELECT {select} FROM ems_v1.master_production_entry {where}")
            load_sql = text(f"SELECT {select} FROM ems_v1.master_load_entry {where}")
            heat_sql = text(f"SELECT {select} FROM ems_v1.master_heat_entry {where}")

            data['dieselLists'] = cnx.execute(diesel_sql).fetchall()
            data['productionLists'] = cnx.execute(production_sql).fetchall()
            data['loadLists'] = cnx.execute(load_sql).fetchall()
            data['heatLists'] = cnx.execute(heat_sql).fetchall()
            print(diesel_sql)
        else:  
            return JSONResponse({"iserror": True, "message": "entry_date is required"})
        # formatted_date = entry_date.strftime('%d/%m/%Y')
        # print("Formatted date:", formatted_date)
        if pdf == 'excel':
            manual_wisereport(data, entry_date)

            file_path = os.path.join(base_path, f"manual_wise_report-{entry_date}.xlsx")
            results = f"http://{request.headers['host']}/attachment/manual_wise_report-{entry_date}.xlsx"
            
            if os.path.exists(file_path):
                return JSONResponse({"iserror": False, "message": "data return successfully","file_url": results})
            else:
                return JSONResponse({"iserror": False, "message": "data return successfully","file_url": None}) 
        else:
            return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message}) 
    
@app.post("/manual_chart_analysis/")
async def manual_chart_analysis(from_date: str = Form(''), 
                                to_date : str = Form(''),
                                cnx: Session = Depends(get_db)):  

    try:

        where = ''
        select = '*'   
             
        if from_date != '':
            from_date = parse_date(from_date)
            where += f"WHERE entry_date >= CONVERT(date, '{from_date}', 120) "
		
        if to_date != '':
            to_date = parse_date(to_date)
            where += f"AND entry_date <= CONVERT(date, '{to_date}', 120)"

        data = {}  

        diesel_sql = text(f"SELECT {select} FROM ems_v1.master_diesel_entry {where}")
        production_sql = text(f"SELECT {select} FROM ems_v1.master_production_entry {where}")
        load_sql = text(f"SELECT {select} FROM ems_v1.master_load_entry {where}")
        heat_sql = text(f"SELECT {select} FROM ems_v1.master_heat_entry {where}")
     
        data['dieselLists'] = cnx.execute(diesel_sql).fetchall()
        data['productionLists'] = cnx.execute(production_sql).fetchall()
        data['loadLists'] = cnx.execute(load_sql).fetchall()
        data['heatLists'] = cnx.execute(heat_sql).fetchall()

        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message}) 

# def functionality_wise_report(datas):
#     try:
#         workbook = openpyxl.Workbook()
#         workbook.remove(workbook.active)  # Remove the default sheet created
#         f_count = 0
#         style = {'alignment': Alignment(horizontal='center')}
        
#         border = Border(
#             left=Side(style='thin', color='000000'),
#             right=Side(style='thin', color='000000'),
#             top=Side(style='thin', color='000000'),
#             bottom=Side(style='thin', color='000000')
#         )
#         styles = style1 = {
#             'fill': PatternFill(fill_type='solid', fgColor='90d973'),
#             'font': Font(bold=True, color='000000', size=30, name='Verdana'),
#             'border': border,
#             'alignment': Alignment(horizontal='center')
#         }
#         style1 = {
#             'fill': PatternFill(fill_type='solid', fgColor='90d973'),
#             'font': Font(bold=True, color='000000', size=10, name='Verdana'),
#             'border': border,
#             'alignment': Alignment(horizontal='center')
#         }

#         style2 = {
#             'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
#             'font': Font(bold=True, color='000000', size=10, name='Verdana'),
#             'border': border,
#             'alignment': Alignment(horizontal='center')
#         }
        
#         kwh1 = 0
#         kwh2 = 0
#         capacity1 = 0
#         capacity2 = 0
#         production1 = 0
#         production2 = 0
#         # for row in datas["res"]:
#         #         date = datas["from_date"]
#         #         print(date)
#         #         sheet = workbook.create_sheet(title=row['name'])
#         #         image_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "jsw.jpeg"))
#         #         img = Image(image_path)
#         #         img.anchor = "A1"  # Set the anchor point to A1
                
#         #         img_width_pixels, img_height_pixels = img.width, img.height
#         #         # Set the desired image size
#         #         desired_width_pixels = 150
#         #         desired_height_pixels = 45
#         #         # Calculate the scaling factor for width and height
#         #         width_scale = desired_width_pixels / img_width_pixels
#         #         height_scale = desired_height_pixels / img_height_pixels
#         #         # Scale the image size
#         #         img.width = int(img_width_pixels * width_scale)
#         #         img.height = int(img_height_pixels * height_scale)
#         #         sheet.column_dimensions['A'].width = img.width / 10
#         #         sheet.row_dimensions[1].height = img.height / 2
#         #         sheet.row_dimensions[2].height = img.height / 2

#         #         sheet.merge_cells('A1:A2')  # Merge cells A1 to A2
#         #         sheet.add_image(img)
                
#         #         sheet['B2'].value = 'SMS Power Consumption Report - ' + row['name'] + ' Wise '
#         #         sheet.merge_cells('B2:H2')
#         #         sheet['B2'].alignment = style['alignment']
#         #         sheet.merge_cells('A3:D3')
#         #         sheet['A3'].value = 'SMS#1'
#         #         sheet['A3'].alignment = style['alignment']
#         #         sheet['A4'].value = 'Area'
#         #         sheet['A4'].alignment = style['alignment']
#         #         sheet['B4'].value = 'Feeder'
#         #         sheet['B4'].alignment = style['alignment']
#         #         sheet['C4'].value = 'Capacity'
#         #         sheet['C4'].alignment = style['alignment']
#         #         sheet['D4'].value = 'Consumption(kWh)'
#         #         sheet['D4'].alignment = style['alignment']
#         #         sheet.merge_cells('E3:H3')
#         #         sheet['E3'].value = 'SMS#2'
#         #         sheet['E3'].alignment = style['alignment']
#         #         sheet['E4'].value = 'Area'
#         #         sheet['E4'].alignment = style['alignment']
#         #         sheet['F4'].value = 'Feeder'
#         #         sheet['F4'].alignment = style['alignment']
#         #         sheet['G4'].value = 'Capacity'
#         #         sheet['G4'].alignment = style['alignment']
#         #         sheet['H4'].value = 'Consumption(kWh)'
#         #         sheet['H4'].alignment = style['alignment']
#         #         sheet['B1'].value = f'date:{date}'
#         #         sheet['B1'].alignment = style['alignment']

#         #         row_count = 5
#         #         for rows in row['data']:
#         #             sheet['A' + str(row_count)].value = rows.area1
#         #             sheet['A' + str(row_count)].alignment = style['alignment'] 
#         #             sheet['B' + str(row_count)].value = rows.machine1
#         #             sheet['B' + str(row_count)].alignment = style['alignment']
                    
#         #             if rows.machine1:
#         #                 sheet['D' + str(row_count)].value = rows.kwh1
#         #                 sheet['D' + str(row_count)].alignment = style['alignment']
                        
#         #                 kwh1= rows.kwh1
#         #             if rows.machine1:
#         #                 sheet['C' + str(row_count)].value = rows.capacity1
#         #                 sheet['C' + str(row_count)].alignment = style['alignment']

#         #             sheet['E' + str(row_count)].value = rows.area2
#         #             sheet['E' + str(row_count)].alignment = style['alignment']
                    
#         #             sheet['F' + str(row_count)].value = rows.machine2
#         #             sheet['F' + str(row_count)].alignment = style['alignment']

#         #             if rows.machine2:
#         #                 sheet['H' + str(row_count)].value = rows.kwh2
#         #                 sheet['H' + str(row_count)].alignment = style['alignment']
                        
#         #                 kwh2 = rows.kwh2
#         #             if rows.machine2:
#         #                 sheet['G' + str(row_count)].value = rows.capacity2
#         #                 sheet['G' + str(row_count)].alignment = style['alignment']
                        
#         #             row_count += 1
                
#         #         for production in datas["productionLists"]:
#         #             print(123456789)  # Check if the loop is iterating as expected
#         #             if production["production_name"] == 'SMS1 Production':
#         #                 production1 = production["production_value"]
#         #             elif production["production_name"] == 'SMS2 Production':
#         #                 production2 = production["production_value"]
    
#         #         add_row = row_count + 4
#         #         sheet['A' + str(add_row)].value = 'SMS1 Emergency Power Consumption'
                
#         #         sheet['D' + str(add_row)].value = '=SUM(D4:D' + str(row_count) + ')'
#         #         sheet['D' + str(add_row)].alignment = style['alignment']
                
#         #         sheet['E' + str(add_row)].value = 'kWh'
#         #         sheet['E' + str(add_row)].alignment = style['alignment']
            
#         #         sheet['F' + str(add_row)].value = 'SMS1 Production'
                
#         #         sheet['G' + str(add_row)].value = int(production1)
#         #         sheet['G' + str(add_row)].alignment = style['alignment']
            
#         #         sheet['H' + str(add_row)].value = 'MT'
                
#         #         add_row1 = add_row + 1
#         #         sheet['A' + str(add_row1)].value = 'SMS2 Emergency Power Consumption' 
#         #         sheet['D' + str(add_row1)].value = '=SUM(H4:H' + str(row_count) + ')'
#         #         sheet['D' + str(add_row1)].alignment = style['alignment']
#         #         sheet['E' + str(add_row1)].value = 'kWh'
#         #         sheet['E' + str(add_row1)].alignment = style['alignment'] 
#         #         sheet['F' + str(add_row1)].value = 'SMS2 Production'
#         #         sheet['G' + str(add_row1)].value = int(production2)
#         #         sheet['G' + str(add_row1)].alignment = style['alignment'] 
#         #         sheet['H' + str(add_row1)].value = 'MT'
                
#         #         total_production = int(production1)+ int(production2)
#         #         print("total_production",total_production)
#         #         add_row2 = add_row1 + 1
#         #         sheet['A' + str(add_row2)].value = 'Total Emergency Power Consumption' 
#         #         total_kwh1 = sheet['D' + str(add_row)].value
#         #         total_kwh2 = sheet['D' + str(add_row1)].value
#         #         sheet['D' + str(add_row2)].value = '=SUM(D4:D' + str(row_count) + ')+SUM(H4:H' + str(row_count) + ')'
#         #         sheet['D' + str(add_row2)].alignment = style['alignment'] 
#         #         sheet['E' + str(add_row2)].value = 'kWh'
#         #         sheet['E' + str(add_row2)].alignment = style['alignment']
#         #         sheet['F' + str(add_row2)].value = 'Total Production'
#         #         sheet['G' + str(add_row2)].value = total_production
#         #         sheet['G' + str(add_row2)].alignment = style['alignment']
#         #         sheet['H' + str(add_row2)].value = 'MT'
#         #         print("production1",production1)
#         #         print("production2",production2)
#         #         print("kwh1",kwh1)
#         #         # if rows.machine1:
#         #         #     kwh1 = rows.kwh
#         #         # if rows.machine2:
#         #         #     kwh2 = rows.kwh2
#         #         add_row3 = add_row2 + 1
#         #         add_row3 = add_row2 + 1
#         #         if int(production1) == 0:
#         #             sms1_cons = 0
#         #         else:
#         #             sms1_cons = kwh1 / int(production1)

#         #         if int(production2) == 0:
#         #             sms2_cons = 0
#         #         else:
#         #             sms2_cons = kwh2 / int(production2)
#         #         print("sms1_cons",sms1_cons)
#         #         sheet['A' + str(add_row3)].value = 'SMS1 Emergency Power Consumption(kWh/MT)' 
#         #         sheet['D' + str(add_row3)].value = sms1_cons
#         #         sheet['D' + str(add_row3)].alignment = style['alignment']
#         #         sheet['D' + str(add_row3)].fill = style1['fill']
#         #         sheet['D' + str(add_row3)].font = style1['font']
#         #         sheet['E' + str(add_row3)].value = 'SMS2 Emergency Power Consumption(kWh/MT)'
#         #         sheet['H' + str(add_row3)].value = sms2_cons
#         #         sheet['H' + str(add_row3)].alignment = style['alignment']
#         #         sheet['H' + str(add_row3)].fill = style1['fill']
#         #         sheet['H' + str(add_row3)].font = style1['font']

#         #         add_row4 = add_row3 + 1
                
#         #         # Evaluate the formula and get the calculated value for total_consumption1
#         #         total_consumption1 = sheet['D' + str(add_row2)].value
#         #         print("total_consumption1",total_consumption1)
#         #                     # Get the value of total_consumption2 directly
#         #         total_consumption2 = sheet['G' + str(add_row2)].value
#         #         print("total_consumption2:", total_consumption2)

#         #         if total_consumption2 == 0:
#         #             total_cons = 0
#         #         else:
#         #             total_cons = f"=(SUM(D4:D{row_count})+SUM(H4:H{row_count}))/{total_consumption2}"

#         #         sheet['A' + str(add_row4)].value = 'SMS Power Consumption (kWh/MT) = '
#         #         sheet.merge_cells('E' + str(add_row4) + ':H' + str(add_row4))

#         #         merged_cell = sheet['E' + str(add_row4)]
#         #         merged_cell.alignment = style['alignment']
#         #         merged_cell.fill = style2['fill']
#         #         merged_cell.font = style2['font']

#         #         sheet['E' + str(add_row4)].value = total_cons
#         #         sheet['E' + str(add_row4)].alignment = style['alignment'] 
#         #         rows_count = 1
#         #         row_range = sheet.iter_rows(min_row=rows_count, max_row=add_row4, min_col=1, max_col=9)
#         #         for row in row_range:
#         #             for cell in row:
#         #                 cell.border = border

#         #         sheet.column_dimensions['A'].width = 40
#         #         sheet.column_dimensions['B'].width = 32
#         #         sheet.column_dimensions['D'].width = 20
#         #         sheet.column_dimensions['E'].width = 25
#         #         sheet.column_dimensions['F'].width = 20
#         #         sheet.column_dimensions['H'].width = 20
                
#         #         f_count += 1
        
#         file_name = 'functionality_wise_report.xlsx'
#         file_path = os.path.join(base_path, file_name)
#         workbook.save(file_path)
#         print(file_name)
       
#     except Exception as e:
#         error_type = type(e).__name__
#         error_line = traceback.extract_tb(e.__traceback__)[0].lineno
#         error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
#         error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
#         createFolder("Log/", "Issue in returning data " + error_message)

def functionality_wise_report(datas,from_date):
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet created
        f_count = 0
        style = {'alignment': Alignment(horizontal='center')}
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        styles = style1 = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        style1 = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }

        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        
        kwh1 = 0
        kwh2 = 0
        capacity1 = 0
        capacity2 = 0
        production1 = 0
        production2 = 0
        if len(datas) == 0:
            row_count = 7
            sheet = workbook.create_sheet()
            # image_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "jsw.jpeg"))
            image_path = os.path.abspath(__file__)+"//jsw.jpeg"
            img = Image(image_path)
            img.anchor = "A1"  # Set the anchor point to A1
            
            img_width_pixels, img_height_pixels = img.width, img.height
            # Set the desired image size
            desired_width_pixels = 150
            desired_height_pixels = 45
            # Calculate the scaling factor for width and height
            width_scale = desired_width_pixels / img_width_pixels
            height_scale = desired_height_pixels / img_height_pixels
            # Scale the image size
            img.width = int(img_width_pixels * width_scale)
            img.height = int(img_height_pixels * height_scale)
            sheet.column_dimensions['A'].width = img.width / 10
            sheet.row_dimensions[1].height = img.height / 2
            sheet.row_dimensions[2].height = img.height / 2
            sheet.merge_cells('A1:A2')  # Merge cells A1 to A2
            sheet.add_image(img)
            sheet['I' + str(row_count)].value = 'No Data'
            sheet['I' + str(row_count)].alignment = styles['alignment'] 
        else:
            for row in datas["res"]:
                date = datas["from_date"]
                print(date)
                sheet = workbook.create_sheet(title=row['name'])
                # image_path = os.path.abspath(os.path.join(os.path.dirname(__file__),"jsw.jpeg"))
                script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))

                # Combine the script directory with the file name to create the file path
                file_name = "jsw.jpeg"
                image_path = os.path.join(script_dir, file_name)
                img = Image(image_path)
                img.anchor = "A1"  # Set the anchor point to A1
                
                img_width_pixels, img_height_pixels = img.width, img.height
                # Set the desired image size
                desired_width_pixels = 150
                desired_height_pixels = 45
                # Calculate the scaling factor for width and height
                width_scale = desired_width_pixels / img_width_pixels
                height_scale = desired_height_pixels / img_height_pixels
                # Scale the image size
                img.width = int(img_width_pixels * width_scale)
                img.height = int(img_height_pixels * height_scale)
                sheet.column_dimensions['A'].width = img.width / 10
                sheet.row_dimensions[1].height = img.height / 2
                sheet.row_dimensions[2].height = img.height / 2

                sheet.merge_cells('A1:A2')  # Merge cells A1 to A2
                sheet.add_image(img)
                
                sheet['B2'].value = 'SMS Power Consumption Report - ' + row['name'] + ' Wise '
                sheet.merge_cells('B2:H2')
                sheet['B2'].alignment = style['alignment']
                sheet.merge_cells('A3:D3')
                sheet['A3'].value = 'SMS#1'
                sheet['A3'].alignment = style['alignment']
                sheet['A4'].value = 'Area'
                sheet['A4'].alignment = style['alignment']
                sheet['B4'].value = 'Feeder'
                sheet['B4'].alignment = style['alignment']
                sheet['C4'].value = 'Capacity'
                sheet['C4'].alignment = style['alignment']
                sheet['D4'].value = 'Consumption(kWh)'
                sheet['D4'].alignment = style['alignment']
                sheet.merge_cells('E3:H3')
                sheet['E3'].value = 'SMS#2'
                sheet['E3'].alignment = style['alignment']
                sheet['E4'].value = 'Area'
                sheet['E4'].alignment = style['alignment']
                sheet['F4'].value = 'Feeder'
                sheet['F4'].alignment = style['alignment']
                sheet['G4'].value = 'Capacity'
                sheet['G4'].alignment = style['alignment']
                sheet['H4'].value = 'Consumption(kWh)'
                sheet['H4'].alignment = style['alignment']
                sheet['B1'].value = f'date:{date}'
                sheet['B1'].alignment = style['alignment']
                
                row_count = 5
                for rows in row['data']:
                    sheet['A' + str(row_count)].value = rows.area1
                    sheet['A' + str(row_count)].alignment = style['alignment'] 
                    sheet['B' + str(row_count)].value = rows.machine1
                    sheet['B' + str(row_count)].alignment = style['alignment']
                    
                    if rows.machine1:
                        sheet['D' + str(row_count)].value = rows.kwh1
                        sheet['D' + str(row_count)].alignment = style['alignment']
                        
                        kwh1= rows.kwh1
                        print("13245612345",kwh1)
                    if rows.machine1:
                        sheet['C' + str(row_count)].value = rows.capacity1
                        sheet['C' + str(row_count)].alignment = style['alignment']

                    sheet['E' + str(row_count)].value = rows.area2
                    sheet['E' + str(row_count)].alignment = style['alignment']
                    
                    sheet['F' + str(row_count)].value = rows.machine2
                    sheet['F' + str(row_count)].alignment = style['alignment']

                    if rows.machine2:
                        sheet['H' + str(row_count)].value = rows.kwh2
                        sheet['H' + str(row_count)].alignment = style['alignment']
                        
                        kwh2 = rows.kwh2
                    if rows.machine2:
                        sheet['G' + str(row_count)].value = rows.capacity2
                        sheet['G' + str(row_count)].alignment = style['alignment']
                        
                    row_count += 1

                for production in datas["productionLists"]:
                    if production.production_name == 'SMS1 Production':
                        production1 = production.production_value
                    elif production.production_name == 'SMS2 Production':
                        production2 = production.production_value

                add_row = row_count + 4
                sheet['A' + str(add_row)].value = 'SMS1 Emergency Power Consumption'
                
                sheet['D' + str(add_row)].value = '=SUM(D4:D' + str(row_count) + ')'
                formula_cell = sheet['D' + str(add_row)]
                formula_value = formula_cell.value
                sheet['D' + str(add_row)].alignment = style['alignment']
                
                sheet['E' + str(add_row)].value = 'kWh'
                sheet['E' + str(add_row)].alignment = style['alignment']
            
                sheet['F' + str(add_row)].value = 'SMS1 Production'
                
                sheet['G' + str(add_row)].value = int(production1)
                sheet['G' + str(add_row)].alignment = style['alignment']
            
                sheet['H' + str(add_row)].value = 'MT'
                
                add_row1 = add_row + 1
                sheet['A' + str(add_row1)].value = 'SMS2 Emergency Power Consumption' 
                sheet['D' + str(add_row1)].value = '=SUM(H4:H' + str(row_count) + ')'
                formula_cell2 = sheet['D' + str(add_row1)]
                formula_value2 = formula_cell2.value
                sheet['D' + str(add_row1)].alignment = style['alignment']
                sheet['E' + str(add_row1)].value = 'kWh'
                sheet['E' + str(add_row1)].alignment = style['alignment'] 
                sheet['F' + str(add_row1)].value = 'SMS2 Production'
                sheet['G' + str(add_row1)].value = int(production2)
                sheet['G' + str(add_row1)].alignment = style['alignment'] 
                sheet['H' + str(add_row1)].value = 'MT'
                
                total_production = int(production1)+ int(production2)
                print("total_production",total_production)
                add_row2 = add_row1 + 1
                sheet['A' + str(add_row2)].value = 'Total Emergency Power Consumption' 
                total_kwh1 = sheet['D' + str(add_row)].value
                total_kwh2 = sheet['D' + str(add_row1)].value
                sheet['D' + str(add_row2)].value = '=SUM(D4:D' + str(row_count) + ')+SUM(H4:H' + str(row_count) + ')'
                sheet['D' + str(add_row2)].alignment = style['alignment'] 
                sheet['E' + str(add_row2)].value = 'kWh'
                sheet['E' + str(add_row2)].alignment = style['alignment']
                sheet['F' + str(add_row2)].value = 'Total Production'
                sheet['G' + str(add_row2)].value = total_production
                sheet['G' + str(add_row2)].alignment = style['alignment']
                sheet['H' + str(add_row2)].value = 'MT'
                print("production1",production1)
                print("production2",production2)
                print("kwh1",formula_value)
                add_row3 = add_row2 + 1
                add_row3 = add_row2 + 1
                if production1 == 0:
                    sms1_cons = 0
                else:
                    sms1_cons = f"=({formula_value[1:]} / {production1})"

                    print("sms1_cons",sms1_cons)
                if production2 == 0:
                    sms2_cons = 0
                else:
                    sms2_cons = f"=({formula_value2[1:]} / {production2})"

                sheet['A' + str(add_row3)].value = 'SMS1 Emergency Power Consumption(kWh/MT)' 
                sheet['D' + str(add_row3)].value = sms1_cons
                sheet['D' + str(add_row3)].alignment = style['alignment']
                sheet['D' + str(add_row3)].fill = style1['fill']
                sheet['D' + str(add_row3)].font = style1['font']
                sheet['E' + str(add_row3)].value = 'SMS2 Emergency Power Consumption(kWh/MT)'
                sheet['H' + str(add_row3)].value = sms2_cons
                sheet['H' + str(add_row3)].alignment = style['alignment']
                sheet['H' + str(add_row3)].fill = style1['fill']
                sheet['H' + str(add_row3)].font = style1['font']

                add_row4 = add_row3 + 1
                
                # Evaluate the formula and get the calculated value for total_consumption1
                total_consumption1 = sheet['D' + str(add_row2)].value
                print("total_consumption1",total_consumption1)
                            # Get the value of total_consumption2 directly
                total_consumption2 = sheet['G' + str(add_row2)].value
                print("total_consumption2:", total_consumption2)

                if total_consumption2 == 0:
                    total_cons = 0
                else:
                    total_cons = f"=(SUM(D4:D{row_count})+SUM(H4:H{row_count}))/{total_consumption2}"

                sheet['A' + str(add_row4)].value = 'SMS Power Consumption (kWh/MT) = '
                sheet.merge_cells('E' + str(add_row4) + ':H' + str(add_row4))

                merged_cell = sheet['E' + str(add_row4)]
                merged_cell.alignment = style['alignment']
                merged_cell.fill = style2['fill']
                merged_cell.font = style2['font']

                sheet['E' + str(add_row4)].value = total_cons
                sheet['E' + str(add_row4)].alignment = style['alignment'] 
                rows_count = 1
                row_range = sheet.iter_rows(min_row=rows_count, max_row=add_row4, min_col=1, max_col=9)
                for row in row_range:
                    for cell in row:
                        cell.border = border

                sheet.column_dimensions['A'].width = 40
                sheet.column_dimensions['B'].width = 32
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width = 25
                sheet.column_dimensions['F'].width = 20
                sheet.column_dimensions['H'].width = 20
                
                f_count += 1
            
        file_name = f'functionality_wise_report-{from_date}.xlsx'
        file_path = os.path.join(base_path, file_name)
        workbook.save(file_path)
        print(file_name)
       
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[-1].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)

@app.post("/functionality_report/")
async def functionality_report(request:Request,
                               function_id :str = Form(''),
                               from_date: str = Form(''), 
                               cnx: Session = Depends(get_db)):  

        
    if from_date == '':
            return JSONResponse({"iserror": False, "message": "from_date is required"})
        
    try:
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
       
        sql= text(f"select * from ems_v1.master_production_entry where FORMAT(entry_date,'dd-MM-yyyy')='{from_date}'")
        datas1 = cnx.execute(sql).fetchall()    
        print(sql)
        res = []
        datas = ''
    
        function_id1 = 0
        where = ''
        date = from_date
        from_date = parse_date(from_date)
        month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
        tbl_name_month = f"{month_year}"
        if function_id != '':
            where = f'and function_id = {function_id}'
        
        query = f'''select * from ems_v1.master_function where status = 'active'{where} '''
        dt = cnx.execute(query).fetchall()
        if len(dt)>0:
            for i in dt:
                function_id1 = i["function_id"]
            
                sql1=text(f'''  SELECT
                            min(fd.function_name)as function_name,
                            min(fd.function_id)as function_id,
                            min(fd.function_code)as function_code,
                            min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
                            min(fd.function_id)as opt_id,
                            min(fd.function_name)as opt_actual,
                            min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
                            count(cp.machine_id)as meter_count,
                            min(fd.image)as image
                        from
                            ems_v1.master_function fd
                            left join ems_v1.master_machine mm ON mm.function_id= fd.function_id and mm.status='active'
                            left join ems_v1.current_power cp ON cp.machine_id=mm.machine_id
                        where fd.status<>'delete' and fd.function_id  = '{function_id1}'
                        group by fd.function_id
                        order by fd.function_id''')
                data1= cnx.execute(sql1).fetchall()
                
                where_1 = f"where FORMAT(cp.mill_date, 'yyyy-MM-dd HH:mm:ss') = '{from_date}'"
    
                sql2 = text(f"""
                        SELECT 
                    isnull((select department_name from ems_v1.dbo.master_department where department_id=m1.department_id),'') as zone1,
                    isnull((select department_name from ems_v1.dbo.master_department where department_id=m2.department_id),'') as zone2,
                    isnull((select shed_name from ems_v1.dbo.master_shed where shed_id=m1.shed_id),'') as area1,
                    isnull((select shed_name from ems_v1.dbo.master_shed where shed_id=m2.shed_id),'') as area2,
                    isnull((select function_name from ems_v1.dbo.master_function where function_id=m1.function_id),'') as function1,
                    isnull((select function_name from ems_v1.dbo.master_function where function_id=m2.function_id),'') as function2,
                    isnull(m1.capacity_id,'')as capacity1,
                    isnull(m2.capacity_id,'') as capacity2,
                    c.*
                from
                (
                    select
                        min(case when p.department_id = 1 then p.machine_name  end) as machine1,
                        sum(case when p.department_id = 1 then p.kwh/1000  end) as kwh1,
                        min(case when p.department_id = 2 then p.machine_name  end) as machine2,
                        sum(case when p.department_id = 2 then p.kwh/1000  end) as kwh2
                    from (
                        select
                            min(md.department_id) as department_id,
                            min(md.department_name) as department_name,
                            min(mf.function_id) as function_id,
                            min(mf.function_name) as function_name,
                            min(mm.machine_id) as machine_id,
                            min(mm.machine_name) as machine_name,
                            sum(cp.kwh) as kwh,
                            ROW_NUMBER() over (partition by min(md.department_name) order by min(mm.machine_name)) as sno
                        from 
                            ems_v1_completed.dbo.power_{tbl_name_month} as cp
                            inner join ems_v1.master_machine mm on mm.machine_id = cp.machine_id
                            inner join ems_v1.master_department md on md.department_id = mm.department_id
                            inner join ems_v1.master_function mf on mf.function_id = mm.function_id
                        {where_1} and cp.status = '0' and  mm.function_id = {function_id1}
                        group by mm.machine_id
                        
                        ) as p
                    group by p.sno
                ) as c
                left join ems_v1.master_machine m1 on m1.machine_name=c.machine1 and m1.status='active'
                left join ems_v1.master_machine m2 on m2.machine_name=c.machine2 and m2.status='active'
                """)
                print(sql2)
                createFolder("Log/", "sql query" + str(sql2))
                data2= cnx.execute(sql2).fetchall()
                
                if function_id == 'all' or function_id == '':
                    for row in data1:
                        res.append({"name": row["function_name"], "data": data2})
                else:
                            
                    res.append({"name":"", "data": data2})
                curtime1 = from_date.strftime("%d-%m-%Y %H:%M:%S")
                from_date1 = curtime1[:11]
                print("from_date1",curtime1)
                datas = {
                        "productionLists": datas1,
                        "res": res,
                        "from_date": from_date1,
                        "excel_filename": ""
                    }
        createFolder("Log/", "functionality report datas" + str(datas))
        print("datas1",datas1)

        functionality_wise_report(datas,date)
        file_path = os.path.join(base_path, f"functionality_wise_report-{date}.xlsx")
        results = f"http://{request.headers['host']}/attachment/functionality_wise_report-{date}.xlsx"

        if os.path.exists(file_path):

           return JSONResponse({"iserror": False, "message": "data return successfully","data": results})
        else:
           
            return JSONResponse({"iserror": False, "message": "data return successfully","data": ""})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message}) 
    
@app.post("/get_sld_datas/")
async def get_sld_datas(sld: int = Form(''), 
                        cnx: Session = Depends(get_db)):  

    try:
        if sld == '':  
            return JSONResponse({"iserror": False, "message": "sld is required"})
        query = text(f'''SELECT * FROM ems_v1.master_shifts WHERE status = 'active' ''')
        data1 = cnx.execute(query).fetchall()
        mill_date = date.today()
        mill_shift = 0        
    
        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]            
              print(mill_date)
        where = ''
        orderby = ''
        if sld == 1:
            print(1)
            where = f'and mm.machine_id IN (38,39,42,43,40,44,45,41)'
            orderby = f'''CASE mm.machine_id
                            WHEN 38 THEN 0
                            WHEN 39 THEN 1
                            WHEN 42 THEN 2
                            WHEN 43 THEN 3
                            WHEN 40 THEN 4
                            WHEN 44 THEN 5
                            WHEN 45 THEN 6
                            WHEN 41 THEN 7
                            ELSE 99
                        END'''
        elif sld == 2:
            where = f'and mm.machine_id IN (53,54,46,47,48,49,55,60,64)'
            orderby = f'''CASE mm.machine_id
                            WHEN 53 THEN 0
                            WHEN 54 THEN 1
                            WHEN 46 THEN 2
                            WHEN 47 THEN 3
                            WHEN 48 THEN 4
                            WHEN 49 THEN 5
                            WHEN 55 THEN 6
                            WHEN 60 THEN 7
                            WHEN 64 THEN 8
                            ELSE 99
                        END'''
        elif sld == 3:
            where = f'and mm.machine_id IN (68,67,65,66)'
            orderby = f'''CASE mm.machine_id
                            WHEN 68 THEN 0
                            WHEN 67 THEN 1
                            WHEN 65 THEN 2
                            WHEN 66 THEN 3
                            ELSE 99
                        END'''
        elif sld == 4:
            where = f'and mm.machine_id IN (61,63)'
            orderby = f'''CASE mm.machine_id
                            WHEN 61 THEN 0
                            WHEN 63 THEN 1
                            ELSE 99
                        END'''
        sql = f'''
            select
                min(mc.company_code) AS company_code,
                min(mc.company_name) AS company_name,
                min(mb.branch_code) AS branch_code,
                min(mb.branch_name) AS branch_name,
                min(md.department_code) AS department_code,
                min(md.department_name) As department_name,
                min(ms.shed_code) AS shed_code,
                min(ms.shed_name) AS shed_name,
                min(mmt.machinetype_code) AS machinetype_code,
                min(mmt.machinetype_name) AS machinetype_name,
                min(mf.function_code) AS function_code,
                min(mf.function_name) As function_name,
                min(mm.machine_code) AS machine_code,
                min(mm.machine_name) AS machine_name,
                count(mm.machine_name) AS machine_count,
                min(cp.power_id) as power_id,
                min(cp.company_id) as company_id,
                min(cp.branch_id) as branch_id,
                min(cp.department_id) as department_id,
                min(cp.shed_id) as shed_id,
                min(cp.machinetype_id) as machinetype_id,
                min(mm.function_id) as function_id,
                min(cp.machine_id) as machine_id,
                min(cp.design_id) as design_id,
                min(cp.beam_id) as beam_id,
                min(cp.date_time) as date_time,
                min(cp.date_time1) as date_time1,
                min(cp.mill_date) as mill_date,
                min(cp.mill_shift) as mill_shift,
                ROUND(SUM(case when mmf.vln_avg = '*' then cp.vln_avg * mmf.vln_avg_value when  mmf.vln_avg = '/' then cp.vln_avg / mmf.vln_avg_value else cp.vln_avg end ),2) AS vln_avg,
                ROUND(SUM(case when mmf.r_volt = '*' then cp.r_volt * mmf.r_volt_value when  mmf.r_volt = '/' then cp.r_volt / mmf.r_volt_value else cp.r_volt end ),2) AS r_volt,
                ROUND(SUM(case when mmf.y_volt = '*' then cp.y_volt * mmf.y_volt_value when  mmf.y_volt = '/' then cp.y_volt / mmf.y_volt_value else cp.y_volt end ),2) AS y_volt,
                ROUND(SUM(case when mmf.b_volt = '*' then cp.b_volt * mmf.b_volt_value when  mmf.b_volt = '/' then cp.b_volt / mmf.b_volt_value else cp.b_volt end ),2) AS b_volt,
                ROUND(SUM(case when mmf.vll_avg = '*' then cp.vll_avg * mmf.vll_avg_value when  mmf.vll_avg = '/' then cp.vll_avg / mmf.vll_avg_value else cp.vll_avg end ),2) AS vll_avg,
                ROUND(SUM(case when mmf.ry_volt = '*' then cp.ry_volt * mmf.ry_volt_value when  mmf.ry_volt = '/' then cp.ry_volt / mmf.ry_volt_value else cp.ry_volt end ),2) AS ry_volt,
                ROUND(SUM(case when mmf.yb_volt = '*' then cp.yb_volt * mmf.yb_volt_value when  mmf.yb_volt = '/' then cp.yb_volt / mmf.yb_volt_value else cp.yb_volt end ),2) AS yb_volt,
                ROUND(SUM(case when mmf.br_volt = '*' then cp.br_volt * mmf.br_volt_value when  mmf.br_volt = '/' then cp.br_volt / mmf.br_volt_value else cp.br_volt end ),2) AS br_volt,
                ROUND(SUM(case when mmf.br_volt = '*' then cp.br_volt * mmf.br_volt_value when  mmf.br_volt = '/' then cp.br_volt / mmf.br_volt_value else cp.br_volt end ),2) AS br_volt,
                ROUND(SUM(case when mmf.t_current = '*' then cp.t_current * mmf.t_current_value when  mmf.t_current = '/' then cp.t_current / mmf.t_current_value else cp.t_current end ),2) AS t_current,
                ROUND(SUM(case when mmf.r_current = '*' then cp.r_current * mmf.r_current_value when  mmf.r_current = '/' then cp.r_current / mmf.r_current_value else cp.r_current end ),2) AS r_current,
                ROUND(SUM(case when mmf.y_current = '*' then cp.y_current * mmf.y_current_value when  mmf.y_current = '/' then cp.y_current / mmf.y_current_value else cp.y_current end ),2) AS y_current,
                ROUND(SUM(case when mmf.b_current = '*' then cp.b_current * mmf.b_current_value when  mmf.b_current = '/' then cp.b_current / mmf.b_current_value else cp.b_current end ),2) AS b_current,
                ROUND(SUM(case when mmf.t_watts = '*' then cp.t_watts * mmf.t_watts_value when  mmf.t_watts = '/' then cp.t_watts / mmf.t_watts_value else cp.t_watts end ),2) AS t_watts,
                ROUND(SUM(case when mmf.r_watts = '*' then cp.r_watts * mmf.r_watts_value when  mmf.r_watts = '/' then cp.r_watts / mmf.r_watts_value else cp.r_watts end ),2) AS r_watts,
                ROUND(SUM(case when mmf.y_watts = '*' then cp.y_watts * mmf.y_watts_value when  mmf.y_watts = '/' then cp.y_watts / mmf.y_watts_value else cp.y_watts end ),2) AS y_watts,
                ROUND(SUM(case when mmf.b_watts = '*' then cp.b_watts * mmf.b_watts_value when  mmf.b_watts = '/' then cp.b_watts / mmf.b_watts_value else cp.b_watts end ),2) AS b_watts,
                ROUND(SUM(case when mmf.t_var = '*' then cp.t_var * mmf.t_var_value when  mmf.t_var = '/' then cp.t_var / mmf.t_var_value else cp.t_var end ),2) AS t_var,
                ROUND(SUM(case when mmf.r_var = '*' then cp.r_var * mmf.r_var_value when  mmf.r_var = '/' then cp.r_var / mmf.r_var_value else cp.r_var end ),2) AS r_var,
                ROUND(SUM(case when mmf.y_var = '*' then cp.y_var * mmf.y_var_value when  mmf.y_var = '/' then cp.y_var / mmf.y_var_value else cp.y_var end ),2) AS y_var,
                ROUND(SUM(case when mmf.b_var = '*' then cp.b_var * mmf.b_var_value when  mmf.b_var = '/' then cp.b_var / mmf.b_var_value else cp.b_var end ),2) AS b_var,
                ROUND(SUM(case when mmf.t_voltampere = '*' then cp.t_voltampere * mmf.t_voltampere_value when  mmf.t_voltampere = '/' then cp.t_voltampere / mmf.t_voltampere_value else cp.t_voltampere end ),2) AS t_voltampere,
                ROUND(SUM(case when mmf.r_voltampere = '*' then cp.r_voltampere * mmf.r_voltampere_value when  mmf.r_voltampere = '/' then cp.r_voltampere / mmf.r_voltampere_value else cp.r_voltampere end ),2) AS r_voltampere,
                ROUND(SUM(case when mmf.y_voltampere = '*' then cp.y_voltampere * mmf.y_voltampere_value when  mmf.y_voltampere = '/' then cp.y_voltampere / mmf.y_voltampere_value else cp.y_voltampere end ),2) AS y_voltampere,
                ROUND(SUM(case when mmf.b_voltampere = '*' then cp.b_voltampere * mmf.b_voltampere_value when  mmf.b_voltampere = '/' then cp.b_voltampere / mmf.b_voltampere_value else cp.b_voltampere end ),2) AS b_voltampere,
                ROUND(SUM(case when mmf.avg_powerfactor = '*' then cp.avg_powerfactor * mmf.avg_powerfactor_value when  mmf.avg_powerfactor = '/' then cp.avg_powerfactor / mmf.avg_powerfactor_value else cp.avg_powerfactor end ),2) AS avg_powerfactor,
                ROUND(SUM(case when mmf.r_powerfactor = '*' then cp.r_powerfactor * mmf.r_powerfactor_value when  mmf.r_powerfactor = '/' then cp.r_powerfactor / mmf.r_powerfactor_value else cp.r_powerfactor end ),2) AS r_powerfactor,
                ROUND(SUM(case when mmf.y_powerfactor = '*' then cp.y_powerfactor * mmf.y_powerfactor_value when  mmf.y_powerfactor = '/' then cp.y_powerfactor / mmf.y_powerfactor_value else cp.y_powerfactor end ),2) AS y_powerfactor,
                ROUND(SUM(case when mmf.b_powerfactor = '*' then cp.b_powerfactor * mmf.b_powerfactor_value when  mmf.b_powerfactor = '/' then cp.b_powerfactor / mmf.b_powerfactor_value else cp.b_powerfactor end ),2) AS b_powerfactor,
                ROUND(SUM(case when mmf.powerfactor = '*' then cp.powerfactor * mmf.powerfactor_value when  mmf.powerfactor = '/' then cp.powerfactor / mmf.powerfactor_value else cp.powerfactor end ),2) AS powerfactor,
                    
                ROUND(SUM(case when mmf.kvah = '*' then cp.kvah * mmf.kvah_value when  mmf.kvah = '/' then cp.kvah / mmf.kvah_value else cp.kvah end ),2) AS kvah,
                ROUND(SUM(case when mmf.kw = '*' then cp.t_watts * mmf.kw_value when  mmf.kw = '/' then cp.t_watts / mmf.kw_value else cp.t_watts end ),2) AS kw,
                ROUND(SUM(case when mmf.kvar = '*' then cp.kvar * mmf.kvar_value when  mmf.kvar = '/' then cp.kvar / mmf.kvar_value else cp.kvar end ),2) AS kvar,
                ROUND(SUM(case when mmf.power_factor = '*' then cp.power_factor * mmf.power_factor_value when  mmf.power_factor = '/' then cp.power_factor / mmf.power_factor_value else cp.power_factor end ),2) AS power_factor,
                ROUND(SUM(case when mmf.kva = '*' then cp.kva * mmf.kva_value when  mmf.kva = '/' then cp.kva / mmf.kva_value else cp.kva end ),2) AS kva,
                ROUND(SUM(case when mmf.frequency = '*' then cp.frequency * mmf.frequency_value when  mmf.frequency = '/' then cp.frequency / mmf.frequency_value else cp.frequency end ),2) AS frequency,
                min(cp.machine_status) as machine_status,
                min(cp.status) as status,
                min(cp.created_on) as created_on,
                min(cp.created_by) as created_by,
                min(cp.modified_on) as modified_on,
                min(cp.modified_by) as modified_by,
                    
                ROUND(SUM(case when mmf.machine_kWh = '*' then cp.machine_kWh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.machine_kWh / mmf.machine_kWh_value else cp.machine_kWh end ),2) AS machine_kWh,
                ROUND(SUM(case when mmf.machine_kWh = '*' then cp.master_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.master_kwh / mmf.machine_kWh_value else cp.master_kwh end ),2) AS master_kwh,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kWh,
                     
                min(mm.ip_address) as ip_address,
                min(mm.port) as port,
                CASE WHEN min(cp.date_time) <= DATEADD(minute, -2, getdate()) THEN 'S' ELSE 'N' END as nocom,       
                ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_1,
                ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_2,
                ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_3,
                ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),2) AS start_kwh_1,
                ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),2) AS start_kwh_2,
                ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),2) AS start_kwh_3,     
                ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),2) AS end_kwh_1,
                ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),2) AS end_kwh_2,
                ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),2) AS end_kwh_3
            from
                ems_v1.current_power cp
                INNER JOIN ems_v1.master_machine mm ON cp.machine_id = mm.machine_id
                INNER JOIN ems_v1.master_company mc ON mm.company_id = mc.company_id
                INNER JOIN ems_v1.master_branch mb ON mm.branch_id = mb.branch_id
                INNER JOIN ems_v1.master_department md ON mm.department_id = md.department_id
                INNER JOIN ems_v1.master_shed ms ON mm.shed_id = ms.shed_id
                INNER JOIN ems_v1.master_machinetype mmt ON mm.machinetype_id = mmt.machinetype_id 
                LEFT JOIN ems_v1.master_function mf ON mm.function_id = mf.function_id
                LEFT JOIN ems_v1.master_machine_factor mmf ON mm.machine_id = mmf.machine_id
            where 
                mm.status = 'active' and
                cp.mill_date = '{mill_date}' and 
                cp.mill_shift = {mill_shift} {where}
            group by mm.machine_id
            order by min({orderby})'''
        print(sql)
        result = cnx.execute(sql).fetchall()
        createFolder("Log/", "Issue in returning data " + str(sql))
        return JSONResponse({"iserror": False, "message": "data return successfully","data": jsonable_encoder(result)})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message}) 

def power_report(data,previous_mill_date,filename):
    try:
        file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "power_report_template.xlsx"))
          
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        sheet.title = 'EMS'# Remove the default sheet created
  
        style = {'alignment': Alignment(horizontal='center')}
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        styles = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        style1 = {
            
            'font': Font(bold=True, color='006400', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }

        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        row_count = 5

        # Create a set to store unique function names
        function_names = set()

        # Iterate over the data to collect unique function names
        for row in data["res"]:
            function_names.add(row.function_name)

        # Iterate over each unique function name
        for function_name in function_names:
            # Display the function name in the sheet
            sheet['A' + str(row_count)].value = function_name
            sheet['A' + str(row_count)].alignment = style1['alignment']
            sheet['A' + str(row_count)].font = style1['font']
            row_count += 1
            
            # Iterate over the data again to find matching records for the current function
            for row in data["res"]:
                if row.function_name == function_name:
                    sheet['A' + str(row_count)].value = row.mill_date
                    sheet['B' + str(row_count)].value = row.machine_name
                    sheet['C' + str(row_count)].value = row.c_kwh
                    sheet['D' + str(row_count)].value = row.kwh
                    row_count += 1

            row_count += 1  # Add an empty row between each function's details

        file_name = 'total_power_report.xlsx'
        file_path = os.path.join(filename, file_name)
        workbook.save(file_path)
        print(file_name)
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)

def power_report_dtl(groupby,groupids,machine_ids,report_type,from_date,to_date,shift_id,report_type_for,cnx):
    try:
        sql = "select * from ems_v1.master_shifts"
        result = cnx.execute(sql).fetchall()      
        mill_date = ''
        mill_shift = 0
        no_of_shifts = 0        
        for row in result:
            mill_date = row.mill_date
            mill_shift = row.mill_shift
            no_of_shifts = row.no_of_shifts
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
 
        if mill_date == "":
            return JSONResponse({"iserror": True, "message": "Mill Date, Shift Missing. Check Current Shifts Table"})       
        day_kwh_tbl = f'''select machine_id,mill_date,kwh from ems_v1.dbo.current_power'''
        # mill_date = parse_date(mill_date)
        month_year=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""       
        tbl_name_month = f"{month_year}"
        if len(result)>0:

            day_kwh_tbl = f" union all select machine_id,mill_date,kwh from ems_v1_completed.dbo.power_{tbl_name_month}"

        is_cur = 0

        if datetime.datetime.strptime(from_date, "%d-%m-%Y") <= mill_date and datetime.datetime.strptime(to_date, "%d-%m-%Y") >= mill_date:
            is_cur = mill_shift
         
        group_by = ''
        order_by = ''
        where = ''
        tbl_name = ''
	
        if shift_id != '':
	
            where += f" and cp.mill_shift={shift_id}"

        if len(groupids) == len(groupids.replace('all', '')) and len(groupids) != 0:
            if groupby == "function":
                where += f" and mm.{groupby}_id in ({groupids})"
            else:
                where += f" and cp.{groupby}_id in ({groupids})"


        if len(machine_ids) == len(machine_ids.replace('all', '')) and len(machine_ids) != 0:
            where += f" and cp.machine_id in ({machine_ids})"

        if report_type == "detail":
            if groupby=="function":
			
                group_by = f"mm.{groupby}_id ,cp.machine_id"
                order_by = f"mm.{groupby}_id ,min(mm.machine_order)"
            else:
                group_by = f"cp.mill_date ,cp.{groupby}_id ,cp.machine_id"
                order_by = f"cp.mill_date ,cp.{groupby}_id ,min(mm.machine_order)"
        else:
            if groupby=="function":
			
                group_by = f"mm.{groupby}_id ,cp.machine_id"
                order_by = f"mm.{groupby}_id ,min(mm.machine_order)"            
            else:
                if report_type_for == 'machine':
                    group_by = f"cp.{groupby}_id ,cp.machine_id"
                    order_by = f"cp.{groupby}_id ,min(mm.machine_order)"
                else:
                    group_by = "cp.{groupby}_id ,cp.mill_date"
                    order_by = "cp.{groupby}_id ,cp.mill_date"                  
	
        tbl_name += f'(SELECT company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,kwh,machine_kwh as end_kwh,master_kwh as start_kwh,mill_date,mill_shift FROM ems_v1.dbo.current_power '

        tmpdt = datetime.datetime.strptime(from_date, '%d-%m-%Y').strftime('%Y-%m-%d')
        j = 0

        for i in range(6):  # For Tmp Loop
            if i == 5:
                i = 1
            tmpdt = (datetime.datetime.strptime(tmpdt, '%Y-%m-%d') + relativedelta(months=j)).strftime('%Y-%m-%d')
            tbl_name += f''' UNION ALL 
                SELECT company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,kwh,machine_kwh as end_kwh,master_kwh as start_kwh,mill_date,mill_shift FROM ems_v1_completed.dbo.power_{tmpdt[5:7]}{tmpdt[0:4]}'''
            j = 1
            print(tmpdt)
            print(to_date)
            if tmpdt[5:7] == to_date[3:5] and tmpdt[0:4] == to_date[6:10]:
                break

        tbl_name += " ) as cp"
        sql = f'''SELECT 
			min(md.department_id) as department_id,
			min(md.department_code) as department_code,
			min(md.department_name) as department_name,
			min(ms.shed_id) as shed_id,
			min(ms.shed_code) as shed_code,
			min(ms.shed_name) as shed_name,
			min(mt.machinetype_id) as machinetype_id,
			min(mt.machinetype_code) as machinetype_code,
			min(mt.machinetype_name) as machinetype_name,
			FORMAT(min(cp.mill_date),'dd-MM-yyyy') as mill_date,
			min(mm.machine_code) as machine_code,
			min(mm.machine_name) as machine_name,
			min(mf.function_id) as function_id,
			min(mf.function_name) as function_name,
			min(mf.function_code) as function_code,
			ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kWh,       
			ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_1,
            ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_2,
            ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_3,
            ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.start_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.start_kwh / mmf.machine_kwh_value else cp.start_kwh end ELSE 0 END),2) AS start_kwh_1,
            ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.start_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.start_kwh / mmf.machine_kwh_value else cp.start_kwh end ELSE 0 END),2) AS start_kwh_2,
            ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.start_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.start_kwh / mmf.machine_kwh_value else cp.start_kwh end ELSE 0 END),2) AS start_kwh_3,     
            ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.end_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.end_kwh / mmf.machine_kwh_value else cp.end_kwh end ELSE 0 END),2) AS end_kwh_1,
            ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.end_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.end_kwh / mmf.machine_kwh_value else cp.end_kwh end ELSE 0 END),2) AS end_kwh_2,
            ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.end_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.end_kwh / mmf.machine_kwh_value else cp.end_kwh end ELSE 0 END),2) AS end_kwh_3,
			case when {is_cur} <> 0 and '{mill_date}' = FORMAT(min(cp.mill_date),'yyyy-MM-dd') then
				ROUND(SUM(CASE WHEN cp.mill_shift = {is_cur} THEN case when mmf.machine_kwh = '*' then cp.end_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.end_kwh / mmf.machine_kwh_value else cp.end_kwh end ELSE 0 END),2)
			else
				ROUND(SUM(CASE WHEN cp.mill_shift = {no_of_shifts} THEN case when mmf.machine_kwh = '*' then cp.end_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.end_kwh / mmf.machine_kwh_value else cp.end_kwh end ELSE 0 END),2)
			end as c_kwh
		FROM 
			{tbl_name} ,
			ems_v1.master_machine mm,
			ems_v1.master_machinetype mt,
			ems_v1.master_shed ms,
			ems_v1.master_department md,
			ems_v1.master_function mf,
			ems_v1.master_machine_factor mmf                 
		WHERE cp.machine_id = mm.machine_id and mm.machinetype_id=mt.machinetype_id and mm.shed_id=ms.shed_id and mm.department_id=md.department_id and mm.function_id=mf.function_id and mm.status='active' and FORMAT(cp.mill_date ,'dd-MM-yyyy')>='{from_date}' and FORMAT(cp.mill_date , 'dd-MM-yyyy')<='{to_date}' {where}
		GROUP BY {group_by}
		ORDER BY {order_by} '''
       
        createFolder("Log/", "query " + str(sql))

        result = cnx.execute(sql).fetchall()
        return result
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)

def auto_download_main(filename,cnx):
    try:
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            ems_v1.master_shifts'''
        shifttimings = cnx.execute(sql).fetchall()
        previous_mill_date = date.today()
        for row in shifttimings:
            previous_mill_date = row.previous_mill_date
        from_date = previous_mill_date
        to_date = previous_mill_date
        shift_id = ''
        pdf1 = 'excel'
        group_by = 'department'
        group_ids = ''
        machine_ids = ''
        report_type = 'detail'
        report_type_for = 'machine'
        rpt_field = 'kwh,c_kwh'
        pdf = 'print'
        data = {}
        head = ""
        if group_by == "shed":
            head = "Area"
        elif group_by == "machinetype":
            head = "Location"
        elif group_by == "department":
            head = "Department"       
        elif group_by == "function":
            head = "Function"
        title = head + "wise Summary Report"
        if report_type == "detail":
            title = head + "wise Detail Report"

        title += " From " + from_date + " To " + to_date
        if shift_id != '':
            title += " Shift " + shift_id

        res = power_report_dtl(group_by,group_ids,machine_ids,report_type,from_date,to_date,shift_id,report_type_for,cnx)
        createFolder("Log/","res"+str(res))
        data['res'] = res
        data['title'] = title
        data['head'] = head
        data['pdf'] = pdf
        data['group_by'] = group_by
        data['rpt_field'] = rpt_field

        data['res'] = res
        power_report(data,previous_mill_date,filename)
        
        data['excel_filename'] = filename
        createFolder("Log/", "data " + str(data)) 

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)    

def functionalityy_wise_report(datas,filename):
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet created
        f_count = 0
        style = {'alignment': Alignment(horizontal='center')}
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        styles = style1 = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=30, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        style1 = {
            'fill': PatternFill(fill_type='solid', fgColor='90d973'),
            'font': Font(bold=True, color='000000', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }

        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=10, name='Verdana'),
            'border': border,
            'alignment': Alignment(horizontal='center')
        }
        
        
        for row in datas["res"]:
            date = datas["from_date"]
            print(date)
            sheet = workbook.create_sheet(title=row['name'])
            image_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "jsw.jpeg"))
            img = Image(image_path)
            img.anchor = "A1"  # Set the anchor point to A1
                
            img_width_pixels, img_height_pixels = img.width, img.height
            # Set the desired image size
            desired_width_pixels = 150
            desired_height_pixels = 45
            # Calculate the scaling factor for width and height
            width_scale = desired_width_pixels / img_width_pixels
            height_scale = desired_height_pixels / img_height_pixels
            # Scale the image size
            img.width = int(img_width_pixels * width_scale)
            img.height = int(img_height_pixels * height_scale)
            sheet.column_dimensions['A'].width = img.width / 10
            sheet.row_dimensions[1].height = img.height / 2
            sheet.row_dimensions[2].height = img.height / 2

            sheet.merge_cells('A1:A2')  # Merge cells A1 to A2
            sheet.add_image(img)
                
            sheet['B2'].value = 'SMS Power Consumption Report - ' + row['name'] + ' Wise '
            sheet.merge_cells('B2:H2')
            sheet['B2'].alignment = style['alignment']
            sheet.merge_cells('A3:D3')
            sheet['A3'].value = 'SMS#1'
            sheet['A3'].alignment = style['alignment']
            sheet['A4'].value = 'Area'
            sheet['A4'].alignment = style['alignment']
            sheet['B4'].value = 'Feeder'
            sheet['B4'].alignment = style['alignment']
            sheet['C4'].value = 'Capacity'
            sheet['C4'].alignment = style['alignment']
            sheet['D4'].value = 'Consumption(kWh)'
            sheet['D4'].alignment = style['alignment']
            sheet.merge_cells('E3:H3')
            sheet['E3'].value = 'SMS#2'
            sheet['E3'].alignment = style['alignment']
            sheet['E4'].value = 'Area'
            sheet['E4'].alignment = style['alignment']
            sheet['F4'].value = 'Feeder'
            sheet['F4'].alignment = style['alignment']
            sheet['G4'].value = 'Capacity'
            sheet['G4'].alignment = style['alignment']
            sheet['H4'].value = 'Consumption(kWh)'
            sheet['H4'].alignment = style['alignment']
            sheet['B1'].value = f'date:{date}'
            sheet['B1'].alignment = style['alignment']
            kwh1 = 0
            kwh2 = 0
            capacity1 = 0
            capacity2 = 0
                
            row_count = 5
            for rows in row['data']:
                sheet['A' + str(row_count)].value = rows.area1
                sheet['A' + str(row_count)].alignment = style['alignment'] 
                sheet['B' + str(row_count)].value = rows.machine1
                sheet['B' + str(row_count)].alignment = style['alignment']
                
                if rows.machine1:
                    sheet['D' + str(row_count)].value = rows.kwh1
                    sheet['D' + str(row_count)].alignment = style['alignment']
                    
                    kwh1= rows.kwh1
                if rows.machine1:
                    sheet['C' + str(row_count)].value = rows.capacity1
                    sheet['C' + str(row_count)].alignment = style['alignment']
                sheet['E' + str(row_count)].value = rows.area2
                sheet['E' + str(row_count)].alignment = style['alignment']
                
                sheet['F' + str(row_count)].value = rows.machine2
                sheet['F' + str(row_count)].alignment = style['alignment']
                if rows.machine2:
                    sheet['H' + str(row_count)].value = rows.kwh2
                    sheet['H' + str(row_count)].alignment = style['alignment']
                    
                    kwh2 = rows.kwh2
                if rows.machine2:
                    sheet['G' + str(row_count)].value = rows.capacity2
                    sheet['G' + str(row_count)].alignment = style['alignment']
                    
                row_count += 1
            production1 = 0
            production2 = 0
            for production in datas["productionLists"]:
                if production.production_name == 'SMS1 Production':
                    production1 = production.production_value
                elif production.production_name == 'SMS2 Production':
                    production2 = production.production_value
            add_row = row_count + 4
            sheet['A' + str(add_row)].value = 'SMS1 Emergency Power Consumption'
            
            sheet['D' + str(add_row)].value = '=SUM(D4:D' + str(row_count) + ')'
            sheet['D' + str(add_row)].alignment = style['alignment']
            
            sheet['E' + str(add_row)].value = 'kWh'
            sheet['E' + str(add_row)].alignment = style['alignment']
        
            sheet['F' + str(add_row)].value = 'SMS1 Production'
            
            sheet['G' + str(add_row)].value = int(production1)
            sheet['G' + str(add_row)].alignment = style['alignment']
        
            sheet['H' + str(add_row)].value = 'MT'
            
            add_row1 = add_row + 1
            sheet['A' + str(add_row1)].value = 'SMS2 Emergency Power Consumption' 
            sheet['D' + str(add_row1)].value = '=SUM(H4:H' + str(row_count) + ')'
            sheet['D' + str(add_row1)].alignment = style['alignment']
            sheet['E' + str(add_row1)].value = 'kWh'
            sheet['E' + str(add_row1)].alignment = style['alignment'] 
            sheet['F' + str(add_row1)].value = 'SMS2 Production'
            sheet['G' + str(add_row1)].value = int(production2)
            sheet['G' + str(add_row1)].alignment = style['alignment'] 
            sheet['H' + str(add_row1)].value = 'MT'
            
            total_production = int(production1)+ int(production2)
            print("total_production",total_production)
            add_row2 = add_row1 + 1
            sheet['A' + str(add_row2)].value = 'Total Emergency Power Consumption' 
            total_kwh1 = sheet['D' + str(add_row)].value
            total_kwh2 = sheet['D' + str(add_row1)].value
            sheet['D' + str(add_row2)].value = '=SUM(D4:D' + str(row_count) + ')+SUM(H4:H' + str(row_count) + ')'
            sheet['D' + str(add_row2)].alignment = style['alignment'] 
            sheet['E' + str(add_row2)].value = 'kWh'
            sheet['E' + str(add_row2)].alignment = style['alignment']
            sheet['F' + str(add_row2)].value = 'Total Production'
            sheet['G' + str(add_row2)].value = total_production
            sheet['G' + str(add_row2)].alignment = style['alignment']
            sheet['H' + str(add_row2)].value = 'MT'
            print("production1",production1)
            print("production2",production2)
            print("kwh1",kwh1)
            add_row3 = add_row2 + 1
            add_row3 = add_row2 + 1
            if int(production1) == 0:
                sms1_cons = 0
            else:
                sms1_cons = kwh1 / int(production1)
            if int(production2) == 0:
                sms2_cons = 0
            else:
                sms2_cons = kwh2 / int(production2)
            sheet['A' + str(add_row3)].value = 'SMS1 Emergency Power Consumption(kWh/MT)' 
            sheet['D' + str(add_row3)].value = sms1_cons
            sheet['D' + str(add_row3)].alignment = style['alignment']
            sheet['D' + str(add_row3)].fill = style1['fill']
            sheet['D' + str(add_row3)].font = style1['font']
            sheet['E' + str(add_row3)].value = 'SMS2 Emergency Power Consumption(kWh/MT)'
            sheet['H' + str(add_row3)].value = sms2_cons
            sheet['H' + str(add_row3)].alignment = style['alignment']
            sheet['H' + str(add_row3)].fill = style1['fill']
            sheet['H' + str(add_row3)].font = style1['font']
            add_row4 = add_row3 + 1
            
            # Evaluate the formula and get the calculated value for total_consumption1
            total_consumption1 = sheet['D' + str(add_row2)].value
            print("total_consumption1",total_consumption1)
                        # Get the value of total_consumption2 directly
            total_consumption2 = sheet['G' + str(add_row2)].value
            print("total_consumption2:", total_consumption2)
            if total_consumption2 == 0:
                total_cons = 0
            else:
                total_cons = f"=(SUM(D4:D{row_count})+SUM(H4:H{row_count}))/{total_consumption2}"
            sheet['A' + str(add_row4)].value = 'SMS Power Consumption (kWh/MT) = '
            sheet.merge_cells('E' + str(add_row4) + ':H' + str(add_row4))
            merged_cell = sheet['E' + str(add_row4)]
            merged_cell.alignment = style['alignment']
            merged_cell.fill = style2['fill']
            merged_cell.font = style2['font']
            sheet['E' + str(add_row4)].value = total_cons
            sheet['E' + str(add_row4)].alignment = style['alignment'] 
            rows_count = 1
            row_range = sheet.iter_rows(min_row=rows_count, max_row=add_row4, min_col=1, max_col=9)
            for row in row_range:
                for cell in row:
                    cell.border = border
            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 32
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 25
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['H'].width = 20
            
            f_count += 1
            
        file_name = 'functionality_wise_report.xlsx'
        file_path = os.path.join(filename, file_name)
        workbook.save(file_path)
        print(file_name)
       
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)

def auto_download_functionality(filename,cnx):
    try:
                
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            ems_v1.dbo.master_shifts
        '''
        shifttimings = cnx.execute(sql).fetchall()
        previous_mill_date = date.today()
        for row in shifttimings:
            previous_mill_date = row.previous_mill_date
		
        function_id = ''
        from_date = previous_mill_date
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
       
        sql1= f'''select * from ems_v1.master_production_entry where FORMAT(entry_date,'dd-MM-yyyy')='{from_date}' '''
        print("sql1",sql1)
        datas1 = cnx.execute(sql1).fetchall() 
        print(datas1)  

        res = []
        datas = ''
    
        function_id1 = 0
        where = ''
        from_date = parse_date(from_date)
        month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
        tbl_name_month = f"{month_year}"
        # if function_id != '':
        #     where = f'and function_id = {function_id}'
        # query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
        # query1 = cnx.execute(query1).fetchall()
        # if len(query1)>0:  
        query = f'''select * from ems_v1.dbo.master_function where status = 'active'{where} '''
        dt = cnx.execute(query).fetchall()
        if len(dt)>0:
            for i in dt:
                function_id1 = i.function_id
                
                sql1=f'''  SELECT
                                min(fd.function_name)as function_name,
                                min(fd.function_id)as function_id,
                                min(fd.function_code)as function_code,
                                min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
                                min(fd.function_id)as opt_id,
                                min(fd.function_name)as opt_actual,
                                min(CONCAT(fd.function_code,' - ',fd.function_name)) AS function_actual,
                                count(cp.machine_id)as meter_count,
                                min(fd.image)as image
                            from
                                ems_v1.master_function fd
                                left join ems_v1.master_machine mm ON mm.function_id= fd.function_id and mm.status='active'
                                left join ems_v1.current_power cp ON cp.machine_id=mm.machine_id
                            where fd.status<>'delete' and fd.function_id  = '{function_id1}'
                            group by fd.function_id
                            order by fd.function_id'''
                data1= cnx.execute(sql1).fetchall()
                
                where_1 = f"where FORMAT(cp.mill_date, 'yyyy-MM-dd HH:mm:ss') = '{from_date}'"
    
                sql2 = f"""
                            SELECT 
                        isnull((select department_name from ems_v1.master_department where department_id=m1.department_id),'') as zone1,
                        isnull((select department_name from ems_v1.master_department where department_id=m2.department_id),'') as zone2,
                        isnull((select shed_name from ems_v1.master_shed where shed_id=m1.shed_id),'') as area1,
                        isnull((select shed_name from ems_v1.master_shed where shed_id=m2.shed_id),'') as area2,
                        isnull((select function_name from ems_v1.master_function where function_id=m1.function_id),'') as function1,
                        isnull((select function_name from ems_v1.master_function where function_id=m2.function_id),'') as function2,
                        isnull(m1.capacity_id,'')as capacity1,
                        isnull(m2.capacity_id,'') as capacity2,
                        c.*
                    from
                    (
                        select
                            min(case when p.department_id = 1 then p.machine_name  end) as machine1,
                            sum(case when p.department_id = 1 then p.kwh/1000  end) as kwh1,
                            min(case when p.department_id = 2 then p.machine_name  end) as machine2,
                            sum(case when p.department_id = 2 then p.kwh/1000  end) as kwh2
                        from (
                            select
                                min(md.department_id) as department_id,
                                min(md.department_name) as department_name,
                                min(mf.function_id) as function_id,
                                min(mf.function_name) as function_name,
                                min(mm.machine_id) as machine_id,
                                min(mm.machine_name) as machine_name,
                                sum(cp.kwh) as kwh,
                                ROW_NUMBER() over (partition by min(md.department_name) order by min(mm.machine_name)) as sno
                            from 
                                ems_v1_completed.dbo.power_{tbl_name_month} as cp
                                inner join ems_v1.master_machine mm on mm.machine_id = cp.machine_id
                                inner join ems_v1.master_department md on md.department_id = mm.department_id
                                inner join ems_v1.master_function mf on mf.function_id = mm.function_id
                            {where_1} and cp.status = '0' and  mm.function_id = {function_id1}
                            group by mm.machine_id
                            
                            ) as p
                        group by p.sno
                    ) as c
                    left join ems_v1.master_machine m1 on m1.machine_name=c.machine1 and m1.status='active'
                    left join ems_v1.master_machine m2 on m2.machine_name=c.machine2 and m2.status='active'

                    """
                print("sql2",sql2)
                    
                data2= cnx.execute(sql2).fetchall()
                    
                if function_id == 'all' or function_id == '':

                    for row in data1:
                        res.append({"name": row.function_name, "data": data2})
                else:
                                
                    res.append({"name":"", "data": data2})
                curtime1 = from_date.strftime("%d-%m-%Y %H:%M:%S")
                from_date1 = curtime1[:11]
                print("from_date1",curtime1)
                datas = {
                        "productionLists": datas1,
                        "res": res,
                        "from_date": from_date1,
                        "excel_filename": filename
                    }
                    
        print("datas...",datas)
        functionalityy_wise_report(datas,filename)
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        
def manual_wise_report(data, entry_date,filename):
    # file_path = f'{static_dir}/manual_wise_report_template.xlsx'
    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "manual_wise_report_template.xlsx"))
          
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS'

    # Set the entry date
    sheet['B3'] = f"Date: {entry_date}"
    sheet['B3'].font = Font(bold=True, name='Calibri', size=11)

    # Set the headers
    sheet['B6'] = 'DG Name'
    sheet['C6'] = 'Consumption (Litre)'

    sheet['E6'] = 'Production Zone'
    sheet['F6'] = 'Production(MT)'

    sheet['H6'] = 'Production Zone'
    sheet['I6'] = 'Production(MT)'

    sheet['E14'] = 'Area'
    sheet['F14'] = 'Heats(Nos)'

    sheet['B14'] = 'DG Name'
    sheet['C14'] = 'Time(Min)'

    # Populate the data
    total_consumption = 0
    total_production = 0 
    total_power = 0
    total_heats = 0
    total_running_time  = 0

    row_index = 7

    for item in data['dieselLists']:
        dg_name = str(item.dg_name)
        if item.diesel_litre == '':
            diesel_litre = 0
        else:
            diesel_litre = float(item.diesel_litre)
        total_consumption += diesel_litre

        sheet.cell(row=row_index, column=2).value = dg_name
        sheet.cell(row=row_index, column=3).value = diesel_litre
        row_index += 1

    # Add total consumption row
    total_row_index = row_index
    sheet.cell(row=11, column=2).value = 'Total Consumption'
    sheet.cell(row=11, column=3).value = total_consumption

    # Auto-adjust column widths
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

    row_index = 7
    for item in data['productionLists']:
       
        production_name = str(item.production_name)
        if item.production_value=='':
            production_value = 0
        else:
            production_value = float(item.production_value)
        total_production += production_value

        sheet.cell(row=row_index, column=5).value = production_name
        sheet.cell(row=row_index, column=6).value = production_value
        row_index += 1

    # Add total production row
    total_row_index = row_index
    sheet.cell(row=11, column=5).value = 'Total Production'
    sheet.cell(row=11, column=6).value = total_production
    # Auto-adjust column widths
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20

    row_index = 7
    for item in data['loadLists']:
        
        load_name = str(item.load_name)
        if item.load_value =='':
            load_value = 0
        else:
            load_value = float(item.load_value)
        total_power += load_value
        sheet.cell(row=row_index, column=8).value = load_name
        sheet.cell(row=row_index, column=9).value = load_value
        row_index += 1
        
        sheet.cell(row=23, column=8).value = 'Total Power'
        sheet.cell(row=23, column=9).value = total_power

        # Auto-adjust column widths
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20

    row_index = 15
    for item in data['heatLists']:
        
        heat_name = str(item.heat_name)
        if item.heat_value == '':
            heat_value = 0
        else:
            heat_value = float(item.heat_value)
        total_heats += heat_value
        sheet.cell(row=row_index, column=5).value = heat_name
        sheet.cell(row=row_index, column=6).value = heat_value
        row_index += 1
        
        sheet.cell(row=23, column=5).value = 'Total Heats'
        sheet.cell(row=23, column=6).value = total_heats

        # Auto-adjust column widths
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20

    row_index = 15
    for item in data['dieselLists']:
        
        dg_name = str(item.dg_name)
        if item.dg_runtime == '':
            dg_runtime = 0
        else:
            dg_runtime = float(item.dg_runtime)
        total_running_time += dg_runtime
        sheet.cell(row=row_index, column=2).value = dg_name
        sheet.cell(row=row_index, column=3).value = dg_runtime
        row_index += 1

        sheet.cell(row=23, column=2).value = 'Total Running Time '
        sheet.cell(row=23, column=3).value = total_running_time

        # Auto-adjust column widths
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20

    file_name = 'manual_wise_report.xlsx'
    file_path = os.path.join(filename, file_name)
    workbook.save(file_path)

def auto_download_manual(filename,cnx):
    try:
        data = {}
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            ems_v1.master_shifts
        '''
        shifttimings = cnx.execute(sql).mappings().all()
        from_date = date.today()
        for row in shifttimings:
            from_date = row.previous_mill_date
        pdf = 'excel'
        data['entry_date'] = from_date
        
        where = f"WHERE FORMAT(entry_date, 'dd-MM-yyyy') = '{from_date}'"

        diesel_sql = f"SELECT * FROM ems_v1.master_diesel_entry {where}"
        production_sql = f"SELECT * FROM ems_v1.master_production_entry {where}"
        load_sql = f"SELECT * FROM ems_v1.master_load_entry {where}"
        heat_sql = f"SELECT * FROM ems_v1.master_heat_entry {where}"
    
        data['dieselLists'] = cnx.execute(diesel_sql).fetchall()
        data['productionLists'] = cnx.execute(production_sql).fetchall()
        data['loadLists'] = cnx.execute(load_sql).fetchall()
        data['heatLists'] = cnx.execute(heat_sql).fetchall()
        manual_wise_report(data,from_date,filename)

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)

@app.post("/auto_download_main_manual_functionality/")
def auto_download_main_manual_functionality(request:Request,cnx: Session = Depends(get_db)):
    try:
        result1 =''
        result2 = ''
        result3 = ''
        sql = f'''
        SELECT
            FORMAT(mill_date,'dd-MM-yyyy') as mill_date,
            FORMAt(DATEADD(minute,5, shift1_start_time),'HH:mm') as shift_1,
            FORMAt(DATEADD(DAY,-1, mill_date),'dd-MM-yyyy') as previous_mill_date
        FROM
            ems_v1.master_shifts'''
        shifttimings = cnx.execute(sql).fetchall()
        previous_mill_date = date.today()
        
        for row in shifttimings:
            previous_mill_date = row.previous_mill_date
            shift_1 = row.shift_1
        drive_name = os.path.dirname(os.path.abspath(sys.argv[0]))[0]
        createFolder("auto_download/","drive_name...... "+str(drive_name))
        if not os.path.isdir(drive_name + ":\\AIC"):
            os.mkdir(drive_name + ":\\AIC")
        if not os.path.isdir(drive_name + ":\\AIC\\Power_Report_v2"):
            os.mkdir(drive_name + ":\\AIC\\Power_Report_v2")
        if not os.path.isdir(drive_name + ":\\AIC\\Power_Report_v2\\" + previous_mill_date):
            os.mkdir(drive_name + ":\\AIC\\Power_Report_v2\\" + previous_mill_date)

        filename = drive_name + ":\\AIC\\Power_Report_v2\\" + previous_mill_date
        print("shift_1",shift_1)
        print(type(shift_1))
        print("filename...",filename)
        file_path = os.path.join(filename)
        print(shift_1 == datetime.datetime.now().strftime("%H:%M"))
        if shift_1 == datetime.datetime.now().strftime("%H:%M"):
            auto_download_main(filename,cnx)
            auto_download_functionality(filename,cnx)
            auto_download_manual(filename,cnx)
            app.mount(f"/{previous_mill_date}", StaticFiles(directory=filename), name=f"{previous_mill_date}")
            
            result1 = f"http://{request.headers['host']}/{previous_mill_date}/total_power_report.xlsx"
                
            result2 = f"http://{request.headers['host']}/{previous_mill_date}/functionality_wise_report.xlsx"
                
            result3 = f"http://{request.headers['host']}/{previous_mill_date}/manual_wise_report.xlsx"

        if os.path.exists(file_path):   
            return JSONResponse({"iserror": False, "message": "data return successfully","data1":result1,"data2":result2,"data3":result3 })
        else:
            
            return JSONResponse({"iserror": False, "message": "data return successfully","data": ""})

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)

def generate_excel_report(result, month_year, report_type, report_for):
    print("result",result)
    if report_type == "date":
        file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "MonthWiseReport_templete.xlsx"))
        # file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "moth_wise_report.xlsx"))
          
        # file_path = f'{static_dir}/performanceReport_templete.xlsx'
    elif report_type == "shift":
        file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "MonthWiseReport_shift_templete.xlsx"))
        # file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "shift_wise_template.xlsx"))
          
        # file_path = f'{static_dir}/performanceReport_shift_templete.xlsx'
    workbook = Workbook()
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS' 
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')  
    cell = "C2"
    if report_for == '6to6':
        time = '06:00 to 06:00'
    else:
        time = '12:00 to 12:00'
    data = f"MONTH WISE ENERGY CONSUMPTION REPORT FOR KWH - {month_year} ({time})"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=25)
    sheet[cell].font = font
    
    # Set the border style
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Set the minimum column width
    minimum_column_width = 15
    cell = "B4"
    sheet[cell] = "Machine Code"
    cell = "C4"
    sheet[cell] = "Machine Name"
    sheet.cell(row=4, column=2, value="Machine Code").fill = fill_cyan
    sheet.cell(row=4, column=3, value="Machine Name").fill = fill_cyan
    if result == []:
        print(1)
        cell = "O10"
        data = "No Data"
        sheet[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        sheet[cell].alignment = alignment

        sheet.column_dimensions[cell[0]].width = len(data) + 2  # Adjust column width

        font = Font(name='Calibri', size=25)
        sheet[cell].font = font
    # Iterate over the data and set the values and formatting
    for i, data in enumerate(result, start=6):  # Assuming machine_code starts from row 7
        machine_code = data["machine_code"]
        machine_name = data["machine_name"]
        print(machine_name)       
        sheet.cell(row=i, column=2, value=machine_code).alignment = Alignment(horizontal="center")
        # Create a new Alignment object with text wrapping enabled for machine_name
        alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.cell(row=i, column=3, value=machine_name).alignment = alignment
        # Adjust column size based on the maximum text length in columns C and D
        machine_name_length = len(machine_name) + 40
        # print(machine_name_length)
        column_letter_c = get_column_letter(3)  # Column C
        column_width_c = max(machine_name_length, sheet.column_dimensions[column_letter_c].width)
        # print(column_width_c)
        sheet.column_dimensions[column_letter_c].width = column_width_c
        if report_type == "date":

            month, year = map(int, month_year.split('-'))
            days_in_month = calendar.monthrange(year, month)[1]

            start_row = 4
            start_col = 4
            
            fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')

            for day in range(1, days_in_month + 1):
                cell = sheet.cell(row=start_row, column=start_col)
                cell.value = f"{day:02d}-{month_year}"
                cell.fill = fill_cyan
                
                start_col += 1
            for j in range(1, 32):
                column_letter = get_column_letter(j + 3)  
                cell = sheet.cell(row=i, column=j + 3)
                cell.value = data.get(f"d{j}", "")
                cell.alignment = Alignment(horizontal="center")
                if cell.value == 0:
                    cell.value = ""
            # Apply border to cells
            row_range = sheet[f"A{i}:AH{i}"]
            for row in row_range:
                for cell in row:
                    cell.border = border
        
        elif report_type == "shift":

            month, year = map(int, month_year.split('-'))
            days_in_month = calendar.monthrange(year, month)[1]

            start_row = 4
            start_col = 4

            fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')

            for day in range(1, days_in_month + 1):
                cell = sheet.cell(row=start_row, column=start_col)
                cell.value = f"{day:02d}-{month_year}"
                cell.fill = fill_cyan

                shifts = ['s1', 's2', 's3']

                for shift in range(1, 4):
                    # Write the shift label (s1, s2, s3) in cells D5, E5, F5 for each day
                    sheet.cell(row=5, column=start_col + shift - 1, value=shifts[shift - 1])
                    sheet.cell(row=5, column=start_col + shift - 1).fill = fill_cyan
                    sheet.cell(row=5, column=start_col + shift - 1).alignment = Alignment(horizontal="center")

                    cell = sheet.cell(row=start_row + shift, column=start_col + shift)
                    cell.value = data.get(f"d{day}_s{shift}", "")
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == 0:
                        cell.value = ""

                start_col += 3
            # Adjust column size based on the maximum text length vertically
            for shift in range(1, 4):
                for j in range(1, 32):
                    column_letter = get_column_letter((shift - 1) * 31 + j + 3)  # Assuming data columns start from column E (column index 5)
                    cell = sheet.cell(row=i, column=(shift - 1) * 31 + j + 3)
                    cell.value = data.get(f"ds{shift}_{j}", "")
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == 0:
                        cell.value = ""

                    # Adjust column size based on the maximum text length vertically
                    cell_text_length = len(str(cell.value))
                    column_width = max(cell_text_length, sheet.column_dimensions[column_letter].width)
                    sheet.column_dimensions[column_letter].width = column_width
                
                row_range = sheet[f"A{start_row}:CR{start_row + i}"]
                for row in row_range:
                    for cell in row:
                        cell.border = border
        
    file_name = f'MonthWiseReport-{month_year}.xlsx'
    file_path = os.path.join(base_path, file_name)
    workbook.save(file_path)

@app.post("/performance_report/")
async def performance_report(request:Request,
                             machine_id: str = Form(''),
                             month_year: str = Form(None),
                             report_for: str = Form(None),
                             report_type: str = Form(None),                                                      
                             cnx: Session = Depends(get_db)):
    
    if month_year is None:
        return JSONResponse({"iserror": True, "message": "month year is required"})

    if report_type is None:
        return JSONResponse({"iserror": True, "message": "report type is required"})
    
    if report_for is None:
        return JSONResponse({"iserror": True, "message": "report_for is required"})
    
    try:
        groupby = ""
        where = ""
        if machine_id == "":
            pass
        else:
            machine_id = machine_id.split(',')
            where += f" and mm.machine_id IN ({','.join(machine_id)})"
        if report_for == '6to6':
            if month_year is not None:
                month, year = month_year.split('-')
                tbl_name = f"ems_v1_completed.power_{month}{year} cp"   

            if month_year !='':
                query = f"""SELECT table_name FROM information_schema.tables WHERE table_schema = 'ems_v1_completed' AND table_name = 'power_{month_year}'"""
                print(query)
                result_query = cnx.execute(query).mappings().all()
                if len(result_query)>0:
                    pass
                else:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
        else:
            if month_year is not None:
                month, year = month_year.split('-')
                tbl_name = f"ems_v1_completed.power_{month}{year}_12 cp"
            if report_type not in ['date']:
                return JSONResponse({"iserror": True, "message": "Invalid report type"})

        if report_type not in ['date', 'shift']:
                return JSONResponse({"iserror": True, "message": "Invalid report type"})

        if report_type == 'date':
            day = "CONCAT('d', DAY(cp.mill_date))"
            
        elif report_type == 'shift':
            day = "CONCAT(CONCAT('ds', cp.mill_shift), '_', DAY(cp.mill_date))"
            groupby = ",cp.mill_shift"

        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                {day} AS day,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh
            FROM
                {tbl_name}
                INNER JOIN ems_v1.master_machine mm ON mm.machine_id = cp.machine_id
                left join ems_v1.master_machine_factor mmf on mmf.machine_id = mm.machine_id
            WHERE
                1=1 {where} and FORMAT(cp.mill_date, 'MM-yyyy') = '{month_year}' 
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                DAY(cp.mill_date)
                {groupby}               
        ''')
        rslt = cnx.execute(query).mappings().all()
        print(query)
        if rslt is not None:
            output = {}

            if report_type == 'date':
                output_keys = [f'd{day}' for day in range(1, 32)]
            elif report_type == 'shift':
                output_keys = [f'ds{shift}_{day}' for day in range(1, 32) for shift in range(1, 4)]
            
            for row in rslt:
                machine_code = row.machine_code
                machine_name = row.machine_name
                day = row.day
                kwh = row.kwh
            
                if machine_code not in output:
                    output[machine_code] = {
                        'machine_code': machine_code,
                        'machine_name': machine_name
                    }
                    for key in output_keys:
                        output[machine_code][key] = 0
            
                output[machine_code][day] = kwh
                result=list(output.values())
            if rslt == []:
                result = []
               
            # machine = {"result" : data["machine_name"],
            #            "result1" : data[]}
            generate_excel_report(result, month_year,report_type, report_for)
            # process_data(month_year, result)
        file_path = os.path.join(base_path, f"MonthWiseReport-{month_year}.xlsx")
        results = f"http://{request.headers['host']}/attachment/MonthWiseReport-{month_year}.xlsx"

        if os.path.exists(file_path):

            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": results})
        else:
            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": None})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

def generate_year_wise_excel_report(result, year):
    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "YearWiseReport_templete.xlsx"))
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS'
    
    cell = "B1"
    data = f"YEAR WISE ENERGY CONSUMPTION REPORT FOR KWH - {year}"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=13)
    sheet[cell].font = font

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    minimum_column_width = 15
    cell = "A3"
    sheet[cell] = "Machine Code"
    cell = "B3"
    sheet[cell] = "Machine Name"
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
    sheet.cell(row=3, column=1, value="Machine Code").fill = fill_cyan
    sheet.cell(row=3, column=2, value="Machine Name").fill = fill_cyan

    # Populate month-year headers in row 3 starting from column C
    col_index = 3
    for month_year in result[0].keys():
        if month_year not in ["machine_code", "machine_name"]:
            sheet.cell(row=3, column=col_index, value=month_year).fill = fill_cyan
            col_index += 1

    for row, data in enumerate(result, start=4):
        machine_code = data["machine_code"]
        machine_name = data["machine_name"]
        sheet.cell(row=row, column=1, value=machine_code).alignment = Alignment(horizontal="center")
        alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.cell(row=row, column=2, value=machine_name).alignment = alignment
        # machine_name_length = len(machine_name)
        # column_letter_c = get_column_letter(3)
        # column_width_c = max(machine_name_length, sheet.column_dimensions[column_letter_c].width)
        # sheet.column_dimensions[column_letter_c].width = column_width_c

        col_indexs = 1
        for kwh_value in data.values():
            if isinstance(kwh_value, (int, float)):
                if kwh_value == 0:
                    kwh_value = ''
                sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
            col_indexs += 1

    file_name = f'YearWiseReport-{year}.xlsx'
    file_path = os.path.join(base_path, file_name)
    workbook.save(file_path)

@app.post("/year_wise_report/")
async def year_wise_report(request: Request,
                           machine_id: str = Form(''),
                           year: str = Form(""),
                           cnx: Session = Depends(get_db)):

    if year  == '':
        return JSONResponse({"iserror": True, "message": "year is required"})

    try:
        groupby = ""
        where = ""
        if machine_id == "":
            pass
        else:
            machine_id = machine_id.split(',')
            where += f" and mm.machine_id IN ({','.join(machine_id)})"
        mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        tables_to_union = []
        for month in range(4, 13):
            month_year = f"{mill_month[month]}{year}"
            print(month_year)
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = 'ems_v1_completed' AND TABLE_NAME = 'power_{month_year}' """
            result_query = cnx.execute(query).mappings().all()

            if len(result_query) > 0:
                tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}")
        
        next_year = int(year) + 1
        mill_month = {1: "01", 2: "02", 3: "03"}

        for month in range(1, 4):
            month_year = f"{mill_month[month]}{next_year}"
            print(month_year)
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = 'ems_v1_completed' AND TABLE_NAME = 'power_{month_year}' """
            result_query = cnx.execute(query).mappings().all() 
            print("result_query",result_query)
            if len(result_query) > 0:
                tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}")
            tables_union_query = " UNION ALL ".join(tables_to_union)
            print("tables_union_query",tables_union_query)

        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh,
                FORMAT(min(cp.mill_date), 'MM-yyyy') AS mill_date
            FROM
                ({tables_union_query}) cp
                INNER JOIN ems_v1.master_machine mm ON mm.machine_id = cp.machine_id
                LEFT JOIN ems_v1.master_machine_factor mmf ON mmf.machine_id = mm.machine_id
            WHERE
                1=1 {where}
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                MONTH(cp.mill_date), 
                YEAR(cp.mill_date) 
            ORDER BY 
                min(cp.mill_date),  
                min(cp.machine_id)
        ''')

        print(query)
        rslt = cnx.execute(query).mappings().all()
        if len(rslt)>0:
          output = {}  # Initialize the output dictionary
        
        output_keys = [
            f"04-{year}", f"05-{year}", f"06-{year}",
            f"07-{year}", f"08-{year}", f"09-{year}",
            f"10-{year}", f"11-{year}", f"12-{year}",
            f"01-{next_year}", f"02-{next_year}", f"03-{next_year}"
        ]
        
        for row in rslt:
            machine_code = row['machine_code']
            machine_name = row['machine_name']
            mill_date = row['mill_date']
            kwh = row['kwh']
            
            if machine_code not in output:
                output[machine_code] = {
                    'machine_code': machine_code,
                    'machine_name': machine_name
                }
                for key in output_keys:
                    output[machine_code][key] = 0
            
            output[machine_code][mill_date] = kwh
        
        result = list(output.values())
               
        generate_year_wise_excel_report(result, year)
            # process_data(month_year, result)
        file_path = os.path.join(base_path, f"YearWiseReport-{year}.xlsx")
        results = f"http://{request.headers['host']}/attachment/YearWiseReport-{year}.xlsx"

        if os.path.exists(file_path):

            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": results})
        else:
            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": None})
        # return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": jsonable_encoder(result)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})
  
@app.post("/order_wise/")
async def order_wise(table_name :str = Form(''),
                     obj :str = Form(''),
                     cnx: Session = Depends(get_db)):

    try:
       
        if obj !='':
            obj_data = json.loads(obj)
            for row in obj_data:
                id = row["id"]
                sno = row["sno"]
                if table_name == 'zone':
                    sql = text(f" update ems_v1.master_department set department_order = {sno} where department_id = {id} ")
                    
                if table_name == 'area':
                    sql = text(f" update ems_v1.master_shed set shed_order = {sno} where shed_id = {id} ")
                
                if table_name == 'location':
                    sql = text(f" update ems_v1.master_machinetype set machinetype_order = {sno} where machinetype_id = {id} ")
                    
                if table_name == 'function_1':
                    sql = text(f" update ems_v1.master_function set function_order = {sno} where function_id = {id}")
                    
                if table_name == 'function_2':
                    sql = text(f" update ems_v1.master_function set function_order = {sno} where function_id = {id}")
                    
                if table_name == 'meter':
                    sql = text(f" update ems_v1.master_machine set machine_order = {sno} where machine_id = {id} ")

                cnx.execute(sql)
                cnx.commit()
        return JSONResponse({"iserror":False,"message":"order update successfully","data":''})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_master_machine_factor/")
async def get_master_machine_factor(id :str = Form(''),
                                    machine_id :str = Form(''),
                                    cnx: Session = Depends(get_db)):

    try:
        where = ''
        if machine_id != '':
            where += f'and  mf.machine_id = {machine_id}'
        if id != '':
            where += f'and  mf.id = {id}'

        sql = text(f'''
                select 
                   mf.*,
                   mm.machine_code,
                   mm.machine_name
                from 
                   ems_v1.master_machine_factor mf,
                   ems_v1.master_machine mm
                where mf.machine_id = mm.machine_id {where}''')
        print(sql)
        data = cnx.execute(sql).mappings().all()
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/update_master_machine_factor/")
async def update_master_machine_factor(machine_id :str = Form(''),
                                       obj :str = Form(''),
                                       cnx: Session = Depends(get_db)):

    try:
        if obj != '':
            obj_data = json.loads(obj)
            sel = {}
            for data in obj_data:
                for key, value in data.items():
                    sel[key] = value
                sql = text(f'''UPDATE ems_v1.master_machine_factor SET {', '.join([f"{key} = '{value}'" for key, value in sel.items()])} WHERE machine_id = '{machine_id}' ''')
                cnx.execute(sql)
                cnx.commit()
                createFolder("Log/","data.. "+str(sql))
        return JSONResponse({"iserror":False,"message":"data update successfully","data":''})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

