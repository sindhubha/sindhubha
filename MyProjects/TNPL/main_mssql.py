from fastapi import FastAPI, Depends, Form, File, UploadFile,Body
from typing import Optional, List, Dict, Any
from datetime import datetime, date, timedelta
import sqlalchemy
import uvicorn
from sqlalchemy.orm import Session
from sqlalchemy.sql import text
from sqlalchemy.exc import SQLAlchemyError
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse
from fastapi.requests import Request 
import traceback
from fastapi.staticfiles import StaticFiles
from datetime import datetime
import os
import json
from dateutil import parser
from fastapi.middleware.cors import CORSMiddleware
import random
from log_folder import createFolder
import shutil
from mssql_connection import get_db
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pathlib import Path
import time
import calendar
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import re
import sys
from starlette.middleware.base import BaseHTTPMiddleware, RequestResponseEndpoint
from openpyxl.drawing.image import Image
app = FastAPI()

class TimingMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        start_time = time.time()

        response = await call_next(request)

        current_time = time.time()
        duration = (current_time - start_time) * 1000
        response.headers["Runtime"] = f"{duration} ms" 
        return response

# Add the TimingMiddleware to the FastAPI app
app.add_middleware(TimingMiddleware)
app.add_middleware(CORSMiddleware,
                   allow_origins= ['*'],
                   allow_credentials = True,
                   allow_methods = ["*"],
                   allow_headers = ["*"]
)    

def parse_date(from_date):
    date_from = from_date.split("-")
    if len(date_from[0]) <= 2:
        if int(date_from[0]) > 12:
            from_date = parser.parse(from_date).strftime("%Y-%m-%d %H:%M:%S")
        else:
            from_date = parser.parse(from_date).strftime("%Y-%d-%m %H:%M:%S")
        from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S")
    else:
        from_date = parser.parse(from_date).strftime("%Y-%m-%d %H:%M:%S")
        from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S")
    return from_date 

def save_image(image, dir_name):
    random_number = random.randint(1, 100000)
    now = datetime.datetime.now()
    extension = os.path.splitext(image.filename)[1].lower()
    image_file_name = f"{random_number}_{now.strftime('%Y_%m_%d_%H_%M_%S')}{extension}"
    
    if not os.path.exists(f"{dir_name}"):
        os.makedirs(dir_name)

    with open(os.path.join(f"{dir_name}", image_file_name), "wb") as buffer:
        shutil.copyfileobj(image.file, buffer)

    return image_file_name

@app.post("/login/")
async def employee_login(username:str=Form(),
                         password:str=Form(),
                         cnx: Session = Depends(get_db)
                        ):
    try:
        query=text(f'''
                SELECT 
                   me.*,
                   mc.company_name as company_name,
                   mc.company_code as company_code,
                   mb.branch_name as branch_name,
                   mb.branch_code as branch_code,
                   md.department_name as department_name,
                   md.department_code as department_code,
                   ms.shed_name as shed_name,
                   ms.shed_code as shed_code,
                   mmt.machinetype_name as machinetype_name,
                   mmt.machinetype_code as machinetype_code
                FROM 
                   [ems_v1].[dbo].[master_employee] me
                   LEFT JOIN [ems_v1].[dbo].[master_company] mc on me.company_id=mc.company_id
                   LEFT JOIN [ems_v1].[dbo].[master_branch] mb on me.branch_id=mb.branch_id
                   LEFT JOIN [ems_v1].[dbo].[master_department] md on me.department_id=md.department_id
                   LEFT JOIN [ems_v1].[dbo].[master_shed] ms on me.shed_id=ms.shed_id
                   LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt on me.machinetype_id=mmt.machinetype_id
                WHERE 
                    me.employee_code='{username}' AND 
                    me.password_login=HASHBYTES('MD5', '{password}') AND me.is_login='yes' AND me.status = 'active' ''')
        data = cnx.execute(query).mappings().all()        
        if len(data) > 0:
            # createFolder("Log/","Response Sent Successfully ")
            return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
        else:
            # createFolder("Log/","username and password incorrect ")
            return JSONResponse({"iserror": True, "message": "invalid username or password"})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})      

base_path = Path(__file__).parent / "attachment"
if not os.path.exists(base_path):
    os.makedirs(base_path)  
# Mount static directory for serving image files
app.mount("/attachment", StaticFiles(directory=base_path), name="attachment")
print(base_path)

@app.post("/get_company_list/{company_id}")
@app.post("/get_company_list/")
async def get_company_list(request: Request,
                           company_id: Optional[str] = None,
                           cnx: Session = Depends(get_db)):
    try:
        # base_path = os.path.abspath(os.path.dirname(__file__))
        where = ""
        if company_id is not None:
            where = f" and mc.company_id = {company_id}"
        query = text(f'''
                    SELECT
                    	mc.*, 
                    	CONCAT('/attachment/company_logo/',mc.logo) AS logo,
                    	ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                    	ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM
                    	[ems_v1].[dbo].[master_company] as mc 
                    	left join  [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mc.created_by
                    	left join  [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mc.modified_by
                    WHERE mc.status !='delete' {where}
                    ''')
        print(query)
        # query = text(f'''SELECT * FROM master_company WHERE status='active' {where}''')
        data = cnx.execute(query).mappings().all()
        results = []
        for i in data:
            new_logo = dict(i)
            if new_logo["logo"] is not None:
                new_logo["logo"] = "http://" + request.headers["host"] + new_logo['logo']
            results.append(new_logo)
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(results)})
      
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})      

@app.post("/save_company_details/")
async def save_company_details(company_id: int = Form(None),                                 
                               company_code: str = Form(None),
                               company_name: str = Form(None),
                               address: str = Form(None),
                               city: str = Form(None),
                               state:str = Form(None),
                               country:str= Form(None),
                               pincode:str = Form(None),
                               phone_number :str=Form(None),
                               email_id: str=Form(None),
                               website:str=Form(None),
                               logo: UploadFile = File(''),
                               old_logo :str = Form(''),
                               user_login_id: str=Form(None),
                               cnx: Session = Depends(get_db)):
   
    if company_code == None:  
        return JSONResponse({"iserror": True, "message": " company_code is required"})

    if company_name == None:  
        return JSONResponse({"iserror": True, "message": " company_name is required"})
    
    try:
        # if logo == "":
        #     filename = ''
        # else:
        if old_logo == "":
            filename = save_image(logo, f"{base_path}/company_logo/")
        else:
            filename = old_logo
       
        # base_path = os.path.abspath(os.path.dirname(__file__)) + "/attachment/company_logo/"
        # if not os.path.exists(base_path):
        #     os.makedirs(base_path)
        # random_number = random.randint(1, 100000) 
        # extension = os.path.splitext(logo.filename)[1].lower()
        # # generate filename
        # filename = f"{random_number}_{datetime.datetime.now().strftime('%y_%m_%d_%H_%M_%S')}{extension}"
        # # save image to disk
        # with open(base_path + filename, "wb") as f:
        #     f.write(await logo.read())
        print(filename)         
        if company_id is not None:
            query = text(f"""
                UPDATE  [ems_v1].[dbo].[master_company] 
                SET company_name = '{company_name}', company_code = '{company_code}', address = '{address}',
                city = '{city}', state = '{state}', country = '{country}', pincode = '{pincode}',
                phone_number = '{phone_number}', email_id = '{email_id}', website = '{website}',
                modified_by = '{user_login_id}', modified_on = GETDATE(), logo = '{filename}'
                WHERE company_id = {company_id}
            """)   
            cnx.execute(query)
            cnx.commit() 
        else:
            select_query = text(f'''select * from  [ems_v1].[dbo].[master_company] where company_code = '{company_code}' and status!='delete' ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data) > 0:
                return JSONResponse({"iserror":True, "message":"company_code already exists", "data":""})

            query = text(f"""
                INSERT INTO  [ems_v1].[dbo].[master_company] (company_name, company_code, address, city, state,
                country, pincode, phone_number, email_id, website, created_on, created_by, logo)
                
                VALUES ('{company_name}', '{company_code}', '{address}', '{city}', '{state}',
                '{country}', '{pincode}', '{phone_number}', '{email_id}', '{website}', GETDATE(), 
                '{user_login_id}', '{filename}')
            """)
            cnx.execute(query)        
            company_id = cnx.execute(text("SELECT SCOPE_IDENTITY()")).first()[0]
            print(query)
            cnx.commit()

            sql = text(f'''
            INSERT INTO  [ems_v1].[dbo].[power_report_fields_original] (report_id, field_code, field_name, is_show, slno, field_name_display, company_id,unit)
                SELECT pr.report_id, pfo.field_code, pfo.field_name, pfo.is_show, pfo.slno, pfo.field_name_display, {company_id}, pfo.unit
                FROM (SELECT DISTINCT report_id FROM [ems_v1].[dbo].[power_report] where report_type = 'report') pr
                CROSS JOIN [ems_v1].[dbo].[power_report_fields_original] pfo
                where pfo.company_id = 0   
            order by
                pr.report_id
        ''')
            cnx.execute(sql)
            cnx.commit()
            sql = text(f'''
            INSERT INTO  [ems_v1].[dbo].[power_report_fields_original] (report_id, field_code, field_name, is_show, slno, field_name_display, company_id,unit)
                SELECT pr.report_id, pfo.field_code, pfo.field_name, pfo.is_show, pfo.slno, pfo.field_name_display, {company_id}, pfo.unit
                FROM (SELECT DISTINCT report_id FROM [ems_v1].[dbo].[power_report] where report_type = 'dashboard') pr
                CROSS JOIN [ems_v1].[dbo].[power_report_fields_original] pfo
                where pfo.company_id = 0 and  pfo.report_type = 'dashboard'
            order by
                pr.report_id
        ''')
            cnx.execute(sql)
            cnx.commit()

        return JSONResponse({"iserror": False, "message": "data save successfully", "data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_company/")
async def remove_company(company_id: int = Form(None),
                         status : str = Form(None),
                         cnx: Session = Depends(get_db)):
    try:    
        if company_id == None:  
            return JSONResponse({"iserror": True, "message": " company_id is required"})
                        
        if company_id is not None:
            if status is not None:
                query = text(f" UPDATE [ems_v1].[dbo].[master_company] SET status = '{status}' WHERE company_id = '{company_id}' ")
                cnx.execute(query)

            else:
                query = text(f" UPDATE [ems_v1].[dbo].[master_company] SET status = 'delete' WHERE company_id = '{company_id}' ")                
                cnx.execute(query)
            cnx.commit()
            createFolder("Log/","query execute sucessfully")

        return JSONResponse({"iserror": False, "message": "status update successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f'''{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}'''
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_branch_list/{branch_id}")
@app.post("/get_branch_list/")
async def branch_list(branch_id: Optional[int] = None, 
                      company_id: int = Form(None),
                      cnx: Session = Depends(get_db)):
    try:        
        where = ""
        if branch_id is not None:
            where = f" and mb.branch_id = {branch_id}"
        
        if company_id is not None:
            where = f" and mb.company_id = {company_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.*,
                       ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM [ems_v1].[dbo].[master_branch] mb
                       left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mb.created_by
                       left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mb.modified_by
                       INNER JOIN [ems_v1].[dbo].[master_company] mc ON mb.company_id = mc.company_id
                    WHERE mb.status != 'delete' {where}
                    ''')            
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_branch_details/")
async def save_branch_details(company_id: int = Form(None),
                              branch_id:int = Form(None),
                              branch_code: str = Form(None),
                              branch_name: str = Form(None),                              
                              address: str = Form(None),
                              city: str = Form(None),
                              state:str = Form(None),
                              country:str= Form(None),
                              pincode:str = Form(None),
                              phone_number :str=Form(None),
                              email_id: str=Form(None),
                              website:str=Form(None),
                              user_login_id : str = Form(None),
                              is_alarm_sound : str = Form(''),
                              cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": True, "message": " company_id is required"})
     
    if branch_code == None:  
        return JSONResponse({"iserror": True, "message": " branch_code is required"})

    if branch_name == None:  
        return JSONResponse({"iserror": True, "message": " branch_name is required"})

    try:    
                    
        if branch_id is not None:
            query = text(f"""
                UPDATE [ems_v1].[dbo].[master_branch] 
                SET company_id = {company_id}, branch_name = '{branch_name}', branch_code = '{branch_code}',
                address = '{address}', city = '{city}', state = '{state}', country = '{country}', pincode = '{pincode}',
                phone_number = '{phone_number}', email_id = '{email_id}',is_alarm_sound = '{is_alarm_sound}',
                website = '{website}', modified_on = GETDATE(), modified_by = '{user_login_id}'
                WHERE branch_id = {branch_id}
            """)         
    
        else:            
            select_query = text(f'''select * from [ems_v1].[dbo].[master_branch] where branch_code = '{branch_code}'  and status != 'delete'  ''')
            data =cnx.execute(select_query).mappings().all()
            
            if len(data)>0:
                return JSONResponse({"iserror":True,"message":"branch_code already exists ","data":""})
            
            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_branch] (company_id, branch_name, branch_code, created_on, 
                created_by, address, city, state, country, pincode, phone_number, email_id, website , is_alarm_sound)
                
                VALUES ({company_id},'{branch_name}', '{branch_code}', GETDATE(), '{user_login_id}', '{address}',
                '{city}','{state}','{country}','{pincode}','{phone_number}','{email_id}','{website}','{is_alarm_sound}')
            """)        
        cnx.execute(query)
        cnx.commit()

        sql = text(f''' select * from [ems_v1].[dbo].[master_branch] where company_id = '{company_id}' ''')
        data1 = cnx.execute(sql).mappings().all()
        if len(data1)>0:
            sql1= text(f''' update [ems_v1].[dbo].[master_company] set is_assign = 'yes' where company_id = '{company_id}' ''')
            cnx.execute(sql1)
            cnx.commit()
                
        return JSONResponse({"iserror": False, "message": "data save successfully", "data":" "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_branch/")
async def remove_branch(branch_id: int = Form(None),
                        status: str = Form(None),
                        cnx: Session = Depends(get_db)):
    try:
        if branch_id == None:  
            return JSONResponse({"iserror": True, "message": " branch_id is required"})
        
        if branch_id is not None:
            if status is not None:
                query = text(f" UPDATE [ems_v1].[dbo].[master_branch] SET status = '{status}' WHERE branch_id = '{branch_id}' ")
            else:
                query = text(f" UPDATE [ems_v1].[dbo].[master_branch] SET status = 'delete' WHERE branch_id = '{branch_id}' ")                
            cnx.execute(query)
            cnx.commit()
        
        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_branch] WHERE company_id = (SELECT company_id FROM [ems_v1].[dbo].[master_branch] WHERE branch_id = '{branch_id}') AND status != 'delete' ''')
        result = cnx.execute(query).mappings().all()           
        # If no active branches are left, update the is_assign status of the company to "no"
        if result == []:
            query = text(f'''UPDATE [ems_v1].[dbo].[master_company] SET is_assign = 'no' WHERE company_id = (SELECT company_id FROM [ems_v1].[dbo].[master_branch] WHERE branch_id = '{branch_id}')''')
            cnx.execute(query)
            cnx.commit()  
                
        return JSONResponse({"iserror":False,"message":" status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/alarm_status/")
async def alarm_status(branch_id: int = Form(''),
                       alarm_status: str = Form(''),
                       cnx: Session = Depends(get_db)):
    try:
        if branch_id == '':  
            return JSONResponse({"iserror": True, "message": " branch_id is required"})
        
        if alarm_status == '':  
            return JSONResponse({"iserror": True, "message": " alarm_status is required"})
            
        query = text(f" UPDATE [ems_v1].[dbo].[master_branch] SET is_alarm_status = '{alarm_status}' WHERE branch_id = '{branch_id}' ")
        
        cnx.execute(query)
        cnx.commit()
        
        return JSONResponse({"iserror":False,"message":" status updated successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_department_list/{department_id}")
@app.post("/get_department_list/")
async def department_list(department_id: Optional[int] = None , 
                          cnx: Session = Depends(get_db)):
    try:        
        where = ""
        if department_id is not None:
            where = f" and md.department_id = {department_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_name AS branch_name,
                       mb.branch_code AS branch_code,
                       md.*,
                       ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM [ems_v1].[dbo].[master_department] md
                       left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=md.created_by
                       left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=md.modified_by
                       INNER JOIN [ems_v1].[dbo].[master_company] mc ON md.company_id = mc.company_id
                       INNER JOIN [ems_v1].[dbo].[master_branch] mb ON  md.branch_id = mb.branch_id
                    WHERE md.status != 'delete' {where}''')
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_department_details/")
async def save_department_details(company_id: int = Form(None),
                                  branch_id: int = Form(None),
                                  department_id: int = Form(None),                              
                                  department_code: str = Form(None),
                                  department_name: str = Form(None),
                                  user_login_id: str = Form(None),
                                  cnx: Session = Depends(get_db)):                              
    if company_id == None:  
        return JSONResponse({"iserror": True, "message": " company_id is required"})
     
    if branch_id == None:  
        return JSONResponse({"iserror": True, "message": " branch_id is required"})
    
    if department_code == None:  
        return JSONResponse({"iserror": True, "message": " department_code is required"})

    if department_name == None:  
        return JSONResponse({"iserror": True, "message": " department_name is required"})

    try:        
        if department_id is not None:
            query = text(f"""
                UPDATE [ems_v1].[dbo].[master_department] 
                SET company_id = {company_id}, department_name = '{department_name}', department_code = '{department_code}',
                modified_on = GETDATE(), modified_by = '{user_login_id}', branch_id = {branch_id}
                WHERE department_id = {department_id}
            """)
            
        else:
            select_query = text(f'''select * from [ems_v1].[dbo].[master_department] where department_code = '{department_code}' and status != 'delete'  ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data)>0:
                return JSONResponse({"iserror": True, "message": "department code already exists", "data": ""})

            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_department] (company_id, department_name, department_code,
                created_on, created_by, branch_id)
                VALUES ({company_id},'{department_name}', '{department_code}', GETDATE(), '{user_login_id}', 
                {branch_id})
            """)
        cnx.execute(query)
        cnx.commit()
        
        sql = text(f''' select * from [ems_v1].[dbo].[master_department] where branch_id = '{branch_id}' ''')
        data1 = cnx.execute(sql).mappings().all()
        if len(data1)>0:
            sql1= text(f''' update ems_v1.dbo.master_branch set is_assign = 'yes' where branch_id = '{branch_id}' ''')
            cnx.execute(sql1)
            cnx.commit()
            
        return JSONResponse({"iserror": False, "message": "data save successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_department/")
async def remove_department(department_id: int = Form(None), 
                            status : str = Form(None),
                            cnx: Session = Depends(get_db)):
    try:
        if department_id == None:  
            return JSONResponse({"iserror": True, "message": " department_id is required"})
        
        if department_id is not None:
            if status is not None:        
                query = text(f'''UPDATE [ems_v1].[dbo].[master_department] SET status = '{status}' WHERE department_id = {department_id}''')        

            else:
                query = text(f'''UPDATE [ems_v1].[dbo].[master_department] SET status = 'delete' WHERE department_id = '{department_id}' ''')
            cnx.execute(query)
            cnx.commit()
        
        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_department] WHERE branch_id = (SELECT branch_id FROM [ems_v1].[dbo].[master_department] WHERE department_id = '{department_id}') AND status != 'delete' ''')
        result = cnx.execute(query).mappings().all()           
        if result == []:
            query = text(f'''UPDATE [ems_v1].[dbo].[master_branch] SET is_assign = 'no' WHERE branch_id = (SELECT branch_id FROM [ems_v1].[dbo].[master_department] WHERE department_id = '{department_id}')''')
            cnx.execute(query)
            cnx.commit()  
       
        return JSONResponse({"iserror":False,"message":" status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_shed_list/{shed_id}")
@app.post("/get_shed_list/")
async def shed_list(shed_id: Optional[int] = None, 
                    department_id : int = Form(None),
                    cnx: Session = Depends(get_db)):
    try:        
        where = ""
        if shed_id is not None:
            where = f" and ms.shed_id = {shed_id}"
        if department_id is not None:
            where = f" and ms.department_id = {department_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_name AS branch_name,
                       mb.branch_code AS branch_code,
                       md.department_name AS department_name ,
                       md.department_code AS department_code,
                       ms.*,
                       ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM [ems_v1].[dbo].[master_shed] ms
                       left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=ms.created_by
                       left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=ms.modified_by
                       INNER JOIN [ems_v1].[dbo].[master_company] mc ON ms.company_id = mc.company_id
                       INNER JOIN [ems_v1].[dbo].[master_branch] mb ON  ms.branch_id = mb.branch_id
                       INNER JOIN [ems_v1].[dbo].[master_department] md ON ms.department_id = md.department_id
                    WHERE ms.status != 'delete' {where}''')       
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_shed_details/")
async def save_shed_details(company_id: int = Form(None), 
                            branch_id:int = Form(None),
                            department_id: int = Form(None),
                            shed_id : int = Form(None),                            
                            shed_code: str = Form(None),
                            shed_name: str = Form(None),
                            user_login_id : str = Form(None),                           
                            cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": True, "message": " company_id is required"})
     
    if branch_id == None:  
        return JSONResponse({"iserror": True, "message": " branch_id is required"})
    
    if department_id == None:  
        return JSONResponse({"iserror": True, "message": " department_id is required"}) 
    
    if shed_code == None:  
        return JSONResponse({"iserror": True, "message": " shed_code is required"})

    if shed_name == None:  
        return JSONResponse({"iserror": True, "message": " shed_name is required"})
    
    try:        
        if shed_id is not None:
            query =text(f"""
                UPDATE [ems_v1].[dbo].[master_shed]
                SET company_id = {company_id}, shed_name = '{shed_name}', shed_code = '{shed_code}',
                branch_id = {branch_id}, department_id = {department_id}, modified_on = GETDATE(),
                modified_by = '{user_login_id}'
                WHERE shed_id = {shed_id}
            """)        
  
        else:
            select_query = text(f'''select * from [ems_v1].[dbo].[master_shed] where shed_code = '{shed_code}' and status != 'delete' ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data) > 0:
              return JSONResponse({"iserror": True, "message": "shed_code already exists","data": " " })

            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_shed] (
                    company_id, shed_name, shed_code, branch_id, department_id, created_on, created_by
                )
                VALUES (
                    {company_id},'{shed_name}', '{shed_code}', {branch_id}, {department_id}, GETDATE(), '{user_login_id}'
                )
            """)    
        cnx.execute(query)
        cnx.commit()        
        sql = text(f''' select * from [ems_v1].[dbo].[master_shed] where department_id = '{department_id}' ''')
        data1 = cnx.execute(sql).mappings().all()
        if len(data1)>0:
            sql1= text(f''' update ems_v1.dbo.master_department set is_assign = 'yes' where department_id = '{department_id}' ''')
            cnx.execute(sql1)
            cnx.commit()
            
        return JSONResponse({"iserror": False, "message": "data save successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_shed/")
async def remove_shed(shed_id: int = Form(None), 
                      status : str = Form(None),
                      cnx: Session = Depends(get_db)):
    try:

        if shed_id == None:  
            return JSONResponse({"iserror": True, "message": " shed_id is required"})
        
        if shed_id is not None:
            if status is not None:        
                query = text(f'''UPDATE [ems_v1].[dbo].[master_shed]  SET status = '{status}' WHERE shed_id = {shed_id}''')        

            else:
                query = text(f'''UPDATE [ems_v1].[dbo].[master_shed]  SET status = 'delete' WHERE shed_id = {shed_id}''')                    
            cnx.execute(query)
            cnx.commit()
        
        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_shed] WHERE department_id = (SELECT department_id FROM [ems_v1].[dbo].[master_shed] WHERE shed_id = '{shed_id}') AND status != 'delete' ''')
        result = cnx.execute(query).mappings().all()
        if result == []:
            query = text(f'''UPDATE [ems_v1].[dbo].[master_department] SET is_assign = 'no' WHERE department_id = (SELECT department_id FROM [ems_v1].[dbo].[master_shed] WHERE shed_id = '{shed_id}')''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        print(error_filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)       
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_machinetype_list/{machinetype_id}")
@app.post("/get_machinetype_list/")
async def machinetype_list(machinetype_id: Optional[int] = None,
                           department_id : int = Form(None),
                           cnx: Session = Depends(get_db)):
    try:        
        where = ""
        if machinetype_id is not None:
            where = f" and mmt.machinetype_id = {machinetype_id}"
            
        if department_id is not None:
            where = f" and mmt.department_id = {department_id}"

        query = text(f'''
                    SELECT 
                       mc.company_code AS company_code,
                       mc.company_name AS company_name,
                       mb.branch_name AS branch_name,
                       mb.branch_code AS branch_code,
                       md.department_name AS department_name ,
                       md.department_code AS department_code,
                       mmt.*,
                       ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                       ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM [ems_v1].[dbo].[master_machinetype]  mmt
                       left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mmt.created_by
                       left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mmt.modified_by
                       INNER JOIN [ems_v1].[dbo].[master_company]  mc ON mmt.company_id = mc.company_id
                       INNER JOIN [ems_v1].[dbo].[master_branch]  mb ON  mmt.branch_id = mb.branch_id
                       INNER JOIN [ems_v1].[dbo].[master_department]  md ON mmt.department_id = md.department_id
                    WHERE 
                       mmt.status != 'delete'{where}
                    ''')  
        print(query)      
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_machinetype_details/")
async def save_machinetype_details(company_id: int = Form(None),
                                   branch_id:int = Form(None),
                                   department_id: int = Form(None), 
                                   machinetype_id : int = Form(None),                                   
                                   machinetype_code: str = Form(None),
                                   machinetype_name: str = Form(None),
                                   user_login_id : str = Form(None),                  
                                   cnx: Session = Depends(get_db)):
    if company_id == None:  
        return JSONResponse({"iserror": True, "message": " company_id is required"})
     
    if branch_id == None:  
        return JSONResponse({"iserror": True, "message": " branch_id is required"})
    
    if department_id == None:  
        return JSONResponse({"iserror": True, "message": " department_id is required"}) 
    
    if machinetype_code == None:  
        return JSONResponse({"iserror": True, "message": " machinetype_code is required"}) 
    
    if machinetype_name == None:  
        return JSONResponse({"iserror": True, "message": " machinetype_name is required"}) 
        
    try:        
        if machinetype_id is not None:
            query =text(f"""
                UPDATE [ems_v1].[dbo].[master_machinetype]
                SET company_id = {company_id}, machinetype_name = '{machinetype_name}', machinetype_code = '{machinetype_code}',
                branch_id = {branch_id}, department_id = {department_id}, modified_on = GETDATE(),
                modified_by = '{user_login_id}'
                WHERE machinetype_id = {machinetype_id}
            """)
         
        else:
            select_query = text(f'''select * from [ems_v1].[dbo].[master_machinetype] where machinetype_code = '{machinetype_code}' and status != 'delete' ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data) > 0:
              return JSONResponse({"iserror": True, "message": "machinetype_code already exists","data": " " })

            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_machinetype] (
                    company_id, machinetype_name, machinetype_code, branch_id, department_id,  created_on, created_by
                )
                VALUES (
                    {company_id},'{machinetype_name}', '{machinetype_code}', {branch_id}, {department_id}, GETDATE(), '{user_login_id}'
                )
            """)    
        cnx.execute(query)
        cnx.commit()
        
        sql = text(f''' select * from [ems_v1].[dbo].[master_machinetype] where department_id = '{department_id}' ''')
        data1 = cnx.execute(sql).mappings().all()
        
        if len(data1)>0:
            sql1= text(f''' update [ems_v1].[dbo].[master_department] set is_assign = 'yes' where department_id = '{department_id}' ''')
            cnx.execute(sql1)
            cnx.commit()
            
        return JSONResponse({"iserror": False, "message": "data save successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_machinetype/")
async def remove_machinetype(machinetype_id: int = Form(None),
                             status : str = Form(None), 
                             cnx: Session = Depends(get_db)):    
    try:
        if machinetype_id == None:  
            return JSONResponse({"iserror": True, "message": " machinetype_id is required"})
        
        if machinetype_id is not None:
            if status is not None:        
                query = text(f'''UPDATE [ems_v1].[dbo].[master_machinetype]  SET status = '{status}' WHERE machinetype_id = {machinetype_id}''')        

            else:
                query = text(f'''UPDATE [ems_v1].[dbo].[master_machinetype]  SET status = 'delete' WHERE machinetype_id = {machinetype_id}''')        

            cnx.execute(query)
            cnx.commit()
        
        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_machinetype] WHERE department_id = (SELECT department_id FROM [ems_v1].[dbo].[master_machinetype] WHERE machinetype_id = '{machinetype_id}') AND status != 'delete' ''')
        result = cnx.execute(query).mappings().all()
        if result == []:
            query = text(f'''UPDATE [ems_v1].[dbo].[master_department] SET is_assign = 'no' WHERE department_id = (SELECT department_id FROM [ems_v1].[dbo].[master_machinetype] WHERE machinetype_id = '{machinetype_id}')''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":" status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_function_list/{function_id}")
@app.post("/get_function_list/")
async def function_list(request:Request,
                        function_id:int = None,
                        function_type : str = Form(''),
                        cnx:Session = Depends(get_db)):
    try:
        # base_path = os.path.abspath(os.path.dirname(__file__))
                         
        where=""
        if function_id is not None:
            where += f"and mf.function_id = {function_id}"
        if function_type !='':
            where += f"and mf.function_type = '{function_type}'"
            
        query =text( f''' 
                    SELECT 
                    	mf.* , 
                    	CONCAT('/attachment/images/',mf.image) AS image,
                        ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                    	ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM
                    	[ems_v1].[dbo].[master_function] as mf 
                    	left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mf.created_by
                    	left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mf.modified_by
                    WHERE
                    	mf.status !='delete' {where} 
                    ''')
        print(query)
        data=cnx.execute(query).mappings().all()
        results = []
        for i in data:
            new_img = dict(i)
            if new_img["image"] is not None:
                new_img["image"] = "http://" + request.headers["host"] + new_img['image']
            results.append(new_img)

        return JSONResponse({"iserror":False, "message": " data return successfully","data":jsonable_encoder(results)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_function_details/")
async def save_function_details(function_id: int = Form(None),        
                                function_code: str = Form(None),
                                function_name: str = Form(None),
                                function_type : str = Form(''),
                                image: UploadFile = File(''),
                                old_image : str = Form(''),
                                user_login_id: str = Form(None),
                                cnx: Session = Depends(get_db)): 
    if function_code == None:  
        return JSONResponse({"iserror": True, "message": " function_code is required"}) 
    if function_name == None:  
        return JSONResponse({"iserror": True, "message": " function_name is required"})  
  
    try:
        # if image == '':
        #     filename = ''
        # else:
        if old_image == "":
            filename = save_image(image, f"{base_path}/images/")
        else:
            filename = old_image
            
        if function_id is not None:            
            query = text(f"""
                UPDATE [ems_v1].[dbo].[master_function] SET function_name = '{function_name}', 
                function_code = '{function_code}',
                image = '{filename}',
                modified_on = GETDATE(), 
                modified_by = '{user_login_id}', function_type = '{function_type}' WHERE function_id = {function_id}
            """)
            
        else:
           
            select_query = text(f'''select * from [ems_v1].[dbo].[master_function] where function_code = '{function_code}' and status != 'delete'  ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data) > 0:
                return JSONResponse({"iserror": True, "message": "function code already exists", "data": ""})
            
            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_function] (function_name, function_code, image, created_on, created_by, function_type)
                VALUES ('{function_name}', '{function_code}', '{filename}', GETDATE(), '{user_login_id}','{function_type}')
            """)
        cnx.execute(query)
        cnx.commit()        
        
        return JSONResponse({"iserror": False, "message": "data save successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_function/")
async def remove_function(function_id: int = Form(None),
                          status : str = Form(None),
                          cnx: Session = Depends(get_db)):    
    try:
        if function_id == None:  
            return JSONResponse({"iserror": True, "message": " function_id is required"})
        
        if function_id is not None:
            if status is not None:        
                query = text(f'''UPDATE [ems_v1].[dbo].[master_function] SET status = '{status}' WHERE function_id = {function_id}''')        
            
            else: 
                query = text(f'''UPDATE [ems_v1].[dbo].[master_function] SET status = 'delete' WHERE function_id = {function_id}''')        
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_converter_list/{converter_id}")
@app.post("/get_converter_list/")
async def converter_list(converter_id:int = None, 
                         cnx:Session = Depends(get_db)):
    try:       
        where=" "
        if converter_id is not None:
            where = f"and mcd.converter_id = {converter_id}"
            
        query =text( f''' 
                    SELECT 
                    	mcd.*,
                    	ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                    	ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                    FROM  
                    	[ems_v1].[dbo].[master_converter_detail] mcd
                    	left join ems_v1.dbo.master_employee cu on cu.employee_id=mcd.created_by
                    	left join ems_v1.dbo.master_employee mu on mu.employee_id=mcd.modified_by
                    WHERE 
                    	mcd.status !='delete'
                    	{where} 
                    ''')
        print(query)
        data=cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror":False, "message": " data return successfully","data":jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_converter_details/")
async def save_converter_details(converter_id:int=Form(None),
                                 converter_name:str = Form(None),
                                 ip_address:str=Form(None),
                                 port_no:str=Form(None),
                                 user_login_id:str=Form(None),
                                 cnx:Session=Depends(get_db)):  
    if converter_name == None:  
        return JSONResponse({"iserror": True, "message": " converter_name is required"}) 
      
    if ip_address == None:  
        return JSONResponse({"iserror": True, "message": " ip_address is required"})
    
    if port_no == None:  
        return JSONResponse({"iserror": True, "message": " port_no is required"})  
    try:        
        if converter_id is not None:
            query =text(f"""
                UPDATE [ems_v1].[dbo].[master_converter_detail]
                SET converter_name = '{converter_name}', ip_address = '{ip_address}',
                port_no = {port_no},  modified_on = GETDATE(),
                modified_by = '{user_login_id}'
                WHERE converter_id = {converter_id}
            """)
         
        else:          
            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_converter_detail] (
                     converter_name, ip_address, port_no, created_on, created_by
                )
                VALUES (
                    '{converter_name}', '{ip_address}', {port_no},  GETDATE(), '{user_login_id}'
                )
            """)    
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data save successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_converter_details/")
async def remove_converter_detail(converter_id: int = Form(None), 
                                  status : str = Form(None),
                                  cnx: Session = Depends(get_db)):    
    try:
        if converter_id == None:  
            return JSONResponse({"iserror": True, "message": " converter_id is required"})
        
        if converter_id is not None: 
            if status is not None:       
                query = text(f'''UPDATE [ems_v1].[dbo].[master_converter_detail]  SET status = '{status}' WHERE converter_id = {converter_id}''')       

            else:
                query = text(f'''UPDATE [ems_v1].[dbo].[master_converter_detail]  SET status = 'delete' WHERE converter_id = {converter_id}''')       
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_machine_list/")
async def machine_list(company_id : int = Form(None),
                       machine_id: int = Form(None),
                       type_value: str = Form(None),
                       type_id : str = Form(None), 
                       is_critical : str = Form(None),
                       model_name: str = Form(None),
                       department_id : int = Form(None),
                       shed_id : int = Form(None),
                       machinetype_id : int = Form(None),
                       function_id : int = Form(None),
                       function2_id :str = Form(''),
                       cnx: Session = Depends(get_db)):
    try:     
        
        print("machine_id",machine_id)
        if type_id is not None:
            value = type_id.split(",")
            if len(value) > 1:
                values = tuple(value)
                type_id = ",".join(values)
            else:
                type_id = value[0]
                
        where = ""
        where1 = ""
        if company_id is not None:
            where += f" and mm.company_id = {company_id}" 
        if machine_id is not None:
            where += f" and mm.machine_id = {machine_id}" 
            where1 += f" and ms.machine_id = {machine_id}" 
             
        if type_value is not None and type_id is not None:
            if type_value == 'zone':
                where += f" and mm.department_id in ({','.join(str(x) for x in value)})"
          
            elif type_value == 'area':
                where += f" and mm.shed_id in ({','.join(str(x) for x in value)})"

            elif type_value == 'location':
                where += f" and mm.machinetype_id in ({','.join(str(x) for x in value)})"

            elif type_value == 'function':
                where += f" and mm.function_id in ({','.join(str(x) for x in value)})"
            
            elif type_value == 'function_1':
                where += f" and mm.function_id in ({','.join(str(x) for x in value)})"
                
            elif type_value == 'function_2':
                where += f" and mm.function2_id in ({','.join(str(x) for x in value)})"
                
        if is_critical=="yes" or is_critical=="no"  :
            where += f" and mm.major_nonmajor = '{is_critical}' "   
        
        if model_name is not None:
            where += f" and mm.model_name" 
        
        if department_id is not None and department_id != 0:
            where += f" and mm.department_id = {department_id}"
            
        if shed_id is not None and shed_id != 0:
            where +=f" and mm.shed_id = {shed_id}"
        
        if machinetype_id is not None and machinetype_id != 0:
            where += f" and mm.machinetype_id = {machinetype_id}"
            
        if function_id is not None:
            where += f" and mm.function_id = {function_id}"  

        if function2_id !='':
            where += f" and mm.function2_id = {function2_id}"                 
        
        query = text(f"""
                    SELECT 
                        mc.company_code AS company_code,
                        mc.company_name AS company_name,
                        mb.branch_code AS branch_code,
                        mb.branch_name AS branch_name,
                        md.department_code AS department_code,
                        md.department_name AS department_name,
                        ms.shed_code AS shed_code,
                        ms.shed_name AS shed_name,
                        mmt.machinetype_code AS machinetype_code,                        
                        mmt.machinetype_name AS machinetype_name,
                        mf.function_name  AS function1_name,
                        mf.function_code  AS function1_code,
                        mff.function_name  AS function2_name,
                        mff.function_code  AS function2_code,
                        mcd.converter_name AS converter_name,
                        mm.*,
                        ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                    ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM 
                        [ems_v1].[dbo].[master_machine] mm
                        left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mm.created_by
	                    left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mm.modified_by
                        INNER JOIN [ems_v1].[dbo].[master_company] mc ON mm.company_id = mc.company_id
                        INNER JOIN [ems_v1].[dbo].[master_branch] mb ON mm.branch_id = mb.branch_id
                        INNER JOIN [ems_v1].[dbo].[master_department] md ON mm.department_id = md.department_id
                        INNER JOIN [ems_v1].[dbo].[master_shed] ms ON mm.shed_id = ms.shed_id
                        INNER JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mm.machinetype_id = mmt.machinetype_id
                        LEFT JOIN [ems_v1].[dbo].[master_function] mf ON mm.function_id = mf.function_id
                        LEFT JOIN [ems_v1].[dbo].[master_function] mff ON mm.function2_id = mff.function_id
                        LEFT JOIN [ems_v1].[dbo].[master_converter_detail] mcd ON mm.converter_id = mcd.converter_id
                    WHERE 
                        mm.status != 'delete' {where} """)
        print(query)
        data = cnx.execute(query).mappings().all()

        query2 = text(f"""
            SELECT                
                ms.*,
                IsNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user
            FROM 
                ems_v1.dbo.master_scaling_value ms
                inner join ems_v1.dbo.master_machine mm on mm.machine_id = ms.machine_id
                left join ems_v1.dbo.master_employee cu on cu.employee_id=ms.created_by
            WHERE 
                mm.status != 'delete' {where1}
        """)
        print(query2)
        data2 = cnx.execute(query2).mappings().all()

        
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data),"data2":jsonable_encoder(data2)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_machine_details/")
async def save_machine_details(company_id: int = Form(None),
                               branch_id:int = Form(None),
                               department_id: int = Form(None),
                               shed_id: int = Form(None), 
                               machinetype_id: int = Form(None),
                               function_id:int = Form(''),
                               converter_id: int = Form(None),
                               machine_id : int = Form(None),
                               machine_code: str = Form(None),
                               machine_name: str = Form(None),
                               ip_address : str = Form(None),
                               port :str = Form(None),
                               major_nonmajor : str = Form(None), 
                               model_name : str = Form(None),                                                
                               energy_selection : str = Form(''), 
                               energy_selection1 : str = Form(''), 
                               IMEI : int = Form(None),  
                               user_login_id : str = Form(None), 
                               function2_id : int = Form(''), 
                               source :str = Form(''),                  
                               obj :str = Form(''),                  
                               cnx: Session = Depends(get_db)):
    type = ''
    scaling1_min = 0
    scaling2_min = 0
    scaling1_max = 0
    scaling2_max = 0
    if company_id == None:  
        return JSONResponse({"iserror": True, "message": " company_id is required"}) 
    
    if branch_id == None:  
        return JSONResponse({"iserror": True, "message": " branch_id is required"})    
    
    if department_id == None:  
        return JSONResponse({"iserror": True, "message": " department_id is required"})   
    
    if shed_id == None:  
        return JSONResponse({"iserror": True, "message": " shed_id is required"})
    
    if machinetype_id == None:  
        return JSONResponse({"iserror": True, "message": " machinetype_id is required"}) 
    
    if function_id == None:  
        return JSONResponse({"iserror": True, "message": " function_id is required"})    
    
    if converter_id == None:  
        return JSONResponse({"iserror": True, "message": " converter_id is required"}) 
       
    if machine_name == None:
        return JSONResponse({"iserror": True, "message": " machine_name is required"})
     
    if machine_code == None:  
        return JSONResponse({"iserror": True, "message": " machine_code is required"}) 
    
    if ip_address == None:  
        return JSONResponse({"iserror": True, "message": " ip_address is required"}) 
    
    if port == None:  
        return JSONResponse({"iserror": True, "message": " port is required"})
         
    if major_nonmajor == None:  
        return JSONResponse({"iserror": True, "message": " major_nomajor is required"})
    
    if model_name == None:  
        return JSONResponse({"iserror": True, "message": " model_name is required"})
    
    if IMEI == None:  
        return JSONResponse({"iserror": True, "message": " IMEI is required"})
    
    if function_id == '' and function2_id == '':  
        return JSONResponse({"iserror": True, "message": " function id is required"})
    
    mill_date = ''
    mill_shift =''
    print(function_id)         
    try:                       
        if machine_id is None:
            
            select_query = text(f'''select * from [ems_v1].[dbo].[master_machine] where machine_code = '{machine_code}' and status != 'delete' ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data) > 0:
              return JSONResponse({"iserror": True, "message": "machine_code already exists","data": " " })

            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_machine] (
                    company_id, machine_name, machine_code, branch_id, department_id, shed_id, converter_id, function_id,machinetype_id,
                    ip_address, port, created_on, created_by, major_nonmajor, model_name, energy_selection, IMEI, energy_selection1, function2_id,source
                )
                VALUES (
                    {company_id},'{machine_name}', '{machine_code}', {branch_id}, {department_id}, {shed_id}, {converter_id}, '{function_id}',
                    {machinetype_id}, '{ip_address}',{port}, GETDATE(), '{user_login_id}', '{major_nonmajor}', '{model_name}','{energy_selection}', {IMEI},'{energy_selection1}','{function2_id}',
                    '{source}'
                )
            """)
            
                
        if machine_id is not None: 

            sql = text(f'''INSERT INTO [ems_v1].[dbo].[master_machine_history] (
                    company_id, machine_name, machine_code, branch_id, department_id, shed_id, converter_id, function_id,machinetype_id,
                    ip_address, port, modified_on, modified_by, major_nonmajor, model_name, energy_selection, IMEI,energy_selection1,function2_id,source
                )
                VALUES (
                    {company_id},'{machine_name}', '{machine_code}', {branch_id}, {department_id}, {shed_id}, {converter_id}, {function_id},
                    {machinetype_id}, '{ip_address}',{port}, GETDATE(), '{user_login_id}', '{major_nonmajor}', '{model_name}','{energy_selection}',{IMEI},'{energy_selection1}','{function2_id}',
                    '{source}'

                ) 
                ''')
            cnx.execute(sql)
            cnx.commit()
            createFolder("Log/"," current power" +str(sql))
            
            query =text(f"""
                UPDATE [ems_v1].[dbo].[master_machine]
                SET company_id = {company_id}, machinetype_id = {machinetype_id}, machine_code = '{machine_code}',machine_name = '{machine_name}',
                branch_id = {branch_id}, shed_id = {shed_id},converter_id = {converter_id}, department_id = {department_id},function_id = '{function_id}',
                ip_address = '{ip_address}', port = '{port}', modified_on = GETDATE(), modified_by = '{user_login_id}', 
                major_nonmajor = '{major_nonmajor}', model_name = '{model_name}', energy_selection = '{energy_selection}', IMEI = {IMEI},energy_selection1='{energy_selection1}',function2_id='{function2_id}',
                source = '{source}'
                WHERE machine_id = '{machine_id}'
            """)           
        cnx.execute(query)
        
        if machine_id is not None:
            insert_id = machine_id
        else:
            insert_id = cnx.execute(text("SELECT SCOPE_IDENTITY()")).first()[0]
            
        print("insert_id",insert_id)
        cnx.commit()        
        # insert_id = cnx.execute("SELECT LAST_INSERT_ID()").first()[0]
        if insert_id is not None: 
            if obj!= '':
                del_query = text(f"delete from ems_v1.dbo.master_scaling_value where machine_id = {insert_id}")
                cnx.execute(del_query)
                cnx.commit()
                scaling_data = json.loads(obj)
                for i in scaling_data:
                    scaling1_min=i["scaling1_min"]
                    print("scaling1_min",scaling1_min)
                    scaling1_max=i["scaling1_max"]
                    scaling2_min=i["scaling2_min"]
                    scaling2_max=i["scaling2_max"]
                    type=i["type"]
                    scl_query = text(f'''insert into ems_v1.dbo.master_scaling_value (machine_id,scaling1_min,scaling1_max,scaling2_min,scaling2_max,type)
                                     values({insert_id},'{scaling1_min}','{scaling1_max}','{scaling2_min}','{scaling2_max}','{type}')''')
                    print(scl_query)
                    cnx.execute(scl_query)
                    cnx.commit()

            query1 = text(f'''select * from ems_v1.dbo.master_machine_factor where machine_id = {insert_id}''')
            record = cnx.execute(query1).mappings().all() 
            if len(record) == 0:
                query2 = text(f'''insert into  ems_v1.dbo.master_machine_factor 
                (machine_id,machine_kWh ,machine_kWh_value,kWh,kWh_value,r_volt,r_volt_value,y_volt,y_volt_value,b_volt,b_volt_value,ry_volt,ry_volt_value,
                yb_volt,yb_volt_value,br_volt,br_volt_value,vll_avg,vll_avg_value,vln_avg,vln_avg_value,r_current,r_current_value,y_current,y_current_value,
                b_current,b_current_value,t_current,t_current_value,frequency,frequency_value,r_watts,r_watts_value,y_watts,y_watts_value,b_watts,b_watts_value,
                t_watts,t_watts_value,kw,kw_value,r_powerfactor,r_powerfactor_value,y_powerfactor,y_powerfactor_value,b_powerfactor,b_powerfactor_value,avg_powerfactor,
                avg_powerfactor_value,powerfactor,powerfactor_value,power_factor,power_factor_value,r_var,r_var_value,y_var,y_var_value,b_var,b_var_value,t_var,t_var_value,
                r_voltampere,r_voltampere_value,y_voltampere,y_voltampere_value,b_voltampere,b_voltampere_value,t_voltampere,t_voltampere_value,kvah,kvah_value,kvar,kvar_value,kva,kva_value,company_id, department_id, shed_id, machinetype_id)
                values({insert_id},'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,'*',1,{company_id},{department_id},{shed_id},{machinetype_id})

                ''')
                cnx.execute(query2)
                cnx.commit()
             
            sql = text(f''' select * from [ems_v1].[dbo].[current_power] where machine_id = {insert_id}''') 
            data = cnx.execute(sql).mappings().all()

            if len(data)==0:                       
                sql1 = text(f"select * from [ems_v1].[dbo].[master_machine] where machine_id = {insert_id}")

                data1 = cnx.execute(sql1).mappings().all()
                for row in data1:
                    machine_id = row["machine_id"]
                    company_id = row["company_id"]
                    branch_id = row["branch_id"]
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]  

                sql2= text(f" select * from [ems_v1].[dbo].[master_shifts]  where company_id = {company_id} and branch_id = {branch_id} AND status = 'active' ")
                data2 = cnx.execute(sql2).mappings().all()
                
                
                if len(data2)>0:
                    for row in data2:
                        mill_date = row["mill_date"]
                        mill_shift = row["mill_shift"]  
                    if source == 'water':
                        sql = text(f''' select * from [ems_v1].[dbo].[current_water] where water_meter_id = {insert_id}''') 
                        data_w = cnx.execute(sql).mappings().all()
                        if len(data_w)==0:
                            sql3 = text(f'''
                                        INSERT INTO [ems_v1].[dbo].[current_water] (water_meter_id,mill_date, mill_shift,created_on,created_by)
                                        VALUES ({machine_id},'{mill_date}', '{mill_shift}',getdate(),{user_login_id})
                                        ''')  
                            cnx.execute(sql3)
                            cnx.commit()
                    else:
                        sql3 = text(f'''
                                    INSERT INTO [ems_v1].[dbo].[current_power] (machine_id, date_time, date_time1,
                                    mill_date, mill_shift,company_id, branch_id, department_id, shed_id, machinetype_id)
                                    VALUES ({machine_id}, GETDATE(), GETDATE(), '{mill_date}', '{mill_shift}',{company_id},
                                    {branch_id}, {department_id}, {shed_id}, {machinetype_id})
                                    ''')  
                        cnx.execute(sql3)
                        cnx.commit()
                        createFolder("Log/"," current power" +str(sql3))
        
        return JSONResponse({"iserror": False, "message": "data save successfully", "data": " "})    
    except Exception as e:
        # error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f" occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_machine_details/")
async def remove_converter_detail(machine_id: str = Form(None),
                                  status : str = Form(None),
                                  cnx: Session = Depends(get_db)):    
    try:
        if machine_id == None:  
            return JSONResponse({"iserror": True, "message": " machine_id is required"})
        
        if machine_id is not None:
            if status is not None:        
                query = text(f'''UPDATE [ems_v1].[dbo].[master_machine]  SET status = '{status}' WHERE machine_id = {machine_id}''')
                cnx.execute(query)
                cnx.commit()

            else:
                query = text(f'''UPDATE [ems_v1].[dbo].[master_machine]  SET status = 'delete' WHERE machine_id = {machine_id}''')
                cnx.execute(query)
                cnx.commit()
                sql = text(f"delete [ems_v1].[dbo].[master_scaling_value] where machine_id = {machine_id}")
                cnx.execute(sql)
                cnx.commit()
           
        
        return JSONResponse({"iserror":False,"message":" status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_meter_group_list/{meter_group_id}")   
@app.post("/get_meter_group_list/")
async def meter_group_list(meter_group_id: int = None, 
                           cnx: Session = Depends(get_db)):
    try:
        where = ''
        if meter_group_id is not None:
            where = f" AND mmg.meter_group_id = {meter_group_id}"   
        # where += f" and mm.machine_id in ({','.join(str(x) for x in machine_id)})  
        query = text(f"""
            SELECT                
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                (CASE 
                WHEN group_type='Zone' THEN (SELECT department_name FROM ems_v1.dbo.master_department WHERE department_id=type_id)
                WHEN group_type='Area' THEN (SELECT shed_name FROM ems_v1.dbo.master_shed WHERE shed_id=type_id)
                WHEN group_type='Location' THEN (SELECT machinetype_name FROM ems_v1.dbo.master_machinetype WHERE machinetype_id=type_id)
                WHEN group_type='Function' THEN (SELECT function_name FROM ems_v1.dbo.master_function WHERE function_id=type_id)
                WHEN group_type='Function_1' THEN (SELECT function_name FROM ems_v1.dbo.master_function WHERE function_id=type_id)
                WHEN group_type='Function_2' THEN (SELECT function_name FROM ems_v1.dbo.master_function WHERE function_id=type_id)
                END) AS type_name,
                mmg.*,
                '' AS machine_dtl,
                ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	            ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

            FROM 
                [ems_v1].[dbo].[master_meter_group] mmg
                left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mmg.created_by
	            left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mmg.modified_by                
                INNER JOIN [ems_v1].[dbo].[master_machine] mm ON (mmg.machine_id LIKE CONCAT('%,', CAST(mm.machine_id AS VARCHAR)) OR 
                 
                mmg.machine_id = CAST(mm.machine_id AS VARCHAR))
            WHERE 
                mmg.status != 'delete'{where}
        """)
        print(query)
        data = cnx.execute(query).mappings().all()
        result = []
        for row in data:
            machine_id_list = row["machine_id"].split(",")   # Split comma-separated machine IDs into a list
            machine_dtl = ""
            for machine_id in machine_id_list:                             
                sub_query = text(f"SELECT * FROM [ems_v1].[dbo].[master_machine] WHERE machine_id = {machine_id}")
                sub_data = cnx.execute(sub_query).mappings().all()
                for sub_row in sub_data:
                    if machine_dtl != "":
                        machine_dtl += '\n' 
                    machine_dtl += f'''{sub_row['machine_code']} - {sub_row['machine_name']} '''  
                    print(machine_dtl)          
            new_row = dict(row)
            new_row["machine_dtl"] = machine_dtl
            result.append(new_row)
            
        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(result)})
    except Exception as e:        
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_meter_group/")
async def save_meter_group(machine_id:str =Form(None),
                           meter_group_id:int=Form(None),
                           group_type:str = Form(None),
                           type_id:str=Form(None),
                           user_login_id:str=Form(None),
                           cnx:Session=Depends(get_db)):  
    if group_type == None:
        return JSONResponse({"iserror": True, "message": " group_type is required"})
    
    if type_id == None:
        return JSONResponse({"iserror": True, "message": " type_id is required"})
    
    if machine_id == None:
        return JSONResponse({"iserror": True, "message": " machine_id is required"})
        
    try:
        if machine_id is not None:
                value = machine_id.split(",")
                if len(value) > 1:
                    values = tuple(value)
                    machine_id = ",".join(values)
                else:
                    machine_id = value[0]  
                       
        if type_id is not None:
                value = type_id.split(",")
                if len(value) > 1:
                    values = tuple(value)
                    type_id = ",".join(values)
                else:
                    type_id = value[0]    
                                      
        if meter_group_id is not None:
         query =text(f"""
            UPDATE [ems_v1].[dbo].[master_meter_group]
            SET group_type = '{group_type}', type_id = '{type_id}',
            machine_id = '{machine_id}',  modified_on = GETDATE(),
            modified_by = '{user_login_id}'
            WHERE meter_group_id = {meter_group_id} 
        """)
         
        else:
            if machine_id is not None:
                value = machine_id.split(",")
                if len(value) > 1:
                    values = tuple(value)
                    machine_id = ",".join(values)
                else:
                    machine_id = value[0]          
             
            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_meter_group] (
                group_type, type_id, machine_id, created_on, created_by
                )
                VALUES (
                    '{group_type}', '{type_id}', '{machine_id}',  GETDATE(), '{user_login_id}'
                )
            """)
        print(query)
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data save successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_meter_group/")
async def remove_meter_group(meter_group_id: int = Form(None), 
                             status : str = Form(None),
                             cnx: Session = Depends(get_db)):    
    try:
        if meter_group_id == None:  
            return JSONResponse({"iserror": True, "message": " meter_group_id is required"})
        
        if meter_group_id is not None:    
            if status is not None:               
                query = text(f'''UPDATE [ems_v1].[dbo].[master_meter_group]  SET status = '{status}' WHERE meter_group_id = {meter_group_id}''')        

            else:
                 query = text(f'''UPDATE [ems_v1].[dbo].[master_meter_group]  SET status = 'delete' WHERE meter_group_id = {meter_group_id}''')        
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_shift_list/{shift_id}")
@app.post("/get_shift_list/")
async def shift_list(shift_id: Optional[int] = None, 
                     cnx: Session = Depends(get_db)):
    try:        
        where = ""
        if shift_id is not None:
            where = f" and ms.shift_id = {shift_id}"

        query = text(f"""
                    SELECT 
                        mc.company_code AS company_code,
                        mc.company_name AS company_name,
                        mb.branch_code AS branch_code,
                        mb.branch_name AS branch_name,
                        FORMAT(ms.shift1_start_time, 'h:mm:ss tt') AS shift1_time,
                        FORMAT(ms.shift2_start_time, 'h:mm:ss tt') AS shift2_time,
                        FORMAT(ms.shift3_start_time, 'h:mm:ss tt') AS shift3_time,
                        ms.*,
                        ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                    ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                    FROM 
                        [ems_v1].[dbo].[master_shifts] ms
                        left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=ms.created_by
	                    left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=ms.modified_by
                        INNER JOIN [ems_v1].[dbo].[master_company] mc ON ms.company_id = mc.company_id
                        INNER JOIN [ems_v1].[dbo].[master_branch] mb ON ms.branch_id = mb.branch_id
                    WHERE 
                        ms.status != 'delete' {where}

                    """)
        data = cnx.execute(query).mappings().all()

        return JSONResponse({"iserror": False, "message": "data return successfully", "data": jsonable_encoder(data)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_shifts_details/")
async def save_shifts_details(company_id: int = Form(None),
                              branch_id:int = Form(None),
                              shift_id : str = Form(None),
                              no_of_shifts:int= Form(None),
                              shift1_start_time:str=Form(None),
                              shift2_start_time:str=Form(None),
                              shift3_start_time:str=Form(None),
                              user_login_id : str = Form(None),                             
                              cnx: Session = Depends(get_db)):
    
    if company_id == None:  
        return JSONResponse({"iserror": True, "message": " company_id is required"})
    
    if branch_id == None:  
        return JSONResponse({"iserror": True, "message": " branch_id is required"}) 
    
    if no_of_shifts == None:         
        return JSONResponse({"iserror": True, "message": " no_of_shifts is required"}) 
    
    if no_of_shifts == 1 and (shift1_start_time is None):
            return JSONResponse({"iserror": True, "message": "shift1_start_time is required"})        
        
    if no_of_shifts == 2 and (shift1_start_time is None or shift2_start_time is None) :
        return JSONResponse({"iserror": True, "message": "shift1_start_time and shift2_start_time is required"})
    
    if no_of_shifts == 3 and (shift1_start_time is None or shift2_start_time is None or shift3_start_time is None):
        return JSONResponse({"iserror": True, "message": "shift1_start_time and shift2_start_time and shift3_start_time are required"})
    
    if shift2_start_time is None:
        shift2_start_time = ''
        
    if shift3_start_time is None:
        shift3_start_time = '' 
           
    try:      
        if shift_id is not None: 
            # Update existing shift record
            query = text(f"""
                UPDATE [ems_v1].[dbo].[master_shifts]
                SET company_id = {company_id}, 
                    branch_id = {branch_id}, 
                    no_of_shifts = {no_of_shifts},
                    shift1_start_time = '{shift1_start_time}',
                    shift2_start_time = '{shift2_start_time}',
                    shift3_start_time = '{shift3_start_time}',
                    modified_on = GETDATE(),
                    modified_by = '{user_login_id}'
                WHERE shift_id = {shift_id}
            """)  
  
        else:            
            query = text(f"""
                INSERT INTO [ems_v1].[dbo].[master_shifts] (
                     company_id, branch_id, no_of_shifts, shift1_start_time, shift2_start_time, shift3_start_time, created_on, created_by
                )
                VALUES (
                   {company_id}, {branch_id}, {no_of_shifts}, 
                    '{shift1_start_time}',
                    '{shift2_start_time}',
                    '{shift3_start_time}',
                    GETDATE(), '{user_login_id}'
                )
            """)
        print(query)
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data save successfully", "data": " "})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/remove_shift_details/")
async def remove_shift_detail(shift_id: int = Form(None), 
                              status : str = Form(None),
                              cnx: Session = Depends(get_db)):    
    try:
        if shift_id == None:  
            return JSONResponse({"iserror": True, "message": " shift_id is required"})
        if shift_id is not None:
            if status is not None:        
                query = text(f'''UPDATE [ems_v1].[dbo].[master_shifts]  SET status = '{status}' WHERE shift_id = {shift_id}''')

            else:
                query = text(f'''UPDATE [ems_v1].[dbo].[master_shifts]  SET status = 'delete' WHERE shift_id = {shift_id}''')       
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully ","data":""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_employee_list/")
async def get_employeelist(request:Request,
                           employee_id:str=Form(None),
                           cnx: Session = Depends(get_db)):   
    
    try:    
        where = ""
        if employee_id is not None:
            where = text(f"and me.employee_id = '{employee_id}'")

        query=text(f'''
                    SELECT
                        me.*,
                        mc.company_name AS company_name,
                        mb.branch_name AS branch_name,
                        ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                    ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user,
                        md.department_name as department_name,
                        md.department_code as department_code,
                        ms.shed_name as shed_name,
                        ms.shed_code as shed_code,
                        mmt.machinetype_name as machinetype_name,
                        mmt.machinetype_code as machinetype_code,
                        HASHBYTES('MD5', me.password_login) password_login     
                        

                    FROM    
                        [ems_v1].[dbo].[master_employee] me
                        left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=me.created_by
	                    left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=me.modified_by
                        left JOIN [ems_v1].[dbo].[master_company] mc ON mc.company_id=me.company_id
                        left JOIN [ems_v1].[dbo].[master_branch] mb ON mb.branch_id=me.branch_id                        
                        left JOIN [ems_v1].[dbo].[master_department] md ON md.department_id=me.department_id                        
                        left JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id=me.shed_id                        
                        left JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id=me.machinetype_id                        
                        WHERE me.status!='delete'and me.employee_type != 'admin' {where} ''')      
        print(query)
        data = cnx.execute(query).mappings().all()
        createFolder("Log/","Query executed successfully for  employee list")
        return JSONResponse({"iserror": False, "message": "data return successfully", "data":jsonable_encoder(data)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_employee_detail/")
async def save_employee_detail(employee_id:str=Form(None),
                               company_id:str=Form(''),
                               branch_id:str=Form(''),
                               department_id:str=Form(''),
                               shed_id:str=Form(''),
                               machinetype_id:str=Form(''),
                               employee_code:str=Form(None),
                               employee_name:str=Form(None),
                               employee_type:str=Form(None),
                               mobileno:str=Form(None),
                               email:str=Form(None),
                               password_login:str=Form(None),
                               is_login:str=Form(None),
                               login_id:str=Form(None),
                               cnx: Session = Depends(get_db)):
    
    if company_id == '':
        return JSONResponse({"iserror":True,"message":"company id is required"}) 
    
    if employee_code == None:
        return JSONResponse({"iserror":True,"message":"employee code is required"})
    
    if employee_name == None:
        return JSONResponse({"iserror":True,"message":"employee name is required"})
    
    if employee_type == None:
        return JSONResponse({"iserror":True,"message":"employee type is required"})
    
    if password_login == None:
        return JSONResponse({"iserror":True,"message":" login password is required"})
    
    if is_login == None:
        return JSONResponse({"iserror":True,"message":" is login is required"})    
    
    if mobileno is None:
        mobileno = ''

    if email is None:
        email=''
    try:
        
        if employee_id is not None:
            query =text(f'''update  [ems_v1].[dbo].[master_employee] set company_id = '{company_id}',branch_id = '{branch_id}',  
                       employee_name = '{employee_name}',employee_code = '{employee_code}',employee_type = '{employee_type}',
                       mobileno = '{mobileno}',email = '{email}',password_login= HASHBYTES('MD5', '{password_login}'),is_login='{is_login}',
                       department_id = '{department_id}',shed_id = '{shed_id}', machinetype_id = '{machinetype_id}',
                       modified_on = GETDATE(),modified_by='{login_id}' where employee_id = '{employee_id}'
                       ''')
            
        else:
            select_query = text(f'''select * from [ems_v1].[dbo].[master_employee] where employee_code = '{employee_code}' and status != 'delete' ''')
            data1 = cnx.execute(select_query).mappings().all()

            if len(data1)>0:
                return JSONResponse({"iserror":True,"message":"employee code already exists "})

            query= text(f'''insert into [ems_v1].[dbo].[master_employee] (company_id,branch_id,employee_name,
                       employee_code,employee_type,mobileno,email,password_login,is_login,created_on,created_by,department_id, shed_id, machinetype_id )
                       values('{company_id}','{branch_id}' ,'{employee_name}','{employee_code}','{employee_type}',
                       '{mobileno}','{email}', HASHBYTES('MD5', '{password_login}'),'{is_login}',GETDATE(),'{login_id}','{department_id}','{shed_id}','{machinetype_id}') ''')
        print(query)
        cnx.execute(query)
        cnx.commit()            

        createFolder("Log/","employee "+str(query))
        return JSONResponse({"iserror":False,"message":"data saved successfully","data":""})
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/remove_employee_detail/")
async def remove_employee_detail(employee_id:str=Form(None),
                                 status : str = Form(None),
                                 cnx: Session = Depends(get_db)):
    
    try:
        if employee_id == None:  
            return JSONResponse({"iserror": True, "message": " employee_id is required"})
    
        if employee_id is not None:
            if status is not None:
                query = text(f''' update [ems_v1].[dbo].[master_employee] set status = '{status}' where employee_id = '{employee_id}' ''')

            else:
                 query = text(f''' update [ems_v1].[dbo].[master_employee] set status = 'delete' where employee_id = '{employee_id}' ''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully","data":""}) 
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
 
@app.post("/current_power/")
async def current_power(company_id : str = Form(''),
                        department_id :str = Form(''),
                        shed_id :str = Form(''),
                        machinetype_id :str = Form(''),
                        function_id : str = Form(''),
                        machine_id : str = Form (''),
                        group_for : str = Form(''),
                        groupby : str = Form(''),
                        period_id: str = Form(''),
                        from_date: str = Form(''),
                        to_date: str = Form(''),                      
                        shift_id: str = Form(''),
                        limit_report_for = Form(''),
                        limit_exception_for:str = Form(''),
                        limit_order_by : str = Form(''),
                        limit_operation_value : str = Form(''),
                        is_critical :str = Form(''),
                        converter_id :int = Form(''),  
                        report_for : str = Form(''), 
                        is_function : str = Form(''),  
                        function_type : str = Form(''), 
                        reportfor:str = Form(''), 
                        employee_id : int = Form(''),                
                        is_main_meter : str = Form(''),                
                        cnx: Session = Depends(get_db)):
    sql = ''
    if period_id == '':
            return JSONResponse({"iserror": True, "message": "period id is required"})
    
    if groupby =='':
        return JSONResponse({"iserror":True, "message": "groupby is required"}) 
    
    if group_for =='':
        return JSONResponse({"iserror":True, "message": "group_for is required"}) 
    try:
        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
     
        machine_id = id(machine_id)
        company_id = id(company_id)
        department_id = id(department_id)
        shed_id = id(shed_id)
        machinetype_id = id(machinetype_id)
        function_id = id(function_id)
        data2 = ''
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="[ems_v1_completed].[dbo]."           
        where = "" 
        group_by = ""
        order_by = ""  
        function_where = ''

        if employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]
                    print("department_id",department_id)

        if  company_id == '' or company_id == "0":
            pass
        else:
            where += f" and  mm.company_id in ({company_id})" 
    
        if department_id == '' or department_id == "0":
            pass
        else:
            where += f" and  mm.department_id in ({department_id})"          
            
        if shed_id == '' or shed_id == "0":
            pass
        else:
            where += f" and mm.shed_id in ({shed_id})"
            
        if machinetype_id == '' or machinetype_id == "0":
            pass
        else:
            where += f" and mm.machinetype_id in ({machinetype_id})"
            
        if function_id == '':
            pass
        else:
            if function_type =='':
                # function_where = f" mm.function_id = mf.function_id"
                where += f"and  mm.function_id in ({function_id})"
            else:
                if function_type == 'function_1':
                    where += f"and  mm.function_id in ({function_id})"
                else:
                    where += f"and  mm.function2_id in ({function_id})"

        if machine_id == 'all' or machine_id == '':
            pass
        else:
            where += f" and mm.machine_id in ({machine_id})"
            
        if converter_id == '':
            pass
        else:
            where += f" and mm.converter_id = {converter_id}"
        
        if function_type !='':
            where += f" and mf.function_type = '{function_type}'"
            if function_type == 'function_2':
                function_where += f" mm.function2_id = mf.function_id"  
            else:
                function_where += f" mm.function_id = mf.function_id"
        else:
            function_where += f" mm.function_id = mf.function_id"
            
        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_shifts] WHERE status = 'active' ''')
        data1 = cnx.execute(query).mappings().all()
        mill_date = date.today()
        mill_shift = 0
        no_of_shifts = 3
        group_id = ""
        group_by_poll = ''
        group_code = ""
        group_name = ""

        month_year = ""
        table_name = ''
        where_mm='' 
        poll_duration = ''
        join = ''

        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]  
              no_of_shifts = shift_record["no_of_shifts"]          
        if reportfor == '12to12':
            if period_id != 'sel_date' and period_id != 'from_to':
                return JSONResponse({"iserror": True, "message": "invalid period id"}) 
            
            if period_id == "sel_date":            
                if from_date == '':
                    return JSONResponse({"iserror": True, "message": "from date is required"})    
                
                from_date = parse_date(from_date)
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12'"""
                result_query = cnx.execute(query).mappings().all()
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "12to12 table not available..."})
                    
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'polling_data_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                
                table_name=f"  {completed_db}[power_{month_year}_12] as cp "       
                where += f''' and cp.mill_date = '{from_date}' '''

            elif period_id == "from_to":            
                if from_date == '':
                    return JSONResponse({"iserror": True, "message": "from date is required"})
                if to_date == '':
                    return JSONResponse({"iserror": True, "message": "to_date is required"})  
                        
                from_date = parse_date(from_date)
                to_date =  parse_date(to_date)
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
            
                where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' ''' 
                if from_date.month == to_date.month:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12'"""
                    result_query = cnx.execute(query).mappings().all()
                    if len(result_query) == 0:
                        return JSONResponse({"iserror": True, "message": "12to12 table not available..."})    
                
                    table_name=f"  {completed_db}[power_{month_year}_12] as cp "
                else:
                    field_name = 'power_id,company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,design_id,beam_id,date_time,date_time1,mill_date,mill_shift,vln_avg,r_volt,y_volt,b_volt,vll_avg,ry_volt,yb_volt,br_volt,t_current,r_current,y_current,b_current,t_watts,r_watts,y_watts,b_watts,t_var,r_var,y_var,b_var,t_voltampere,r_voltampere,y_voltampere,b_voltampere,avg_powerfactor,r_powerfactor,y_powerfactor,b_powerfactor,powerfactor,kWh,kvah,kw,kvar,power_factor,kva,frequency,machine_status,status,created_on,created_by,modified_on,modified_by,machine_kWh,master_kwh,reverse_machine_kwh,reverse_master_kwh,reverse_kwh'
                    
                    from_month = from_date.month
                    to_month = to_date.month
                    month_year_range = [
                        (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]

                    union_queries = []

                    for month_year in month_year_range:
                        query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12'"""
                        result_query = cnx.execute(query).mappings().all()
                        print(query)
                        if len(result_query) > 0:
                            table_name = f"[ems_v1_completed].[dbo].[power_{month_year}_12]"
                            union_queries.append(f"SELECT {field_name} FROM {table_name}")
                    if len(union_queries) == 0:
                        return JSONResponse({"iserror": True, "message": "12to12 table not available..."})    
                
                    subquery_union = " UNION ALL ".join(union_queries)
                    table_name = f"( {subquery_union}) cp"
        
        else:

            if period_id == "cur_shift":       
                where += f''' and cp.mill_date = '{mill_date}' AND cp.mill_shift = '{mill_shift}' '''              
                table_name = "[ems_v1].[dbo].[current_power] cp"  

            elif period_id == "#cur_shift":
                where += f''' and cp.mill_date = '{mill_date}' AND cp.mill_shift = '{mill_shift}' '''              
                table_name = "[ems_v1].[dbo].[current_power] cp" 

            elif period_id == "sel_shift":                  
                if from_date == '':
                    return JSONResponse({"iserror": True, "message": "from date is required"})
                if shift_id == '':
                    return JSONResponse({"iserror": True, "message": "shift_id is required"}) 
                
                from_date = parse_date(from_date)          
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                table_name=f"  {completed_db}[power_{month_year}] as cp" 

                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
                if groupby == 'machine':
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'polling_data_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    if len(result_query) > 0:
                        group_by_poll = ''
                        where_poll = ''
                        machine = ''

                        if machine_id != '' and machine_id != 'all':
                            machine_id = id(machine_id)
                            machine = f" and cpd.machine_id in ({machine_id})"

                        if report_for == 'detail' or report_for == '':
                            where_poll = ' and cpd.mill_date = cp.mill_date and cpd.mill_shift = cp.mill_shift'
                            group_by_poll = " ,cpd.mill_date , cpd.mill_shift"
                        
                        if report_for == 'summary':
                            where_poll = ' and cpd.mill_date = cp.mill_date'
                            group_by_poll = " ,cpd.mill_date"
                        if groupby == 'machine':
                            join = f'''left join (select 
                                        cpd.machine_id,
                                        min(cpd.mill_date) mill_date,
                                        min(cpd.mill_shift) mill_shift,
                                        SUM(CASE WHEN cpd.machine_status = 0 THEN cpd.poll_duration ELSE 0 END) AS off_time,
                                        SUM(CASE WHEN cpd.machine_status = 1 THEN cpd.poll_duration ELSE 0 END) AS idle_time,
                                        SUM(CASE WHEN cpd.machine_status = 2 THEN cpd.poll_duration ELSE 0 END) AS on_load_time
                                    from
                                        polling_data_{month_year} cpd
                                    where cpd.mill_date = '{from_date}' AND cpd.mill_shift = '{shift_id}' {machine} 
                                    group by cpd.machine_id {group_by_poll}) as cpd 
                                    on cpd.machine_id = cp.machine_id {where_poll}'''
                    else:
                        return JSONResponse({"iserror": True, "message": "polling_data not available..."})
                
                where += f''' and cp.mill_date = '{from_date}' AND cp.mill_shift = '{shift_id}' '''   

            elif period_id == "#sel_shift":                 
                if mill_shift == 1:
                    shift_id = no_of_shifts
                    from_date = parse_date(mill_date) - timedelta(days=1)

                else:
                    shift_id = int(mill_shift) - 1
                    from_date = mill_date                      
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                table_name=f"  {completed_db}[power_{month_year}] as cp" 

                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
                where += f''' and cp.mill_date = '{from_date}' AND cp.mill_shift = '{shift_id}' '''   
            
            elif period_id == "sel_date":            
                if from_date == '':
                    return JSONResponse({"iserror": True, "message": "from date is required"})    
                
                from_date = parse_date(from_date)
                
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                table_name=f"  {completed_db}[power_{month_year}] as cp "       
                where += f''' and cp.mill_date = '{from_date}' '''

                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."}) 
                   
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'polling_data_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                if len(result_query) > 0:
                    group_by_poll = ''
                    where_poll = ''
                    machine = ''

                    if machine_id != '' and machine_id != 'all':
                        machine_id = id(machine_id)
                        machine = f" and cpd.machine_id in ({machine_id})"
                    if report_for == 'detail' or report_for == '':
                        where_poll = ' and cpd.mill_date = cp.mill_date and cpd.mill_shift = cp.mill_shift'
                        group_by_poll = " ,cpd.mill_date , cpd.mill_shift"
                    
                    if report_for == 'summary':
                        where_poll = ' and cpd.mill_date = cp.mill_date'
                        group_by_poll = " ,cpd.mill_date"
                    if groupby == 'machine':
                        join = f'''left join (select 
                                    cpd.machine_id,
                                    min(cpd.mill_date) mill_date,
                                    min(cpd.mill_shift) mill_shift,
                                    SUM(CASE WHEN cpd.machine_status = 0 THEN cpd.poll_duration ELSE 0 END) AS off_time,
                                    SUM(CASE WHEN cpd.machine_status = 1 THEN cpd.poll_duration ELSE 0 END) AS idle_time,
                                    SUM(CASE WHEN cpd.machine_status = 2 THEN cpd.poll_duration ELSE 0 END) AS on_load_time
                                from
                                    polling_data_{month_year} cpd
                                where cpd.mill_date = '{from_date}' {machine} 
                                group by cpd.machine_id {group_by_poll}) as cpd 
                                on cpd.machine_id = cp.machine_id {where_poll}'''
                else:
                    return JSONResponse({"iserror": True, "message": "polling_data not available..."}) 
                
            elif period_id == "#sel_date":             
                from_date = mill_date - timedelta(days=1)
                # from_date = parse_date(from_date)
                
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
                table_name=f"  {completed_db}[power_{month_year}] as cp "
                where += f''' and cp.mill_date = '{from_date}' '''

                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
            elif period_id  == "#this_week":
                dt = mill_date
                from_date=dt-timedelta(dt.weekday()+1)
                to_date = mill_date

                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
                table_name=f"  {completed_db}[power_{month_year}] as cp "
                where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''

                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
            elif period_id == "#this_month":
                from_date = mill_date.replace(day=1)
                to_date = mill_date

                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
                table_name=f"  {completed_db}[power_{month_year}] as cp "
                where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''

                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
            elif period_id == "from_to":            
                if from_date == '':
                    return JSONResponse({"iserror": True, "message": "from date is required"})
                if to_date == '':
                    return JSONResponse({"iserror": True, "message": "to_date is required"})  
                        
                from_date = parse_date(from_date)
                to_date =  parse_date(to_date)
                month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
                where += f''' and cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''
                if shift_id != ""and shift_id !="all":                
                    where += f''' and cp.mill_shift = '{shift_id}' ''' 

                if from_date.month == to_date.month and from_date.year == to_date.year:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    print(query)
                    if len(result_query) == 0:
                        return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
                    table_name=f"  {completed_db}[power_{month_year}] as cp "
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'polling_data_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    if len(result_query) > 0:
                        group_by_poll = ''
                        where_poll = ''
                        machine = ''

                        if machine_id != '' and machine_id != 'all':
                            machine_id = id(machine_id)
                            machine = f" and cpd.machine_id in ({machine_id})"
                        if report_for == 'detail' or report_for == '':
                            where_poll = ' and cpd.mill_date = cp.mill_date and cpd.mill_shift = cp.mill_shift'
                            group_by_poll = " ,cpd.mill_date , cpd.mill_shift"
                        
                        if report_for == 'summary':
                            where_poll = ' and cpd.mill_date = cp.mill_date'
                            group_by_poll = " ,cpd.mill_date"
                        if groupby == 'machine':
                            join = f'''left join (select 
                                        cpd.machine_id,
                                        min(cpd.mill_date) mill_date,
                                        min(cpd.mill_shift) mill_shift,
                                        SUM(CASE WHEN cpd.machine_status = 0 THEN cpd.poll_duration ELSE 0 END) AS off_time,
                                        SUM(CASE WHEN cpd.machine_status = 1 THEN cpd.poll_duration ELSE 0 END) AS idle_time,
                                        SUM(CASE WHEN cpd.machine_status = 2 THEN cpd.poll_duration ELSE 0 END) AS on_load_time
                                    from
                                        polling_data_{month_year} cpd
                                    where cpd.mill_date = '{from_date}'  {machine}
                                    group by cpd.machine_id {group_by_poll}) as cpd 
                                    on cpd.machine_id = cp.machine_id {where_poll}'''
                    else:
                        return JSONResponse({"iserror": True, "message": "polling_data not available..."})    
                else:
                    field_name = 'power_id,company_id,branch_id,department_id,shed_id,machinetype_id,machine_id,design_id,beam_id,date_time,date_time1,mill_date,mill_shift,vln_avg,r_volt,y_volt,b_volt,vll_avg,ry_volt,yb_volt,br_volt,t_current,r_current,y_current,b_current,t_watts,r_watts,y_watts,b_watts,t_var,r_var,y_var,b_var,t_voltampere,r_voltampere,y_voltampere,b_voltampere,avg_powerfactor,r_powerfactor,y_powerfactor,b_powerfactor,powerfactor,kWh,kvah,kw,kvar,power_factor,kva,frequency,machine_status,status,created_on,created_by,modified_on,modified_by,machine_kWh,master_kwh,reverse_machine_kwh,reverse_master_kwh,reverse_kwh'
                    
                    from_month = from_date.month
                    to_month = to_date.month
                    # month_year_range = [
                    # f"{mill_month[month]}{str(from_date.year)}" for month in range(from_month, to_month + 1)
                    # ]
                    month_year_range = [
                        (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                    print(month_year_range)
                    union_queries = []
                    join = []

                    for month_year in month_year_range:
                        month_year_poll = ''
                        query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                        result_query = cnx.execute(query).mappings().all()
                        print(query)
                        if len(result_query) > 0:
                            table_name = f"[ems_v1_completed].[dbo].[power_{month_year}]"
                            union_queries.append(f"SELECT {field_name} FROM {table_name}")
                        
                        query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'polling_data_{month_year}'"""
                        result_query = cnx.execute(query).mappings().all()
                        if len(result_query) > 0:
                            print("result_query",result_query)
                            join_p = f"ems_v1_completed.dbo.polling_data_{month_year} "
                            join.append(f"select machine_status,poll_duration, mill_date, mill_shift, machine_id from {join_p}")

                    if len(union_queries) == 0:
                        return JSONResponse({"iserror": True, "message": "power table not available..."})    
                
                    if len(join) == 0:
                        return JSONResponse({"iserror": True, "message": "polling_data not available..."})    

                    subquery_union = " UNION ALL ".join(union_queries)
                    join = " UNION ALL ".join(join)
                    table_name = f"( {subquery_union}) cp"
                    month_year_poll = f"( {join})"    
                    group_by_poll = ''
                    where_poll = ''
                    machine = ''

                    if machine_id != '' and machine_id != 'all':
                        machine_id = id(machine_id)
                        machine = f" and cpd.machine_id in ({machine_id})"
                    if report_for == 'detail' or report_for == '':
                        where_poll = ' and cpd.mill_date = cp.mill_date and cpd.mill_shift = cp.mill_shift'
                        group_by_poll = " ,cpd.mill_date , cpd.mill_shift"
                    
                    if report_for == 'summary':
                        where_poll = ' and cpd.mill_date = cp.mill_date'
                        group_by_poll = " ,cpd.mill_date"
                    if groupby == 'machine':
                        join = f'''left join (select 
                                        cpd.machine_id,
                                        min(cpd.mill_date) mill_date,
                                        min(cpd.mill_shift) mill_shift,
                                        SUM(CASE WHEN cpd.machine_status = 0 THEN cpd.poll_duration ELSE 0 END) AS off_time,
                                        SUM(CASE WHEN cpd.machine_status = 1 THEN cpd.poll_duration ELSE 0 END) AS idle_time,
                                        SUM(CASE WHEN cpd.machine_status = 2 THEN cpd.poll_duration ELSE 0 END) AS on_load_time
                                    from
                                        {month_year_poll} cpd
                                    where cpd.mill_date = '{from_date}' {machine}  
                                    group by cpd.machine_id {group_by_poll}) as cpd 
                                    on cpd.machine_id = cp.machine_id {where_poll}'''
                    
            else:
                return JSONResponse({"iserror": True, "message": "invalid period id"}) 
            
        if limit_report_for == "exception" :
           
            if limit_exception_for == "kwh" or limit_exception_for == 'kWh':
                order_by += "kWh"
            if limit_order_by == "asc":
                order_by += " "+limit_order_by +","
            else:
                order_by += " "+limit_order_by +","

        if groupby != '' and groupby == "company":
            group_by += " mm.company_id "
            order_by += " mm.company_id "
            group_id = '''min(mc.company_id) AS group_id '''
            group_code = '''min(mc.company_code) AS group_code ,'''
            group_name = '''min(mc.company_name) AS group_name'''       
            
        if groupby !='' and groupby == "department":
            group_by += " mm.department_id "
            order_by += " min(md.department_order)"
            group_id = '''min(md.department_id) AS group_id '''
            group_code = '''min(md.department_code) AS group_code ,'''
            group_name = '''min(md.department_name) AS group_name'''        
            
        if groupby !='' and groupby == "shed":
            group_by += "  mm.shed_id "
            order_by += "  min(ms.shed_order)"
            group_id = ''' min(ms.shed_id) AS group_id '''
            group_code = ''' min(ms.shed_code) AS group_code ,'''
            group_name = ''' min(ms.shed_name) AS group_name'''
            
        if groupby !='' and groupby == "machinetype":
            group_by += " mm.machinetype_id"
            order_by += " min(mmt.machinetype_order)"
            group_id = '''min(mmt.machinetype_id) AS group_id '''
            group_code = '''min(mmt.machinetype_code) AS group_code ,'''
            group_name = '''min(mmt.machinetype_name) AS group_name'''
            
        if groupby !='' and groupby == "function":    
            order_by += " min(mf.function_order)"
            group_id = '''min(mf.function_id) AS group_id '''
            group_code = '''min(mf.function_code) AS group_code ,'''
            group_name = '''min(mf.function_name) AS group_name'''       
            if function_type !='':
                if function_type == 'function_1':
                    group_by += " mm.function_id" 
                else:
                    group_by += " mm.function2_id"     
            else:
                group_by += " mm.function_id"

            if is_function !="":
                group_by += " ,mm.machine_id"
                order_by += " ,min(mm.machine_order)"

        if groupby !='' and groupby == "converter":           
            group_by += " mm.converter_id"
            order_by += " mm.converter_id"
            group_id = '''min(mcd.converter_id) AS group_id '''
            group_code = ''
            group_name = '''min(mcd.converter_name) AS group_name'''

            if is_function !="":
                group_by += " ,mm.machine_id"
                order_by += " ,min(mm.machine_order)"
            
        if groupby !='' and groupby == "machine":             
            group_by += " mm.machine_id"
            group_by_poll += " mm.machine_id"
            order_by += " min(mm.machine_order)"
            group_id = '''min(mm.machine_id) AS group_id '''
            group_code = '''min(mm.machine_code) AS group_code ,'''
            group_name = '''min(mm.machine_name) AS group_name'''  
           
        if limit_operation_value !='' and limit_operation_value != '0':          
            order_by += " " +"OFFSET 0 ROWS FETCH NEXT"+' '+ limit_operation_value+" "+"ROWS ONLY"
    
        if is_critical == "yes" or is_critical == "no"  :
            where += f" and mm.major_nonmajor = '{is_critical}' "   

        data2 = '' 
        data3 = ''
        where_group_for = ""
        where_main = ''  
        # if group_for == "exception" and machine_id != 'all' and machine_id!= "":
        if group_for == "exception": 
            if groupby == "department":
                where_group_for += "and group_type = 'zone' " 
                if department_id != 'all' and department_id !='' and department_id != '0':
                    where_group_for += f"and type_id = '{department_id}'"

            elif groupby == "shed":
                where_group_for += "and group_type = 'area' "
                if shed_id != 'all' and shed_id !='' and shed_id != '0':
                    where_group_for += f"and type_id = '{shed_id}'"

            elif groupby == "machinetype":
                where_group_for += "and group_type = 'location' "
                if machinetype_id != 'all' and machinetype_id !='' and machinetype_id != '0':
                    where_group_for += f"and type_id = '{machinetype_id}'"

            elif groupby == "function":
                if function_type == 'function_1':
                    where_group_for += "and group_type = 'function_1' "

                elif function_type == 'function_2':
                    where_group_for += "and group_type = 'function_2' "
                
                else:
                    where_group_for += "and group_type = 'function' "

                if function_id != 'all' and function_id !='' and function_id != '0':
                    where_group_for += f"and type_id = '{function_id}'"

            elif groupby == 'machine' and is_main_meter == 'yes': # main meter list 
                if department_id != 'all' and department_id !='' and department_id != '0':
                    where_group_for += "and group_type = 'zone' " 
                    where_group_for += f"and type_id = '{department_id}'"
                    
                if shed_id != 'all' and shed_id !='' and shed_id != '0':
                    where_group_for += "and group_type = 'area' " 
                    where_group_for += f"and type_id = '{shed_id}'" 

                if machinetype_id != 'all' and machinetype_id !='' and machinetype_id != '0':
                    where_group_for += "and group_type = 'location' " 
                    where_group_for += f"and type_id = '{machinetype_id}'"

                if function_id != 'all' and function_id !='' and function_id != '0':
                    if function_type!= '':
                        if function_type == 'function_1':
                            where_group_for += "and group_type = 'function_1' " 
                        if function_type == 'function_2':
                            where_group_for += "and group_type = 'function_2' " 
                    where_group_for += f"and type_id = '{function_id}'"

            sql = text(f'''SELECT * FROM ems_v1.dbo.master_meter_group where status = 'active' {where_group_for} ''') 
            print(sql)
            data2 = cnx.execute(sql).fetchall()

            machine_id = []  
            type_id = []
            if len(data2) > 0:
                for record in data2:
                    machine_id.append(record["machine_id"]) 
                    type_id.append(record["type_id"]) 
                if len(machine_id)>0:   
                    where_main = f" and mm.machine_id in ({','.join(str(x) for x in machine_id)})"
               
            # where_m = ''
            # if type_id != [] and is_main_meter != 'yes':
            #     print("type_id",type_id)
            #     if groupby == "department":
            #         where_m = f"and department_id NOT IN ({','.join(str(x) for x in type_id)})"
            #         print(where_m)
            #     if groupby == "shed":
            #         where_m = f"and shed_id not in ({','.join(str(x) for x in type_id)})"

            #     if groupby == "machinetype":
            #         where_m = f"and machinetype_id not in ({','.join(str(x) for x in type_id)})"

            #     if groupby == "function":
            #         if function_type == 'function_2':
            #             where_m = f"and function2_id not in ({','.join(str(x) for x in type_id)})"
            #         else:
            #             where_m = f"and function_id not in ({','.join(str(x) for x in type_id)})"

            #     sql = text(f"select * from ems_v1.dbo.master_machine where status = 'active'{where_m} ")
            #     rec = cnx.execute(sql).mappings().all()
            
            #     if len(rec)>0:
            #         machine = []
            #         for row in rec:
            #             machine.append(row["machine_id"]) 
            #         where_mm = f" and mm.machine_id in ({','.join(str(x) for x in machine)})"

            #     if where_main != '' and where_mm != '': 
            #         where_main = f" and (mm.machine_id in ({','.join(str(x) for x in machine_id)}) or mm.machine_id in ({','.join(str(x) for x in machine)})) "
            #         where_mm = ''

            # if groupby == 'machine' and is_main_meter == 'yes' and where_main == '' and where_mm == '': 
            #     where += f" and mm.machine_id  = 0"

        if report_for == 'detail' or report_for == '':
            group_by = " cp.mill_date , cp.mill_shift," + group_by
            order_by = " cp.mill_date, cp.mill_shift," + order_by
        
        if report_for == 'summary':
            group_by = " cp.mill_date," + group_by
            order_by = " cp.mill_date," + order_by  
                       
        if group_by != "":
            group_by = f"group by {group_by} "    
        if order_by != "":
            order_by = f"order by {order_by}"

        poll_duration = '''
                            '' as off_time,
                            '' AS idle_time,
                            '' AS on_load_time,

                                            ''' 
        if reportfor != '12to12':
            if groupby == 'machine':
                if period_id == "cur_shift":
                    poll_duration = f'''CONVERT(TIME, DATEADD(SECOND,sum(cp.off_time),'00:00:00')) as off_time,
                                        CONVERT(TIME, DATEADD(SECOND,sum(cp.idle_time),'00:00:00')) as idle_time,
                                        CONVERT(TIME, DATEADD(SECOND,sum(cp.on_load_time),'00:00:00')) as on_load_time,'''
                
                else:
                    poll_duration = '''
                                CONVERT(TIME, DATEADD(SECOND,sum(cpd.off_time),'00:00:00')) as off_time,
                                CONVERT(TIME, DATEADD(SECOND,sum(cpd.idle_time),'00:00:00')) as idle_time,
                                CONVERT(TIME, DATEADD(SECOND,sum(cpd.on_load_time),'00:00:00')) as on_load_time,
                                '''

        
    #    ROUND(sum(CASE WHEN mm.energy_selection = 'wh' THEN CAST(cp.kWh AS DECIMAL(18, 2))/1000 ELSE cp.kWh END),2) AS kWh,
        query = text(f'''
                SELECT                       
                    min(mc.company_code) AS company_code,
                    min(mc.company_name) AS company_name,
                    min(mb.branch_code) AS branch_code,
                    min(mb.branch_name) AS branch_name,
                    min(md.department_code) AS department_code,
                    min(md.department_name) As department_name,
                    min(ms.shed_code) AS shed_code,
                    min(ms.shed_name) AS shed_name,
                    min(mmt.machinetype_code) AS machinetype_code,
                    min(mmt.machinetype_name) AS machinetype_name,
                    min(mf.function_name) AS function_name,
                    min(mf.function_code) AS function_code,
                    min(mm.machine_code) AS machine_code,
                    min(mm.machine_name) AS machine_name,
                    count(mm.machine_name) AS machine_count,
                    min(cp.power_id) as power_id,
                    min(mm.company_id) as company_id,
                    min(mm.branch_id) as branch_id,
                    min(md.department_id) as department_id,
                    min(mm.shed_id) as shed_id,
                    min(mm.machinetype_id) as machinetype_id,
                    min(mf.function_id) AS function_id,
                    min(mm.machine_id) as machine_id,
                    min(cp.design_id) as design_id,
                    min(cp.beam_id) as beam_id,
                    min(cp.date_time) as date_time,
                    min(cp.date_time1) as date_time1,
                    min(cp.mill_date) as mill_date,
                    min(cp.mill_shift) as mill_shift,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.vln_avg = '*' then cp.vln_avg * mmf.vln_avg_value when  mmf.vln_avg = '/' then cp.vln_avg / mmf.vln_avg_value else cp.vln_avg end ),2)  AS DECIMAL(18, 2)),0) AS vln_avg,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.r_volt = '*' then cp.r_volt * mmf.r_volt_value when  mmf.r_volt = '/' then cp.r_volt / mmf.r_volt_value else cp.r_volt end ),2)  AS DECIMAL(18, 2)),0) AS r_volt,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.y_volt = '*' then cp.y_volt * mmf.y_volt_value when  mmf.y_volt = '/' then cp.y_volt / mmf.y_volt_value else cp.y_volt end ),2)  AS DECIMAL(18, 2)),0) AS y_volt,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.b_volt = '*' then cp.b_volt * mmf.b_volt_value when  mmf.b_volt = '/' then cp.b_volt / mmf.b_volt_value else cp.b_volt end ),2)  AS DECIMAL(18, 2)),0) AS b_volt,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.vll_avg = '*' then cp.vll_avg * mmf.vll_avg_value when  mmf.vll_avg = '/' then cp.vll_avg / mmf.vll_avg_value else cp.vll_avg end ),2)  AS DECIMAL(18, 2)),0) AS vll_avg,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.ry_volt = '*' then cp.ry_volt * mmf.ry_volt_value when  mmf.ry_volt = '/' then cp.ry_volt / mmf.ry_volt_value else cp.ry_volt end ),2)  AS DECIMAL(18, 2)),0) AS ry_volt,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.yb_volt = '*' then cp.yb_volt * mmf.yb_volt_value when  mmf.yb_volt = '/' then cp.yb_volt / mmf.yb_volt_value else cp.yb_volt end ),2)  AS DECIMAL(18, 2)),0) AS yb_volt,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.br_volt = '*' then cp.br_volt * mmf.br_volt_value when  mmf.br_volt = '/' then cp.br_volt / mmf.br_volt_value else cp.br_volt end ),2)  AS DECIMAL(18, 2)),0) AS br_volt,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.r_current = '*' then cp.r_current * mmf.r_current_value when  mmf.r_current = '/' then cp.r_current / mmf.r_current_value else cp.r_current end ),2)  AS DECIMAL(18, 2)),0) AS r_current,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.y_current = '*' then cp.y_current * mmf.y_current_value when  mmf.y_current = '/' then cp.y_current / mmf.y_current_value else cp.y_current end ),2)  AS DECIMAL(18, 2)),0) AS y_current,
                    ISNULL(TRY_CAST(ROUND(AVG(CASE WHEN mmf.b_current = '*' THEN cp.b_current * mmf.b_current_value WHEN mmf.b_current = '/' THEN cp.b_current / mmf.b_current_value ELSE cp.b_current END ), 2)  AS DECIMAL(18, 2)),0) AS b_current,
                    ISNULL(TRY_CAST(ROUND(AVG(
                        CASE 
                            WHEN mmf.t_current = '*' THEN ((cp.r_current + cp.y_current + cp.b_current) / 3) * mmf.t_current_value 
                            WHEN mmf.t_current = '/' THEN ((cp.r_current + cp.y_current + cp.b_current) / 3) / mmf.t_current_value 
                            ELSE ((cp.r_current + cp.y_current + cp.b_current) / 3) 
                        END 
                    ), 2)  AS DECIMAL(18, 2)),0) AS avg_current,
                    ISNULL(TRY_CAST(ROUND(AVG(
                        CASE 
                            WHEN mmf.t_current = '*' THEN (cp.r_current + cp.y_current + cp.b_current) * mmf.t_current_value 
                            WHEN mmf.t_current = '/' THEN (cp.r_current + cp.y_current + cp.b_current) / mmf.t_current_value 
                            ELSE (cp.r_current + cp.y_current + cp.b_current) 
                        END 
                    ), 2)  AS DECIMAL(18, 2)),0) AS t_current,                    
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.t_watts = '*' then cp.t_watts * mmf.t_watts_value when  mmf.t_watts = '/' then cp.t_watts / mmf.t_watts_value else cp.t_watts end ),2)  AS DECIMAL(18, 2)),0) AS t_watts,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.r_watts = '*' then cp.r_watts * mmf.r_watts_value when  mmf.r_watts = '/' then cp.r_watts / mmf.r_watts_value else cp.r_watts end ),2)  AS DECIMAL(18, 2)),0) AS r_watts,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.y_watts = '*' then cp.y_watts * mmf.y_watts_value when  mmf.y_watts = '/' then cp.y_watts / mmf.y_watts_value else cp.y_watts end ),2)  AS DECIMAL(18, 2)),0) AS y_watts,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.b_watts = '*' then cp.b_watts * mmf.b_watts_value when  mmf.b_watts = '/' then cp.b_watts / mmf.b_watts_value else cp.b_watts end ),2)  AS DECIMAL(18, 2)),0) AS b_watts,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.t_var = '*' then cp.t_var * mmf.t_var_value when  mmf.t_var = '/' then cp.t_var / mmf.t_var_value else cp.t_var end ),2)  AS DECIMAL(18, 2)),0) AS t_var,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.r_var = '*' then cp.r_var * mmf.r_var_value when  mmf.r_var = '/' then cp.r_var / mmf.r_var_value else cp.r_var end ),2)  AS DECIMAL(18, 2)),0) AS r_var,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.y_var = '*' then cp.y_var * mmf.y_var_value when  mmf.y_var = '/' then cp.y_var / mmf.y_var_value else cp.y_var end ),2)  AS DECIMAL(18, 2)),0) AS y_var,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.b_var = '*' then cp.b_var * mmf.b_var_value when  mmf.b_var = '/' then cp.b_var / mmf.b_var_value else cp.b_var end ),2)  AS DECIMAL(18, 2)),0) AS b_var,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.t_voltampere = '*' then cp.t_voltampere * mmf.t_voltampere_value when  mmf.t_voltampere = '/' then cp.t_voltampere / mmf.t_voltampere_value else cp.t_voltampere end ),2)  AS DECIMAL(18, 2)),0) AS t_voltampere,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.r_voltampere = '*' then cp.r_voltampere * mmf.r_voltampere_value when  mmf.r_voltampere = '/' then cp.r_voltampere / mmf.r_voltampere_value else cp.r_voltampere end ),2)  AS DECIMAL(18, 2)),0) AS r_voltampere,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.y_voltampere = '*' then cp.y_voltampere * mmf.y_voltampere_value when  mmf.y_voltampere = '/' then cp.y_voltampere / mmf.y_voltampere_value else cp.y_voltampere end ),2)  AS DECIMAL(18, 2)),0) AS y_voltampere,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.b_voltampere = '*' then cp.b_voltampere * mmf.b_voltampere_value when  mmf.b_voltampere = '/' then cp.b_voltampere / mmf.b_voltampere_value else cp.b_voltampere end ),2)  AS DECIMAL(18, 2)),0) AS b_voltampere,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.avg_powerfactor = '*' then cp.avg_powerfactor * mmf.avg_powerfactor_value when  mmf.avg_powerfactor = '/' then cp.avg_powerfactor / mmf.avg_powerfactor_value else cp.avg_powerfactor end ),2)  AS DECIMAL(18, 2)),0) AS avg_powerfactor,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.r_powerfactor = '*' then cp.r_powerfactor * mmf.r_powerfactor_value when  mmf.r_powerfactor = '/' then cp.r_powerfactor / mmf.r_powerfactor_value else cp.r_powerfactor end ),2)  AS DECIMAL(18, 2)),0) AS r_powerfactor,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.y_powerfactor = '*' then cp.y_powerfactor * mmf.y_powerfactor_value when  mmf.y_powerfactor = '/' then cp.y_powerfactor / mmf.y_powerfactor_value else cp.y_powerfactor end ),2)  AS DECIMAL(18, 2)),0) AS y_powerfactor,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.b_powerfactor = '*' then cp.b_powerfactor * mmf.b_powerfactor_value when  mmf.b_powerfactor = '/' then cp.b_powerfactor / mmf.b_powerfactor_value else cp.b_powerfactor end ),2)  AS DECIMAL(18, 2)),0) AS b_powerfactor,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.powerfactor = '*' then cp.powerfactor * mmf.powerfactor_value when  mmf.powerfactor = '/' then cp.powerfactor / mmf.powerfactor_value else cp.powerfactor end ),2)  AS DECIMAL(18, 2)),0) AS powerfactor,
                    
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.kvah = '*' then cp.kvah * mmf.kvah_value when  mmf.kvah = '/' then cp.kvah / mmf.kvah_value else cp.kvah end ),2)  AS DECIMAL(18, 2)),0) AS kvah,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.kw = '*' then cp.t_watts * mmf.kw_value when  mmf.kw = '/' then cp.t_watts / mmf.kw_value else cp.t_watts end ),2)  AS DECIMAL(18, 2)),0) AS kw,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.kw = '*'  and cp.t_watts >0 then cp.t_watts * mmf.kw_value when  mmf.kw = '/' and cp.t_watts >0 then cp.t_watts / mmf.kw_value else 0 end ),2)  AS DECIMAL(18, 2)),0) AS import_kw,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.kw = '*'  and cp.t_watts <0 then cp.t_watts * mmf.kw_value when  mmf.kw = '/' and cp.t_watts <0 then cp.t_watts / mmf.kw_value else 0 end ),2)  AS DECIMAL(18, 2)),0) AS export_kw,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.kvar = '*' then cp.kvar * mmf.kvar_value when  mmf.kvar = '/' then cp.kvar / mmf.kvar_value else cp.kvar end ),2)  AS DECIMAL(18, 2)),0) AS kvar,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.power_factor = '*' then cp.power_factor * mmf.power_factor_value when  mmf.power_factor = '/' then cp.power_factor / mmf.power_factor_value else cp.power_factor end ),2)  AS DECIMAL(18, 2)),0) AS power_factor,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.kva = '*' then cp.kva * mmf.kva_value when  mmf.kva = '/' then cp.kva / mmf.kva_value else cp.kva end ),2)  AS DECIMAL(18, 2)),0) AS kva,
                    ISNULL(TRY_CAST(ROUND(AVG(case when mmf.frequency = '*' then cp.frequency * mmf.frequency_value when  mmf.frequency = '/' then cp.frequency / mmf.frequency_value else cp.frequency end ),2)  AS DECIMAL(18, 2)),0) AS frequency,
                    min(cp.machine_status) as machine_status,
                    min(cp.status) as status,
                    min(cp.created_on) as created_on,
                    min(cp.created_by) as created_by,
                    min(cp.modified_on) as modified_on,
                    min(cp.modified_by) as modified_by,
                    
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.machine_kWh = '*' then cp.machine_kWh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.machine_kWh / mmf.machine_kWh_value else cp.machine_kWh end ),0) AS DECIMAL(18, 2)),0) AS machine_kWh,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.machine_kWh = '*' then cp.master_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.master_kwh / mmf.machine_kWh_value else cp.master_kwh end ),0) AS DECIMAL(18, 2)),0) AS master_kwh,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),0) AS DECIMAL(18, 2)),0) AS kWh,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.machine_kWh = '*' then cp.reverse_machine_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.reverse_machine_kwh / mmf.machine_kWh_value else cp.reverse_machine_kwh end ),0) AS DECIMAL(18, 2)),0) AS reverse_machine_kwh,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.machine_kWh = '*' then cp.reverse_master_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.reverse_master_kwh / mmf.machine_kWh_value else cp.reverse_master_kwh end ),0) AS DECIMAL(18, 2)),0) AS reverse_master_kwh,
                    ISNULL(TRY_CAST(ROUND(SUM(case when mmf.kWh = '*' then cp.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then cp.reverse_kwh / mmf.kWh_value else cp.reverse_kwh end ),0) AS DECIMAL(18, 2)),0) AS reverse_kwh,
                    
                    min(mm.ip_address) as ip_address,
                    min(mm.port) as port,
                    CASE WHEN min(cp.date_time) <= DATEADD(minute, -2, getdate()) THEN 'S' ELSE 'N' END as nocom,       
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS kwh_1,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS kwh_2,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS kwh_3,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS start_kwh_1,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS start_kwh_2,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS start_kwh_3,     
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS end_kwh_1,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS end_kwh_2,
                    ISNULL(TRY_CAST(ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),0) AS DECIMAL(18, 2)),0) AS end_kwh_3,
                    {poll_duration}
                    {group_id},
                    {group_code}
                    {group_name}                       
                FROM 
                    {table_name}                       
                    INNER JOIN [ems_v1].[dbo].[master_machine] mm ON cp.machine_id = mm.machine_id
                    INNER JOIN [ems_v1].[dbo].[master_company] mc ON mm.company_id = mc.company_id
                    INNER JOIN [ems_v1].[dbo].[master_branch] mb ON mm.branch_id = mb.branch_id
                    INNER JOIN [ems_v1].[dbo].[master_department] md ON mm.department_id = md.department_id
                    INNER JOIN [ems_v1].[dbo].[master_shed] ms ON mm.shed_id = ms.shed_id
                    INNER JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mm.machinetype_id = mmt.machinetype_id 
                    LEFT JOIN [ems_v1].[dbo].[master_function] mf ON {function_where}
                    LEFT JOIN [ems_v1].[dbo].[master_converter_detail] mcd ON mm.converter_id = mcd.converter_id 
                    LEFT JOIN [ems_v1].[dbo].[master_machine_factor] mmf ON mm.machine_id = mmf.machine_id 
                    {join}                  
                WHERE  
                    cp.status = '0' and mm.status = 'active'
                    {where} {where_main}            
                    {group_by}
                    {order_by}
                ''') 
        createFolder("Current_power_log/","current_power api query "+str(query))
        data = cnx.execute(query).mappings().all()
        return JSONResponse({"iserror":False, "message":"data return successfully", "data" : jsonable_encoder(data)})    
    
    except Exception as e:
        traceback.print_exc()      
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Current_power_log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
       
@app.post("/alarm_report/")
async def alarm_report(company_id : int = Form(''),
                       branch_id : int = Form (''),
                       department_id : str = Form (''),
                       shed_id : str = Form (''),
                       machinetype_id : str = Form (''),
                       machine_id : str = Form (None),
                       report_for : str = Form(None),
                       period_id: str = Form(None),
                       from_date: str = Form(None),
                       to_date: str = Form(None),                      
                       shift_id: int = Form(None),  
                       employee_id : int = Form(''),                                         
                       order : str = Form(''),                                         
                       cnx: Session = Depends(get_db)):
    
    if period_id is None:
            return JSONResponse({"iserror": True, "message": "period id is required"})
    
    if  employee_id != '':
        query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
        res = cnx.execute(query).mappings().all()
        if len(res)>0:
            for row in res:
                department_id = row["department_id"]
                shed_id = row["shed_id"]
                machinetype_id = row["machinetype_id"]

    start_time = '1900-01-01 00:00:00'

    alarm_status = ''
    is_alarm_sound = ''
    is_alarm_status= ''

    try:    
        where = ""         
        query = text(f'''SELECT * FROM ems_v1.dbo.master_shifts WHERE status = 'active' ''')
        data1 = cnx.execute(query).fetchall()
        mill_date = date.today()
        mill_shift = 0        
    
        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]            
              print(mill_date)

        order_by = '' 
        if order != '':
            order_by += f' order by pa.start_time {order}'
        
        if period_id == "cur_shift":          
            where += f'''  pa.mill_date = '{mill_date}' AND pa.mill_shift = '{mill_shift}' ''' 
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time) '''
            
        elif period_id == "sel_shift":            
            if from_date is None:
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if shift_id is None:
                return JSONResponse({"iserror": True, "message": "shift_id is required"})                            
            where += f'''  pa.mill_date = '{parse_date(from_date)}' AND pa.mill_shift = '{shift_id}' '''
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time) ''' 

        elif period_id == "sel_date":            
            if from_date is None:
                 return JSONResponse({"iserror": True, "message": "from date is required"})                     
            where += f'''pa.mill_date = '{parse_date(from_date)}' '''
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time)'''
            
        elif period_id == "from_to":            
            if from_date is None:
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date is None:
                 return JSONResponse({"iserror": True, "message": "to_date is required"})            
                    
            where += f''' pa.mill_date  >= '{parse_date(from_date)}' and pa.mill_date <= '{parse_date(to_date)}' '''
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time) '''
            if shift_id is not None:                
                where += f''' and pa.mill_shift = '{shift_id}' ''' 
                       
        elif period_id == "live_alarm":

            if branch_id == '':
                return JSONResponse({"iserror": True, "message": "branch id is required"})
    
            if company_id == '':
                return JSONResponse({"iserror": True, "message": "company id is required"})
    
            where += f''' pa.start_time <> '1900-01-01 00:00:00'  and pa.stop_time is Null ''' 
            duration = f'''DATEDIFF(second, pa.start_time, getdate())''' 

            sql = text(f'''SELECT TOP 1 FORMAT(start_time, 'yyyy-MM-dd HH:mm:ss') as start_time FROM ems_v1.dbo.present_alarm ORDER BY start_time DESC''')
            data = cnx.execute(sql).fetchall()
            for i in data:
                start_time = i['start_time']
                createFolder("Alarm_Log/","Alarm Time "+start_time)
            

            sql1= text(f''' UPDATE ems_v1.dbo.master_company
            SET alarm_status = CASE WHEN alarm_last_time >= '{start_time}' THEN alarm_status ELSE 1 END,
                alarm_last_time = '{start_time}'
            WHERE company_id = '{company_id}'
        ''')
    
            cnx.execute(sql1)
            cnx.commit()

            query2=text(f'''select * from ems_v1.dbo.master_company where company_id = {company_id} and alarm_status = 1''')
            data2 = cnx.execute(query2).fetchall() 
            if len(data2)>0:
                for row in data2 :
                    alarm_status = row["alarm_status"]

            query3 = text(f" select * from ems_v1.dbo.master_branch where branch_id = {branch_id} ")
            data3 = cnx.execute(query3).mappings().all()
            if len(data3)>0:
                for i in data3:
                    is_alarm_sound = i["is_alarm_sound"]
                    is_alarm_status = i["is_alarm_status"]

            data2 = {"alarm_status": alarm_status,"is_alarm_sound":is_alarm_sound,"is_alarm_status": is_alarm_status}
            
        else:
             return JSONResponse({"iserror": True, "message": "invalid period id"})   
    
        if machine_id is not None and machine_id != 'all':
            where += f" and mm.machine_id = '{machine_id}' "   

        if company_id !='' and company_id !="0" and company_id != None:
            where += f" and mc.company_id = '{company_id}' "

        if department_id !='' and department_id !="0" and department_id != None:
            where += f" and md.department_id = '{department_id}' " 

        if shed_id !='' and  shed_id !="0" and shed_id != None:
            where += f" and ms.shed_id = '{shed_id}' "   

        if machinetype_id !='' and machinetype_id!="0" and machinetype_id != None:
            where += f" and mmt.machinetype_id = '{machinetype_id}' "  

        groupby = ""  
        if report_for == "summary":
            sql = text(f'''
                SELECT 
                    min(ma.alarm_name) as alarm_name,		    
                    min(pa.parameter_name) as parameter_name,
                    min(mm.machine_name) as machine_name,
                    sum({duration}) as duration
                ''')
            groupby = f'''group by ma.alarm_name '''
            
        else:
            sql = text(f'''
                SELECT 
                    mm.machine_code,
                    mm.machine_name,
                    ma.alarm_name,
                    ma.parameter_name,
                    pa.start_time,
                    pa.stop_time,
                    {duration} as duration,
                    pa.description,
                    mc.company_name,
                    mc.company_code,
                    mb.branch_name,
                    mb.branch_code,
                    md.department_name,
                    md.department_code,
                    ms.shed_name,
                    ms.shed_code,
                    mmt.machinetype_name,
                    mmt.machinetype_code

                ''')
        query = text(f''' 
            {sql}
            FROM 
                ems_v1.dbo.master_alarm_target ma 
                INNER JOIN ems_v1.dbo.present_alarm pa on pa.alarm_target_id = ma.alarm_target_id
                INNER JOIN ems_v1.dbo.master_machine mm on pa.machine_id = mm.machine_id
                LEFT JOIN [ems_v1].[dbo].[master_company] mc on mc.company_id=ma.company_id
                LEFT JOIN [ems_v1].[dbo].[master_branch] mb on mb.branch_id=ma.branch_id
                LEFT JOIN [ems_v1].[dbo].[master_department] md on md.department_id=ma.department_id
                LEFT JOIN [ems_v1].[dbo].[master_shed] ms on ms.shed_id=ma.shed_id
                LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt on mmt.machinetype_id=ma.machinetype_id
            WHERE  {where} 
            {groupby} 
            {order_by}
         ''')
        # print(query)
        data = cnx.execute(query).fetchall() 
                
        return JSONResponse({"iserror":False, "message":"data return successfully", "data" : jsonable_encoder(data), "data1":jsonable_encoder(data2)})    
    except Exception as e:            
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Alarm_Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/alarm_list/")
async def get_alarmlist(company_id: str=Form(''),
                        alarm_target_id: str=Form(None),
                        alarm_type : str = Form(None),
                        branch_id : str = Form(''),
                        department_id : str = Form(''),
                        shed_id : str = Form(''),
                        machinetype_id : str = Form(''),
                        cnx: Session = Depends(get_db)):
    try:       
        where = ""
        if alarm_target_id is not  None:
            where += f"and at.alarm_target_id = '{alarm_target_id}' "
            
        if alarm_type is not None:
            where += f" and at.alarm_type= '{alarm_type}' "

        if company_id !='' and company_id != "0" and company_id != None:
            where += f" and at.company_id= '{company_id}' "

        if branch_id !='' and branch_id !="0" and branch_id != None:
            where += f" and at.branch_id= '{branch_id}' "

        if department_id !='' and department_id!= "0" and department_id != None:
            where += f" and at.department_id= '{department_id}' "

        if shed_id !='' and shed_id != "0" and shed_id != None:
            where += f" and at.shed_id= '{shed_id}' "

        if machinetype_id !='' and machinetype_id !="0" and machinetype_id != None:
            where += f" and at.machinetype_id= '{machinetype_id}' "

        query=text(f''' 
                SELECT 
                    at.*, 
                    '' as machine_dtl,
                    mb.branch_code,
                    mb.branch_name,
                    md.department_code,
                    md.department_name,
                    ms.shed_code,
                    ms.shed_name,
                    mmt.machinetype_code,
                    mmt.machinetype_name,
                    ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                FROM 
                    ems_v1.dbo.master_alarm_target at
                    left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=at.created_by
	                left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=at.modified_by
                    left join [ems_v1].[dbo].[master_company] mc on mc.company_id=at.company_id
                    left join [ems_v1].[dbo].[master_branch] mb on mb.branch_id=at.branch_id
                    left join [ems_v1].[dbo].[master_department] md on md.department_id=at.department_id
                    left join [ems_v1].[dbo].[master_shed] ms on ms.shed_id=at.shed_id
                    left join [ems_v1].[dbo].[master_machinetype] mmt on mmt.machinetype_id=at.machinetype_id 
                WHERE 
                    at.status <> 'delete'
                    {where} 
                ''')
        
        # createFolder("Alarm_Log/","Issue in returning data "+str(query))
        data = cnx.execute(query).mappings().all()
        result = []
        for row in data:
            machine_id_list = row["machine_id"].strip(",").split(",")     
            machine_dtl = ""
            for machine_id in machine_id_list:                             
                sub_query = text(f"SELECT * FROM ems_v1.dbo.master_machine WHERE machine_id = {machine_id}")
                sub_data = cnx.execute(sub_query).mappings().all()
                for sub_row in sub_data:
                    if machine_dtl != "":
                        machine_dtl += '\n' 
                    machine_dtl += f'''{sub_row['machine_name']}''' 
                    print(machine_dtl)           
            new_row = dict(row)
            new_row["machine_dtl"] = machine_dtl
            result.append(new_row)            
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(result)})        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Alarm_Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/save_alarm_detail/")
async def save_alarm_detail(alarm_target_id:str=Form(None),
                            company_id:str=Form(''),
                            branch_id:str=Form(''),
                            department_id : str = Form(''),
                            shed_id : str = Form(''),
                            machinetype_id : str = Form(''),
                            parameter_name:str=Form(None),
                            machine_id:str=Form(''),
                            alarm_name:str=Form(None),
                            alarm_type:str=Form(None),
                            alarm_duration:int=Form(None),
                            conditions:int=Form(''),
                            color_1:str=Form(None),
                            color_2:str=Form(None),
                            color_3:str=Form(None),
                            login_id:str=Form(None),
                            cnx: Session = Depends(get_db)):
 
    # if machine_id == None:
    #     return JSONResponse({"iserror":True,"message":"machine id is required"}) 
    
    if company_id == '':
        return JSONResponse({"iserror":True,"message":"company id is required"}) 
    
    if branch_id == '':
        return JSONResponse({"iserror":True,"message":"branch_id is required"}) 
    
    # if department_id == '':
    #     return JSONResponse({"iserror":True,"message":"department_id is required"}) 
    
    # if shed_id == '':
    #     return JSONResponse({"iserror":True,"message":"shed_id is required"}) 
    
    # if machinetype_id == '':
    #     return JSONResponse({"iserror":True,"message":"machinetype_id is required"}) 
    
    if parameter_name == None:
        return JSONResponse({"iserror":True,"message":"parameter name is required"}) 
      
    if alarm_name == None:
        return JSONResponse({"iserror":True,"message":"alarm name is required"}) 
    
    if alarm_type == None:
        return JSONResponse({"iserror":True,"message":"alarm type is required"}) 
    
    if alarm_type == "time_based":
        if alarm_duration == None:
            return JSONResponse({"iserror":True,"message":"alarm duration is required"})
        if conditions == '':
            return JSONResponse({"iserror":True,"message":"conditions is required"})
        
        color_1 = 0
        color_2 = 0
        color_3 = 0
    else:
        alarm_duration = 0
        
        if color_1 == None:
            return JSONResponse({"iserror":True,"message":"color_1 is required"})
        
        if color_2 == None:
            return JSONResponse({"iserror":True,"message":"color_2 is required"})
        
        if color_3 == None:
            return JSONResponse({"iserror":True,"message":"color_3 is required"})
        
    try: 
        where =''
        if company_id!='' and company_id!= 0 :
            where += f'where mm.company_id = {company_id}'
        
        if branch_id!='' and company_id!= 0:
            where += f'and mm.branch_id = {branch_id}'
        
        if department_id!='' and company_id!= 0:
            where += f'and mm.department_id = {department_id}'
        
        if shed_id!='' and company_id!= 0:
            where += f'and mm.shed_id = {shed_id}'
        
        if machinetype_id!='' and company_id!= 0:
            where += f'and mm.machinetype_id = {machinetype_id}'

        if machine_id == '' or machine_id == 'all':
            query = text(f'''
                        select 
                            DISTINCT mm.machine_id 
                        from 
                            ems_v1.dbo.master_machine mm
                        {where}          
            ''') 
            data=cnx.execute(query).mappings().all()

            machine_id = []  
            if len(data) > 0:
                for record in data:
                    machine_id.append(str(record["machine_id"]))  
            machine_id = ",".join(machine_id)  
        print(machine_id)

        if machine_id is not None and machine_id != '':
            value = machine_id.split(",")
            if len(value) > 1:
                values = ",".join(value)  
                machine_id = f",{values},"  
            else:
                machine_id = f",{value[0]},"
          

        if alarm_target_id is not None:
            query =text(f'''UPDATE  ems_v1.dbo.master_alarm_target SET machine_id='{machine_id}',parameter_name='{parameter_name}',
                       alarm_name='{alarm_name}',color_1='{color_1}',color_2='{color_2}',color_3='{color_3}',
                       modified_on = GETDATE(),modified_by='{login_id}', alarm_duration = {alarm_duration},alarm_type = '{alarm_type}',company_id = {company_id},
                       department_id = '{department_id}', shed_id = '{shed_id}', machinetype_id = '{machinetype_id}',branch_id='{branch_id}', conditions = '{conditions}'
                       where alarm_target_id = '{alarm_target_id}'   
                       ''')
            cnx.execute(query)
            cnx.commit()

        else:
            
            select_query = text(f'''select * from ems_v1.dbo.master_alarm_target where alarm_name = '{alarm_name}' and status != 'delete' ''')
            data1 = cnx.execute(select_query).mappings().all()

            if len(data1)>0:
                return JSONResponse({"iserror":True,"message":"alarm name already exists "})
            
            query= text(f'''INSERT INTO ems_v1.dbo.master_alarm_target (machine_id,parameter_name,alarm_name,color_1,color_2,color_3,
                       created_on,created_by, alarm_duration, alarm_type ,company_id,department_id, shed_id, machinetype_id,branch_id,conditions)
                       VALUES ('{machine_id}','{parameter_name}','{alarm_name}','{color_1}', '{color_2}','{color_3}',
                       GETDATE(),'{login_id}', '{alarm_duration}', '{alarm_type}',{company_id},'{department_id}','{shed_id}','{machinetype_id}','{branch_id}','{conditions}') ''')
            cnx.execute(query)
            cnx.commit()

        return JSONResponse({"iserror":False,"message":"data save successfully","data":""})
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Alarm_Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})    
    
@app.post("/remove_alarm_detail/")
async def remove_alarm_detail(alarm_target_id:str=Form(None),
                              status : str = Form(None),
                              cnx: Session = Depends(get_db)):

    try:
        if alarm_target_id is not None:
            if status is not None:
                query = text(f''' UPDATE [ems_v1].[dbo].[master_alarm_target] SET status = '{status}' where alarm_target_id = '{alarm_target_id}' ''' )           
                cnx.execute(query)
                cnx.commit()
                return JSONResponse({"iserror":False,"message":" status updated successfully","data":""}) 
            else: 
                query = text(f''' UPDATE [ems_v1].[dbo].[master_alarm_target] SET status = 'delete' where alarm_target_id = '{alarm_target_id}' ''' )           
                cnx.execute(query)
                cnx.commit()
        
                createFolder("Alarm_Log/","Query executed successfully for remove alarm")
                return JSONResponse({"iserror":False,"message":"deleted successfully","data":""}) 
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Alarm_Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/get_power_report_name/")
async def get_power_report_name(cnx: Session = Depends(get_db)):

    try:        
        query= text(f'''select * from [ems_v1].[dbo].[power_report] where status = 'active' ''')
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/get_power_report_fields/")
async def get_power_report_fields(company_id : int = Form(None),
                                  report_id:int=Form(None),
                                  cnx: Session = Depends(get_db)):
    if report_id == None:
        return JSONResponse({"iserror":True,"message":"report_id is required"}) 
    
    try:
        if report_id is not None:
            if report_id == 0:
                query = text(f'''
                            SELECT 
                                min(report_field_id) as report_field_id,
                                min(report_id) as report_id,
                                min(field_code) as field_code,
                                min(field_name) as field_name,
                                min(is_show) as is_show,
                                min(slno) as slno,
                                min(field_name_display) as field_name_display,
                                min(unit) as unit,
                                min(company_id) as company_id
                            FROM 
                                [ems_v1].[dbo].[power_report_fields_original] 
                            WHERE company_id = {company_id} 
                            group by 
                                field_code 
                            order by
                                 slno
                             ''')

            else:
                query = text(f'''SELECT * FROM [ems_v1].[dbo].[power_report_fields_original] where report_id = {report_id} and company_id = {company_id} order by slno''')

        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)}) 
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})  
    
# [{"field_name_display": "Y_Volts", "report_field_id": "10", "slno": "4", "field_code": "y_volt"}]
@app.post("/update_power_report_fields/")
async def update_power_report_fields(company_id : int = Form(None),
                                     report_id:int=Form(None),
                                     obj: str = Form(None),
                                     cnx: Session = Depends(get_db)):
    print(obj)
    if company_id == None:
        return JSONResponse({"iserror":True,"message":"company_id is required"}) 
    
    query = text(f'''UPDATE [ems_v1].[dbo].[power_report_fields_original] 
                                SET is_show = 'no' 
                                WHERE company_id = {company_id} and report_id = '{report_id}' ''')
    print(query)
    cnx.execute(query)
    cnx.commit()  

    try:
        if obj is not None:
            obj_data = json.loads(obj)
            for row in obj_data:
                field_name_display = row["field_name_display"]
                report_field_id = row["report_field_id"]
                slno = row["slno"]
                field_code = row["field_code"]
                unit = row["field_unit"]
                
                if report_id == 0:                       
                    
                    sql = text(f'''UPDATE [ems_v1].[dbo].[power_report_fields_original] 
                                   SET field_name_display = '{field_name_display}', slno = '{slno}', unit = '{unit}'
                                   WHERE company_id = {company_id} and field_code = '{field_code}' ''')
                
                else:
                    sql = text(f'''UPDATE [ems_v1].[dbo].[power_report_fields_original] 
                                   SET is_show = 'yes', field_name_display = '{field_name_display}', slno = '{slno}' 
                                   WHERE company_id = {company_id} and report_field_id = {report_field_id} ''')
                    print(sql)
                    
                cnx.execute(sql)
                cnx.commit()
        return JSONResponse({"iserror":False,"message":"data save successfully","data":''}) 
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})    
    
@app.post("/communication_status/")
async def communication_status(machinetype_id:int=Form(None),
                               cnx: Session = Depends(get_db)):
    
    try:
        query = text(f'''
                    select
                        min(c.converter_id) as converter_id,
                        min(c.converter_name) as converter_name
                    from
                        ems_v1.dbo.master_machine mm,
                        ems_v1.dbo.master_converter_detail c
                    where
                        mm.converter_id = c.converter_id and mm.machinetype_id = '{machinetype_id}'
                        group by c.converter_id
                     ''')
        print(query)
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror":False,"message":"data retrun successfully","data":jsonable_encoder(data)}) 
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/model/")
async def model_list(cnx: Session = Depends(get_db)):

    try:        
        query= text(f'''
                     SELECT
                        mm.*,
                        ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                        ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user

                    FROM
                        [ems_v1].[dbo].[master_model] mm

                        left join ems_v1.dbo.master_employee cu on cu.employee_id=mm.created_by
                        left join ems_v1.dbo.master_employee mu on mu.employee_id=mm.modified_by
                    WHERE
                        mm.status != 'delete'
                    ''')
        print(query)
        data = cnx.execute(query).mappings().all()
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/save_model/")
async def save_model(model_id :int =Form(None),
                     model_name : str = Form(None),
                     user_login_id : str = Form(None),
                     cnx: Session = Depends(get_db)):
    try:
        
        if model_id is None:
            select_query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_model] WHERE model_name = '{model_name}' and status != 'delete' ''')
            data1 = cnx.execute(select_query).mappings().all()

            if len(data1)>0:
                return JSONResponse({"iserror":True,"message":"model name already exists "})
            
            query = text(f'''
                        INSERT INTO [ems_v1].[dbo].[master_model] (model_name, created_on, created_by)
                        VALUES ('{model_name}', GETDATE() , '{user_login_id}')
                        ''')
        else:
            
            query = text(f'''
                        UPDATE [ems_v1].[dbo].[master_model] SET model_name = '{model_name}', modified_on = GETDATE(),
                        modified_by = '{user_login_id}'
                        WHERE model_id = {model_id}
                        ''')
        cnx.execute(query)
        cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"data save successfully","data":""})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/remove_model/")
async def remove_model(model_id :int =Form(None),
                       status : str = Form(None),
                       cnx: Session = Depends(get_db)):
    try:
        if model_id is not None: 
            if status is not  None:    
                query = text(f'''
                            UPDATE [ems_v1].[dbo].[master_model] SET status = '{status}'
                            WHERE model_id = {model_id}
                            ''')
            else:
                query = text(f'''
                            UPDATE [ems_v1].[dbo].[master_model] SET status = 'delete'
                            WHERE model_id = {model_id}
                            ''')
            cnx.execute(query)
            cnx.commit()
        
        return JSONResponse({"iserror":False,"message":"status update successfully","data":""})   
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/employeelist_userrights/")
async def employeelistuser(employee_id:str=Form(None),
                           is_login:str=Form(None),
                           cnx: Session = Depends(get_db)):    
    
    if is_login == None:
        return JSONResponse({"iserror":True,"message":" is login is required"})
    
    try: 
        where = ""
        if employee_id is not None:
            where = text(f"and employee_id = '{employee_id}' ")

        query=text(f'''SELECT * FROM  [ems_v1].[dbo].[master_employee] WHERE status='active' and employee_type <> 'admin' {where} ''')
        data1 = cnx.execute(query).mappings().all()
        createFolder("Log/","Query executed successfully for  user rights employee list")
        return JSONResponse({"iserror":False,"message":"data return succesfully","data":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/user_menu_list/")
async def employeelistuser(employee_id:str=Form(None),
                           cnx: Session = Depends(get_db)):
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})
    
    try:
        query1=text(f''' 
                    SELECT 
                        ms.*,
                        ISNULL (u.id,0) AS u_r_id,
                        ISNULL (u.add_op,'')AS add_opp,
                        ISNULL (u.edit_op,'')AS edit_opp,
                        ISNULL (u.delete_op,'')AS delete_opp
                    FROM
                        [ems_v1].[dbo].[menu_mas] ms
                        LEFT JOIN 
                        (select * from [ems_v1].[dbo].[user_rights] where userid={employee_id}) As u
                        ON u.menu_id=ms.menu_id
                        WHERE ms.status='active' 
                        ORDER BY ms.slno
			  ''')
        print(query1)  
        data1 = cnx.execute(query1).mappings().all()
        return JSONResponse({"iserror":False,"message":"data return successfully","data1":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
# menu='[{"menu_id":1,"add_op": "yes","edit_op": "yes","delete_op": "yes"},{"menu_id":2,"add_op": "yes","edit_op": "yes","delete_op": "yes"}]'
@app.post("/save_userrights_detail/")
async def save_userrights_detail(employee_id:str=Form(None),
                                 menu:str=Form(None),
                                 cnx: Session = Depends(get_db)):
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"user id is required"}) 
    
    if menu == None:
        return JSONResponse({"iserror":True,"message":"menu is required"})
    
    try:

        if employee_id is not None:
                  del_query=text(f'''DELETE FROM [ems_v1].[dbo].[user_rights] WHERE userid='{employee_id}' ''')
                  cnx.execute(del_query)
                  cnx.commit()

                  user_dict = json.loads(menu)
                  for i in user_dict:
                        menu_id=i['menu_id']
                        add_op = i['add_op']
                        edit_op = i['edit_op']
                        delete_op=i['delete_op']
                        query=text(f'''insert into [ems_v1].[dbo].[user_rights](menu_id,add_op,edit_op,delete_op,userid)
                                values('{menu_id}','{add_op}','{edit_op}','{delete_op}','{employee_id}') ''')        
                        cnx.execute(query)
                        cnx.commit()    
        return JSONResponse({"iserror":False,"message":"Data Saved Succesfully","data":""})
    
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/sidebar_list/")
async def sidebarlistuser(employee_id:str=Form(None),
                           cnx: Session = Depends(get_db)):
    
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})
    
    try:
        query=text(f''' select * from [ems_v1].[dbo].[master_employee] where employee_id={employee_id}''')
        data=cnx.execute(query).mappings().all()
        print(data)
        
        if len(data) > 0 :
            for record in data:
                employee_type=record['employee_type']

        if employee_type == 'admin':
            query1=text(f'''
                        select * 
                        from [ems_v1].[dbo].[menu_mas]  
                        where status='active' 
                        order by slno
                        ''')
            # print(query1)
        else:
            query1=text(f''' SELECT 
                            ms.*,
                            ISNULL(u.id, 0) AS u_r_id,
                            ISNULL(u.add_op, '') AS add_opp,
                            ISNULL(u.edit_op, '') AS edit_opp,
                            ISNULL(u.delete_op, '') AS delete_opp
                        FROM
                            [ems_v1].[dbo].[menu_mas] ms,
                            [ems_v1].[dbo].[user_rights] u
                        WHERE
                            ms.status = 'active'
                            AND ms.menu_id = u.menu_id
                            AND u.userid = {employee_id}
                        ORDER BY ms.slno
                            
			  ''')
        # print(query1)  
        data1 = cnx.execute(query1).mappings().all()
        # createFolder("Log/","Query executed successfully for  sidebar list")
        return JSONResponse({"iserror":False,"message":"data return successfully","data1":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/operations/")
async def operations(employee_id:str=Form(None),
                     menu_id:str=Form(None),
                     cnx: Session = Depends(get_db)):
    
    if employee_id == None:
        return JSONResponse({"iserror":True,"message":"employee id is required"})
    
    where=""
    if menu_id is not None:
        where+=f''' and menu_id='{menu_id}' '''
    
    try:
        query=text(f'''SELECT 
                        u.*,
                        e.employee_type
                  FROM 
                        [ems_v1].[dbo].[user_rights] u
                        inner join [ems_v1].[dbo].[master_employee] e on e.employee_id=u.userid
                  WHERE
                        u.userid={employee_id} {where}
                        
        ''')
        data1 = cnx.execute(query).mappings().all()
        # createFolder("Log/","Query executed successfully for  sidebar list")
        return JSONResponse({"iserror":False,"message":"data return successfully","data1":jsonable_encoder(data1)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/load_analysis/")
async def load_analysis(period_id: str = Form(''),
                        machine_id: str=Form(''), 
                        from_date: str = Form(''),
                        to_date: str = Form(''),
                        shift_id :int = Form(''),
                        from_time: str=Form(''), 
                        to_time: str=Form(''),  
                        duration :int = Form(''),                      
                        cnx: Session = Depends(get_db)):
    
    if machine_id == '':
        return JSONResponse({"iserror":True,"message":"machine id is required"}) 
    
    if period_id == '':
        return JSONResponse({"iserror":True,"message":"period_id is required"}) 
    
    mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}    
    
    try: 
        where = ''
        table_name = ''
        if duration == "":
            duration = 1
        
        if period_id == 'cur_shift': 
            query=text(f'''SELECT * FROM [ems_v1].[dbo].[master_shifts] WHERE status='active' ''')
            data1 = cnx.execute(query).mappings().all()
            mill_date = date.today()
            mill_shift = 0       
            if len(data1) > 0:
                for shift_record in data1:
                    mill_date = shift_record["mill_date"]
                    mill_shift = shift_record["mill_shift"]  
                        
            table_name = 'ems_v1.dbo.current_power_analysis cp'
            where += f"cp.mill_date = '{mill_date}' and cp.mill_shift ='{mill_shift}' "

        elif period_id == 'sel_shift' or period_id == 'sel_date':
            if from_date == '':
                return JSONResponse({"iserror":True,"message":"date is required"}) 
            
            mill_date=parse_date(from_date)             
            month_year=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""
            table_name=f"[ems_v1_completed].[dbo].[power_analysis_{month_year}]" 
            where += f"cp.mill_date = '{mill_date}' "

            field_name = 'id,machine_id, created_on, mill_date, mill_shift, t_current, r_current, y_current, b_current, vll_avg, ry_volt, yb_volt, br_volt, vln_avg, r_volt, y_volt, b_volt, t_watts, kWh, kvah, kw, kvar, power_factor, r_watts, kva, y_watts, b_watts, avg_powerfactor, r_powerfactor, y_powerfactor, b_powerfactor, powerfactor, kwh_actual, frequency, t_voltampere, r_voltampere, y_voltampere, b_voltampere, t_var, r_var, y_var, b_var, master_kwh, machine_kwh'
            table_name = f'(select {field_name} from [ems_v1].[dbo].[current_power_analysis] UNION All select {field_name} from {table_name})cp'

            if period_id == 'sel_shift':
                if shift_id == '':
                    return JSONResponse({"iserror":True,"message":"shift is required"}) 
                where += f" and cp.mill_shift ='{shift_id}' " 
            if month_year !='':
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_analysis_{month_year}'"""
                print(query)
                result_query = cnx.execute(query).mappings().all()
                if len(result_query)>0:
                    pass
                else:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    
   
        elif period_id == "from_to":            
            if from_date == '':
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date == '':
                return JSONResponse({"iserror": True, "message": "to_date is required"})  
                    
            from_date = parse_date(from_date)
            to_date =  parse_date(to_date)
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
        
            where += f'''  cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''
            
            if shift_id != "":                
                where += f''' and cp.mill_shift = '{shift_id}' ''' 
            field_name = 'id,machine_id, created_on, mill_date, mill_shift, t_current, r_current, y_current, b_current, vll_avg, ry_volt, yb_volt, br_volt, vln_avg, r_volt, y_volt, b_volt, t_watts, kWh, kvah, kw, kvar, power_factor, r_watts, kva, y_watts, b_watts, avg_powerfactor, r_powerfactor, y_powerfactor, b_powerfactor, powerfactor, kwh_actual, frequency, t_voltampere, r_voltampere, y_voltampere, b_voltampere, t_var, r_var, y_var, b_var, master_kwh, machine_kwh'
            
            if from_date.month == to_date.month:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_analysis_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    
       
                table_name=f"[ems_v1_completed].[dbo].[power_analysis_{month_year}]" 
                table_name = f'(select {field_name} from [ems_v1].[dbo].[current_power_analysis] UNION All select {field_name} from {table_name})cp'
            else:
                from_month = from_date.month
                to_month = to_date.month
                month_year_range = [
                        (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                union_queries = []

                for month_year in month_year_range:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_analysis_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    print(query)
                    if len(result_query) > 0:
                        table_name = f"[ems_v1_completed].[dbo].[power_analysis_{month_year}]"
                        union_queries.append(f"SELECT {field_name} FROM {table_name}")

                if len(union_queries) == 0:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    

                subquery_union = " UNION ALL ".join(union_queries)
                table_name = f"(SELECT {field_name} FROM [ems_v1].[dbo].[current_power_analysis] UNION ALL {subquery_union}) cp"

        if from_time !='':
            where += f" and FORMAT(cp.created_on ,'HH:mm:ss')>='{from_time}' "
    
        if to_time !='':
            where += f" and FORMAT(cp.created_on ,'HH:mm:ss')<='{to_time}' "
                
        query=text(f'''
            SELECT *
            FROM (
                SELECT 
                (ROW_NUMBER() OVER (ORDER BY cp.machine_id) - 1) % {duration} + 1 AS slno,
			    mm.machine_code,
			    mm.machine_name,
			    cp.machine_id,
                FORMAT(cp.created_on, 'yyyy-MM-ddTHH:mm:ss') as date_time,
			    cp.mill_date,
			    cp.mill_shift,
			    ROUND((case when mmf.t_current = '*' then cp.t_current * mmf.t_current_value  when mmf.t_current = '/' then cp.t_current / mmf.t_current_value else cp.t_current end),2) AS t_current,
                ROUND((case when mmf.r_current = '*' then cp.r_current * mmf.r_current_value  when mmf.r_current = '/' then cp.r_current / mmf.r_current_value else cp.r_current end),2) AS r_current,
			    ROUND((case when mmf.y_current = '*' then cp.y_current * mmf.y_current_value  when mmf.y_current = '/' then cp.y_current / mmf.y_current_value else cp.y_current end),2) AS y_current,
			    ROUND((case when mmf.b_current = '*' then cp.b_current * mmf.b_current_value  when mmf.b_current = '/' then cp.b_current / mmf.b_current_value else cp.b_current end),2) AS b_current,
			    ROUND((case when mmf.vll_avg = '*' then cp.vll_avg * mmf.vll_avg_value  when mmf.vll_avg = '/' then cp.vll_avg / mmf.vll_avg_value else cp.vll_avg end),2) AS vll_avg,
			    ROUND((case when mmf.ry_volt = '*' then cp.ry_volt * mmf.ry_volt_value  when mmf.ry_volt = '/' then cp.ry_volt / mmf.ry_volt_value else cp.ry_volt end),2) AS ry_volt,
			    ROUND((case when mmf.yb_volt = '*' then cp.yb_volt * mmf.yb_volt_value  when mmf.yb_volt = '/' then cp.yb_volt / mmf.yb_volt_value else cp.yb_volt end),2) AS yb_volt,
			    ROUND((case when mmf.br_volt = '*' then cp.br_volt * mmf.br_volt_value  when mmf.br_volt = '/' then cp.br_volt / mmf.br_volt_value else cp.br_volt end),2) AS br_volt,
			    ROUND((case when mmf.vln_avg = '*' then cp.vln_avg * mmf.vln_avg_value  when mmf.vln_avg = '/' then cp.vln_avg / mmf.vln_avg_value else cp.vln_avg end),2) AS vln_avg,
			    ROUND((case when mmf.r_volt = '*' then cp.r_volt * mmf.r_volt_value  when mmf.r_volt = '/' then cp.r_volt / mmf.r_volt_value else cp.r_volt end),2) AS r_volt,
			    ROUND((case when mmf.y_volt = '*' then cp.y_volt * mmf.y_volt_value  when mmf.y_volt = '/' then cp.y_volt / mmf.y_volt_value else cp.y_volt end),2) AS y_volt,
			    ROUND((case when mmf.b_volt = '*' then cp.b_volt * mmf.b_volt_value  when mmf.b_volt = '/' then cp.b_volt / mmf.b_volt_value else cp.b_volt end),2) AS b_volt,
			    ROUND((case when mmf.t_watts = '*' then cp.t_watts * mmf.t_watts_value  when mmf.t_watts = '/' then cp.t_watts / mmf.t_watts_value else cp.t_watts end),2) AS t_watts,
                ROUND((case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end),2) AS kWh,
			    ROUND((case when mmf.kvah = '*' then cp.kvah * mmf.kvah_value  when mmf.kvah = '/' then cp.kvah / mmf.kvah_value else cp.kvah end),2) AS kvah,
			    ROUND((case when mmf.kw = '*' then cp.t_watts * mmf.kw_value  when mmf.kw = '/' then cp.t_watts / mmf.kw_value else cp.t_watts end),2)  AS kw,
			    ROUND((case when mmf.kw = '*' and cp.t_watts >0 then cp.t_watts * mmf.kw_value  when mmf.kw = '/' and cp.t_watts >0 then cp.t_watts / mmf.kw_value else 0 end),2)  AS import_kw,
			    ROUND((case when mmf.kw = '*' and cp.t_watts <0 then cp.t_watts * mmf.kw_value  when mmf.kw = '/' and cp.t_watts <0 then cp.t_watts / mmf.kw_value else 0 end),2)  AS export_kw,
			    ROUND((case when mmf.kvar = '*' then cp.kvar * mmf.kvar_value  when mmf.kvar = '/' then cp.kvar / mmf.kvar_value else cp.kvar end),2) AS kvar,
			    ROUND((case when mmf.power_factor = '*' then cp.power_factor * mmf.power_factor_value  when mmf.power_factor = '/' then cp.power_factor / mmf.power_factor_value else cp.power_factor end),2) AS power_factor,
			    ROUND((case when mmf.r_watts = '*' then cp.r_watts * mmf.r_watts_value  when mmf.r_watts = '/' then cp.r_watts / mmf.r_watts_value else cp.r_watts end),2) AS r_watts,
			    ROUND((case when mmf.kva = '*' then cp.kva * mmf.kva_value  when mmf.kva = '/' then cp.kva / mmf.kva_value else cp.kva end),2) AS kva,
			    ROUND((case when mmf.y_watts = '*' then cp.y_watts * mmf.y_watts_value  when mmf.y_watts = '/' then cp.y_watts / mmf.y_watts_value else cp.y_watts end),2) AS y_watts,
			    ROUND((case when mmf.b_watts = '*' then cp.b_watts * mmf.b_watts_value  when mmf.b_watts = '/' then cp.b_watts / mmf.b_watts_value else cp.b_watts end),2) AS b_watts,
			    ROUND((case when mmf.avg_powerfactor = '*' then cp.avg_powerfactor * mmf.avg_powerfactor_value  when mmf.avg_powerfactor = '/' then cp.avg_powerfactor / mmf.avg_powerfactor_value else cp.avg_powerfactor end),2) AS avg_powerfactor,
			    ROUND((case when mmf.r_powerfactor = '*' then cp.r_powerfactor * mmf.r_powerfactor_value  when mmf.r_powerfactor = '/' then cp.r_powerfactor / mmf.r_powerfactor_value else cp.r_powerfactor end),2) AS r_powerfactor,
			    ROUND((case when mmf.y_powerfactor = '*' then cp.y_powerfactor * mmf.y_powerfactor_value  when mmf.y_powerfactor = '/' then cp.y_powerfactor / mmf.y_powerfactor_value else cp.y_powerfactor end),2) AS y_powerfactor,
			    ROUND((case when mmf.b_powerfactor = '*' then cp.b_powerfactor * mmf.b_powerfactor_value  when mmf.b_powerfactor = '/' then cp.b_powerfactor / mmf.b_powerfactor_value else cp.b_powerfactor end),2) AS b_powerfactor,
			    ROUND((case when mmf.powerfactor = '*' then cp.powerfactor * mmf.powerfactor_value  when mmf.powerfactor = '/' then cp.powerfactor / mmf.powerfactor_value else cp.powerfactor end),2) AS powerfactor,
			    ROUND((case when mmf.kWh = '*' then cp.kwh_actual * mmf.kWh_value  when mmf.kWh = '/' then cp.kwh_actual / mmf.kWh_value else cp.kwh_actual end),2) AS kwh_actual,
			    ROUND((case when mmf.frequency = '*' then cp.frequency * mmf.frequency_value  when mmf.frequency = '/' then cp.frequency / mmf.frequency_value else cp.frequency end),2) AS frequency,
			    ROUND((case when mmf.t_voltampere = '*' then cp.t_voltampere * mmf.t_voltampere_value  when mmf.t_voltampere = '/' then cp.t_voltampere / mmf.t_voltampere_value else cp.t_voltampere end),2) AS t_voltampere,
			    ROUND((case when mmf.r_voltampere = '*' then cp.r_voltampere * mmf.r_voltampere_value  when mmf.r_voltampere = '/' then cp.r_voltampere / mmf.r_voltampere_value else cp.r_voltampere end),2) AS r_voltampere,
			    ROUND((case when mmf.y_voltampere = '*' then cp.y_voltampere * mmf.y_voltampere_value  when mmf.y_voltampere = '/' then cp.y_voltampere / mmf.y_voltampere_value else cp.y_voltampere end),2) AS y_voltampere,
			    ROUND((case when mmf.b_voltampere = '*' then cp.b_voltampere * mmf.b_voltampere_value  when mmf.b_voltampere = '/' then cp.b_voltampere / mmf.b_voltampere_value else cp.b_voltampere end),2) AS b_voltampere,
			    ROUND((case when mmf.t_var = '*' then cp.t_var * mmf.t_var_value  when mmf.t_var = '/' then cp.t_var / mmf.t_var_value else cp.t_var end),2) AS t_var,
			    ROUND((case when mmf.r_var = '*' then cp.r_var * mmf.r_var_value  when mmf.r_var = '/' then cp.r_var / mmf.r_var_value else cp.r_var end),2) AS r_var,
			    ROUND((case when mmf.y_var = '*' then cp.y_var * mmf.y_var_value  when mmf.y_var = '/' then cp.y_var / mmf.y_var_value else cp.y_var end),2) AS y_var,
			    ROUND((case when mmf.b_var = '*' then cp.b_var * mmf.b_var_value  when mmf.b_var = '/' then cp.b_var / mmf.b_var_value else cp.b_var end),2) AS b_var,
                ROUND((case when mmf.machine_kWh = '*' then cp.master_kwh * mmf.machine_kWh_value  when mmf.machine_kWh = '/' then cp.master_kwh / mmf.machine_kWh_value else cp.master_kwh end),2) AS master_kwh,
                ROUND((case when mmf.machine_kWh = '*' then cp.machine_kWh * mmf.machine_kWh_value  when mmf.machine_kWh = '/' then cp.machine_kWh / mmf.machine_kWh_value else cp.machine_kWh end),2) AS machine_kWh
		    from
                {table_name}   

		        inner join [ems_v1].[dbo].[master_machine] mm on mm.machine_id=cp.machine_id
                left join [ems_v1].[dbo].[master_machine_factor] mmf on mmf.machine_id=mm.machine_id
                
		    where 
                cp.machine_id in ({machine_id}) and {where} ) AS subquery
            WHERE
                slno = 1
		    order by machine_id, date_time                                
            ''')  
    
        createFolder("Load_analysis_log/","query executed successfully in load analysis..."+str(query))
        data=cnx.execute(query).mappings().all()
        label = {}
        machine_data = {}
        org_data = []
        for d in data:
            machine_id = d['machine_id']
            machine_name = d['machine_name']
            if machine_id not in label:        
                label[machine_id] = machine_name
            if machine_id not in machine_data:
                machine_data[machine_id] = []

            # set machine_data for machine_id
            temp = {
                'date_time': d['date_time'],
                't_current': d['t_current'],
                'r_current': d['r_current'],
                'y_current': d['y_current'],
                'b_current': d['b_current'],
                'vll_avg': d['vll_avg'],
                'ry_volt': d['ry_volt'],
                'yb_volt': d['yb_volt'],
                'br_volt': d['br_volt'],
                'vln_avg': d['vln_avg'],
                'r_volt': d['r_volt'],
                'y_volt': d['y_volt'],
                'b_volt': d['b_volt'],
                't_watts': d['t_watts'],
                'kWh': d['kWh'],
                'kvah': d['kvah'],
                'kw': d['kw'],
                'import_kw': d['import_kw'],
                'export_kw': d['export_kw'],
                'kvar': d['kvar'],
                'power_factor': d['power_factor'],
                'r_watts': d['r_watts'],
                'kva': d['kva'],
                'y_watts': d['y_watts'],
                'b_watts': d['b_watts'],
                'avg_powerfactor': d['avg_powerfactor'],
                'r_powerfactor': d['r_powerfactor'],
                'y_powerfactor': d['y_powerfactor'],
                'b_powerfactor': d['b_powerfactor'],
                'powerfactor': d['powerfactor'],
                'kwh_actual': d['kwh_actual'],
                'frequency': d['frequency'],
                't_voltampere': d['t_voltampere'],
                'r_voltampere': d['r_voltampere'],
                'y_voltampere': d['y_voltampere'],
                'b_voltampere': d['b_voltampere'],
                't_var': d['t_var'],
                'r_var': d['r_var'],
                'y_var': d['y_var'],
                'b_var': d['b_var'],
                'master_kwh':d['master_kwh'],
                'machine_kWh':d['machine_kWh']
            }

            machine_data[machine_id].append(temp)

        for key, value in machine_data.items():
            org_data.append({'label': label[key], 'data': value})

        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(org_data),"data1":jsonable_encoder(data)}) 

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Load_analysis_log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/change_password/")
async def change_password(employee_id:str=Form(None),
                          old_password:str=Form(None),
                          new_password:str=Form(None),
                          retype_password:str=Form(None),
                          cnx: Session = Depends(get_db)):  
    
    try:
        sql = text(f'''select * from [ems_v1].[dbo].[master_employee] where employee_id = {employee_id} and password_login = HASHBYTES('MD5', '{old_password}') ''')
        data = cnx.execute(sql).mappings().all()          

        if len(data) == 0:            
           return JSONResponse({"iserror":True,"message":"incorrect user id or password"})

        else:
            if new_password != retype_password:
                return JSONResponse({"iserror":True,"message":"retype password is incorrect"})

            query=text(f'''update [ems_v1].[dbo].[master_employee] set password_login = HASHBYTES('MD5', '{new_password}') where employee_id ='{employee_id}' ''')
            cnx.execute(query)
            cnx.commit()

        return JSONResponse({"iserror":False,"message":"password changed successfully","data":""}) 
    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
static_dir = Path(__file__).parent 

def generate_excel_report(result, month_year, report_type, report_for):
    print("result",result)
    if report_type == "date":
        file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "MonthWiseReport_templete.xlsx"))
        # file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "moth_wise_report.xlsx"))
          
        # file_path = f'{static_dir}/performanceReport_templete.xlsx'
    elif report_type == "shift":
        file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "MonthWiseReport_shift_templete.xlsx"))
        # file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..","..", "shift_wise_template.xlsx"))
          
        # file_path = f'{static_dir}/performanceReport_shift_templete.xlsx'
    workbook = Workbook()
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS' 
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')  
    cell = "C2"
    if report_for == '6to6':
        time = '06:00 to 06:00'
    else:
        time = '12:00 to 12:00'
    data = f"MONTH WISE ENERGY CONSUMPTION REPORT FOR KWH - {month_year} ({time})"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=25)
    sheet[cell].font = font
    
    # Set the border style
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Set the minimum column width
    minimum_column_width = 15
    cell = "B4"
    sheet[cell] = "Machine Code"
    cell = "C4"
    sheet[cell] = "Machine Name"
    sheet.cell(row=4, column=2, value="Machine Code").fill = fill_cyan
    sheet.cell(row=4, column=3, value="Machine Name").fill = fill_cyan
    if result == []:
        print(1)
        cell = "O10"
        data = "No Data"
        sheet[cell] = data

        alignment = Alignment(horizontal="center", vertical="center")
        sheet[cell].alignment = alignment

        sheet.column_dimensions[cell[0]].width = len(data) + 2  # Adjust column width

        font = Font(name='Calibri', size=25)
        sheet[cell].font = font
    # Iterate over the data and set the values and formatting
    for i, data in enumerate(result, start=6):  # Assuming machine_code starts from row 7
        machine_code = data["machine_code"]
        machine_name = data["machine_name"]
        print(machine_name)       
        sheet.cell(row=i, column=2, value=machine_code).alignment = Alignment(horizontal="center")
        # Create a new Alignment object with text wrapping enabled for machine_name
        alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.cell(row=i, column=3, value=machine_name).alignment = alignment
        # Adjust column size based on the maximum text length in columns C and D
        machine_name_length = len(machine_name) + 40
        # print(machine_name_length)
        column_letter_c = get_column_letter(3)  # Column C
        column_width_c = max(machine_name_length, sheet.column_dimensions[column_letter_c].width)
        # print(column_width_c)
        sheet.column_dimensions[column_letter_c].width = column_width_c
        if report_type == "date":

            month, year = map(int, month_year.split('-'))
            days_in_month = calendar.monthrange(year, month)[1]

            start_row = 4
            start_col = 4
            
            fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')

            for day in range(1, days_in_month + 1):
                cell = sheet.cell(row=start_row, column=start_col)
                cell.value = f"{day:02d}-{month_year}"
                cell.fill = fill_cyan
                
                start_col += 1
            for j in range(1, 32):
                column_letter = get_column_letter(j + 3)  
                cell = sheet.cell(row=i, column=j + 3)
                cell.value = data.get(f"d{j}", "")
                cell.alignment = Alignment(horizontal="center")
                if cell.value == 0:
                    cell.value = ""
            
            row_range = sheet[f"A{i}:AH{i}"]
            for row in row_range:
                for cell in row:
                    cell.border = border
        
        elif report_type == "shift":

            month, year = map(int, month_year.split('-'))
            days_in_month = calendar.monthrange(year, month)[1]

            start_row = 4
            start_col = 4

            fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')

            for day in range(1, days_in_month + 1):
                cell = sheet.cell(row=start_row, column=start_col)
                cell.value = f"{day:02d}-{month_year}"
                cell.fill = fill_cyan

                shifts = ['s1', 's2', 's3']

                for shift in range(1, 4):
                    # Write the shift label (s1, s2, s3) in cells D5, E5, F5 for each day
                    sheet.cell(row=5, column=start_col + shift - 1, value=shifts[shift - 1])
                    sheet.cell(row=5, column=start_col + shift - 1).fill = fill_cyan
                    sheet.cell(row=5, column=start_col + shift - 1).alignment = Alignment(horizontal="center")

                    cell = sheet.cell(row=start_row + shift, column=start_col + shift)
                    cell.value = data.get(f"d{day}_s{shift}", "")
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == 0:
                        cell.value = ""

                start_col += 3
            # Adjust column size based on the maximum text length vertically
            for shift in range(1, 4):
                for j in range(1, 32):
                    column_letter = get_column_letter((shift - 1) * 31 + j + 3)  # Assuming data columns start from column E (column index 5)
                    cell = sheet.cell(row=i, column=(shift - 1) * 31 + j + 3)
                    cell.value = data.get(f"ds{shift}_{j}", "")
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == 0:
                        cell.value = ""

                    # Adjust column size based on the maximum text length vertically
                    cell_text_length = len(str(cell.value))
                    column_width = max(cell_text_length, sheet.column_dimensions[column_letter].width)
                    sheet.column_dimensions[column_letter].width = column_width
                
                row_range = sheet[f"A{start_row}:CR{start_row + i}"]
                for row in row_range:
                    for cell in row:
                        cell.border = border
        
    file_name = f'MonthWiseReport-{month_year}.xlsx'
    file_path = os.path.join(base_path, file_name)
    workbook.save(file_path)

@app.post("/performance_report/")
async def performance_report(request:Request,
                             machine_id: str = Form(''),
                             month_year: str = Form(None),
                             report_for: str = Form(None),
                             report_type: str = Form(None),                                                      
                             employee_id: str = Form(''), 
                             is_critical : str = Form(''),                                                
                             cnx: Session = Depends(get_db)):
    
    if month_year is None:
        return JSONResponse({"iserror": True, "message": "month year is required"})

    if report_type is None:
        return JSONResponse({"iserror": True, "message": "report type is required"})
    
    if report_for is None:
        return JSONResponse({"iserror": True, "message": "report_for is required"})
    
    try:
        groupby = ""
        where = ""
        result_query = ''
        department_id = ''
        shed_id =''
        machinetype_id = ''
        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
     
        machine_id = id(machine_id)
        if is_critical != '':
           where += f"and mm.major_nonmajor = '{is_critical}'"

        if machine_id == "" or machine_id == 'all':
            pass
        else:
            where += f" and mm.machine_id IN ({machine_id})"

        if  employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]

        if department_id !='' and department_id !=0:
            where += f" and md.department_id ={department_id}"
        if shed_id !='' and shed_id != 0:
            where += f" and ms.shed_id ={shed_id}"
        if machinetype_id !='' and machinetype_id!= 0:
            where += f" and mmt.machinetype_id ={machinetype_id}"

        if report_for == '6to6':
            month, year = month_year.split('-')
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month}{year}'"""
            result_query = cnx.execute(query).mappings().all()
            if len(result_query)==0:
                return JSONResponse({"iserror": True, "message": "power table not available..."})
            tbl_name = f"ems_v1_completed.dbo.power_{month}{year} cp"       
        else:
            month, year = month_year.split('-')
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month}{year}_12'"""
            result_query = cnx.execute(query).mappings().all()
            if len(result_query)==0:
                return JSONResponse({"iserror": True, "message": "power table not available..."})    
            tbl_name = f"ems_v1_completed.dbo.power_{month}{year}_12 cp"
            if report_type not in ['date']:
                return JSONResponse({"iserror": True, "message": "Invalid report type"})

        if report_type not in ['date', 'shift']:
                return JSONResponse({"iserror": True, "message": "Invalid report type"})

        if report_type == 'date':
            day = "CONCAT('d', DAY(cp.mill_date))"
            
        elif report_type == 'shift':
            day = "CONCAT(CONCAT('ds', cp.mill_shift), '_', DAY(cp.mill_date))"
            groupby = ",cp.mill_shift"

        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                {day} AS day,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh
            FROM
                {tbl_name}
                INNER JOIN ems_v1.dbo.master_machine mm ON mm.machine_id = cp.machine_id
                left join ems_v1.dbo.master_machine_factor mmf on mmf.machine_id = mm.machine_id
                LEFT JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id = mm.shed_id                   
                LEFT JOIN [ems_v1].[dbo].[master_department] md ON md.department_id = mm.department_id                   
                LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id = mm.machinetype_id                   
            WHERE
                1=1 and  mm.status = 'active' {where} and FORMAT(cp.mill_date, 'MM-yyyy') = '{month_year}' 
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                DAY(cp.mill_date)
                {groupby}               
        ''')
        print(query)
        rslt = cnx.execute(query).mappings().all()
        
        if rslt is not None:
            output = {}

            if report_type == 'date':
                output_keys = [f'd{day}' for day in range(1, 32)]
            elif report_type == 'shift':
                output_keys = [f'ds{shift}_{day}' for day in range(1, 32) for shift in range(1, 4)]
            
            for row in rslt:
                machine_code = row.machine_code
                machine_name = row.machine_name
                day = row.day
                kwh = row.kwh
            
                if machine_code not in output:
                    output[machine_code] = {
                        'machine_code': machine_code,
                        'machine_name': machine_name
                    }
                    for key in output_keys:
                        output[machine_code][key] = 0
            
                output[machine_code][day] = kwh
                result=list(output.values())
            if rslt == []:
                result = []
               
            # machine = {"result" : data["machine_name"],
            #            "result1" : data[]}
            generate_excel_report(result, month_year,report_type, report_for)
            # process_data(month_year, result)
        file_path = os.path.join(base_path, f"MonthWiseReport-{month_year}.xlsx")
        results = f"http://{request.headers['host']}/attachment/MonthWiseReport-{month_year}.xlsx"

        if os.path.exists(file_path):

            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": results})
        else:
            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": None})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})
    
# [{"machine_id":"1","kWh": "34658","mill_date": "2023-05-26 ","mill_shift": "2"}]
@app.post("/manual_entry/")
async def save_manual_entry(obj: str = Form(None),   
                            reportfor :str =Form(None),                          
                            cnx: Session = Depends(get_db)):
    try:
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="[ems_v1_completed].[dbo]."    
        if obj == None:
            return JSONResponse({"iserror":True,"message":"obj is required"})
        report = ''
        if reportfor == '12to12':
            report =f'_12'

        if obj != '':
            user_dict = json.loads(obj)
            for i in user_dict:
                machine_id = i['machine_id']
                kWh = i['kWh']
                mill_date = i['mill_date']
                mill_shift = i['mill_shift']
                month_year=f"""{mill_month[parse_date(mill_date).month]}{str(parse_date(mill_date).year)}"""
                table_name=f"  {completed_db}[power_{month_year}]"           
                query = text(f'''
                    UPDATE {table_name}{report}
                    SET kWh = {kWh}
                    WHERE machine_id = '{machine_id}' and mill_date = '{mill_date}' and mill_shift = '{mill_shift}' ''')
                createFolder("Manual_Entry_Log/", "query " + str(query))
                cnx.execute(query)
                cnx.commit()

        return JSONResponse({"iserror": False, "message": "parameter saved sucessfully", "data": ''})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Manual_Entry_Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/machine_history_list/")
async def machine_history_list(machine_id: str = Form(None), 
                               company_id: str = Form(''),                             
                               department_id: str = Form(''),                             
                               shed_id: str = Form(''),                             
                               machinetype_id: str = Form(''),                             
                               cnx: Session = Depends(get_db)):
    try:
        where = ''
        if machine_id is not None:
            where = f' and mh.machine_id = {machine_id}'

        if company_id != '':
            where = f' and mh.company_id = {company_id}'

        if department_id != '':
            where = f' and mh.department_id = {department_id}'

        if shed_id != '':
            where = f' and mh.shed_id = {shed_id}'

        if machinetype_id != '':
            where = f' and mh.machinetype_id = {machinetype_id}'

        query = text(f'''
                    SELECT 
                        mc.company_code AS company_code,
                        mc.company_name AS company_name,
                        mb.branch_code AS branch_code,
                        mb.branch_name AS branch_name,
                        md.department_code AS department_code,
                        md.department_name AS department_name,
                        ms.shed_code AS shed_code,
                        ms.shed_name AS shed_name,
                        mmt.machinetype_code AS machinetype_code,                        
                        mmt.machinetype_name AS machinetype_name,
                        mf.function_code AS function_code,
                        mf.function_name AS function_name,
                        mcd.converter_name AS converter_name,
                        mh.*,
                        ISNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
	                    ISNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
                    FROM 
                        [ems_v1].[dbo].[master_machine_history] mh
                        left join [ems_v1].[dbo].[master_employee] cu on cu.employee_id=mh.created_by
	                    left join [ems_v1].[dbo].[master_employee] mu on mu.employee_id=mh.modified_by
                        INNER JOIN [ems_v1].[dbo].[master_company] mc ON mh.company_id = mc.company_id
                        INNER JOIN [ems_v1].[dbo].[master_branch] mb ON mh.branch_id = mb.branch_id
                        INNER JOIN [ems_v1].[dbo].[master_department] md ON mh.department_id = md.department_id
                        INNER JOIN [ems_v1].[dbo].[master_shed] ms ON mh.shed_id = ms.shed_id
                        INNER JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mh.machinetype_id = mmt.machinetype_id
                        INNER JOIN [ems_v1].[dbo].[master_function] mf ON mh.function_id = mf.function_id
                        INNER JOIN [ems_v1].[dbo].[master_converter_detail] mcd ON mh.converter_id = mcd.converter_id                   
                    WHERE mh.status = 'active'  {where}''')
        data = cnx.execute(query).mappings().all()
        return JSONResponse({"iserror": False, "message": "data return sucessfully", "data": jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/update_alarm_popup_status/")
async def update_alarm_popup_status(company_id: str = Form(None),                             
                            cnx: Session = Depends(get_db)):
    
    try:

        query = text(f'''Update [ems_v1].[dbo].[master_company] set alarm_status = 0 where company_id = {company_id}''')
        cnx.execute(query)
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "sucessfully updated", "data": ''})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/energy_calculation_list/")
async def energy_calculation_list(cnx: Session = Depends(get_db)):
    
    try:

        query = text(f'''select * from [ems_v1].[dbo].[master_energy_calculations]''')
        data = cnx.execute(query).mappings().all()
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data return sucessfully", "data":jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/save_energy_calculation/")
async def save_energy_calculation(obj :str = Form(""),
                                  cnx: Session = Depends(get_db)):
    
    try:
        print(obj)
        del_query=text(f'''DELETE FROM [ems_v1].[dbo].[master_energy_calculations]''')
        cnx.execute(del_query)
        cnx.commit()
        obj_data = json.loads(obj)
        if obj !="":
            for row in obj_data:
                s_no = row["s_no"]
                group_name = row["group_name"]
                function_name = row["function_name"]
                formula1 = row["formula1"]
                formula2 = row["formula2"]
                parameter = row["parameter"]
                roundoff_value = row["roundoff_value"]
                loss_percentage = row["loss_percentage"]
                query = text(f'''INSERT INTO [ems_v1].[dbo].[master_energy_calculations] 
                                (s_no,group_name,function_name,formula1,formula2,parameter,roundoff_value,loss_percentage)
                                values({s_no},'{group_name}','{function_name}','{formula1}','{formula2}','{parameter}','{roundoff_value}','{loss_percentage}')''')

                cnx.execute(query)
                cnx.commit()
        return JSONResponse({"iserror": False, "message": "data save sucessfully", "data":""})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/group_name_list/")
async def group_name_list(cnx: Session = Depends(get_db)):
    
    try:
        query = text(f'''select * from [ems_v1].[dbo].[master_group_name] where status!='delete' ''')
        data = cnx.execute(query).mappings().all()
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data return sucessfully", "data":jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/save_group_name/")
async def save_group_name(id :int = Form(""),
                          group_name :str = Form(""),
                          user_login_id : str = Form(""),
                          cnx: Session = Depends(get_db)):

    if group_name == "":
        return JSONResponse({"iserror": True, "message": "group name is required"})

    if user_login_id == "":
        return JSONResponse({"iserror": True, "message": "user_login_id is required"})

    try:
        if id == "":
            select_query = text(f'''select * from [ems_v1].[dbo].[master_group_name]  where group_name = '{group_name}' and status!='delete' ''')
            data = cnx.execute(select_query).mappings().all()

            if len(data) > 0:
                return JSONResponse({"iserror":True, "message":"group name already exists", "data":""})

            query = text(f'''INSERT INTO [ems_v1].[dbo].[master_group_name] 
                         (group_name,created_on,created_by)
                         values('{group_name}',getdate(),'{user_login_id}')''')

        else:
            query = text(f'''UPDATE [ems_v1].[dbo].[master_group_name]
                         SET group_name ='{group_name}',modified_on= getdate(),modified_by = '{user_login_id}' where id = {id} ''')
        cnx.execute(query)
        cnx.commit()
        return JSONResponse({"iserror": False, "message": "data save sucessfully", "data":""})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/remove_group_name/")
async def remove_group_name(id: int = Form(None),
                            status : str = Form(None),
                            cnx: Session = Depends(get_db)):
    try:                            
        if id is not None:
            if status is not None:
                query = text(f" UPDATE [ems_v1].[dbo].[master_group_name]  SET status = '{status}' WHERE id = '{id}' ")
                cnx.execute(query)

            else:
                query = text(f" UPDATE [ems_v1].[dbo].[master_group_name]  SET status = 'delete' WHERE id = '{id}' ")                
                cnx.execute(query)
            cnx.commit()

        return JSONResponse({"iserror": False, "message": "status update successfully", "data": ""})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f'''{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}'''
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

def dailyreport(date,day_month_year_value,report_for,dates,year,next_year):
        try:
            # file_path = f'{static_dir}\daily_report_template.xlsx'
            file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..",  "daily_report_template.xlsx"))
            workbook = Workbook()
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            sheet.title = 'DAILY REPORT'
            border = Border(left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"))
            cell = "F5"
            data = f"Daily Energy Report"
            sheet[cell] = data
            cell = "J6"
            data = f"{date}"
            sheet[cell] = data
            bold_font = Font(bold=True,size = 12)
            sheet[cell].font = bold_font
            center_alignment = Alignment(horizontal="center")
            sheet[cell].alignment = center_alignment
            cell = "H7"
            if report_for == '6to6':
                datas = f'06:00 to 06:00'
            else:
                datas = f'Day'
            sheet[cell] = datas
            bold_font = Font(bold=True,size = 12)
            sheet[cell].font = bold_font
            cell = "J7"
            # parsed_date = datetime.datetime.strptime(date, "%d-%b-%Y")
            # year = parsed_date.year
            fiscal_year = f"{year}-{next_year}"
            data = f" FY {fiscal_year}"
            sheet[cell] = data
            sheet[cell].font = bold_font
            sheet[cell].alignment = center_alignment
            row = 8
            col = 6
            for row_item in day_month_year_value:
                energy_in_units = row_item['func_name']
                day = row_item['formula_d']
                month = row_item['formula_m']
                year = row_item['formula_y']
                roundoff_value = row_item['roundoff_value']
                report_type = row_item['report_type']
                print("roundoff_value",roundoff_value)
                merge_range = f'F{row}:G{row}'
                sheet[f'F{row}']=energy_in_units

                for col in range(6, 8):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = border
                if isinstance(day, (int, float)) and isinstance(month, (int, float)) and isinstance(year, (int, float)):
                    
                    day = float(day)
                    month = float(month)
                    year = float(year)
          
                    day_cell = sheet.cell(row=row, column=8)
                    if roundoff_value == 0:
                        day_cell.value = f'=ABS({day})'
                    else:
                        day_cell.value = f'=MROUND(ABS(({day})),{roundoff_value})'
                    day_cell.number_format = '0.00'
                    day_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    day_cell.border = border

                    month_cell = sheet.cell(row=row, column=9)
                    if roundoff_value == 0:
                        month_cell.value = f'=ABS({month})'
                    else:
                        month_cell.value = f'=MROUND(ABS(({month})),{roundoff_value})'
                    month_cell.number_format = '0.00'
                    month_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    month_cell.border = border

                    year_cell = sheet.cell(row=row, column=10)
                    if roundoff_value == 0:
                        year_cell.value = f'=ABS({year})'
                    else:
                        year_cell.value = f'=MROUND(ABS(({year})),{roundoff_value})'
                    year_cell.number_format = '0.00'
                    year_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    year_cell.border = border

                else:
                    day_cell = sheet.cell(row=row, column=8)
                    if roundoff_value == 0:
                        day_cell.value = f'=ABS({day})'
                    else:
                        day_cell.value = f'=MROUND(ABS(({day})),{roundoff_value})'
                    day_cell.number_format = '0.00'
                    day_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    day_cell.border = border

                    month_cell = sheet.cell(row=row, column=9)
                    if roundoff_value == 0:
                        month_cell.value = f'=ABS({month})'
                    else:
                        month_cell.value = f'=MROUND(ABS(({month})),{roundoff_value})'
                    month_cell.number_format = '0.00'
                    month_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    month_cell.border = border

                    year_cell = sheet.cell(row=row, column=10)
                    if roundoff_value == 0:
                        year_cell.value = f'=ABS({year})'
                    else:
                        year_cell.value = f'=MROUND(ABS(({year})),{roundoff_value})'
                    year_cell.number_format = '0.00'
                    year_cell.alignment = Alignment(horizontal='center', vertical='center') 
                    year_cell.border = border

                sheet.merge_cells(merge_range)
                sheet.row_dimensions[row].height = 30
                row+=1

            last_row = sheet.max_row
            next_row = last_row + 1
            merge_range = f'F{next_row}:J{next_row}'

            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_row, column=6)

            sheet.row_dimensions[next_row].height = 25
            merged_cell.value = "Remarks:"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            for col in range(6, 11):  
                cell = sheet.cell(row=next_row, column=col)
                cell.border = border

            next_rows = last_row + 2
            merge_range = f'F{next_rows}:J{next_rows}'
            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_rows, column=6)
            sheet.row_dimensions[next_rows].height = 25
            for col in range(6, 11):  
                cell = sheet.cell(row=next_rows, column=col)
                cell.border = border

            next_rows1 = last_row + 3
            merge_range = f'F{next_rows1}:J{next_rows1}'
            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_rows1, column=6)
            sheet.row_dimensions[next_rows1].height = 25
            for col in range(6, 11):  
                cell = sheet.cell(row=next_rows1, column=col)
                cell.border = border

            next_rows2 = last_row + 4
            merge_range = f'F{next_rows2}:G{next_rows2}'
            sheet.merge_cells(merge_range)
            merged_cell = sheet.cell(row=next_rows2, column=6)
            sheet.row_dimensions[next_rows2].height = 15
            merged_cell.value = "Shift Incharge-(Elect)"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')

            for col in range(6, 8):  
                cell = sheet.cell(row=next_rows2, column=col)
                cell.border = border
            merge_ranged = f'H{next_rows2}'
            sheet.merge_cells(merge_ranged)
            merged_cell = sheet.cell(row=next_rows2, column=8)
            sheet.row_dimensions[next_rows2].height = 15
            merged_cell.value = "Sectional Incharge-(Elect)"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(8,9):  
                cell = sheet.cell(row=next_rows2, column=col)
                cell.border = border
            merge_rangee = f'I{next_rows2}:J{next_rows2}'
            sheet.merge_cells(merge_rangee)
            merged_cell = sheet.cell(row=next_rows2, column=9)
            sheet.row_dimensions[next_rows2].height = 15
            merged_cell.value = "HOD"
            merged_cell.font = Font(bold=True)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(9, 11):  
                cell = sheet.cell(row=next_rows2, column=col)
                cell.border = border
            file_name = f'DailyReport-{dates}-({report_type}).xlsx'
            print("111111111111",file_name)
            file_path = os.path.join(base_path, file_name)
            workbook.save(file_path)
            return file_path
        
        except Exception as e:
            error_type = type(e).__name__
            error_line = traceback.extract_tb(e.__traceback__)[0].lineno
            error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
            error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
            createFolder("Log/","Issue in returning data "+error_message)
            return JSONResponse({"iserror":True,"message":"error"+error_message})

@app.post("/custom_daily_report/")
async def custom_daily_report_api(request:Request,
                                  date: str=Form(None), 
                                  report_for : str = Form(''),          
                                  loss_record : str = Form(''),          
                                  report_type : str = Form(''),          
                                  cnx: Session = Depends(get_db)):
    try:
        res = ''
        ress = ''
        res_q4 = ''
        formula_d = 0
        formula_m = 0
        formula_y = 0
        func_name = ''
        formula = ''
        roundoff_value = 0
        type = ''
        data_p= ''
        formula_day = 0
        if date == None or date == '':
            return JSONResponse({"iserror": True, "message": "date is required"})

        dates=parse_date(date)
        if report_for == "":
            return JSONResponse({"iserror": True, "message": "report_for is required"})

        from_date_str = dates.strftime("%d-%m-%Y")
        datetime_obj = datetime.datetime.strptime(from_date_str, "%d-%m-%Y")
        f_date = datetime_obj.strftime("%d-%m-%Y") 
        formatted_date = datetime_obj.strftime("%d-%b-%Y") 
        month = f_date[3:5]
        given_year = formatted_date[7:]
        print("month",month)

        if int(month) >=4:
            year = formatted_date[7:]
            next_year = int(given_year) +1
        if int(month) <4:
            year = int(given_year) -1
            next_year =formatted_date[7:]
        print("next_year",next_year)

        query = f'''SELECT * FROM [ems_v1].[dbo].[master_energy_calculations] ORDER BY s_no '''
        result = cnx.execute(query).mappings().all()

        para = ''
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="[ems_v1_completed].[dbo]."    
        table_name = ""
        where = ""
        if report_for == '6to6':
            for rows in result:
                para = rows['parameter']
                    
            if para == 'kw':
                para = "case when mmf.kw = '*' then p.t_watts * mmf.kw_value when  mmf.kw = '/' then p.t_watts / mmf.kw_value else p.t_watts end "

            if para == 'kWh':
                para = "case when mmf.kWh = '*' then p.kWh * mmf.kWh_value when  mmf.kWh = '/' then p.kWh / mmf.kWh_value else p.kWh end "
            
            month_year=f"""{mill_month[dates.month]}{str(dates.year)}"""
            table_name=f"  {completed_db}[power_{month_year}] as p"
            tblname = f'power_{month_year}'
            query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '{tblname}'"""
            
        else:
            para = "case when mmf.kWh = '*' then p.kWh * mmf.kWh_value when  mmf.kWh = '/' then p.kWh / mmf.kWh_value else p.kWh end "
            
            month_year=f"""{mill_month[dates.month]}{str(dates.year)}"""
            table_name=f"  {completed_db}[power_{month_year}_12] as p"
            tblname = f'power_{month_year}_12'
            query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '{tblname}'"""
            
        query1 = cnx.execute(query1).mappings().all()
        where += f'''p.mill_date = '{dates}' ''' 
          
        if report_type == 'with_loss':
            pass
        else:
            if len(query1)>0:     
                
                query2= f''' 
                SELECT 
                    p.machine_id,
                    sum({para}) as kWh,
                    SUM(case when mmf.kWh = '*' then p.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then p.reverse_kwh / mmf.kWh_value else p.reverse_kwh end ) AS reverse_kwh,
                    min(mm.machine_name) as machine_name
                from {table_name} 
                    left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id
                    left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id
                where 
                    {where} 
                group by 
                    p.machine_id 
                order by 
                    p.machine_id '''
                createFolder("Log/","query for day "+str(query2))
                res = cnx.execute(query2).mappings().all()

                dict={}
                dict_r_d={}
                
                for row in res:
                    dict[row['machine_id']] = row['kWh']
                    dict_r_d[row['machine_id']] = row['reverse_kwh']

        if len(query1)>0: 

            query3= f''' 
            SELECT 
                p.machine_id,
                sum({para}) as total ,
                SUM(case when mmf.kWh = '*' then p.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then p.reverse_kwh / mmf.kWh_value else p.reverse_kwh end ) AS reverse_kwh,
                min(mm.machine_name) as machine_name
            from 
                {table_name}  
                left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id
                left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id
            group by 
                p.machine_id 
            order by 
                p.machine_id '''
            ress = cnx.execute(query3).mappings().all()
            createFolder("Log/","query for month "+str(query3))
            dict1={}
            dict_r_m = {}
            
            for row in ress:
                dict1[row['machine_id']] = row['total']
                dict_r_m[row['machine_id']] = row['reverse_kwh']   
        else:
            pass
        
        table_names = []

        for month in range(4, 13):
            month_year = f"{mill_month[month]}{year}"

            if report_for == "12to12":
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all()
            else:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all()
            if len(result_query) > 0:
                table_names.append(f"ems_v1_completed.dbo.power_{month_year}")
        
        for month in range(1, 4):
            month_year = f"{mill_month[month]}{next_year}"

            if report_for == "12to12":
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all() 
            else:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all() 
            if len(result_query) > 0:        
                table_names.append(f"ems_v1_completed.dbo.power_{month_year}")

        if len(table_names)==0:
            return JSONResponse({"iserror": True, "message": "table not available"})
        if report_for =="12to12":
            type = f"_12 p"
        else:
            type = f" p"
        
        union_query = " UNION ALL ".join([f"SELECT p.machine_id, {para} as data, case when mmf.kWh = '*' then p.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then p.reverse_kwh / mmf.kWh_value else p.reverse_kwh end  AS reverse_kwh, mm.machine_name FROM {table_name}{type}  left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id" for table_name in table_names])

        query4 = f"""
                SELECT 
                    pp.machine_id, 
                    SUM(pp.data) as total_kwh ,
                    sum(pp.reverse_kwh) as reverse_kwh ,
                    min(pp.machine_name) as machine_name
                FROM 
                    ({union_query}) AS pp 
                    
                GROUP BY 
                    pp.machine_id 
                ORDER BY 
                    pp.machine_id"""
        createFolder("Log/","query for year "+str(query4))
        res_q4 = cnx.execute(query4).mappings().all()
        
        dict2={}
        dict_r_y = {}
        
        for row in res_q4:
            dict2[row['machine_id']] = row['total_kwh']
            dict_r_y[row['machine_id']] = row['reverse_kwh']
                     
        rows_to_write = []
        records = []

        for rows in result:
            func_name = rows['function_name']
            formula = rows['formula2']
            roundoff_value = rows['roundoff_value']
            if report_type == 'with_loss':
                query2 = f''' select * from ems_v1.dbo.loss_record where date = '{dates}' and report_for = '{report_for}' and energy_in_units = '{func_name}' '''
                print(query2)
                data_p = cnx.execute(query2).fetchone()
                
                if data_p == '' or data_p == None:
                    formula_d = 0
                else:
                    formula_d = data_p.day + data_p.loss_value
            else:
                if len(res) == 0:
                    formula_d = 0
                    formula_day = 0
                else:
                    if func_name == 'Power Import':
                        formula_d = dict[14]
                        formula_day = formula_d

                    elif func_name == 'Power Export':
                        formula_d = dict_r_d[14]
                        formula_day = formula_d
                    else:
                        try:  
                            numbers = re.findall(r'\[(\d+)\]', formula)
                            valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict]
                            numeric_formula = formula
                            for machine_id in valid_ids:
                                numeric_value = dict.get(machine_id, 0)  # Get the value from dict2 or use 0 if not found
                                numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                            createFolder("Log/","numeric_formula"+numeric_formula)
                            formula_d = re.sub(r'dict', '', numeric_formula)
                            createFolder("Log/","formula_d"+formula_d)
                            formula_day = eval(formula_d)
                        except Exception as e:
                            # error_type = type(e).__name__
                            error_line = traceback.extract_tb(e.__traceback__)[0].lineno
                            error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
                            error_message = f" occurred in file {error_filename}, line {error_line}: {str(e)}"
                            createFolder("Log/","Issue in returning data "+error_message)
                            formula_d = 0
                            formula_day = 0
                     
            if len(ress) == 0:
                month_resultsss = 0
                formula_m = 0
            else:
                if func_name == 'Power Import':
                    formula_m = dict1[14]
                elif func_name == 'Power Export':
                    formula_m = dict_r_m[14]
                else:
                    try:
                        numbers = re.findall(r'\[(\d+)\]', formula)

                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict1]
                        numeric_formula = formula
                        for machine_id in valid_ids:
                            numeric_value = dict1.get(machine_id, 0)  # Get the value from dict2 or use 0 if not found
                            numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                        formula_m = re.sub(r'dict', '', numeric_formula)
                    except Exception as e:
                        # error_type = type(e).__name__
                        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
                        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
                        error_message = f" occurred in file {error_filename}, line {error_line}: {str(e)}"
                        formula_m = 0
                        createFolder("Log/","Issue in returning data "+error_message)
                        

            if len(res_q4) == 0:
                year_resultsss = 0
                formula_y = 0
            else:
                if func_name == 'Power Import':
                    formula_y = dict2[14]
                    print("formula_y",formula_y)
                elif func_name == 'Power Export':
                    formula_y = dict_r_y[14]

                else:
                    try:
                        numbers = re.findall(r'\[(\d+)\]', formula)

                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict2]
                        numeric_formula = formula
                        for machine_id in valid_ids:
                        
                            numeric_value = dict2.get(machine_id, 0)  # Get the value from dict2 or use 0 if not found
                            numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                        formula_y = re.sub(r'dict', '', numeric_formula)
                    except Exception as e:
                        # error_type = type(e).__name__
                        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
                        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
                        error_message = f" occurred in file {error_filename}, line {error_line}: {str(e)}"
                        createFolder("Log/","Issue in returning data "+error_message)
                        formula_y = 0
                        
                        
                    
            if loss_record == 'insert':
                records.append({
                                        "func_name": func_name,
                                        'formula_day': formula_day,
                                    })      
            else:
                rows_to_write.append({
                                    "func_name": func_name,
                                    'formula_d': formula_d,
                                    "formula_m": formula_m,
                                    "formula_y": formula_y,
                                    "roundoff_value":roundoff_value,
                                    "report_type":report_type 
                                })
            
        dailyreport(formatted_date,rows_to_write,report_for,date,year,next_year)
        file_path = os.path.join(base_path, f"DailyReport - {date}.xlsx")
        result1 = f"http://{request.headers['host']}/attachment/DailyReport-{date}-({report_type}).xlsx"

        if loss_record == 'insert':
            return records
        else:
            return JSONResponse({"iserror":False,"message":"data returned successfully","data":result1})
        
    except Exception as e:
        # error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f" occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_custom_function_dashboard_detail/")
async def get_custom_function_dashboard_detail(cnx: Session = Depends(get_db)):

    try:
        mill_date = date.today()
        mill_shift = 0
        group_name = ''
        func_name = ''
        formula1 = ''
        results = ''

        sql1 = f"select * from ems_v1.dbo.master_shifts"
        data = cnx.execute(sql1).mappings().all()
        if len(data)>0:
            for row in data:
                mill_date = row["mill_date"]
                mill_shift = row["mill_shift"]

        sql2 = f'SELECT * FROM [ems_v1].[dbo].[master_energy_calculations] ORDER BY group_name, s_no'
        result = cnx.execute(sql2).mappings().all()
        if len(result)>0:
            para = ''
            for rows in result:
                para = rows['parameter']

            if para == 'kw':
                para = "case when mmf.kw = '*' then p.t_watts * mmf.kw_value when  mmf.kw = '/' then p.t_watts / mmf.kw_value else p.t_watts end "

            if para == 'kWh':
                para = "case when mmf.kWh = '*' then p.kWh * mmf.kWh_value when  mmf.kWh = '/' then p.kWh / mmf.kWh_value else p.kWh end "
        
            sql3 = text(f'''
                    select 
                        p.machine_id,
                        min(mm.machine_name) as machine_name,
                        sum({para}) as kWh 
                    from 
                        [ems_v1].[dbo].[current_power] p
                        left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id
                        left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id
                    where 
                        p.mill_date = '{mill_date}' and p.mill_shift = {mill_shift}
                    group by 
                        p.machine_id 
                    order by 
                        p.machine_id''')
            print(sql3)
            res = cnx.execute(sql3).mappings().all()

            machine_id_dict={}
            # dict_tt ={}
            dict={}
            for row in res:
                dict[row['machine_id']] = row['kWh']
                # dict_tt[row['machine_name']] = row['kWh']
                machine_id_dict[row['machine_id']] = row['machine_name']
            datas = []

            for rows in result:
                group_name = rows['group_name']
                func_name = rows['function_name']
                formula = rows['formula2']
                formula1 = rows['formula1']
                
                results = eval(formula, {"dict": dict})
                
                # machine_ids = re.findall(r'dict\[(\d+)\]', formula)
        
                # # Convert extracted IDs to integers and filter out IDs not present in the dict
                # valid_formula_machine_ids = [int(id) for id in machine_ids if int(id) in dict]

                # # Create a tooltip dictionary with valid machine IDs and their kWh values
                # formula_tooltip = {machine_id: dict[machine_id] for machine_id in valid_formula_machine_ids}
                machine_ids = re.findall(r'dict\[(\d+)\]', formula)

                valid_formula_machine_ids = [int(id) for id in machine_ids if int(id) in dict]
                print(valid_formula_machine_ids)
                formula_tooltip = {machine_id_dict[machine_id]: dict[machine_id] for machine_id in valid_formula_machine_ids}
                print(formula_tooltip)
                datas.append({"group_name": group_name,"function_name": func_name,"function_value": results,"formula1": formula1,"tooltip":formula_tooltip})

            return JSONResponse({"iserror":False,"message":"data returned successfully","data":jsonable_encoder(datas)})
        else:
            return JSONResponse({"iserror":False,"message":"data returned successfully","data":''})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/order_wise/")
async def order_wise(table_name :str = Form(''),
                     obj :str = Form(''),
                     cnx: Session = Depends(get_db)):

    try:
       
        if obj !='':
            obj_data = json.loads(obj)
            for row in obj_data:
                id = row["id"]
                sno = row["sno"]
                if table_name == 'zone':
                    sql = text(f" update ems_v1.dbo.master_department set department_order = {sno} where department_id = {id} ")
                    
                if table_name == 'area':
                    sql = text(f" update ems_v1.dbo.master_shed set shed_order = {sno} where shed_id = {id} ")
                
                if table_name == 'location':
                    sql = text(f" update ems_v1.dbo.master_machinetype set machinetype_order = {sno} where machinetype_id = {id} ")
                    
                if table_name == 'function_1':
                    sql = text(f" update ems_v1.dbo.master_function set function_order = {sno} where function_id = {id}")
                    
                if table_name == 'function_2':
                    sql = text(f" update ems_v1.dbo.master_function set function_order = {sno} where function_id = {id}")
                    
                if table_name == 'meter':
                    sql = text(f" update ems_v1.dbo.master_machine set machine_order = {sno} where machine_id = {id} ")

                cnx.execute(sql)
                cnx.commit()
        return JSONResponse({"iserror":False,"message":"order update successfully","data":''})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_master_machine_factor/")
async def get_master_machine_factor(id :str = Form(''),
                                    company_id :str = Form(''),
                                    department_id :str = Form(''),
                                    shed_id :str = Form(''),
                                    machinetype_id :str = Form(''),
                                    machine_id :str = Form(''),
                                    cnx: Session = Depends(get_db)):

    try:
        where = ''
        if company_id != ''  and company_id != "0":
            where += f'and  mc.company_id = {company_id}'

        if department_id != '' and department_id != "0" :
            where += f'and  md.department_id = {department_id}'

        if shed_id != '' and shed_id != "0":
            where += f'and  ms.shed_id = {shed_id}'

        if machinetype_id != '' and machinetype_id !="0":
            where += f'and  mmt.machinetype_id = {machinetype_id}'
            
        if machine_id != '':
            where += f'and  mf.machine_id = {machine_id}'
        
        if id != '':
            where += f'and  mf.id = {id}'

        sql = text(f'''
                select 
                   mf.*,
                   mm.machine_code,
                   mm.machine_name
                from 
                   ems_v1.dbo.master_machine_factor mf,
                   ems_v1.dbo.master_machine mm
                   left join [ems_v1].[dbo].[master_company] mc on mc.company_id=mm.company_id
                   left join [ems_v1].[dbo].[master_branch] mb on mb.branch_id=mm.branch_id
                   left join [ems_v1].[dbo].[master_department] md on md.department_id=mm.department_id
                   left join [ems_v1].[dbo].[master_shed] ms on ms.shed_id=mm.shed_id
                   left join [ems_v1].[dbo].[master_machinetype] mmt on mmt.machinetype_id=mm.machinetype_id
           
                where mf.machine_id = mm.machine_id {where}''')
        print(sql)
        data = cnx.execute(sql).mappings().all()
        
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/update_master_machine_factor/")
async def update_master_machine_factor(machine_id :str = Form(''),
                                       obj :str = Form(''),
                                       cnx: Session = Depends(get_db)):

    try:
        if obj != '':
            obj_data = json.loads(obj)
            sel = {}
            for data in obj_data:
                for key, value in data.items():
                    sel[key] = value
                sql = text(f'''UPDATE ems_v1.dbo.master_machine_factor SET {', '.join([f"{key} = '{value}'" for key, value in sel.items()])} WHERE machine_id = '{machine_id}' ''')
                cnx.execute(sql)
                cnx.commit()
                createFolder("Log/","data.. "+str(sql))
        return JSONResponse({"iserror":False,"message":"data update successfully","data":''})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/get_sld_detail/")
async def get_sld_detail(sld_type :int = Form(''),                                    
                         cnx: Session = Depends(get_db)):

    try:
        where = ''
        if sld_type == 1:
            where =f' where mm.machine_id in (6,7,8,11,12,13,16,17,18,19,20,21,22,23,24,25,26,15)'
        if sld_type == 2:
            where =f' where mm.machine_id in (27,28,30,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,29)'
        if sld_type == 3:
            where =f' where mm.machine_id in (16,6,28,25,26,41,14,15,29)'
        sql = text(f'''
            select 
                min(mc.company_code) AS company_code,
                min(mc.company_name) AS company_name,
                min(mb.branch_code) AS branch_code,
                min(mb.branch_name) AS branch_name,
                min(md.department_code) AS department_code,
                min(md.department_name) As department_name,
                min(ms.shed_code) AS shed_code,
                min(ms.shed_name) AS shed_name,
                min(mmt.machinetype_code) AS machinetype_code,
                min(mmt.machinetype_name) AS machinetype_name,
                min(mf.function_name) AS function_name,
                min(mf.function_code) AS function_code,
                min(mm.machine_code) AS machine_code,
                min(mm.machine_name) AS machine_name,
                count(mm.machine_name) AS machine_count,
                min(cp.power_id) as power_id,
                min(cp.company_id) as company_id,
                min(cp.branch_id) as branch_id,
                min(cp.department_id) as department_id,
                min(cp.shed_id) as shed_id,
                min(cp.machinetype_id) as machinetype_id,
                min(mf.function_id) AS function_id,
                min(cp.machine_id) as machine_id,
                min(cp.design_id) as design_id,
                min(cp.beam_id) as beam_id,
                min(cp.date_time) as date_time,
                min(cp.date_time1) as date_time1,
                min(cp.mill_date) as mill_date,
                min(cp.mill_shift) as mill_shift,
                ROUND(AVG(case when mmf.vln_avg = '*' then cp.vln_avg * mmf.vln_avg_value when  mmf.vln_avg = '/' then cp.vln_avg / mmf.vln_avg_value else cp.vln_avg end ),2) AS vln_avg,
                ROUND(AVG(case when mmf.r_volt = '*' then cp.r_volt * mmf.r_volt_value when  mmf.r_volt = '/' then cp.r_volt / mmf.r_volt_value else cp.r_volt end ),2) AS r_volt,
                ROUND(AVG(case when mmf.y_volt = '*' then cp.y_volt * mmf.y_volt_value when  mmf.y_volt = '/' then cp.y_volt / mmf.y_volt_value else cp.y_volt end ),2) AS y_volt,
                ROUND(AVG(case when mmf.b_volt = '*' then cp.b_volt * mmf.b_volt_value when  mmf.b_volt = '/' then cp.b_volt / mmf.b_volt_value else cp.b_volt end ),2) AS b_volt,
                ROUND(AVG(case when mmf.vll_avg = '*' then cp.vll_avg * mmf.vll_avg_value when  mmf.vll_avg = '/' then cp.vll_avg / mmf.vll_avg_value else cp.vll_avg end ),2) AS vll_avg,
                ROUND(AVG(case when mmf.ry_volt = '*' then cp.ry_volt * mmf.ry_volt_value when  mmf.ry_volt = '/' then cp.ry_volt / mmf.ry_volt_value else cp.ry_volt end ),2) AS ry_volt,
                ROUND(AVG(case when mmf.yb_volt = '*' then cp.yb_volt * mmf.yb_volt_value when  mmf.yb_volt = '/' then cp.yb_volt / mmf.yb_volt_value else cp.yb_volt end ),2) AS yb_volt,
                ROUND(AVG(case when mmf.br_volt = '*' then cp.br_volt * mmf.br_volt_value when  mmf.br_volt = '/' then cp.br_volt / mmf.br_volt_value else cp.br_volt end ),2) AS br_volt,
                ROUND(AVG(case when mmf.r_current = '*' then cp.r_current * mmf.r_current_value when  mmf.r_current = '/' then cp.r_current / mmf.r_current_value else cp.r_current end ),2) AS r_current,
                ROUND(AVG(case when mmf.y_current = '*' then cp.y_current * mmf.y_current_value when  mmf.y_current = '/' then cp.y_current / mmf.y_current_value else cp.y_current end ),2) AS y_current,
                ROUND(AVG(case when mmf.b_current = '*' then cp.b_current * mmf.b_current_value when  mmf.b_current = '/' then cp.b_current / mmf.b_current_value else cp.b_current end ),2) AS b_current,
                ROUND(AVG(case when mmf.t_current = '*' then cp.t_current * mmf.t_current_value when  mmf.t_current = '/' then cp.t_current / mmf.t_current_value else cp.t_current end ),2) AS t_current,
                ROUND(AVG(case when mmf.t_watts = '*' then cp.t_watts * mmf.t_watts_value when  mmf.t_watts = '/' then cp.t_watts / mmf.t_watts_value else cp.t_watts end ),2) AS t_watts,
                ROUND(AVG(case when mmf.r_watts = '*' then cp.r_watts * mmf.r_watts_value when  mmf.r_watts = '/' then cp.r_watts / mmf.r_watts_value else cp.r_watts end ),2) AS r_watts,
                ROUND(AVG(case when mmf.y_watts = '*' then cp.y_watts * mmf.y_watts_value when  mmf.y_watts = '/' then cp.y_watts / mmf.y_watts_value else cp.y_watts end ),2) AS y_watts,
                ROUND(AVG(case when mmf.b_watts = '*' then cp.b_watts * mmf.b_watts_value when  mmf.b_watts = '/' then cp.b_watts / mmf.b_watts_value else cp.b_watts end ),2) AS b_watts,
                ROUND(AVG(case when mmf.t_var = '*' then cp.t_var * mmf.t_var_value when  mmf.t_var = '/' then cp.t_var / mmf.t_var_value else cp.t_var end ),2) AS t_var,
                ROUND(AVG(case when mmf.r_var = '*' then cp.r_var * mmf.r_var_value when  mmf.r_var = '/' then cp.r_var / mmf.r_var_value else cp.r_var end ),2) AS r_var,
                ROUND(AVG(case when mmf.y_var = '*' then cp.y_var * mmf.y_var_value when  mmf.y_var = '/' then cp.y_var / mmf.y_var_value else cp.y_var end ),2) AS y_var,
                ROUND(AVG(case when mmf.b_var = '*' then cp.b_var * mmf.b_var_value when  mmf.b_var = '/' then cp.b_var / mmf.b_var_value else cp.b_var end ),2) AS b_var,
                ROUND(AVG(case when mmf.t_voltampere = '*' then cp.t_voltampere * mmf.t_voltampere_value when  mmf.t_voltampere = '/' then cp.t_voltampere / mmf.t_voltampere_value else cp.t_voltampere end ),2) AS t_voltampere,
                ROUND(AVG(case when mmf.r_voltampere = '*' then cp.r_voltampere * mmf.r_voltampere_value when  mmf.r_voltampere = '/' then cp.r_voltampere / mmf.r_voltampere_value else cp.r_voltampere end ),2) AS r_voltampere,
                ROUND(AVG(case when mmf.y_voltampere = '*' then cp.y_voltampere * mmf.y_voltampere_value when  mmf.y_voltampere = '/' then cp.y_voltampere / mmf.y_voltampere_value else cp.y_voltampere end ),2) AS y_voltampere,
                ROUND(AVG(case when mmf.b_voltampere = '*' then cp.b_voltampere * mmf.b_voltampere_value when  mmf.b_voltampere = '/' then cp.b_voltampere / mmf.b_voltampere_value else cp.b_voltampere end ),2) AS b_voltampere,
                ROUND(AVG(case when mmf.avg_powerfactor = '*' then cp.avg_powerfactor * mmf.avg_powerfactor_value when  mmf.avg_powerfactor = '/' then cp.avg_powerfactor / mmf.avg_powerfactor_value else cp.avg_powerfactor end ),2) AS avg_powerfactor,
                ROUND(AVG(case when mmf.r_powerfactor = '*' then cp.r_powerfactor * mmf.r_powerfactor_value when  mmf.r_powerfactor = '/' then cp.r_powerfactor / mmf.r_powerfactor_value else cp.r_powerfactor end ),2) AS r_powerfactor,
                ROUND(AVG(case when mmf.y_powerfactor = '*' then cp.y_powerfactor * mmf.y_powerfactor_value when  mmf.y_powerfactor = '/' then cp.y_powerfactor / mmf.y_powerfactor_value else cp.y_powerfactor end ),2) AS y_powerfactor,
                ROUND(AVG(case when mmf.b_powerfactor = '*' then cp.b_powerfactor * mmf.b_powerfactor_value when  mmf.b_powerfactor = '/' then cp.b_powerfactor / mmf.b_powerfactor_value else cp.b_powerfactor end ),2) AS b_powerfactor,
                ROUND(AVG(case when mmf.powerfactor = '*' then cp.powerfactor * mmf.powerfactor_value when  mmf.powerfactor = '/' then cp.powerfactor / mmf.powerfactor_value else cp.powerfactor end ),2) AS powerfactor,
                
                ROUND(AVG(case when mmf.kvah = '*' then cp.kvah * mmf.kvah_value when  mmf.kvah = '/' then cp.kvah / mmf.kvah_value else cp.kvah end ),2) AS kvah,
                ROUND(SUM(case when mmf.kw = '*' then cp.t_watts * mmf.kw_value when  mmf.kw = '/' then cp.t_watts / mmf.kw_value else cp.t_watts end ),2) AS kw,
                ROUND(AVG(case when mmf.kvar = '*' then cp.kvar * mmf.kvar_value when  mmf.kvar = '/' then cp.kvar / mmf.kvar_value else cp.kvar end ),2) AS kvar,
                ROUND(AVG(case when mmf.power_factor = '*' then cp.power_factor * mmf.power_factor_value when  mmf.power_factor = '/' then cp.power_factor / mmf.power_factor_value else cp.power_factor end ),2) AS power_factor,
                ROUND(AVG(case when mmf.kva = '*' then cp.kva * mmf.kva_value when  mmf.kva = '/' then cp.kva / mmf.kva_value else cp.kva end ),2) AS kva,
                ROUND(AVG(case when mmf.frequency = '*' then cp.frequency * mmf.frequency_value when  mmf.frequency = '/' then cp.frequency / mmf.frequency_value else cp.frequency end ),2) AS frequency,
                min(cp.machine_status) as machine_status,
                min(cp.status) as status,
                min(cp.created_on) as created_on,
                min(cp.created_by) as created_by,
                min(cp.modified_on) as modified_on,
                min(cp.modified_by) as modified_by,
                
                ROUND(SUM(case when mmf.machine_kWh = '*' then cp.machine_kWh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.machine_kWh / mmf.machine_kWh_value else cp.machine_kWh end ),2) AS machine_kWh,
                ROUND(SUM(case when mmf.machine_kWh = '*' then cp.master_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.master_kwh / mmf.machine_kWh_value else cp.master_kwh end ),2) AS master_kwh,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kWh,
                 
                ROUND(SUM(case when mmf.machine_kWh = '*' then cp.reverse_machine_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.reverse_machine_kwh / mmf.machine_kWh_value else cp.reverse_machine_kwh end ),2) AS reverse_machine_kwh,
                ROUND(SUM(case when mmf.machine_kWh = '*' then cp.reverse_master_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.reverse_master_kwh / mmf.machine_kWh_value else cp.reverse_master_kwh end ),2) AS reverse_master_kwh,
                ROUND(SUM(case when mmf.kWh = '*' then cp.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then cp.reverse_kwh / mmf.kWh_value else cp.reverse_kwh end ),2) AS reverse_kwh,
                
                min(mm.ip_address) as ip_address,
                min(mm.port) as port,
                CASE WHEN min(cp.date_time) <= DATEADD(minute, -2, getdate()) THEN 'S' ELSE 'N' END as nocom,       
                ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_1,
                ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_2,
                ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.kWh = '*' then cp.kwh * mmf.kwh_value when  mmf.kwh = '/' then cp.kwh / mmf.kwh_value else cp.kwh end ELSE 0 END),2) AS kwh_3,
                ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),2) AS start_kwh_1,
                ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),2) AS start_kwh_2,
                ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.master_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.master_kwh / mmf.machine_kwh_value else cp.master_kwh end ELSE 0 END),2) AS start_kwh_3,     
                ROUND(SUM(CASE WHEN cp.mill_shift = 1 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),2) AS end_kwh_1,
                ROUND(SUM(CASE WHEN cp.mill_shift = 2 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),2) AS end_kwh_2,
                ROUND(SUM(CASE WHEN cp.mill_shift = 3 THEN case when mmf.machine_kwh = '*' then cp.machine_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.machine_kwh / mmf.machine_kwh_value else cp.machine_kwh end ELSE 0 END),2) AS end_kwh_3                    
            from
                ems_v1.dbo.current_power cp
                INNER JOIN [ems_v1].[dbo].[master_machine] mm ON cp.machine_id = mm.machine_id
                INNER JOIN [ems_v1].[dbo].[master_company] mc ON mm.company_id = mc.company_id
                INNER JOIN [ems_v1].[dbo].[master_branch] mb ON mm.branch_id = mb.branch_id
                INNER JOIN [ems_v1].[dbo].[master_department] md ON mm.department_id = md.department_id
                INNER JOIN [ems_v1].[dbo].[master_shed] ms ON mm.shed_id = ms.shed_id
                INNER JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mm.machinetype_id = mmt.machinetype_id 
                LEFT JOIN [ems_v1].[dbo].[master_function] mf ON mm.function_id = mf.function_id
                LEFT JOIN [ems_v1].[dbo].[master_function] mff ON mm.function2_id = mff.function_id
                LEFT JOIN [ems_v1].[dbo].[master_machine_factor] mmf ON mm.machine_id = mmf.machine_id
            {where} 
            group by cp.machine_id
            ''')
        print(sql)
        data = cnx.execute(sql).mappings().all()
        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(data)})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

@app.post("/energy_calculation_list2/")
async def energy_calculation_list2(cnx: Session = Depends(get_db)):
    
    try:

        query = text(f'''select * from [ems_v1].[dbo].[master_energy_calculations2]''')
        data = cnx.execute(query).mappings().all()
        cnx.commit()

        return JSONResponse({"iserror": False, "message": "data return sucessfully", "data":jsonable_encoder(data)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/save_energy_calculation2/")
async def save_energy_calculation2(obj :str = Form(""),
                                  cnx: Session = Depends(get_db)):
    
    try:
        print(obj)
        del_query=text(f'''DELETE FROM [ems_v1].[dbo].[master_energy_calculations2]''')
        cnx.execute(del_query)
        cnx.commit()
        obj_data = json.loads(obj)
        if obj !="":
            for row in obj_data:
                s_no = row["s_no"]
                group_name = row["group_name"]
                function_name = row["function_name"]
                formula1 = row["formula1"]
                formula2 = row["formula2"]
                parameter = row["parameter"]
                roundoff_value = row["roundoff_value"]
                query = text(f'''INSERT INTO [ems_v1].[dbo].[master_energy_calculations2] 
                                (s_no,group_name,function_name,formula1,formula2,parameter,roundoff_value)
                                values({s_no},'{group_name}','{function_name}','{formula1}','{formula2}','{parameter}','{roundoff_value}')''')

                cnx.execute(query)
                cnx.commit()
        return JSONResponse({"iserror": False, "message": "data save sucessfully", "data":""})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/get_custom_function_dashboard_detail2/")
async def get_custom_function_dashboard_detail2(cnx: Session = Depends(get_db)):

    try:
        mill_date = date.today()
        mill_shift = 0
        group_name = ''
        func_name = ''
        formula1 = ''
        results = ''

        sql1 = f'select * from ems_v1.dbo.master_shifts'
        data = cnx.execute(sql1).mappings().all()
        if len(data)>0:
            for row in data:
                mill_date = row["mill_date"]
                mill_shift = row["mill_shift"]

        sql2 = f'SELECT * FROM [ems_v1].[dbo].[master_energy_calculations2] ORDER BY group_name, s_no'
        result = cnx.execute(sql2).mappings().all()
        if len(result)>0:
            para = ''
            for rows in result:
                para = rows['parameter']

            if para == 'kw':
                para = "case when mmf.kw = '*' then p.t_watts * mmf.kw_value when  mmf.kw = '/' then p.t_watts / mmf.kw_value else p.t_watts end "

            if para == 'kWh':
                para = "case when mmf.kWh = '*' then p.kWh * mmf.kWh_value when  mmf.kWh = '/' then p.kWh / mmf.kWh_value else p.kWh end "
            
            sql3 = text(f'''
                    select 
                        p.machine_id,
                        min(mm.machine_name) as machine_name,
                        sum ({para}) as kWh 
                    from 
                        [ems_v1].[dbo].[current_power] p
                        left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id
                        left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id
                    where 
                        p.mill_date = '{mill_date}' and p.mill_shift = {mill_shift}
                    group by 
                        p.machine_id 
                    order by 
                        p.machine_id''')
            res = cnx.execute(sql3).mappings().all()
            machine_id_dict={}
            dict={}
            for row in res:
                dict[row['machine_id']] = row['kWh']
                machine_id_dict[row['machine_id']] = row['machine_name']
                
            datas = []

            for rows in result:
                group_name = rows['group_name']
                func_name = rows['function_name']
                formula = rows['formula2']
                formula1 = rows['formula1']
                results = eval(formula, {"dict": dict})
                machine_ids = re.findall(r'dict\[(\d+)\]', formula)

                valid_formula_machine_ids = [int(id) for id in machine_ids if int(id) in dict]

                formula_tooltip = {machine_id_dict[machine_id]: dict[machine_id] for machine_id in valid_formula_machine_ids}

                datas.append({"group_name": group_name,"function_name": func_name,"function_value": results,"formula1": formula1,"tooltip":formula_tooltip})

            return JSONResponse({"iserror":False,"message":"data returned successfully","data":jsonable_encoder(datas)})
        else:
            return JSONResponse({"iserror":False,"message":"data returned successfully","data":''})
       
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
                                     
@app.post("/current_power_ie/")
async def current_power_ie(machine_id : str = Form (''),
                           period_id: str = Form(''),
                           from_date: str = Form(''),
                           to_date: str = Form(''),                      
                           shift_id: int = Form(''),
                           report_for : str = Form(''),           
                           employee_id : str = Form(''),           
                           cnx: Session = Depends(get_db)):
    
    if period_id == '':
            return JSONResponse({"iserror": True, "message": "period id is required"})
    
    try:
        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
     
        machine_id = id(machine_id)
        
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}       
        
        where = "" 
        group_by = ''
        order_by = ''
        department_id = ''
        shed_id = ''
        machinetype_id = ''
        if machine_id == 'all' or machine_id == '':
            pass
        else:
            where += f" and mm.machine_id in ({machine_id})"
        if  employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]
                    
        if department_id !='' and department_id !=0:
            where += f" and md.department_id ={department_id}"
        if shed_id !='' and shed_id != 0:
            where += f" and ms.shed_id ={shed_id}"
        if machinetype_id !='' and machinetype_id!= 0:
            where += f" and mmt.machinetype_id ={machinetype_id}"

        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_shifts] WHERE status = 'active' ''')
        data1 = cnx.execute(query).mappings().all()
        mill_date = date.today()
        mill_shift = 0
        table_name = ''
        
        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]  

        field_name_import = '''mill_date, mill_shift, machine_id, master_kwh as start_kwh, machine_kwh as end_kwh, kwh, 'Import' as kwh_type '''
        field_name_export = '''mill_date, mill_shift, machine_id,reverse_master_kwh as start_kwh, reverse_machine_kwh as end_kwh, reverse_kwh as kwh, 'Export' as kwh_type'''       

        if period_id == 'cur_shift':            
            table_name = f'(select {field_name_import} from [ems_v1].[dbo].[current_power] UNION All select {field_name_export} from [ems_v1].[dbo].[current_power])cp'
            where += f" and cp.mill_date = '{mill_date}' and cp.mill_shift ='{mill_shift}' "

        elif period_id == 'sel_shift' or period_id == 'sel_date':
            if from_date == '':
                return JSONResponse({"iserror":True,"message":"date is required"}) 
            
            mill_date=parse_date(from_date)             
            month_year=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""
            table_name=f"[ems_v1_completed].[dbo].[power_{month_year}]" 
            where += f" and cp.mill_date = '{mill_date}' "

            table_name = f'(select {field_name_import} from {table_name} UNION All select {field_name_export} from {table_name})cp'

            if period_id == 'sel_shift':
                if shift_id == '':
                    return JSONResponse({"iserror":True,"message":"shift is required"}) 
                where += f" and cp.mill_shift ='{shift_id}' " 
            
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
            print(query)
            result_query = cnx.execute(query).mappings().all()
            if len(result_query)>0:
                pass
            else:
                return JSONResponse({"iserror": True, "message": "power table not available..."})    
   
        elif period_id == "from_to":            
            if from_date == '':
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date == '':
                return JSONResponse({"iserror": True, "message": "to_date is required"})  
                    
            from_date = parse_date(from_date)
            to_date =  parse_date(to_date)
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
        
            where += f''' and  cp.mill_date  >= '{from_date}' and cp.mill_date <= '{to_date}' '''
            
            if shift_id != "":                
                where += f''' and cp.mill_shift = '{shift_id}' ''' 
            
            if from_date.month == to_date.month:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_analysis_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    
                table_name=f"[ems_v1_completed].[dbo].[power_{month_year}]" 
                table_name = f'(select {field_name_import} from {table_name} UNION All select {field_name_export} from {table_name})cp'
            else:
                from_month = from_date.month
                to_month = to_date.month
                month_year_range = [
                        (from_date + timedelta(days=30 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                union_queries_export = []
                union_queries_import = []

                for month_year in month_year_range:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    print(query)
                    if len(result_query) > 0:
                        table_name = f"[ems_v1_completed].[dbo].[power_{month_year}]"
                        union_queries_export.append(f"SELECT {field_name_export} FROM {table_name}")
                        union_queries_import.append(f"SELECT {field_name_import} FROM {table_name}")

                subquery_union_import = " UNION ALL ".join(union_queries_import)
                subquery_union_export = " UNION ALL ".join(union_queries_export)
                table_name = f"({subquery_union_import} union all {subquery_union_export}) cp"
    
        if report_for == 'detail' or report_for == '':
            group_by = " ,cp.mill_date, cp.mill_shift"
            order_by = " ,cp.mill_date, cp.mill_shift" 
        
        if report_for == 'summary':
            group_by = " ,cp.mill_date" 
            order_by = " ,cp.mill_date"   
                       
        if group_by != "":
            group_by = f"{group_by} "    
        if order_by != "":
            order_by = f"{order_by}"

        query = text(f'''
                select 
                    min(cp.mill_date) mill_date,
                    min(cp.mill_shift) mill_shift,
                    min(cp.machine_id) machine_id,
                    min(mm.machine_name) machine_name,
                    min(mm.machine_code) machine_code,
                    cp.kwh_type,
                    ROUND(SUM(
                        case when cp.kwh_type = 'Export' then
                            case when mmf.reverse_machine_kwh = '*' then cp.start_kwh * mmf.reverse_machine_kwh_value when  mmf.reverse_machine_kwh = '/' then cp.start_kwh / mmf.reverse_machine_kwh_value else cp.start_kwh end 
                        else
                            case when mmf.machine_kwh = '*' then cp.start_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.start_kwh / mmf.machine_kwh_value else cp.start_kwh end 
                        end
                     ),2) AS start_kwh,
                    ROUND(SUM(
                        case when cp.kwh_type = 'Export' then
                            case when mmf.reverse_machine_kwh = '*' then cp.end_kwh * mmf.reverse_machine_kwh_value when  mmf.reverse_machine_kwh = '/' then cp.end_kwh / mmf.reverse_machine_kwh_value else cp.end_kwh end
                        else
                            case when mmf.machine_kwh = '*' then cp.end_kwh * mmf.machine_kwh_value when  mmf.machine_kwh = '/' then cp.end_kwh / mmf.machine_kwh_value else cp.end_kwh end
                        end
                     ),2) AS end_kwh,
                    ROUND(SUM(
                        case when cp.kwh_type = 'Export' then
                            case when mmf.reverse_kwh = '*' then cp.kWh * mmf.reverse_kwh_value when  mmf.reverse_kwh = '/' then cp.kWh / mmf.reverse_kwh_value else cp.kWh end 
                        else
                            case when mmf.kwh = '*' then cp.kWh * mmf.kwh_value when  mmf.kwh = '/' then cp.kWh / mmf.kwh_value else cp.kWh end 
                        end
                     ),2) AS kwh
                     
                from
                    {table_name}
                    INNER JOIN [ems_v1].[dbo].[master_machine] mm ON cp.machine_id = mm.machine_id and mm.import_export = 'yes'
                    left JOIN [ems_v1].[dbo].[master_machine_factor] mmf ON mm.machine_id = mmf.machine_id  
                    LEFT JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id = mm.shed_id                   
                    LEFT JOIN [ems_v1].[dbo].[master_department] md ON md.department_id = mm.department_id                   
                    LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id = mm.machinetype_id 
                    where 1=1 
                    {where}
                    group by mm.machine_id {group_by}, cp.kwh_type
                    order by mm.machine_id {order_by}
                    ''')
   
        createFolder("Log/","current_power_ie api query "+str(query))
        data = cnx.execute(query).mappings().all()
        return JSONResponse({"iserror":False, "message":"data return successfully", "data" : jsonable_encoder(data)})    
    except Exception as e:      
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
 
@app.post("/save_energy_detail/")
async def save_energy_detail(obj :str = Form(''),                                    
                             cnx: Session = Depends(get_db)):
    print(obj)
    try:
        if obj is not None:
            obj_data = json.loads(obj)
            for row in obj_data:
                id = row["id"]
                machine_id = row["machine_id"]  
                mill_date = row["mill_date"]
                initial_kwh = row["initial_kwh"]
                shift1_kwh = row["shift1_kwh"]
                shift2_kwh = row["shift2_kwh"]
                shift3_kwh = row["shift3_kwh"]
                user_login_id = row["user_login_id"]
                if id == '':
                    sql = text(f'''insert into ems_v1.dbo.energy (machine_id,mill_date,initial_kwh,shift1_kwh,shift2_kwh,shift3_kwh,created_on,created_by)
                              values('{machine_id}','{mill_date}','{initial_kwh}','{shift1_kwh}','{shift2_kwh}','{shift3_kwh}',getdate(), '{user_login_id}')''')
                else:
                    sql = text(f'''update ems_v1.dbo.energy set mill_date = '{mill_date}', initial_kwh = '{initial_kwh}',
                              shift1_kwh = '{shift1_kwh}', shift2_kwh = '{shift2_kwh}', shift3_kwh = '{shift3_kwh}', modified_on = getdate(),
                              modified_by = '{user_login_id}',machine_id = '{machine_id}'
                              where id = '{id}' ''')
                cnx.execute(sql)
                cnx.commit()
        return JSONResponse({"iserror":False,"message":"data save successfully","data":''})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})

def generate_year_wise_excel_report(result, year,next_year,report_for):
    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "YearWiseReport_templete.xlsx"))
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'EMS'
    if report_for == '6to6':
        report = '06:00 to 06:00'
    else:
        report = '12:00 to 12:00'
    cell = "B1"
    data = f"YEAR WISE ENERGY CONSUMPTION REPORT FOR KWH - {year}({report})"
    sheet[cell] = data
    font = Font(bold=True, name='Calibri', size=13)
    sheet[cell].font = font

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    minimum_column_width = 15
    cell = "A3"
    sheet[cell] = "Machine Code"
    cell = "B3"
    sheet[cell] = "Machine Name"
    fill_cyan = PatternFill(start_color='309490', end_color='309490', fill_type='solid')
    sheet.cell(row=3, column=1, value="Machine Code").fill = fill_cyan
    sheet.cell(row=3, column=2, value="Machine Name").fill = fill_cyan
    if result == []:
        print(1)
        cell = "E10"
        data = "No Data"
        sheet[cell] = data
        font = Font(bold=True, name='Calibri', size=13)
        sheet[cell].font = font
    else:
    # Populate month-year headers in row 3 starting from column C
        col_index = 3
        for month_year in result[0].keys():
            if month_year not in ["machine_code", "machine_name"]:
                sheet.cell(row=3, column=col_index, value=month_year).fill = fill_cyan
                col_index += 1

        for row, data in enumerate(result, start=4):
            machine_code = data["machine_code"]
            machine_name = data["machine_name"]
            sheet.cell(row=row, column=1, value=machine_code).alignment = Alignment(horizontal="center")
            alignment = Alignment(horizontal="center", wrap_text=True)
            sheet.cell(row=row, column=2, value=machine_name).alignment = alignment
            # machine_name_length = len(machine_name)
            # column_letter_c = get_column_letter(3)
            # column_width_c = max(machine_name_length, sheet.column_dimensions[column_letter_c].width)
            # sheet.column_dimensions[column_letter_c].width = column_width_c

            col_indexs = 1
            for kwh_value in data.values():
                if isinstance(kwh_value, (int, float)):
                    if kwh_value == 0:
                        kwh_value = ''
                    sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
                    cell = sheet.cell(row=row, column=col_indexs, value=kwh_value).alignment = Alignment(horizontal="center")
                    cell.number_format = '0.00'
                col_indexs += 1

    file_name = f'YearWiseReport-{year}-{next_year}.xlsx'
    file_path = os.path.join(base_path, file_name)
    workbook.save(file_path)

@app.post("/year_wise_report/")
async def year_wise_report(request: Request,
                           machine_id: str = Form(''),
                           year: str = Form(""),
                           report_for : str  = Form(""),
                           employee_id: str = Form(""),
                           is_critical: str = Form(""),
                           cnx: Session = Depends(get_db)):

    if year  == '':
        return JSONResponse({"iserror": True, "message": "year is required"})

    try:
        groupby = ""
        where = ""
        result = ''
        output = {}
        department_id = ''
        shed_id = ''
        machinetype_id = ''
        
        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
     
        machine_id = id(machine_id)
        if machine_id == "" or machine_id == 'all':
            pass
        else:
            where += f" and mm.machine_id IN ({machine_id})"

        if  employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]

        if is_critical != '':
            where += f" and mm.major_nonmajor = '{is_critical}'"

        if department_id !='' and department_id !=0:
            where += f" and md.department_id ={department_id}"
        if shed_id !=''and shed_id != 0:
            where += f" and ms.shed_id ={shed_id}"
        if machinetype_id !='' and machinetype_id !=0:
            where += f" and mmt.machinetype_id ={machinetype_id}"

        mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        tables_to_union = []
        for month in range(4, 13):
            month_year = f"{mill_month[month]}{year}"
            print(month_year)
            if report_for == '12to12':
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all()

                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}_12")
                print(month_year)
            else:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all()

                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}")
        
        next_year = int(year) + 1
        mill_month = {1: "01", 2: "02", 3: "03"}

        for month in range(1, 4):
            month_year = f"{mill_month[month]}{next_year}"
            print(month_year)
            if report_for == '12to12':
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all() 
                print("result_query",result_query)
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}_12")
                tables_union_query = " UNION ALL ".join(tables_to_union)
            else:   
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all() 
                print("result_query",result_query)
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}")
                tables_union_query = " UNION ALL ".join(tables_to_union)
                print("tables_union_query",tables_union_query)

        if len(tables_union_query)==0:
            return JSONResponse({"iserror": True, "message": "table not available"})
        
        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh,
                FORMAT(min(cp.mill_date), 'MM-yyyy') AS mill_date
            FROM
                ({tables_union_query}) cp
                INNER JOIN ems_v1.dbo.master_machine mm ON mm.machine_id = cp.machine_id
                LEFT JOIN ems_v1.dbo.master_machine_factor mmf ON mmf.machine_id = mm.machine_id
                LEFT JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id = mm.shed_id                   
                LEFT JOIN [ems_v1].[dbo].[master_department] md ON md.department_id = mm.department_id                   
                LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id = mm.machinetype_id                                      
            WHERE
                1=1  and mm.status = 'active' {where}
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                MONTH(cp.mill_date), 
                YEAR(cp.mill_date) 
            ORDER BY 
                min(cp.mill_date),  
                min(cp.machine_id)
        ''')

        print(query)
        rslt = cnx.execute(query).mappings().all()
        if len(rslt)>0:
          output = {}  # Initialize the output dictionary
        
        output_keys = [
            f"04-{year}", f"05-{year}", f"06-{year}",
            f"07-{year}", f"08-{year}", f"09-{year}",
            f"10-{year}", f"11-{year}", f"12-{year}",
            f"01-{next_year}", f"02-{next_year}", f"03-{next_year}"
        ]
        
        for row in rslt:
            machine_code = row['machine_code']
            machine_name = row['machine_name']
            mill_date = row['mill_date']
            kwh = row['kwh']
            
            if machine_code not in output:
                output[machine_code] = {
                    'machine_code': machine_code,
                    'machine_name': machine_name
                }
                for key in output_keys:
                    output[machine_code][key] = 0
            
            output[machine_code][mill_date] = kwh
        
        result = list(output.values())
               
        generate_year_wise_excel_report(result, year,next_year,report_for)
            # process_data(month_year, result)
        file_path = os.path.join(base_path, f"YearWiseReport-{year}-{next_year}.xlsx")
        results = f"http://{request.headers['host']}/attachment/YearWiseReport-{year}-{next_year}.xlsx"

        if os.path.exists(file_path):

            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": results})
        else:
            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": None})
        # return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": jsonable_encoder(result)})

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})
    
def year_wise_excel_report(machine_data, year,next_year,mill_date,year_record,report_type,res):
    try:
        print("res",res)
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  
        f_count = 0
        sheet = workbook.create_sheet(title="Energy Electrical Report")
        print(year_record)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        style = {'font': Font(color='000000', size=11, name='Calibri'),
                'border': border,
                'alignment': Alignment(vertical='center',horizontal='center')}
            
        style1 = {'font': Font(bold=True,color='000000', size=12, name='Calibri'),
            'border': border,
            'alignment': Alignment(vertical='center',horizontal='center')}
        
        style2 = {
            'fill': PatternFill(fill_type='solid', fgColor='f1ff52'),
            'font': Font(bold=True, color='000000', size=13, name='Calibri'),
            'border': border,
            'alignment': Alignment(vertical='center',horizontal='right',wrap_text=True)
        }
        style3 = {
            'fill': PatternFill(fill_type='solid', fgColor='f6c492'),
            'font': Font(bold=True, color='000000', size=13, name='Calibri'),
            'border': border,
            'alignment': Alignment(vertical='center',horizontal='center',wrap_text=True)
        }
        # script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        script_dir =  os.path.join(os.path.dirname(__file__), "..", "..","..")

        machine_name_to_row_mapping = {}
        if len(res) == 0:
            cell = "O10"
            data = "No Data"
            sheet[cell] = data

            alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell].alignment = alignment

            sheet.column_dimensions[cell[0]].width = len(data) + 2  # Adjust column width

            font = Font(name='Calibri', size=25)
            sheet[cell].font = font
        else:
            file_name = "tnpl.png"
            image_path = os.path.join(script_dir, file_name)
            img = Image(image_path)
            img.anchor = "C3"  
            
            img_width_pixels, img_height_pixels = img.width, img.height
    
            desired_width_pixels = 165
            desired_height_pixels = 40
    
            width_scale = desired_width_pixels / img_width_pixels
            height_scale = desired_height_pixels / img_height_pixels

            img.width = int(img_width_pixels * width_scale)
            img.height = int(img_height_pixels * height_scale)
            # sheet.column_dimensions['C'].width = img.width / 10
            # sheet.row_dimensions[1].height = img.height / 2
            # sheet.row_dimensions[2].height = img.height / 2
            sheet.merge_cells('C3:D4')  
            sheet.add_image(img)
            sheet['C3'].alignment = style['alignment']

            cell = "C5"
            sheet[cell] = "Energy in Units"
            sheet[cell].border = border
            sheet.merge_cells('C5:D6')
            sheet['C5'].alignment = style['alignment']
            sheet['C5'].font = style1['font']
            row_number = 7

            start_year = int(year)
            end_year = start_year + 1
            year_str = str(start_year)
            end_str = str(end_year)
            current_year = mill_date[6:10]
            current_month = mill_date[3:5]
            cur_year = int(mill_date[6:10])
            cur_month = int(mill_date[3:5])
            cur_day= int(mill_date[:2])
            print("cur_day",cur_day)

            if current_year == year:
                financial_year_months = [
            f"{month:02d}{year}" for month in range(4, cur_month)
        ]
            elif current_year == next_year:
                if cur_month <3:
                    financial_year_months = [
                    f"{month:02d}{year_str}" for month in range(4, 13)
                ] + [
                    f"{month:02d}{end_str}" for month in range(1, cur_month)
                ]
                else:
                    financial_year_months = [
                        f"{month:02d}{year_str}" for month in range(4, 13)
                    ] + [
                        f"{month:02d}{end_str}" for month in range(1, 3)
                    ]
            else:
                financial_year_months = [
                        f"{month:02d}{year_str}" for month in range(4, 13)
                    ] + [
                        f"{month:02d}{end_str}" for month in range(1, 3)
                    ]
                
            print("Financial Year Months:", financial_year_months)

            row_number = 6  
            col_number = 5  
            col_number1 = 3  
            row_number1 = 7 
            row_number2 = 7 
            
            # if machine_data == []:
            #     cell = "E10"
            #     data = "No Data"
            #     sheet[cell] = data
            #     font = Font(bold=True, name='Calibri', size=13)
            #     sheet[cell].font = font
            
            for i in financial_year_months:
                cell = sheet.cell(row=row_number, column=col_number, value=i)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, name='Calibri', size=12)
                column_letter = get_column_letter(col_number)
                sheet.column_dimensions[column_letter].width = 15
                col_number += 1
        
            for machine_name, date_values in machine_data["year_record"].items():
                sheet.cell(row=row_number1, column=3, value=machine_name)
                sheet.cell(row=row_number1, column=3).font = style['font']
                sheet.merge_cells(
                    start_row=row_number1, start_column=3, end_row=row_number1, end_column=3 + 1)
                machine_name_to_row_mapping[machine_name] = row_number1 
                row_number1 += 1
                col_number1 = 5  

                machine_name_length = len(machine_name) + 10
                column_letter_c = get_column_letter(3) 
                column_width_c = max(machine_name_length, sheet.column_dimensions[column_letter_c].width)
                sheet.column_dimensions[column_letter_c].width = column_width_c

                for i in financial_year_months:
                    if i in date_values:        
                        cell_value = date_values[i]["formulas"]
                        roundoff_value_month = date_values[i].get("roundoff_value_month", 0) 
                        if roundoff_value_month == 0: # Access roundoff_value_month roundoff_value_month
                            sheet.cell(row=row_number2, column=col_number1, value=f"=ABS({cell_value})")
                            cell = sheet.cell(row=row_number2, column=col_number1, value=f"=MROUND(ABS(({cell_value})),{roundoff_value_month})")
                        else:
                            sheet.cell(row=row_number2, column=col_number1, value=f"=MROUND(ABS(({cell_value})),{roundoff_value_month})")
                            cell = sheet.cell(row=row_number2, column=col_number1, value=f"=MROUND(ABS(({cell_value})),{roundoff_value_month})")
                        cell.number_format = '0.00'
                        sheet.row_dimensions[1].height = 100
                        sheet.column_dimensions["AW"].width = 20
                
                    col_number1 += 1
                row_number2 +=1
        
            rows = row_number2+1
            rowno = row_number2+2
            month_mapping = {
                "01": "Jan",
                "02": "Feb",
                "03": "Mar",
                "04": "Apr",
                "05": "May",
                "06": "June",
                "07": "July",
                "08": "Aug",
                "09": "Sept",
                "10": "Oct",
                "11": "Nov",
                "12": "Dec"
            }
            
            updated_financial_year_months = [
                f"{month_mapping[month[:2]]}'{month[4:]+'(AVG)'}" for month in financial_year_months
            ]

            row = 6
            col = 5
            for i in updated_financial_year_months:
                cell = sheet.cell(row=row, column=col, value=i)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, name='Calibri', size=12)
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 17
                col += 1
            
            month_year = mill_date[3:]
            output_keys = [f'{day}-{month_year}' for day in range(1, cur_day+1)]
            rows= 6
            column=col_number1

            for day_date in output_keys:
                cell = sheet.cell(row=rows, column=column, value=day_date)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, name='Calibri', size=12)
                column_letter = get_column_letter(column)
                sheet.column_dimensions[column_letter].width = 17
                column += 1

            cell = sheet.cell(row=6, column=column, value="Monthly")

            cell.alignment = style['alignment']
            cell.font = Font(bold=True, name='Calibri', size=12)
            cols=column
            cell = sheet.cell(row=6, column=column+1, value=f"FY  {year}-{next_year}")
            cell.alignment = style['alignment']
            cell.font = Font(bold=True, name='Calibri', size=12)
            
            date_to_col = {date: col for col, date in enumerate(output_keys, start=col_number1)}
            
            # for machine_name, date_values in machine_data["current_month_record"].items():
            #     machine_row = machine_name_to_row_mapping[machine_name]
            #     row_formula = []  # To store SUM formula parts for each date

            #     for day_date, value in date_values.items():
            #         formula = value['formula']
            #         roundoff_value_day = value.roundoff_value_day
            #         # roundoff_value_day = date_values[next(iter(date_values))]["roundoff_value_day"]
            #         formatted_date = f"{int(day_date[:2])}-{day_date[3:]}"
            #         col_idx = date_to_col.get(formatted_date, None)

            #         if col_idx is not None:
                        
            #             sheet.cell(row=machine_row, column=col_idx, value=f"=MROUND(({formula}),{roundoff_value_day})")
            #             cell = sheet.cell(row=machine_row, column=col_idx, value=f"=MROUND(({formula}),{roundoff_value_day})")
            #             cell.number_format = '0.00'
            #             row_formula.append(f'{sheet.cell(row=machine_row, column=col_idx).coordinate}')

            #     if row_formula:
            #         sum_formula = '=SUM(' + ':'.join(row_formula) + ')'
            #         sheet.cell(row=machine_row, column=column).value = sum_formula
            #     rows += 1 
            for machine_name, date_values in machine_data["current_month_record"].items():
                machine_row = machine_name_to_row_mapping[machine_name]
                row_formula = [] 

                for day_date, value in date_values.items():
                    formula = value["formula"]
                    roundoff_value_day = value["roundoff_value_day"]  

                    formatted_date = f"{int(day_date[:2])}-{day_date[3:]}"
                    col_idx = date_to_col.get(formatted_date, None)

                    if col_idx is not None:
                        if roundoff_value_day == 0:
                            formula_with_mround = f"=ABS({formula})"
                        else:
                            formula_with_mround = f"=MROUND(ABS(({formula})),{roundoff_value_day})"
                        sheet.cell(row=machine_row, column=col_idx, value=formula_with_mround)
                        cell = sheet.cell(row=machine_row, column=col_idx)
                        cell.number_format = '0.00'

                        row_formula.append(f'{cell.coordinate}')

                if row_formula:
                    sum_formula = '=SUM(' + ':'.join(row_formula) + ')'
                    sheet.cell(row=machine_row, column=column).value = sum_formula

                rows += 1

            row = 7
            # for machine_name, value in year_record.items():# add year record 
            #     cell = sheet.cell(row=row, column=column+1)
            #     cell.value = f"=MROUND(({value}),{roundoff_value_year})"
            #     cell.number_format = '0.00'
            #     cell.alignment = Alignment(horizontal='center', vertical='center')
            #     cell.font = Font(name='Calibri', size=11)
            #     column_letter = get_column_letter(column+1)
            #     sheet.column_dimensions[column_letter].width = 19
            #     row += 1
            for func_name, data in year_record.items():
                formula_y = data['formula_y']
                roundoff_value_year = data['roundoff_value_year']
                cell = sheet.cell(row=row, column=column+1)
                if roundoff_value_year == 0:
                    cell.value = f"=ABS({formula_y})"            
                else:
                    cell.value = f"=MROUND(ABS({formula_y}), {roundoff_value_year})"            
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calibri', size=11)
                column_letter = get_column_letter(column+1)
                sheet.column_dimensions[column_letter].width = 19
                row += 1
            if report_type == '6to6':
                report = '06:00 to 06:00'
            else:
                report = 'Day'
            columns = 3
            cell_value = f"Daily Energy Electrical  Report-({report})"
            sheet.cell(row=4, column=columns+2).value = cell_value
            sheet.cell(row=4, column=columns+2).alignment = style2['alignment']
            sheet.merge_cells(start_row=4, start_column=5, end_row=4, end_column=column+1)
            sheet.alignment = style1['alignment']
            sheet.cell(row=4, column=columns+2).font = style2['font']

            columns = 3
            cell_value = " TNPL Unit - II   - Energy - Electrical"
            sheet.cell(row=3, column=columns+2).value = cell_value
            sheet.cell(row=3, column=columns+2).alignment = style2['alignment']
            sheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=column+1)
            sheet.cell(row=3, column=columns+2).font = style2['font']

            sheet.cell(row=5, column=column+1).value = sheet.cell(row=6, column=column-1).value
            alignment = Alignment(horizontal='center', vertical='center')
            # fill =   PatternFill(fill_type='solid', fgColor='f6c492')
            font = Font(bold = True,size=12)  

            sheet.cell(row=5, column=column + 1).alignment = alignment
            sheet.cell(row=5, column=column + 1).font = font
            column = column+2
            merge_range = f"C{row_number2}:{openpyxl.utils.get_column_letter(column-1)}{row_number2}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=row_number2, column=3)
            sheet.row_dimensions[row_number2].height = 25

            merged_cell.value = "REMARKS: "
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')

            column = column+2
            merge_range = f"C{rowno}:D{rowno}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=rowno, column=3)
            sheet.row_dimensions[rows].height = 25

            merged_cell.value = "Shift Incharge-(Elect)"
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            merge_range = f"E{rowno}:{openpyxl.utils.get_column_letter(column-5)}{rowno}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=rowno, column=5)
            sheet.row_dimensions[rowno].height = 25

            merged_cell.value = "Sectional Incharge-(Elect)"
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')

            merge_range = f"{openpyxl.utils.get_column_letter(column-4)}{rowno}:{openpyxl.utils.get_column_letter(column-3)}{rowno}"
            sheet.merge_cells(merge_range)

            merged_cell = sheet.cell(row=rowno, column=column-4)
            sheet.row_dimensions[rowno].height = 25
            merged_cell.value = "HOD"
            merged_cell.font = Font(bold=True, name='Calibri', size=12)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            rows_count = 3																																										
            row_range = sheet.iter_rows(min_row=rows_count, max_row=rowno, min_col=3, max_col=column-3)
            for row in row_range:
                for cell in row:
                    cell.border = border

        file_name = f'YearReport-{year}-{next_year}.xlsx'
        file_path = os.path.join(base_path, file_name)
        workbook.save(file_path)

    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)
    
@app.post("/year_report/")
async def year_report(request: Request,
                      year: str = Form(""),
                      report_type: str = Form(""),
                      cnx: Session = Depends(get_db)):

    if year  == '':
        return JSONResponse({"iserror": True, "message": "year is required"})
    
    if report_type  == '':
        return JSONResponse({"iserror": True, "message": "report_type is required"})

    if report_type  == '6to6' and report_type  =='12to12':
        return JSONResponse({"iserror": True, "message": "invalid report_type "})
    
    try:
        
        where = ""
        type = ''
        mill_date = date.today()
        func_name = ''
        formula = ''
        machine_name = ''
        
        kwh = 0
        reverse_kwh = 0
        machine_id = 0
        query = f'''SELECT * FROM [ems_v1].[dbo].[master_energy_calculations] ORDER BY s_no '''
        result = cnx.execute(query).mappings().all()
        if len(result)>0:

            query = text(f'''select mill_date from ems_v1.dbo.master_shifts''')
            shift=cnx.execute(query).mappings().all()
    
            for row in shift :
                mill_date = row['mill_date']  
                
            next_year = int(year) + 1
            mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
            tables_to_union = []
            for month in range(4, 13):
                month_year = f"{mill_month[month]}{year}"

                if report_type == "12to12":
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                    result_query = cnx.execute(query).mappings().all()
                else:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                    result_query = cnx.execute(query).mappings().all()
                if len(result_query) > 0:
                    tables_to_union.append(f"{month_year}")
            
            for month in range(1, 4):
                month_year = f"{mill_month[month]}{next_year}"

                if report_type == "12to12":
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                    result_query = cnx.execute(query).mappings().all() 
                else:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                    result_query = cnx.execute(query).mappings().all() 
                if len(result_query) > 0:
                    tables_to_union.append(f"{month_year}")
            print("tables_to_union",tables_to_union)

            if len(tables_to_union)==0:
                return JSONResponse({"iserror": True, "message": "table not available"})
            
            if report_type =="12to12":
                type = f"_12 cp"
            else:
                type = f" cp"

            result_dict= {}
            
            for table_name in tables_to_union:
                query = text(f'''
                SELECT
                    mm.machine_name AS machine_name,
                    min(mm.machine_id) machine_id ,
                    ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh,
                    SUM(case when mmf.kWh = '*' then cp.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then cp.reverse_kwh / mmf.kWh_value else cp.reverse_kwh end ) AS reverse_kwh,
                    FORMAT(min(cp.mill_date), 'dd-MM-yyyy') AS mill_date
                FROM
                    ems_v1_completed.dbo.power_{table_name}{type}
                    INNER JOIN ems_v1.dbo.master_machine mm ON mm.machine_id = cp.machine_id
                    LEFT JOIN ems_v1.dbo.master_machine_factor mmf ON mmf.machine_id = mm.machine_id
                WHERE
                    1=1 {where}
                GROUP BY   
                    mm.machine_name,   
                    DAY(cp.mill_date) 
                ORDER BY 
                    min(cp.mill_date),  
                    min(cp.machine_id)
            ''')
                
                rslt = cnx.execute(query).mappings().all()
                result_dict[table_name] = rslt
            # if len(result_dict[table_name])==0:
            #     return JSONResponse({"iserror": False, "message": "no data available"})
            results = {}
            dict2 = {}
            dict2_r = {}
            roundoff_value_day = 1
            roundoff_value_month = 10
            roundoff_value_year = 1
            formula = 0
            roundoff_values={}

            for table_name, table_data in result_dict.items():
                table_dict = {}
                for row in table_data:
                    machine_id = row['machine_id']
                    mill_date = row['mill_date']
                    kwh = row['kwh']
                    reverse_kwh = row['reverse_kwh']
                
                    if table_name not in dict2:
                        dict2[table_name] = {}
                        dict2_r[table_name] = {}

                    if mill_date not in dict2[table_name]:
                        dict2[table_name][mill_date] = {}
                        dict2_r[table_name][mill_date] = {}

                    if machine_id not in dict2[table_name][mill_date]:
                        dict2[table_name][mill_date][machine_id] = kwh
                        dict2_r[table_name][mill_date][machine_id] = reverse_kwh
                    # print("dict2",dict2)
                    
                    for row in result:
                        func_name = row['function_name']
                        formula = row['formula2']
                        roundoff_value_month = row['roundoff_value']
                        if table_name not in results:
                            results[table_name] = {}

                        if mill_date not in results[table_name]:
                            results[table_name][mill_date] = {}
                        
                        if len(table_data) == 0:
                            formula = 0
                        else:
                            if func_name == 'Power Import':
                                if table_name in dict2 and mill_date in dict2[table_name]:
                                    formula = dict2[table_name][mill_date].get(14, 0)  
                                else:
                                    formula = 0

                            elif func_name == 'Power Export':
                                if table_name in dict2_r and mill_date in dict2_r[table_name]:
                                    formula = dict2_r[table_name][mill_date].get(14, 0)
                                else:
                                    formula = 0
                            else:
                                formula = re.sub(r'dict\[(\d+)\]', lambda match: str(dict2.get(table_name, {}).get(mill_date, {}).get(int(match.group(1)), 0)), formula)
                                formula = eval(formula)
                        results[table_name][mill_date][func_name] = formula
                        roundoff_values[(table_name, mill_date, func_name)] = roundoff_value_month # Add roundoff_value_month

            aggregated_results = {}
        
            for table_name, funcs in results.items():
                for mill_date, formula_result in funcs.items():
                    for func_name, value in formula_result.items():
                        if table_name not in aggregated_results:
                            aggregated_results[table_name] = {}

                        if func_name not in aggregated_results[table_name]:
                            aggregated_results[table_name][func_name] = {"formula": 0.00, "count": 0, "roundoff_value_month": 0}     

                        if aggregated_results[table_name][func_name]["formula"] == 0.00:
                            aggregated_results[table_name][func_name]["formula"] = value
                        else:
                            aggregated_results[table_name][func_name]["formula"] +=value
                        aggregated_results[table_name][func_name]["count"] += 1
                        roundoff_value_month = roundoff_values.get((table_name, mill_date, func_name), 0)
                        aggregated_results[table_name][func_name]["roundoff_value_month"] = roundoff_value_month

            createFolder("YearReport_Log/", "aggregated_results " + f'{aggregated_results}')
            machine_data = {"year_record": {}}

            for table_name, functions in aggregated_results.items():
                for func_name, values in functions.items():
                    kwh = values["formula"]
                    count = values["count"]
                    roundoff_value_month = values["roundoff_value_month"]  
                    
                    if func_name not in machine_data["year_record"]:
                        machine_data["year_record"][func_name] = {}

                    avg_kwh = kwh / count if count > 0 else kwh 
                    machine_data["year_record"][func_name][table_name] = {"roundoff_value_month": roundoff_value_month, "formulas": avg_kwh}
                
            current_month = list(result_dict.keys())[-1]
            machine_data["current_month_record"] = {}

            for row in result_dict[current_month]:
                machine_name = row['machine_name']
                mill_date = row['mill_date']
                kwh = row['kwh']
                reverse_kwh = row['reverse_kwh']

                for row in result:
                    func_name = row['function_name']
                    formula = row['formula2']
                    roundoff_value_day = row['roundoff_value']

                    if current_month not in results:
                        results[current_month] = {}

                    if mill_date not in results[current_month]:
                        results[current_month][mill_date] = {}
                    
                    if roundoff_value_day not in results[current_month][mill_date]:
                        results[current_month][mill_date][roundoff_value_day] = {}
                
                    if len(result_dict[current_month]) == 0:
                        formula_result = 0
                    else:
                        if func_name == 'Power Import':
                            if table_name in dict2 and mill_date in dict2[table_name]:
                                formula = dict2[table_name][mill_date].get(14, 0)
                            else:
                                formula = 0

                        elif func_name == 'Power Export':
                            if table_name in dict2_r and mill_date in dict2_r[table_name]:
                                formula = dict2_r[table_name][mill_date].get(14, 0)
                            else:
                                formula = 0
                        else:
                            formula = re.sub(r'dict\[(\d+)\]', lambda match: str(dict2.get(table_name, {}).get(mill_date, {}).get(int(match.group(1)), 0)), formula)
                            
                    if func_name not in machine_data["current_month_record"]:
                        machine_data["current_month_record"][func_name] = {}
                    
                    machine_data["current_month_record"][func_name][mill_date] = {
                "formula": formula,
                "roundoff_value_day": roundoff_value_day
            }
            para = "case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end "
            union_query = " UNION ALL ".join([f"SELECT cp.machine_id, {para} as data, (case when mmf.kWh = '*' then cp.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then cp.reverse_kwh / mmf.kWh_value else cp.reverse_kwh end ) AS reverse_kwh, mm.machine_name FROM ems_v1_completed.dbo.power_{table_name}{type}  left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=cp.machine_id left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id" for table_name in tables_to_union])
            query4 = f"""
            SELECT 
                pp.machine_id, 
                SUM(pp.data) as total_kwh ,
                SUM(pp.reverse_kwh) AS reverse_kwh,
                min(pp.machine_name) as machine_name
            FROM 
                ({union_query}) AS pp 
                
            GROUP BY 
                pp.machine_id 
            ORDER BY 
                pp.machine_id"""
            createFolder("YearReport_Log/","query for year "+str(query4))
            res_q4 = cnx.execute(query4).mappings().all()
            
            dict_y={}
            dict3 = {}
            dict_r_y = {}
            for row in res_q4:
                dict_y[row['machine_id']] = row['total_kwh']
                dict_r_y[row['machine_id']] = row['reverse_kwh']

            for row in result:
                func_name = row['function_name']
                formula = row['formula2']
                roundoff_value_year = row['roundoff_value']
                if len(res_q4) == 0:
                    formula_y = 0
                else:
                    if func_name == 'Power Import':
                        formula_y = dict_y[14]

                    elif func_name == 'Power Export':
                        formula_y = dict_r_y[14]
                    else:
                        numbers = re.findall(r'\[(\d+)\]', formula)
                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict_y]
                        numeric_formula = formula
                        for machine_id in valid_ids:
                            numeric_value = dict_y.get(machine_id, 0)  
                            numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                        formula_y = re.sub(r'dict', '', numeric_formula)
                        
                # dict3[func_name] = formula_y
                dict3[func_name] = {
                    'formula_y': formula_y,
                    'roundoff_value_year': roundoff_value_year
                }
            createFolder("YearReport_Log/", "machine_data" + f"{machine_data}")
            year_wise_excel_report(machine_data, year,next_year,mill_date,dict3,report_type,res_q4)
            file_path = os.path.join(base_path, f"YearReport-{year}-{next_year}.xlsx")
            results = f"http://{request.headers['host']}/attachment/YearReport-{year}-{next_year}.xlsx"

            if os.path.exists(file_path):

                return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": results})
            else:
                return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": None})
        else:
            return JSONResponse({"iserror": False, "message": "data return sucessfully","file_url": ''})

    except Exception as e:
        # error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f" occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror": True, "message": error_message})

@app.post("/loss_record_list/")
async def loss_record_list(date: str=Form(''), 
                           report_for : str = Form(''),                  
                           cnx: Session = Depends(get_db)):
    try:
        where = ''
        diff = ''
        if date == '':
            return JSONResponse({"iserror":True,"message":"date is required"}) 
        
        date = parse_date(date)
        if report_for != '':
            where += f" and report_for = '{report_for}'"

        sql = f''' 
            select 
                *
            FROM
            	[ems_v1].[dbo].[loss_record] 
                where date = '{date}' {where}'''
        data = cnx.execute(sql).mappings().all()
        
        if len(data)>0:
            r1 = f" select sum(day) as d1 from ems_v1.dbo.loss_record where date = '{date}'  and report_for = '{report_for}' and loss_percentage = '+' " 
            r1 = cnx.execute(r1).fetchone()
        
            r2 = f" select sum(day) as d1 from ems_v1.dbo.loss_record where date = '{date}' and report_for = '{report_for}' and loss_percentage = '-' " 
            r2 = cnx.execute(r2).fetchone()
            
            act_c = r1.d1 - r2.d1
            print(act_c)
            r3 = f" select sum(day) as d1 from ems_v1.dbo.loss_record where date = '{date}' and report_for = '{report_for}' and loss_percentage not in ('-','+','x') " 
            r3 = cnx.execute(r3).fetchone()
            
            meter_c = r3.d1
            diff = act_c - meter_c
    
        return JSONResponse({"iserror": False, "message": "data return sucessfully","data": jsonable_encoder(data),"data2": jsonable_encoder(diff)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)
    
@app.post("/update_loss/")
async def update_loss(obj: str = Form(''),
                      date : str = Form(''),    
                      report_for : str = Form(''),    
                      send_to_erp : str = Form(''),    
                      cnx: Session = Depends(get_db)):
    try:
        if obj == '':
            return JSONResponse({"iserror":True,"message":"obj is required"}) 
        
        if date == '':
            return JSONResponse({"iserror":True,"message":"date is required"}) 
        
        if report_for == '':
            return JSONResponse({"iserror":True,"message":"report_for is required"}) 
        
        obj_data = json.loads(obj)
        for row in obj_data:
            id = row["id"]
            loss_percentage = row["loss_percentage"]
            sql = f'''
                    update
                        ems_v1.dbo.loss_record 
                    set 
                        loss_percentage = '{loss_percentage}'
                    where id = '{id}' '''
            cnx.execute(sql)
            cnx.commit()

        createFolder("loss_l/", f"loss percentage update sucessfully" )    

        r1 = f" select sum(day) as d1 from ems_v1.dbo.loss_record where format(date,'dd-MM-yyyy') = '{date}'  and report_for = '{report_for}' and loss_percentage = '+' " 
        print(r1)
        r1 = cnx.execute(r1).fetchone()
       
        r2 = f" select sum(day) as d1 from ems_v1.dbo.loss_record where format(date,'dd-MM-yyyy') = '{date}' and report_for = '{report_for}' and loss_percentage = '-' " 
        r2 = cnx.execute(r2).fetchone()
        
        act_c = r1.d1 - r2.d1
        createFolder("loss_l/", f"actual consumption : {act_c}" )

        r3 = f" select sum(day) as d1 from ems_v1.dbo.loss_record where format(date,'dd-MM-yyyy') = '{date}' and report_for = '{report_for}' and loss_percentage not in ('-','+','x') " 
        r3 = cnx.execute(r3).fetchone()
        meter_c = r3.d1
        createFolder("loss_l/", f"meter_consumption : {meter_c}" )

        diff = act_c - meter_c
        createFolder("loss_l/", f"loss_value : {diff}" )

        r4 =f" select * from ems_v1.dbo.loss_record where format(date,'dd-MM-yyyy') = '{date}' and report_for = '{report_for}' and loss_percentage not in ('-','+','~','x') " 
        r4 = cnx.execute(r4).fetchall()
        total_loss_value = 0
        for row in r4:
            loss_p = row.loss_percentage
            id   = row.id
            loss_p = float(loss_p)
            loss_value = loss_p * (diff/100)

            sql = f" update ems_v1.dbo.loss_record set loss_value = {loss_value} where id = {id}"
            cnx.execute(sql)
            cnx.commit()
            total_loss_value = total_loss_value + loss_value
        createFolder("loss_l/", f"total_loss_value : {total_loss_value}" )

        diff_r = diff - total_loss_value
        createFolder("loss_l/", f"reminder loss : {diff_r}" )

        sql = f" update ems_v1.dbo.loss_record set loss_value = {diff_r}  where format(date,'dd-MM-yyyy') = '{date}' and report_for = '{report_for}' and loss_percentage = '~' "
        cnx.execute(sql)
        cnx.commit()
        if send_to_erp != '':
            query = f"update ems_v1.dbo.loss_record set  send_to_erp = '{send_to_erp}' where format(date,'dd-MM-yyyy') = '{date}' and report_for = '{report_for}'"
            cnx.execute(query)
            cnx.commit()
        if send_to_erp =='yes':
            return JSONResponse({"iserror": False, "message": "Data Sended to ERP Sucessfully"})
        else:
            return JSONResponse({"iserror": False, "message": "updated sucessfully...","data": ''})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)

@app.post("/polling_time_list/")
async def polling_time_list(machine_id: int=Form(''),                   
                            cnx: Session = Depends(get_db)):
    try:
        where = ''
        if machine_id != '':
            where += f" where machine_id = {machine_id}"
        
        main_query = text(f'''
            SELECT
                min(machine_order) as machine_order,
                machine_state_condition1,
                machine_state_condition2,
                STRING_AGG(CAST(machine_id AS NVARCHAR(MAX)), ',') AS machine_ids
            FROM
                [ems_v1].[dbo].[master_machine]
            
            {where}
            GROUP BY
                machine_state_condition1, 
                machine_state_condition2
            order by machine_order
        ''')

        # Execute the main query
        data = cnx.execute(main_query).mappings().all()

        result = []

        for row in data:
            machine_ids = row["machine_ids"]
            machine_id_list = machine_ids.split(",")
            machine_dtl = []

            for machine_id in machine_id_list:
                sub_query = text(f'''
                    SELECT machine_name
                    FROM [ems_v1].[dbo].[master_machine]
                    WHERE machine_id = {machine_id}
                ''')

                sub_data = cnx.execute(sub_query).mappings().all()

                for sub_row in sub_data:
                    machine_dtl.append(sub_row['machine_name'])

            new_row = dict(row)
            new_row["machine_dtl"] = '\n'.join(machine_dtl)
            result.append(new_row)

        return JSONResponse({"iserror": False, "message": "data return sucessfully","data": jsonable_encoder(result)})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)
  
@app.post("/polling_time_entry/")
async def polling_time_list(machine_id: str=Form(''), 
                            machine_state_condition1 : int = Form(''),               
                            machine_state_condition2 : int = Form(''),               
                            cnx: Session = Depends(get_db)):
    try:
        print(machine_id)
        if machine_id == '':
            return JSONResponse({"iserror":True,"message":"machine_id is required"}) 
        
        obj_data = json.loads(machine_id)
        for row in obj_data:
            machine_id = row["machine_id"]
            query = text(f''' 
                update 
                    ems_v1.dbo.master_machine 
                set 
                    machine_state_condition1 = {machine_state_condition1},
                    machine_state_condition2 = {machine_state_condition2}
                where machine_id = {machine_id}
                ''')
            cnx.execute(query)
            cnx.commit()
            print(query)

        return JSONResponse({"iserror": False, "message": "data saved sucessfully","data": ''})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)
  

@app.post("/get_source_list/")
async def source_listsss(source_id:str = Form(""),                        
                      cnx:Session = Depends(get_db)):
    try:

        where = ""
        if source_id !='':
            where = f" AND s.source_id = {source_id}"   
        # where += f" and mm.source_id in ({','.join(str(x) for x in source_id)})  
        query = text(f"""
            SELECT                
                s.*,
                IsNULL(concat(cu.employee_code,'-',cu.employee_name),'') as created_user,
                IsNULL(concat(mu.employee_code,'-',mu.employee_name),'') as modified_user
            FROM 
                ems_v1.dbo.master_source s
                left join ems_v1.dbo.master_employee cu on cu.employee_id=s.created_by
                left join ems_v1.dbo.master_employee mu on mu.employee_id=s.modified_by                
            WHERE 
                s.status != 'delete'{where}
        """)
    
        data = cnx.execute(query).mappings().all()
       
        return JSONResponse({"iserror": False, "message": "data return sucessfully","data": jsonable_encoder(data)})
    
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Log/", "Issue in returning data " + error_message)
        return JSONResponse({"iserror":True,"message":error_message})
    
@app.post("/save_source/")
async def savesource(source_id: int = Form(""),                                
                     source_name:str= Form(""),
                     user_login_id : str = Form(""),                             
                     cnx: Session = Depends(get_db)):
  
    try:
                
        if source_name == "":
            return JSONResponse({"iserror":True,"message":"Source Name is required"}) 
                
        if source_id == '':
            query = text(f"""
            INSERT INTO ems_v1.master_source (
            source_name,created_on, created_by
            )
            VALUES (
                '{source_name}',  now(), '{user_login_id}'
            )
        """) 
            cnx.execute(query)
            cnx.commit()
        else:
            query =text(f"""
                UPDATE 
                    ems_v1.master_source
                SET 
                    source_name = '{source_name}', 
                    modified_on = NOW(),
                    modified_by = '{user_login_id}'
                    WHERE source_id = {source_id} 
            """)
            
            cnx.execute(query)
            cnx.commit()
        return JSONResponse({"iserror": False, "message": "data saved sucessfully","data": ''})
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)
    
    
@app.post("/remove_source/")
async def remove_source(source_id:str=Form(""),status : str = Form(""),cnx: Session = Depends(get_db)):
    
    if source_id == "":
        return JSONResponse({"iserror":True,"message":"Source ID is required"}) 
    
    try:
       
        if status !='':
            query=f''' Update ems_v1.master_source Set status = '{status}' Where source_id='{source_id}' '''
        else:
            query=f''' Update ems_v1.master_source Set status = 'delete' Where source_id='{source_id}' '''
        cnx.execute(text(query))
        cnx.commit()
        return JSONResponse({"iserror": False, "message": "Status Update sucessfully","data": ''})
        
    except Exception as e:
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("YearReport_Log/", "Issue in returning data " + error_message)

@app.post("/current_water/")
async def current_water_report(company_id : str = Form(''),
                        department_id :str = Form(''),
                        shed_id :str = Form(''),
                        machinetype_id :str = Form(''),
                        function_id : str = Form(''),
                        machine_id : str = Form (''),
                        group_for : str = Form(''),
                        groupby : str = Form(''),
                        period_id: str = Form(''),
                        from_date: str = Form(''),
                        to_date: str = Form(''),                      
                        shift_id: str = Form(''),
                        converter_id :int = Form(''),  
                        report_for : str = Form(''), 
                        is_function : str = Form(''),  
                        function_type : str = Form(''), 
                        employee_id : int = Form(''),                              
                        cnx: Session = Depends(get_db)):

    if period_id == '':
            return JSONResponse({"iserror": True, "message": "period id is required"})
    
    if groupby =='':
        return JSONResponse({"iserror":True, "message": "groupby is required"}) 
    
    if group_for =='':
        return JSONResponse({"iserror":True, "message": "group_for is required"}) 
    try:
        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
     
        machine_id = id(machine_id)
        company_id = id(company_id)
        department_id = id(department_id)
        shed_id = id(shed_id)
        machinetype_id = id(machinetype_id)
        function_id = id(function_id)
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="[ems_v1_completed].[dbo]."           
        where = "" 
        group_by = ""
        order_by = ""  
        function_where = ''

        if employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]
                    print("department_id",department_id)

        if  company_id == '' or company_id == "0":
            pass
        else:
            where += f" and  mm.company_id in ({company_id})" 
    
        if department_id == '' or department_id == "0":
            pass
        else:
            where += f" and  mm.department_id in ({department_id})"          
            
        if shed_id == '' or shed_id == "0":
            pass
        else:
            where += f" and mm.shed_id in ({shed_id})"
            
        if machinetype_id == '' or machinetype_id == "0":
            pass
        else:
            where += f" and mm.machinetype_id in ({machinetype_id})"
            
        if function_id == '':
            pass
        else:
            if function_type =='':
                where += f"and  mm.function_id in ({function_id})"
            else:
                if function_type == 'function_1':
                    where += f"and  mm.function_id in ({function_id})"
                else:
                    where += f"and  mm.function2_id in ({function_id})"

        if machine_id == 'all' or machine_id == '':
            pass
        else:
            where += f" and mm.machine_id in ({machine_id})"
            
        if converter_id == '':
            pass
        else:
            where += f" and mm.converter_id = {converter_id}"
        
        if function_type !='':
            where += f" and mf.function_type = '{function_type}'"
            if function_type == 'function_2':
                function_where += f" mm.function2_id = mf.function_id"  
            else:
                function_where += f" mm.function_id = mf.function_id"
        else:
            function_where += f" mm.function_id = mf.function_id"
            
        query = text(f'''SELECT * FROM [ems_v1].[dbo].[master_shifts_wfm] WHERE status = 'active' ''')
        data1 = cnx.execute(query).mappings().all()
        mill_date = date.today()
        mill_shift = 0
        no_of_shifts = 3
        group_id = ""
        group_code = ""
        group_name = ""

        month_year = ""
        table_name = ''
        join = ''

        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]  
              no_of_shifts = shift_record["no_of_shifts"] 
        table_name = "[ems_v1].[dbo].[current_water] cw" 
        if period_id == "cur_shift":       
            where += f''' and cw.mill_date = '{mill_date}' AND cw.mill_shift = '{mill_shift}' '''              
            table_name = "[ems_v1].[dbo].[current_water] cw"  

        elif period_id == "#cur_shift":
            where += f''' and cw.mill_date = '{mill_date}' AND cw.mill_shift = '{mill_shift}' '''              
            table_name = "[ems_v1].[dbo].[current_water] cw" 

        elif period_id == "sel_shift":                  
            if from_date == '':
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if shift_id == '':
                return JSONResponse({"iserror": True, "message": "shift_id is required"}) 
            
            from_date = parse_date(from_date)          
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}[water_{month_year}] as cw" 

            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
            result_query = cnx.execute(query).mappings().all()
            
            if len(result_query) == 0:
                return JSONResponse({"iserror": True, "message": "power table not available..."})    
                            
            where += f''' and cw.mill_date = '{from_date}' AND cw.mill_shift = '{shift_id}' '''   

        elif period_id == "#sel_shift":                 
            if mill_shift == 1:
                shift_id = no_of_shifts
                from_date = parse_date(mill_date) - timedelta(days=1)

            else:
                shift_id = int(mill_shift) - 1
                from_date = mill_date                      
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}[water_{month_year}] as cw" 

            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
            result_query = cnx.execute(query).mappings().all()
            
            if len(result_query) == 0:
                return JSONResponse({"iserror": True, "message": "water table not available..."})    
            
            where += f''' and cw.mill_date = '{from_date}' AND cw.mill_shift = '{shift_id}' '''   
        
        elif period_id == "sel_date":            
            if from_date == '':
                return JSONResponse({"iserror": True, "message": "from date is required"})    
            
            from_date = parse_date(from_date)
            
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}[water_{month_year}] as cw "       
            where += f''' and cw.mill_date = '{from_date}' '''

            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
            result_query = cnx.execute(query).mappings().all()
            
            if len(result_query) == 0:
                return JSONResponse({"iserror": True, "message": "water table not available..."}) 
                
            
        elif period_id == "#sel_date":             
            from_date = mill_date - timedelta(days=1)
            # from_date = parse_date(from_date)
            
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""
            table_name=f"  {completed_db}[water_{month_year}] as cw "
            where += f''' and cw.mill_date = '{from_date}' '''

            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
            result_query = cnx.execute(query).mappings().all()
            
            if len(result_query) == 0:
                return JSONResponse({"iserror": True, "message": "water table not available..."})    
            
        elif period_id  == "#this_week":
            dt = mill_date
            from_date=dt-timedelta(dt.weekday()+1)
            to_date = mill_date

            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
            table_name=f"  {completed_db}[water_{month_year}] as cw "
            where += f''' and cw.mill_date  >= '{from_date}' and cw.mill_date <= '{to_date}' '''

            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
            result_query = cnx.execute(query).mappings().all()
            
            if len(result_query) == 0:
                return JSONResponse({"iserror": True, "message": "water table not available..."})    
            
        elif period_id == "#this_month":
            from_date = mill_date.replace(day=1)
            to_date = mill_date

            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""            
            table_name=f"  {completed_db}[water_{month_year}] as cw "
            where += f''' and cw.mill_date  >= '{from_date}' and cw.mill_date <= '{to_date}' '''

            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
            result_query = cnx.execute(query).mappings().all()
            
            if len(result_query) == 0:
                return JSONResponse({"iserror": True, "message": "water table not available..."})    
            
        elif period_id == "from_to":            
            if from_date == '':
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date == '':
                return JSONResponse({"iserror": True, "message": "to_date is required"})  
                    
            from_date = parse_date(from_date)
            to_date =  parse_date(to_date)
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
            where += f''' and cw.mill_date  >= '{from_date}' and cw.mill_date <= '{to_date}' '''
            if shift_id != ""and shift_id !="all":                
                where += f''' and cw.mill_shift = '{shift_id}' ''' 

            if from_date.month == to_date.month and from_date.year == to_date.year:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "water table not available..."})    
            
                table_name=f"  {completed_db}[power_{month_year}] as cw "
                    
            else:
                field_name = 'id,water_meter_id, created_on, mill_date, mill_shift, pressure, flow, actula_pressure, avg_pressure'
                
                month_year_range = [
                    (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                    for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                ]
                print(month_year_range)
                union_queries = []
                join = []

                for month_year in month_year_range:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    print(query)
                    if len(result_query) > 0:
                        table_name = f"[ems_v1_completed].[dbo].[water_{month_year}]"
                        union_queries.append(f"SELECT {field_name} FROM {table_name}")
                    
                if len(union_queries) == 0:
                    return JSONResponse({"iserror": True, "message": "power table not available..."})    
            
                subquery_union = " UNION ALL ".join(union_queries)
                table_name = f"( {subquery_union}) cw"
                
                
        else:
            return JSONResponse({"iserror": True, "message": "invalid period id"})     

        if groupby != '' and groupby == "company":
            group_by += " mm.company_id "
            order_by += " mm.company_id "
            group_id = '''min(mc.company_id) AS group_id '''
            group_code = '''min(mc.company_code) AS group_code ,'''
            group_name = '''min(mc.company_name) AS group_name'''       
            
        elif groupby !='' and groupby == "department":
            group_by += " mm.department_id "
            order_by += " min(md.department_order)"
            group_id = '''min(md.department_id) AS group_id '''
            group_code = '''min(md.department_code) AS group_code ,'''
            group_name = '''min(md.department_name) AS group_name'''        
            
        elif groupby !='' and groupby == "shed":
            group_by += "  mm.shed_id "
            order_by += "  min(ms.shed_order)"
            group_id = ''' min(ms.shed_id) AS group_id '''
            group_code = ''' min(ms.shed_code) AS group_code ,'''
            group_name = ''' min(ms.shed_name) AS group_name'''
            
        elif groupby !='' and groupby == "machinetype":
            group_by += " mm.machinetype_id"
            order_by += " min(mmt.machinetype_order)"
            group_id = '''min(mmt.machinetype_id) AS group_id '''
            group_code = '''min(mmt.machinetype_code) AS group_code ,'''
            group_name = '''min(mmt.machinetype_name) AS group_name'''
            
        elif groupby !='' and groupby == "function":    
            order_by += " min(mf.function_order)"
            group_id = '''min(mf.function_id) AS group_id '''
            group_code = '''min(mf.function_code) AS group_code ,'''
            group_name = '''min(mf.function_name) AS group_name'''       
            if function_type !='':
                if function_type == 'function_1':
                    group_by += " mm.function_id" 
                else:
                    group_by += " mm.function2_id"     
            else:
                group_by += " mm.function_id"

            if is_function !="":
                group_by += " ,mm.machine_id"
                order_by += " ,mm.machine_id"

        elif groupby !='' and groupby == "converter":           
            group_by += " mm.converter_id"
            order_by += " mm.converter_id"
            group_id = '''min(mcd.converter_id) AS group_id '''
            group_code = ''
            group_name = '''min(mcd.converter_name) AS group_name'''

            if is_function !="":
                group_by += " ,mm.machine_id"
                order_by += " ,mm.machine_id"
            
        elif groupby !='' and groupby == "machine":             
            group_by += " mm.machine_id"
            order_by += " min(mm.machine_order)"
            group_id = '''min(mm.machine_id) AS group_id '''
            group_code = '''min(mm.machine_code) AS group_code ,'''
            group_name = '''min(mm.machine_name) AS group_name'''  
           
       
        if report_for == 'detail' or report_for == '':
            group_by = " cw.mill_date , cw.mill_shift," + group_by
            order_by = " cw.mill_date, cw.mill_shift," + order_by
        
        if report_for == 'summary':
            group_by = " cw.mill_date," + group_by
            order_by = " cw.mill_date," + order_by  
                       
        if group_by != "":
            group_by = f"group by {group_by} "    
        if order_by != "":
            order_by = f"order by {order_by}"

        
        query = text(f'''
                select 
                    min(mc.company_name) as company_name,
                    min(mc.company_code) as company_code,
                    min(mb.branch_name) as branch_name,
                    min(mb.branch_code) as branch_code,
                    min(md.department_name) as department_name,
                    min(md.department_code) as department_code,
                    min(mmt.machinetype_name) as machinetype_name,
                    min(mmt.machinetype_code) as machinetype_code,
                    min(mf.function_name) as function_name,
                    min(mf.function_code) as function_code,
                    min(mm.machine_name) as machine_name,
                    min(mm.machine_code) as machine_code,
                    min(mm.company_id) as company_id,
                    min(mm.branch_id) as branch_id,
                    min(mm.department_id) as department_id,
                    min(mm.machinetype_id) as machinetype_id,
                    min(mm.function_id) as function_id,
                    min(cw.water_meter_id) as water_meter_id,
                    COUNT(DISTINCT mm.machine_id) machine_count,
                    ISNULL(TRY_CAST(ROUND(
                        SUM(
                            CASE
                                WHEN (msv.scaling1_max - msv.scaling1_min) != 0 and msv.type = 'pressure'  THEN
                                    (((cw.pressure - msv.scaling1_min) / (msv.scaling1_max - msv.scaling1_min)) * (msv.scaling2_max - msv.scaling2_min) + msv.scaling2_min)
                                ELSE
                                    0
                            END
                        ),2)  AS DECIMAL(18, 2)),0) AS pressure_value,
                    ISNULL(TRY_CAST(ROUND(
                        SUM(
                            CASE
                                WHEN (msv.scaling1_max - msv.scaling1_min) != 0 and msv.type = 'flow' THEN
                                    (((cw.flow - msv.scaling1_min) / (msv.scaling1_max - msv.scaling1_min)) * (msv.scaling2_max - msv.scaling2_min) + msv.scaling2_min)
                                ELSE
                                    0
                            END
                        )
                        ,2)  AS DECIMAL(18, 2)),0) AS flow_value,
                    cw.mill_date,
                    cw.mill_shift,
                    CASE WHEN min(cw.modified_on) <= DATEADD(minute, -2, getdate()) THEN 'S' ELSE 'N' END as nocom,
                    min(cw.modified_on) as date_time,     
                    {group_id},
                    {group_code}
                    {group_name} 
                FROM 
                    {table_name}                       
                    INNER JOIN [ems_v1].[dbo].[master_machine] mm ON mm.machine_id = cw.water_meter_id
                    INNER JOIN [ems_v1].[dbo].[master_company] mc ON mm.company_id = mc.company_id
                    INNER JOIN [ems_v1].[dbo].[master_branch] mb ON mm.branch_id = mb.branch_id
                    INNER JOIN [ems_v1].[dbo].[master_department] md ON mm.department_id = md.department_id
                    INNER JOIN [ems_v1].[dbo].[master_shed] ms ON mm.shed_id = ms.shed_id
                    INNER JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mm.machinetype_id = mmt.machinetype_id 
                    LEFT JOIN [ems_v1].[dbo].[master_function] mf ON {function_where}
                    LEFT JOIN [ems_v1].[dbo].[master_converter_detail] mcd ON mm.converter_id = mcd.converter_id 
                    LEFT JOIN [ems_v1].[dbo].[master_machine_factor] mmf ON mm.machine_id = mmf.machine_id 
                    inner JOIN [ems_v1].[dbo].[master_scaling_value] msv ON msv.machine_id = mm.machine_id 
                    {join}                  
                WHERE  
                    mm.status = 'active'
                    {where}         
                    {group_by}
                    {order_by}
                
                ''') 
        createFolder("current_water/","query "+str(query))
        data = cnx.execute(query).mappings().all()
        return JSONResponse({"iserror":False, "message":"data return successfully", "data" : jsonable_encoder(data)})    
    
    except Exception as e:
        traceback.print_exc()      
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Current_power_log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
      
@app.post("/water_analysis/")
async def water_analysis(period_id: str = Form(''),
                         machine_id: str=Form(''), 
                         from_date: str = Form(''),
                         to_date: str = Form(''),
                         shift_id :int = Form(''),
                         from_time: str=Form(''), 
                         to_time: str=Form(''),  
                         duration :int = Form(''),                      
                         cnx: Session = Depends(get_db)):
    
    if machine_id == '':
        return JSONResponse({"iserror":True,"message":"machine id is required"}) 
    
    if period_id == '':
        return JSONResponse({"iserror":True,"message":"period_id is required"}) 
    
    mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}    
    
    try: 
        where = ''
        table_name = ''
        if duration == "":
            duration = 1
        
        if period_id == 'cur_shift': 
            query=text(f'''SELECT * FROM [ems_v1].[dbo].[master_shifts_wfm] WHERE status='active' ''')
            data1 = cnx.execute(query).mappings().all()
            mill_date = date.today()
            mill_shift = 0       
            if len(data1) > 0:
                for shift_record in data1:
                    mill_date = shift_record["mill_date"]
                    mill_shift = shift_record["mill_shift"]  
                        
            table_name = 'ems_v1.dbo.current_water_analysis cw'
            where += f"cw.mill_date = '{mill_date}' and cw.mill_shift ='{mill_shift}' "

        elif period_id == 'sel_shift' or period_id == 'sel_date':
            if from_date == '':
                return JSONResponse({"iserror":True,"message":"date is required"}) 
            
            mill_date=parse_date(from_date)             
            month_year=f"""{mill_month[mill_date.month]}{str(mill_date.year)}"""
            table_name=f"[ems_v1_completed].[dbo].[water_analysis_{month_year}]" 
            where += f"cw.mill_date = '{mill_date}' "

            field_name = 'id,water_meter_id, created_on, mill_date, mill_shift, pressure, flow, actula_pressure, avg_pressure'
            table_name = f'(select {field_name} from [ems_v1].[dbo].[current_water_analysis] UNION All select {field_name} from {table_name})cw'

            if period_id == 'sel_shift':
                if shift_id == '':
                    return JSONResponse({"iserror":True,"message":"shift is required"}) 
                where += f" and cw.mill_shift ='{shift_id}' " 
            if month_year !='':
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_analysis_{month_year}'"""
                print(query)
                result_query = cnx.execute(query).mappings().all()
                if len(result_query)>0:
                    pass
                else:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    
   
        elif period_id == "from_to":            
            if from_date == '':
                return JSONResponse({"iserror": True, "message": "from date is required"})
            if to_date == '':
                return JSONResponse({"iserror": True, "message": "to_date is required"})  
                    
            from_date = parse_date(from_date)
            to_date =  parse_date(to_date)
            month_year=f"""{mill_month[from_date.month]}{str(from_date.year)}"""       
        
            where += f'''  cw.mill_date  >= '{from_date}' and cw.mill_date <= '{to_date}' '''
            
            if shift_id != "":                
                where += f''' and cw.mill_shift = '{shift_id}' ''' 
            field_name = 'id,water_meter_id, created_on, mill_date, mill_shift, pressure, flow, actula_pressure, avg_pressure'
            
            if from_date.month == to_date.month:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_analysis_{month_year}'"""
                result_query = cnx.execute(query).mappings().all()
                print(query)
                if len(result_query) == 0:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    
       
                table_name=f"[ems_v1_completed].[dbo].[water_analysis_{month_year}]" 
                table_name = f'(select {field_name} from [ems_v1].[dbo].[current_water_analysis] UNION All select {field_name} from {table_name})cw'
            else:
                
                month_year_range = [
                        (from_date + timedelta(days=31 * i)).strftime("%m%Y")
                        for i in range((to_date.year - from_date.year) * 12 + to_date.month - from_date.month + 1)
                    ]
                union_queries = []

                for month_year in month_year_range:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'water_analysis_{month_year}'"""
                    result_query = cnx.execute(query).mappings().all()
                    print(query)
                    if len(result_query) > 0:
                        table_name = f"[ems_v1_completed].[dbo].[water_analysis_{month_year}]"
                        union_queries.append(f"SELECT {field_name} FROM {table_name}")

                if len(union_queries) == 0:
                    return JSONResponse({"iserror": True, "message": "analysis table not available..."})    

                subquery_union = " UNION ALL ".join(union_queries)
                table_name = f"(SELECT {field_name} FROM [ems_v1].[dbo].[current_water_analysis] UNION ALL {subquery_union}) cw"

        if from_time !='':
            where += f" and FORMAT(cw.created_on ,'HH:mm:ss')>='{from_time}' "
    
        if to_time !='':
            where += f" and FORMAT(cw.created_on ,'HH:mm:ss')<='{to_time}' "
                
        query=text(f'''
            SELECT *
            FROM (
                SELECT 
                (ROW_NUMBER() OVER (ORDER BY min(cw.water_meter_id)) - 1) % {duration} + 1 AS slno,
			    min(mm.machine_code) machine_code,
			    min(mm.machine_name) machine_name,
			    min(cw.water_meter_id) water_meter_id,
                min(FORMAT(cw.created_on, 'yyyy-MM-ddTHH:mm:ss')) as date_time,
			    min(cw.mill_date) mill_date,
			    min(cw.mill_shift) mill_shift,
			    ISNULL(TRY_CAST(ROUND(
                        SUM(
                            CASE
                                WHEN (msv.scaling1_max - msv.scaling1_min) != 0 and msv.type = 'pressure'  THEN
                                    (((cw.pressure - msv.scaling1_min) / (msv.scaling1_max - msv.scaling1_min)) * (msv.scaling2_max - msv.scaling2_min) + msv.scaling2_min)
                                ELSE
                                    0
                            END
                        ),2)  AS DECIMAL(18, 2)),0) AS pressure_value,
                    ISNULL(TRY_CAST(ROUND(
                        SUM(
                            CASE
                                WHEN (msv.scaling1_max - msv.scaling1_min) != 0 and msv.type = 'flow' THEN
                                    (((cw.flow - msv.scaling1_min) / (msv.scaling1_max - msv.scaling1_min)) * (msv.scaling2_max - msv.scaling2_min) + msv.scaling2_min)
                                ELSE
                                    0
                            END
                        )
                        ,2)  AS DECIMAL(18, 2)),0) AS flow_value,
                avg(cw.actual_pressure) actual_pressure,
                avg(cw.avg_pressure) avg_pressure

		    from
                {table_name}   

		        inner join [ems_v1].[dbo].[master_machine] mm on mm.machine_id=cw.water_meter_id
		        inner join [ems_v1].[dbo].[master_scaling_value] msv on msv.machine_id=mm.machine_id
                
                
		    where 
                cw.water_meter_id in ({machine_id}) and {where} 
                group by mm.machine_id,cw.created_on) AS subquery
            WHERE
                slno = 1
                
		    order by water_meter_id, date_time                                
            ''')  
    
        createFolder("Load_analysis_log/","query executed successfully in load analysis..."+str(query))
        data=cnx.execute(query).mappings().all()
        label = {}
        machine_data = {}
        org_data = []
        for d in data:
            machine_id = d['water_meter_id']
            machine_name = d['machine_name']
            if machine_id not in label:        
                label[machine_id] = machine_name
            if machine_id not in machine_data:
                machine_data[machine_id] = []

            # set machine_data for machine_id
            temp = {
                'date_time': d['date_time'],
                'pressure_value': d['pressure_value'],
                'flow_value': d['flow_value'],
                
            }

            machine_data[machine_id].append(temp)

        for key, value in machine_data.items():
            org_data.append({'label': label[key], 'data': value})

        return JSONResponse({"iserror":False,"message":"data return successfully","data":jsonable_encoder(org_data),"data1":jsonable_encoder(data)}) 

    except Exception as e :
        error_type = type(e).__name__
        error_line = traceback.extract_tb(e.__traceback__)[0].lineno
        error_filename = os.path.basename(traceback.extract_tb(e.__traceback__)[0].filename)
        error_message = f"{error_type} occurred in file {error_filename}, line {error_line}: {str(e)}"
        createFolder("Load_analysis_log/","Issue in returning data "+error_message)
        return JSONResponse({"iserror":True,"message":error_message})
